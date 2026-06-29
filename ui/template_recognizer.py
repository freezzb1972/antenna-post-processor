"""
模板识别对话框
==============
加载 Excel 模板，显示列头检测结果，允许手动修改 col_type，
确认后保存到 config/column_patterns.json。
"""
from __future__ import annotations

import json
import os
import sys
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QComboBox, QDialog, QDialogButtonBox, QFileDialog, QFormLayout,
    QGroupBox, QHBoxLayout, QHeaderView, QLabel, QLineEdit, QMessageBox,
    QPushButton, QSpinBox, QTableWidget, QTableWidgetItem, QVBoxLayout,
)

if TYPE_CHECKING:
    from ui.main_window import MainWindow

from src.excel_reader import normalize_header
from src.column_mapping import classify_header

# ── 所有可选的列类型 ──
ALL_COL_TYPES = [
    ("frequency",           "频率"),
    ("directivity",         "方向性"),
    ("total_efficiency_pct","总效率 (%)"),
    ("total_efficiency_db", "总效率 (dB)"),
    ("efficiency_pct",      "效率 (%)"),
    ("efficiency_db",       "效率 (dB)"),
    ("gain",                "峰值增益"),
    ("trp",                 "TRP"),
    ("nhprp_45",            "NHPRP ±45°"),
    ("nhprp_30",            "NHPRP ±30°"),
    ("nhprp_225",           "NHPRP ±22.5°"),
    ("peak_eirp",           "Peak EIRP"),
    ("ar_single",           "AR 单角度"),
    ("ar_range",            "AR 范围"),
    ("uh_prp",              "上半球 PRP"),
    ("lh_prp",              "下半球 PRP"),
    ("nhprp45_ratio_pct",   "NHPRP45/TRP 比率 (%)"),
    ("nhprp45_ratio_db",    "NHPRP45/TRP 比率 (dB)"),
    ("nhprp30_ratio_pct",   "NHPRP30/TRP 比率 (%)"),
    ("nhprp30_ratio_db",    "NHPRP30/TRP 比率 (dB)"),
    ("nhprp225_ratio_pct",  "NHPRP225/TRP 比率 (%)"),
    ("nhprp225_ratio_db",   "NHPRP225/TRP 比率 (dB)"),
    ("uh_ratio_pct",        "UH/TRP 比率 (%)"),
    ("uh_ratio_db",         "UH/TRP 比率 (dB)"),
    ("lh_ratio_pct",        "LH/TRP 比率 (%)"),
    ("lh_ratio_db",         "LH/TRP 比率 (dB)"),
    ("boresight_phi",       "Boresight Phi"),
    ("boresight_theta",     "Boresight Theta"),
    ("max_power",           "最大功率"),
    ("min_power",           "最小功率"),
    ("avg_gain",            "平均增益"),
    ("avg_power",           "平均功率"),
    ("xpi_boresight",       "XPI Boresight"),
    ("xpi_mean",            "XPI Mean"),
    ("xpi_min",             "XPI Min"),
    ("mismatch_loss_db",    "Mismatch Loss"),
    ("pc_theta_mm",         "Phase Center θ"),
    ("pc_phi_mm",           "Phase Center φ"),
    ("lag_single",          "LAG (单角度)"),
    ("lag_range",           "LAG (范围)"),
    ("unknown",             "未知 (跳过)"),
]


class TemplateRecognizerDialog(QDialog):
    """模板识别对话框 — 检测 + 手动修正 + 保存自定义模式。"""

    def __init__(self, parent: "MainWindow"):
        super().__init__(parent)
        self._mw = parent
        self.setWindowTitle(self.tr("模板识别"))
        self.resize(900, 650)
        self.setMinimumSize(720, 500)
        # 启用最大化按钮，方便全屏查看大量列头参数
        # 注意: setWindowFlags 必须在 show() 之前调用，否则会触发 hide()
        self.setWindowFlags(
            self.windowFlags() | Qt.WindowMaximizeButtonHint)

        self._template_path: str = ""
        self._sheet_names: list = []
        self._header_rows: dict = {}   # sheet_name → row number
        self._column_data: dict = {}   # sheet_name → [(col_letter, raw_header, detected_type), ...]

        self._setup_ui()
        self._load_default_path()

    # ── UI ──────────────────────────────────────────────────────────

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # ── 模板选择 ──
        grp_tpl = QGroupBox(self.tr("模板文件"))
        tpl_layout = QHBoxLayout(grp_tpl)
        self._edit_tpl_path = QLineEdit()
        self._edit_tpl_path.setPlaceholderText(self.tr("选择模板 Excel 文件 (.xlsx)..."))
        self._edit_tpl_path.textChanged.connect(self._on_path_changed)
        tpl_layout.addWidget(self._edit_tpl_path, 1)
        btn_browse = QPushButton(self.tr("浏览..."))
        btn_browse.clicked.connect(self._on_browse)
        tpl_layout.addWidget(btn_browse)
        btn_load = QPushButton(self.tr("检测"))
        btn_load.clicked.connect(self._on_detect)
        tpl_layout.addWidget(btn_load)
        layout.addWidget(grp_tpl)

        # ── 工作表选择 + 标题行 ──
        grp_sheet = QGroupBox(self.tr("工作表 & 标题行"))
        sheet_layout = QHBoxLayout(grp_sheet)
        sheet_layout.addWidget(QLabel(self.tr("工作表:")))
        self._cmb_sheet = QComboBox()
        self._cmb_sheet.currentIndexChanged.connect(self._on_sheet_changed)
        sheet_layout.addWidget(self._cmb_sheet, 1)
        sheet_layout.addWidget(QLabel(self.tr("标题行:")))
        self._spin_header = QSpinBox()
        self._spin_header.setRange(1, 20)
        self._spin_header.setValue(1)
        self._spin_header.valueChanged.connect(self._on_header_row_changed)
        sheet_layout.addWidget(self._spin_header)
        sheet_layout.addStretch()
        layout.addWidget(grp_sheet)

        # ── 列检测结果表 ──
        grp_cols = QGroupBox(self.tr("列头检测结果 (双击类型可修改)"))
        cols_layout = QVBoxLayout(grp_cols)
        self._table = QTableWidget()
        self._table.setColumnCount(4)
        self._table.setHorizontalHeaderLabels(
            [self.tr("列号"), self.tr("原始列头"), self.tr("检测类型"), self.tr("操作")])
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self._table.setColumnWidth(0, 60)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
        self._table.setColumnWidth(2, 180)
        self._table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Fixed)
        self._table.setColumnWidth(3, 120)
        self._table.setSelectionBehavior(QTableWidget.SelectRows)
        self._table.setEditTriggers(QTableWidget.NoEditTriggers)
        cols_layout.addWidget(self._table)

        # 结果摘要
        self._lbl_summary = QLabel("")
        cols_layout.addWidget(self._lbl_summary)
        layout.addWidget(grp_cols, 1)

        # ── 按钮 ──
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        self._btn_save = QPushButton(self.tr("💾 保存到配置文件"))
        self._btn_save.setEnabled(False)
        self._btn_save.clicked.connect(self._on_save)
        btn_layout.addWidget(self._btn_save)
        btn_layout.addWidget(QLabel("   "))
        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(self.close)
        btn_layout.addWidget(btns)
        layout.addLayout(btn_layout)

    # ── 路径 ────────────────────────────────────────────────────────

    def _load_default_path(self):
        mw = self._mw if self._mw else None
        if mw and hasattr(mw, '_current_template_path'):
            tp = mw._current_template_path
            if tp and os.path.isfile(tp):
                self._edit_tpl_path.setText(tp)

    def _on_path_changed(self, _text):
        self._template_path = self._edit_tpl_path.text().strip()
        if os.path.isfile(self._template_path):
            self._btn_save.setEnabled(False)

    def _on_browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, self.tr("选择模板文件"), "",
            self.tr("Excel 文件 (*.xlsx *.xls);;所有文件 (*)"))
        if path:
            self._edit_tpl_path.setText(path)

    # ── 检测 ────────────────────────────────────────────────────────

    def _on_detect(self):
        path = self._edit_tpl_path.text().strip()
        if not path or not os.path.isfile(path):
            QMessageBox.warning(self, self.tr("文件不存在"),
                                self.tr("请选择有效的模板文件。\n{path}").format(path=path))
            return

        try:
            self._scan_template(path)
        except Exception as e:
            QMessageBox.critical(self, self.tr("检测失败"),
                                 self.tr("无法解析模板文件:\n{e}").format(e=e))
            return

        self._populate_sheet_combo()
        self._refresh_table()
        self._btn_save.setEnabled(True)

    def _scan_template(self, path: str):
        """扫描模板所有 sheet 的列头。"""
        wb = openpyxl.load_workbook(path, data_only=True)
        self._sheet_names = []
        self._header_rows = {}
        self._column_data = {}

        try:
            for ws in wb.worksheets:
                name = ws.title
                max_row = ws.max_row or 100
                max_col = ws.max_column or 30

                # 找第一个看起来像标题行的行
                best_row = 1
                best_score = 0
                for row_idx in range(1, min(max_row + 1, 30)):
                    text_cells = 0
                    freq_match = False
                    for c_idx in range(1, max_col + 1):
                        v = ws.cell(row_idx, c_idx).value
                        if v is not None:
                            text_cells += 1
                            t = str(v).strip()
                            if is_frequency_column(t):
                                freq_match = True
                    score = text_cells + (10 if freq_match else 0)
                    if score > best_score:
                        best_score = score
                        best_row = row_idx

                if best_score < 2:
                    continue  # 跳过几乎没有文字的 sheet

                self._sheet_names.append(name)
                self._header_rows[name] = best_row

                # 解析列
                cols = []
                for c_idx in range(1, max_col + 1):
                    raw = str(ws.cell(best_row, c_idx).value or "").strip()
                    if not raw:
                        continue
                    col_letter = openpyxl.utils.get_column_letter(c_idx)
                    ctype = self._classify(raw)
                    cols.append((col_letter, raw, ctype))
                self._column_data[name] = cols
        finally:
            wb.close()

    def _classify(self, raw_header: str) -> str:
        """完整分类入口，委托给 column_mapping.classify_header。"""
        return classify_header(raw_header)

    # ── 表格显示 ─────────────────────────────────────────────────────

    def _populate_sheet_combo(self):
        self._cmb_sheet.blockSignals(True)
        self._cmb_sheet.clear()
        for name in self._sheet_names:
            self._cmb_sheet.addItem(name)
        self._cmb_sheet.blockSignals(False)

    def _on_sheet_changed(self, _idx):
        self._refresh_table()

    def _on_header_row_changed(self, _val):
        # 标题行变化 → 重新检测当前 sheet
        sheet = self._cmb_sheet.currentText()
        if sheet:
            try:
                wb = openpyxl.load_workbook(self._template_path, data_only=True)
                ws = wb[sheet]
                max_col = ws.max_column or 30
                row = self._spin_header.value()
                cols = []
                for c_idx in range(1, max_col + 1):
                    raw = str(ws.cell(row, c_idx).value or "").strip()
                    if not raw:
                        continue
                    col_letter = openpyxl.utils.get_column_letter(c_idx)
                    ctype = self._classify(raw)
                    cols.append((col_letter, raw, ctype))
                self._column_data[sheet] = cols
                self._header_rows[sheet] = row
                wb.close()
                self._refresh_table()
            except Exception as e:
                QMessageBox.warning(self, self.tr("重新检测失败"), str(e))

    def _refresh_table(self):
        sheet = self._cmb_sheet.currentText()
        if not sheet or sheet not in self._column_data:
            self._table.setRowCount(0)
            self._lbl_summary.setText("")
            return

        cols = self._column_data[sheet]
        self._table.setRowCount(len(cols))

        # 构建类型 → 显示名映射
        type_labels = {t: l for t, l in ALL_COL_TYPES}

        for row, (col_letter, raw_header, ctype) in enumerate(cols):
            # 列号
            self._table.setItem(row, 0, QTableWidgetItem(col_letter))
            # 原始列头
            self._table.setItem(row, 1, QTableWidgetItem(raw_header))
            # 检测类型 — 用下拉框
            cmb = QComboBox()
            for t, label in ALL_COL_TYPES:
                cmb.addItem(label, t)
            idx = cmb.findData(ctype)
            if idx >= 0:
                cmb.setCurrentIndex(idx)
            cmb.currentIndexChanged.connect(
                lambda _i, r=row, c=cmb: self._on_type_changed(r, c))
            self._table.setCellWidget(row, 2, cmb)

            # 操作按钮
            btn_widget = QWidget()
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(2, 2, 2, 2)
            btn_layout.setSpacing(4)
            btn_redo = QPushButton(self.tr("重检"))
            btn_redo.setFixedWidth(50)
            btn_redo.clicked.connect(lambda checked, r=row: self._on_recheck_row(r))
            btn_layout.addWidget(btn_redo)
            btn_layout.addStretch()
            self._table.setCellWidget(row, 3, btn_widget)

        # 摘要
        detected = sum(1 for _, _, t in cols if t != "unknown")
        total = len(cols)
        self._lbl_summary.setText(
            self.tr("共 %1 列，识别 %2 列，%3 列未识别").arg(total).arg(detected).arg(total - detected))

        self._table.setRowHeight
        self._table.resizeRowsToContents()

    def _on_type_changed(self, row: int, cmb: QComboBox):
        sheet = self._cmb_sheet.currentText()
        new_type = cmb.currentData()
        if sheet in self._column_data and row < len(self._column_data[sheet]):
            cols = self._column_data[sheet]
            self._column_data[sheet] = [
                cols[i] if i != row else (cols[i][0], cols[i][1], new_type)
                for i in range(len(cols))
            ]

    def _on_recheck_row(self, row: int):
        """重新对单列运行自动分类。"""
        sheet = self._cmb_sheet.currentText()
        if sheet in self._column_data and row < len(self._column_data[sheet]):
            cols = self._column_data[sheet]
            _, raw, _ = cols[row]
            new_type = self._classify(raw)
            self._column_data[sheet] = [
                cols[i] if i != row else (cols[i][0], cols[i][1], new_type)
                for i in range(len(cols))
            ]
            self._refresh_table()

    # ── 保存 ────────────────────────────────────────────────────────

    def _on_save(self):
        """将当前检测结果保存到 config/column_patterns.json。"""
        # 收集所有用户修改过的列头 → 类型映射
        all_cols = {}
        for sheet in self._sheet_names:
            if sheet in self._column_data:
                for _, raw_header, ctype in self._column_data[sheet]:
                    if ctype and ctype != "unknown":
                        all_cols[raw_header] = ctype

        # 生成 patterns 列表（用简单关键词提取）
        import re
        new_patterns = []
        seen_types = set()
        for raw_header, ctype in all_cols.items():
            if ctype in seen_types:
                continue  # 同类只保留一条
            seen_types.add(ctype)
            # 提取关键词: 去掉数字和符号后的纯字母
            norm = normalize_header(raw_header).lower()
            words = re.findall(r"[a-z一-鿿]+", norm)
            # 过滤太短或无意义的词
            stop = {"at", "in", "of", "the", "and", "vs", "over", "deg", "mhz",
                    "ghz", "dbi", "dbm", "mm", "db", "pct", "tot", "rad", "pwr"}
            keywords = [w for w in words if len(w) > 1 and w not in stop]
            if not keywords:
                keywords = [norm[:20]]
            # 对于 pct/db 后缀，添加排除
            negate = []
            if ctype.endswith("_pct"):
                negate = ["db"]
            elif ctype.endswith("_db"):
                negate = ["%", "％", "pct"]
            # 对于 gain 类型排除 average/theta
            if ctype == "gain":
                negate = ["average", "avg", "theta"]
            # TRP 排除 NHPRP
            if ctype == "trp":
                negate = ["nhprp"]
            new_patterns.append({
                "col_type": ctype,
                "label": dict(ALL_COL_TYPES).get(ctype, ctype),
                "keywords": [k.lower() for k in keywords],
                "negate": [n.lower() for n in negate],
            })

        if not new_patterns:
            QMessageBox.information(self, self.tr("无数据"), self.tr("没有识别到任何列类型，请先检测模板。"))
            return

        # 写回 JSON — 打包模式存到 EXE 同目录 config/，开发模式存到项目根目录
        if getattr(sys, 'frozen', False):
            config_dir = Path(sys.executable).parent / "config"
        else:
            config_dir = Path(__file__).resolve().parent.parent / "config"
        config_dir.mkdir(parents=True, exist_ok=True)
        json_path = config_dir / "column_patterns.json"

        try:
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            data = {"version": "1.0", "patterns": []}

        # 合并: 用户手动修改的 pattern 替换同名，保留所有新的
        existing = {p.get("col_type"): p for p in data.get("patterns", [])}
        for p in new_patterns:
            existing[p["col_type"]] = p
        data["patterns"] = list(existing.values())

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        # 让 excel_reader 重新加载
        from src.excel_reader import reload_column_patterns
        reload_column_patterns()

        QMessageBox.information(
            self, self.tr("保存成功"),
            self.tr("已保存 {0} 类列头模式到:\n{1}\n\n"
                    "下次模板解析时将优先使用这些模式。").format(
                len(new_patterns), json_path))
