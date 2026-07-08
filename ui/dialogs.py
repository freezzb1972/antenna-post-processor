"""
设置对话框
=========
从主窗口通过菜单调出的设置对话框。
每个对话框管理一组相关设置,确认后将值写回主窗口。
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import TYPE_CHECKING, Dict, List, Optional

if TYPE_CHECKING:
    from src.lag_config import LagConfig  # noqa: F401

import numpy as np

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QRadioButton,
    QTabWidget,  # noqa: F401

    QApplication, QCheckBox, QComboBox, QDialog, QDialogButtonBox,
    QDoubleSpinBox, QFileDialog, QFormLayout, QGroupBox,
    QHBoxLayout, QHeaderView, QLabel, QLineEdit, QMessageBox,
    QListWidget, QListWidgetItem, QPlainTextEdit, QProgressBar, QPushButton,
    QScrollArea, QSizePolicy, QSpinBox, QSplitter, QTableWidget, QTableWidgetItem,
    QVBoxLayout, QWidget,
)

if TYPE_CHECKING:
    from ui.main_window import MainWindow

from src.scale_manager import ScaleManager
from ui.layout_utils import FlowLayout, auto_size_dialog, wrap_in_scroll


# ═══════════════════════════════════════════════════════════════
# 数据源配置对话框
# ═══════════════════════════════════════════════════════════════

class DataSourceDialog(QDialog):
    """文件设置: 模板、数据文件、输出、匹配"""

    def __init__(self, parent: "MainWindow"):
        super().__init__(parent)
        self._mw = parent
        self.setWindowTitle("数据源配置")
        self.resize(750, 650)
        self._setup_ui()
        self._load_state()
        self._init_presets()
        auto_size_dialog(self, 680, 800)

    def _init_presets(self):
        """填充预设模板下拉列表 (可搜索)。"""
        for mfr in self._mw._tm.manufacturers:
            self._cmb_tpl_mfr.addItem(mfr, mfr)

    def _setup_ui(self):
        # 模板选择 — 厂商+模板搜索下拉, 选中后自动填入路径
        grp_tpl = QGroupBox("选择预设模板")
        tpl_layout = QVBoxLayout(grp_tpl)
        preset_row = QHBoxLayout()
        self._cmb_tpl_mfr = QComboBox()
        self._cmb_tpl_mfr.setEditable(True)
        self._cmb_tpl_mfr.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_tpl_mfr.lineEdit().setPlaceholderText("搜索厂商...")
        self._cmb_tpl_mfr.addItem("")
        self._cmb_tpl_mfr.currentIndexChanged.connect(self._on_tpl_preset_changed)
        self._cmb_tpl = QComboBox()
        self._cmb_tpl.setEditable(True)
        self._cmb_tpl.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_tpl.lineEdit().setPlaceholderText("搜索模板...")
        self._cmb_tpl.setMinimumWidth(130)
        self._cmb_tpl.currentIndexChanged.connect(self._on_tpl_item_selected)
        self._cmb_tpl.addItem("")
        preset_row.addWidget(QLabel("厂商:"))
        preset_row.addWidget(self._cmb_tpl_mfr, 1)
        preset_row.addWidget(QLabel("模板:"))
        preset_row.addWidget(self._cmb_tpl, 2)
        tpl_layout.addLayout(preset_row)
        # 手动路径行 — 预设选中后自动填入, 也可手动浏览
        path_row = QHBoxLayout()
        self._edit_template = QLineEdit(); self._edit_template.setPlaceholderText("选择模板 .xlsx ...")
        btn_tpl = QPushButton("浏览..."); btn_tpl.clicked.connect(self._on_browse_template)
        path_row.addWidget(self._edit_template); path_row.addWidget(btn_tpl)
        # 预览列映射按钮
        self._btn_preview_mapping = QPushButton(self.tr("🔍 预览列映射"))
        self._btn_preview_mapping.clicked.connect(self._on_preview_mapping)
        path_row.addWidget(self._btn_preview_mapping)
        tpl_layout.addLayout(path_row)

        # 数据文件
        grp_data = QGroupBox("数据文件")
        data_layout = QVBoxLayout(grp_data)
        btn_row = QHBoxLayout()
        self._btn_add = QPushButton("📂 添加数据文件..."); self._btn_add.setMinimumHeight(32)
        self._btn_add.clicked.connect(self._on_add_files)
        btn_clear = QPushButton("清除"); btn_clear.setMinimumHeight(32)
        btn_clear.clicked.connect(self._on_clear_files)
        btn_row.addWidget(self._btn_add); btn_row.addWidget(btn_clear); btn_row.addStretch()
        data_layout.addLayout(btn_row)

        self._file_list = QListWidget()
        self._file_list.setMinimumHeight(120)
        self._file_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._file_list.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._file_list.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._file_list.setAlternatingRowColors(True)
        data_layout.addWidget(self._file_list)

        self._match_table = QTableWidget(); self._match_table.setColumnCount(3)
        self._match_table.setHorizontalHeaderLabels(["工作表", "数据文件", "状态"])
        self._match_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self._match_table.setMaximumHeight(150)
        data_layout.addWidget(self._match_table)

        # 匹配状态提示 (自动匹配, 无需手动点击)
        self._lbl_match_info = QLabel("添加数据文件后自动匹配工作表")
        self._lbl_match_info.setStyleSheet("color: #888; padding: 2px 0;")
        data_layout.addWidget(self._lbl_match_info)

        # 输出设置
        grp_out = QGroupBox("输出设置")
        out_layout = QFormLayout(grp_out)
        self._edit_dir = QLineEdit(); self._edit_dir.setPlaceholderText("默认: ./output")
        btn_dir = QPushButton("浏览..."); btn_dir.clicked.connect(self._on_browse_dir)
        dir_row = QHBoxLayout(); dir_row.addWidget(self._edit_dir); dir_row.addWidget(btn_dir)
        out_layout.addRow("输出目录:", dir_row)
        self._edit_name = QLineEdit("antenna_report.xlsx")
        out_layout.addRow("文件名:", self._edit_name)
        self._check_full = QCheckBox("生成完整报告")
        out_layout.addRow("", self._check_full)
        self._edit_report = QLineEdit(); self._edit_report.setPlaceholderText("默认: ./output/full_report.xlsx")
        btn_report = QPushButton("浏览..."); btn_report.clicked.connect(self._on_browse_report)
        rpt_row = QHBoxLayout(); rpt_row.addWidget(self._edit_report); rpt_row.addWidget(btn_report)
        out_layout.addRow("报告路径:", rpt_row)

        # 图表选择
        grp_chart = QGroupBox("输出图表")
        ch_row = QHBoxLayout(grp_chart)
        self._check_chart_eff = QCheckBox("效率曲线"); self._check_chart_eff.setChecked(True)
        self._check_chart_lag = QCheckBox("增益曲线"); self._check_chart_lag.setChecked(True)
        ch_row.addWidget(self._check_chart_eff); ch_row.addWidget(self._check_chart_lag); ch_row.addStretch()

        # 工作表命名
        grp_name = QGroupBox("工作表命名")
        name_row = QHBoxLayout(grp_name)
        name_row.addWidget(QLabel("多数据源时工作表命名方式:"))
        self._cmb_naming_mode = QComboBox()
        self._cmb_naming_mode.setEditable(True)
        self._cmb_naming_mode.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_naming_mode.lineEdit().setPlaceholderText("搜索...")
        self._cmb_naming_mode.addItem("保留原模板工作表名", 0)
        self._cmb_naming_mode.addItem("用数据源文件名替换", 1)
        self._cmb_naming_mode.setToolTip("保留模板原名 或 用数据源文件名命名工作表")
        name_row.addWidget(self._cmb_naming_mode)
        name_row.addStretch()

        # 按钮
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._on_accept); btns.rejected.connect(self.reject)

        wrap_in_scroll(self, [grp_tpl, grp_data, grp_name, grp_out, grp_chart], btns)

    def _load_state(self):
        mw = self._mw
        self._edit_template.setText(mw.ui.editTemplatePath.text())
        self._edit_dir.setText(mw.ui.editOutputDir.text())
        self._edit_name.setText(mw.ui.editOutputName.text())
        self._check_full.setChecked(mw.ui.checkFullReport.isChecked())
        self._edit_report.setText(mw.ui.editFullReportPath.text())
        if hasattr(mw, '_check_chart_eff'):
            self._check_chart_eff.setChecked(mw._check_chart_eff.isChecked())
        if hasattr(mw, '_check_chart_lag'):
            self._check_chart_lag.setChecked(mw._check_chart_lag.isChecked())
        # 工作表命名模式
        if hasattr(mw, '_worksheet_naming_mode') and hasattr(self, '_cmb_naming_mode'):
            idx = self._cmb_naming_mode.findData(mw._worksheet_naming_mode)
            if idx >= 0:
                self._cmb_naming_mode.setCurrentIndex(idx)
        # 多文件列表
        if hasattr(mw, '_data_file_paths') and mw._data_file_paths:
            self._file_list.clear()
            for p in mw._data_file_paths:
                self._file_list.addItem(p)
                self._file_list.item(self._file_list.count()-1).setToolTip(p)
        if hasattr(mw, '_last_matches') and mw._last_matches:
            self._copy_match_table()

    def _copy_match_table(self):
        """从 MainWindow._last_matches 填充对话框的匹配表格。"""
        matches = self._mw._last_matches
        self._match_table.setRowCount(len(matches))
        for r, m in enumerate(matches):
            self._match_table.setItem(r, 0, QTableWidgetItem(m.sheet_name))
            combo = QComboBox()
            combo.addItem(m.file_path or "—")
            if m.file_path:
                combo.setCurrentIndex(0)
            self._match_table.setCellWidget(r, 1, combo)
            status = "✓ 已匹配" if m.file_path else "✗ 未匹配"
            self._match_table.setItem(r, 2, QTableWidgetItem(status))

    def _on_browse_template(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择模板", "",
            "所有支持格式 (*.xlsx *.xls *.csv *.docx);;Excel 新版 (*.xlsx);;Excel 旧版 (*.xls);;CSV (*.csv);;Word (*.docx);;所有文件 (*)")
        if path:
            self._edit_template.setText(path)
            self._run_auto_match()

    def _on_tpl_preset_changed(self, index: int):
        """厂商下拉变化 → 刷新模板下拉列表。"""
        self._cmb_tpl.blockSignals(True)
        self._cmb_tpl.clear()
        self._cmb_tpl.addItem("", "")
        mfr = self._cmb_tpl_mfr.currentData()
        templates = self._mw._tm.get_templates(mfr) if mfr else self._mw._tm.get_all_templates()
        for tpl in templates:
            self._cmb_tpl.addItem(tpl.name, tpl)
        self._cmb_tpl.blockSignals(False)

    def _on_tpl_item_selected(self, index: int):
        tpl = self._cmb_tpl.currentData()
        from src.template_manager import TemplatePreset
        if isinstance(tpl, TemplatePreset):
            self._edit_template.setText(tpl.path)

    def _on_preview_mapping(self):
        """打开列映射预览弹窗。"""
        path = self._edit_template.text().strip()
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, self.tr("提示"), self.tr("请先选择模板文件。"))
            return
        try:
            from src.column_mapping import detect_columns_from_template, get_col_type_labels
            mappings = detect_columns_from_template(path)

            dlg = QDialog(self)
            dlg.setWindowTitle(self.tr("列映射预览"))
            dlg.setMinimumSize(550, 400)
            dl = QVBoxLayout(dlg)

            table = QTableWidget()
            table.setColumnCount(3)
            table.setHorizontalHeaderLabels([self.tr("列"), self.tr("列头文本"), self.tr("类型")])
            table.horizontalHeader().setStretchLastSection(True)
            table.setRowCount(len(mappings))
            for ri, m in enumerate(mappings):
                table.setItem(ri, 0, QTableWidgetItem(m.col_letter))
                table.setItem(ri, 1, QTableWidgetItem(m.raw_header))
                cmb = QComboBox()
                for ct, label in get_col_type_labels(0):
                    cmb.addItem(f"{label} ({ct})", ct)
                idx = cmb.findData(m.detected_type)
                if idx >= 0:
                    cmb.setCurrentIndex(idx)
                table.setCellWidget(ri, 2, cmb)
            dl.addWidget(table)

            btn_save = QPushButton(self.tr("💾 保存为预设"))
            btn_save.clicked.connect(lambda: self._save_mapping_from_preview(
                dlg, path, mappings, table))
            dl.addWidget(btn_save)
            dlg.exec()
        except Exception as e:
            QMessageBox.warning(self, self.tr("识别失败"), str(e))

    def _save_mapping_from_preview(self, dlg, path, mappings, table):
        """从预览弹窗保存列映射为模板预设。"""
        from src.column_mapping import ColumnMapping, TemplatePreset, save_preset
        import os
        updated = []
        for ri in range(table.rowCount()):
            cmb = table.cellWidget(ri, 2)
            new_type = cmb.currentData() if cmb else mappings[ri].detected_type
            m = mappings[ri]
            updated.append(ColumnMapping(
                col_letter=m.col_letter, col_index=ri + 1,
                raw_header=m.raw_header, detected_type=m.detected_type,
                confirmed_type=new_type if new_type != m.detected_type else "",
            ))
        name = os.path.splitext(os.path.basename(path))[0]
        ext = os.path.splitext(path)[1].lstrip(".")
        preset = TemplatePreset(
            name=name, path=path, file_type=ext,
            column_mappings=updated)
        save_preset(preset)
        QMessageBox.information(dlg, self.tr("保存成功"),
            self.tr(f"模板预设已保存: {name}"))
        dlg.accept()

    def _on_add_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "选择数据文件", "",
            "所有支持格式 (*.csv *.xlsx *.xls);;CSV (*.csv);;Excel (*.xlsx *.xls)")
        if paths:
            for p in paths:
                self._file_list.addItem(p)
                self._file_list.item(self._file_list.count()-1).setToolTip(p)
            self._run_auto_match()

    def _on_clear_files(self):
        self._file_list.clear(); self._match_table.setRowCount(0)
        self._lbl_match_info.setText("添加数据文件后自动匹配工作表")

    def _run_auto_match(self):
        """执行自动匹配并更新状态显示。"""
        n_files = self._file_list.count()
        if n_files == 0:
            self._lbl_match_info.setText("添加数据文件后自动匹配工作表")
            return
        self._on_auto_match()
        matched = sum(1 for r in range(self._match_table.rowCount())
                      if self._match_table.item(r, 2) and "✓" in (self._match_table.item(r, 2).text() or ""))
        total = self._match_table.rowCount()
        if total > 0:
            self._lbl_match_info.setText(f"已添加 {n_files} 个文件, 自动匹配 {matched}/{total} 个工作表 — 可手动调整")
        else:
            self._lbl_match_info.setText(f"已添加 {n_files} 个文件, 未匹配到工作表 (请检查模板)")

    def _on_auto_match(self):
        tpl = self._edit_template.text().strip()
        if not tpl:
            QMessageBox.warning(self, "提示", "请先选择模板文件。"); return
        from src.excel_reader import read_template
        from src.sheet_file_matcher import auto_match
        try:
            sheets = read_template(tpl)
            data_files = [self._file_list.item(i).text() for i in range(self._file_list.count())]
            matches = auto_match([s.name for s in sheets], data_files)
            self._match_table.setRowCount(len(matches))
            for i, m in enumerate(matches):
                self._match_table.setItem(i, 0, QTableWidgetItem(m.sheet_name))
                combo = QComboBox(); combo.addItem("—")
                for fp in data_files: combo.addItem(fp)
                if m.file_path:
                    idx = combo.findText(m.file_path)
                    if idx >= 0: combo.setCurrentIndex(idx)
                self._match_table.setCellWidget(i, 1, combo)
                status = "✓ 已匹配" if m.file_path else "未匹配"
                self._match_table.setItem(i, 2, QTableWidgetItem(status))
            matched = sum(1 for m in matches if m.file_path is not None)
            self._lbl_match_info.setText(f"✓ {matched}/{len(matches)} 已匹配")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"读取模板失败: {e}")

    def _on_browse_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择输出目录", self._edit_dir.text() or ".")
        if d: self._edit_dir.setText(d)

    def _on_browse_report(self):
        path, _ = QFileDialog.getSaveFileName(self, "完整报告路径", "full_report.xlsx", "Excel (*.xlsx)")
        if path: self._edit_report.setText(path)

    def _on_accept(self):
        mw = self._mw
        mw.ui.editTemplatePath.setText(self._edit_template.text())
        mw.ui.editOutputDir.setText(self._edit_dir.text())
        mw.ui.editOutputName.setText(self._edit_name.text())
        # 同步到 azimuth_config
        az = getattr(mw, '_azimuth_config', None)
        if az:
            az.excel_output_dir = self._edit_dir.text().strip()
            az.excel_output_filename = self._edit_name.text().strip()
        mw.ui.checkFullReport.setChecked(self._check_full.isChecked())
        mw.ui.editFullReportPath.setText(self._edit_report.text())
        # 持久化模板和输出路径到配置文件
        if hasattr(mw, '_save_template_path'):
            mw._save_template_path(self._edit_template.text())
        # 模板变更后自动识别并应用计算参数
        if hasattr(mw, '_auto_apply_template_params'):
            mw._cached_template_params = set()
            mw._auto_apply_template_params()
        mw._cfg.config.last_output_dir = self._edit_dir.text()
        mw._cfg._dirty = True
        if hasattr(mw, '_check_chart_eff'):
            mw._check_chart_eff.setChecked(self._check_chart_eff.isChecked())
        if hasattr(mw, '_check_chart_lag'):
            mw._check_chart_lag.setChecked(self._check_chart_lag.isChecked())
        # 工作表命名模式
        if hasattr(self, '_cmb_naming_mode'):
            mw._worksheet_naming_mode = self._cmb_naming_mode.currentData() or 0
            if hasattr(mw, '_cmb_naming_mode'):
                idx = mw._cmb_naming_mode.findData(mw._worksheet_naming_mode)
                if idx >= 0:
                    mw._cmb_naming_mode.setCurrentIndex(idx)
        # 更新数据文件列表 (存完整路径) — 陈旧数据保护
        data_files = [self._file_list.item(i).text() for i in range(self._file_list.count())]
        if hasattr(mw, '_data_file_paths'):
            old_count = len(mw._data_file_paths)
            mw._data_file_paths = data_files
            # 如果通过对话框清空/替换了文件，标记非陈旧
            mw._data_stale = False
            if hasattr(mw, '_sync_file_entries'):
                mw._sync_file_entries()
            # 对话框替换了文件列表 → 完整重建 UI 和匹配
            if hasattr(mw, '_file_list_widget'):
                mw._refresh_data_file_ui()
            if hasattr(mw, '_last_matches'):
                mw._last_matches = []
            if hasattr(mw, '_lbl_match_status') and mw._lbl_match_status is not None:
                mw._lbl_match_status.setText("")
            # 重建匹配表
            if data_files and hasattr(mw, '_on_auto_match'):
                mw._on_auto_match()
        self.accept()


# ═══════════════════════════════════════════════════════════════
# 计算参数配置对话框
# ═══════════════════════════════════════════════════════════════

class CalcParamsDialog(QDialog):
    """计算参数配置: 三 Tab 分类 + 双列参数 + 内嵌角度配置 + 算法选项"""

    # ── 参数定义: 三类 Tab 的通用参数 ──
    _COMMON_PARAMS = [
        ("Gain", [
            ("gain", "Peak Gain"),
            ("lag_single", "Gain @ θ（单角度）"),
            ("lag_range", "Gain @ θ Range（范围）"),
        ]),
        ("Directivity", [
            ("directivity", "Directivity (dBi)"),
        ]),
        ("Efficiency", [
            ("efficiency_pct", "Efficiency (%)"),
            ("efficiency_db", "Efficiency (dB)"),
        ]),
        ("Axial Ratio", [
            ("ar_single", "AR @ θ（单角度）"),
            ("ar_range", "AR @ θ Range（范围）"),
        ]),
        ("波束参数", [
            ("boresight_theta", "Boresight θ"),
            ("boresight_phi", "Boresight φ"),
            ("theta_bw", "Theta Beamwidth (3dB)"),
            ("phi_bw", "Phi Beamwidth (3dB)"),
            ("front_back_ratio", "Front/Back Ratio"),
        ]),
        ("功率统计", [
            ("max_power", "Maximum Power"),
            ("min_power", "Minimum Power"),
            ("avg_gain", "Average Gain"),
            ("avg_power", "Average Power"),
            ("max_min_ratio", "Max/Min Ratio"),
            ("max_avg_ratio", "Max/Avg Ratio"),
            ("min_avg_ratio", "Min/Avg Ratio"),
        ]),
        ("交叉极化隔离度 (XPI)", [
            ("xpi_boresight", "XPI @ Boresight (dB)"),
            ("xpi_mean", "XPI Mean (dB)"),
            ("xpi_min", "XPI Min (dB)"),
        ]),
        ("总效率", [
            ("total_efficiency_pct", "Total Efficiency (%)"),
            ("mismatch_loss_db", "Mismatch Loss (dB)"),
        ]),
        ("相位中心", [
            ("pc_theta_mm", "PC Theta (mm)"),
            ("pc_phi_mm", "PC Phi (mm)"),
        ]),
    ]

    # 有源发射特有参数
    _TRP_PARAMS = [
        ("TRP", [
            ("trp", "Total Radiated Power (TRP)"),
            ("peak_eirp", "Peak EIRP"),
        ]),
        ("NHPRP", [
            ("nhprp_45", "NHPRP ±45° (Pi/4)"),
            ("nhprp_30", "NHPRP ±30° (Pi/6)"),
            ("nhprp_225", "NHPRP ±22.5° (Pi/8)"),
            ("nhprp_custom", "NHPRP 自定义角度"),
        ]),
        ("半球 PRP", [
            ("uh_prp", "Upper Hemisphere PRP"),
            ("lh_prp", "Lower Hemisphere PRP"),
            ("prp_120", "PRP (theta 0-120°)"),
        ]),
        ("比率", [
            ("nhprp45_ratio", "NHPRP45 / TRP"),
            ("nhprp30_ratio", "NHPRP30 / TRP"),
            ("nhprp225_ratio", "NHPRP22.5 / TRP"),
            ("uh_ratio", "UHPRP / TRP"),
            ("lh_ratio", "LHPRP / TRP"),
        ]),
    ]

    # 有源接收特有参数
    _TIS_PARAMS = [
        ("TIS", [
            ("tis", "Total Isotropic Sensitivity (TIS)"),
        ]),
        ("NHPIS", [
            ("nhpis_45", "NHPIS ±45° (Pi/4)"),
            ("nhpis_30", "NHPIS ±30° (Pi/6)"),
            ("nhpis_225", "NHPIS ±22.5° (Pi/8)"),
            ("nhpis_custom", "NHPIS 自定义角度"),
        ]),
        ("半球 PIS", [
            ("uh_pis", "Upper Hemisphere PIS"),
            ("lh_pis", "Lower Hemisphere PIS"),
            ("pis_120", "PIS (theta 0-120°)"),
        ]),
        ("比率", [
            ("nhpis45_ratio", "NHPIS45 / TIS"),
            ("nhpis30_ratio", "NHPIS30 / TIS"),
            ("nhpis225_ratio", "NHPIS22.5 / TIS"),
            ("uh_ratio", "UHPIS / TIS"),
            ("lh_ratio", "LHPIS / TIS"),
        ]),
    ]

    def __init__(self, parent: "MainWindow"):
        super().__init__(parent)
        self._mw = parent
        self.setWindowTitle("计算参数配置")
        self.resize(820, 680)

        # ── 状态 (每种测试模式独立存储) ──
        self._template_params: set = set()
        # 三个模式的独立状态: [被动=0, 有源发射=1, 有源接收=2]
        self._mode_states = [
            {"singles": [], "ranges": [], "ar_singles": [], "ar_ranges": [],
             "nh_custom_angles": [], "extrapolate": False, "robust_peak": False,
             "ar_output_db": True,
             "freq_source": "datasource", "trim_start": 0, "trim_end": 0,
             "required": set(), "extra": set()},
            {"singles": [], "ranges": [], "ar_singles": [], "ar_ranges": [],
             "nh_custom_angles": [], "extrapolate": False, "robust_peak": False,
             "ar_output_db": True,
             "freq_source": "datasource", "trim_start": 0, "trim_end": 0,
             "required": set(), "extra": set()},
            {"singles": [], "ranges": [], "ar_singles": [], "ar_ranges": [],
             "nh_custom_angles": [], "extrapolate": False, "robust_peak": False,
             "ar_output_db": True,
             "freq_source": "datasource", "trim_start": 0, "trim_end": 0,
             "required": set(), "extra": set()},
        ]
        self._angle_singles: List[float] = []
        self._angle_ranges: List[tuple] = []
        self._ar_angle_singles: List[float] = []
        self._ar_angle_ranges: List[tuple] = []
        self._nh_custom_angles: List[float] = []
        self._extrapolate: bool = False
        self._robust_peak: bool = False
        self._active_tab: int = 0
        self._test_mode: int = 0  # 0=passive, 1=TRP, 2=TIS
        self._required_params: set = set()
        self._extra_params: set = set()

        # ── 动态 widget 引用（按 tab 切换时重建） ──
        self._left_checkboxes: Dict[str, QCheckBox] = {}
        self._right_checkboxes: Dict[str, QCheckBox] = {}
        self._left_scroll: Optional[QScrollArea] = None
        self._right_scroll: Optional[QScrollArea] = None

        self._setup_ui()
        self._load_state()
        self._rebuild_param_columns()
        # 初始状态: 无源模式, AR按钮和频点行可见
        self._btn_ar_angle.setVisible(True)
        self._freq_widget.setVisible(True)
        auto_size_dialog(self, 780, 650)

    # ═══════════════════════════════════════════════════════════
    # UI 构建
    # ═══════════════════════════════════════════════════════════

    def _setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(8)

        # ── 测试模式选择 (outside scroll) ──
        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("<b>测试模式:</b>"))
        self._cmb_test_mode = QComboBox()
        self._cmb_test_mode.addItem("📡 无源天线", 0)
        self._cmb_test_mode.addItem("📶 有源发射 TRP", 1)
        self._cmb_test_mode.addItem("📻 有源接收 TIS", 2)
        self._cmb_test_mode.currentIndexChanged.connect(self._on_mode_changed)
        mode_row.addWidget(self._cmb_test_mode)
        mode_row.addStretch()
        main_layout.addLayout(mode_row)

        # ── 参数选择区域 (inside scroll) ──
        param_widget = QWidget()
        param_layout = QVBoxLayout(param_widget)
        param_layout.setSpacing(8)

        # 双列参数
        splitter = QHBoxLayout()
        splitter.setSpacing(8)

        # 左列: 报告必需参数
        left_grp = QGroupBox("报告必需参数（模板自动识别，可调整）")
        left_layout = QVBoxLayout(left_grp)
        self._left_scroll = QScrollArea()
        self._left_scroll.setWidgetResizable(True)
        self._left_scroll.setFrameShape(QScrollArea.NoFrame)
        left_layout.addWidget(self._left_scroll)
        splitter.addWidget(left_grp, 1)

        # 右列: 额外报告
        right_grp = QGroupBox("额外报告")
        right_layout = QVBoxLayout(right_grp)
        self._right_scroll = QScrollArea()
        self._right_scroll.setWidgetResizable(True)
        self._right_scroll.setFrameShape(QScrollArea.NoFrame)
        right_layout.addWidget(self._right_scroll)
        splitter.addWidget(right_grp, 1)

        param_layout.addLayout(splitter, 1)

        # ── 角度配置: Gain/AR 切换 ──
        angle_grp = QGroupBox("角度配置 (已移至各参数组内)")
        angle_grp.setVisible(False)  # 角度设置已移至 Gain/AR 组内弹窗
        angle_outer = QVBoxLayout(angle_grp)
        angle_outer.setSpacing(4)

        # 切换按钮
        toggle_row = QHBoxLayout()
        toggle_row.addWidget(QLabel("<b>当前编辑:</b>"))
        self._btn_gain_angle = QPushButton("📡 Gain 角度")
        self._btn_gain_angle.setCheckable(True); self._btn_gain_angle.setChecked(True)
        self._btn_gain_angle.clicked.connect(lambda: self._on_angle_target_changed(0))
        toggle_row.addWidget(self._btn_gain_angle)
        self._btn_ar_angle = QPushButton("🔄 AR 角度")
        self._btn_ar_angle.setCheckable(True)
        self._btn_ar_angle.clicked.connect(lambda: self._on_angle_target_changed(1))
        toggle_row.addWidget(self._btn_ar_angle)
        toggle_row.addStretch()
        angle_outer.addLayout(toggle_row)
        self._active_angle_tab: int = 0

        # 共享角度控件
        self._build_angle_control_widget()
        angle_outer.addWidget(self._angle_ctrl_widget)

        # NHPRP/NHPIS 自定义角度 (TRP/TIS Tab 时显示)
        self._grp_nh = QGroupBox("NHPRP / NHPIS 自定义角度")
        nh_layout = QHBoxLayout(self._grp_nh)
        self._btn_nh_angle = QPushButton("⚙ 自定义角度...")
        self._btn_nh_angle.clicked.connect(self._show_nh_angle_popup)
        nh_layout.addWidget(self._btn_nh_angle)
        self._lbl_nh_angles = QLabel("（默认 45°）")
        nh_layout.addWidget(self._lbl_nh_angles)
        nh_layout.addStretch()
        self._grp_nh.setVisible(False)
        angle_outer.addWidget(self._grp_nh)

        # 已选择区域
        self._selected_widget = QWidget()
        self._selected_layout = QVBoxLayout(self._selected_widget)
        self._selected_layout.setContentsMargins(0, 0, 0, 0)
        self._selected_layout.setSpacing(2)
        angle_outer.addWidget(self._selected_widget)

        param_layout.addWidget(angle_grp)

        # ── 算法选项 ──
        algo_grp = QGroupBox("算法选项")
        algo_layout = QVBoxLayout(algo_grp)
        algo_layout.setSpacing(4)

        self._freq_widget = QWidget()
        freq_row = QHBoxLayout(self._freq_widget)
        freq_row.setContentsMargins(0, 0, 0, 0)
        freq_row.addWidget(QLabel("频点来源:"))
        self._cmb_freq_src = QComboBox()
        self._cmb_freq_src.addItem("新 sheet 频点: 数据源", "datasource")
        self._cmb_freq_src.addItem("新 sheet 频点: 模板", "template")
        self._cmb_freq_src.currentIndexChanged.connect(lambda: self._update_summary())
        freq_row.addWidget(self._cmb_freq_src)
        freq_row.addWidget(QLabel("  去除频点: 前"))
        self._spin_trim_start = QSpinBox()
        self._spin_trim_start.setRange(0, 50); self._spin_trim_start.setFixedWidth(50)
        self._spin_trim_start.valueChanged.connect(lambda: self._update_summary())
        freq_row.addWidget(self._spin_trim_start)
        freq_row.addWidget(QLabel("后"))
        self._spin_trim_end = QSpinBox()
        self._spin_trim_end.setRange(0, 50); self._spin_trim_end.setFixedWidth(50)
        self._spin_trim_end.valueChanged.connect(lambda: self._update_summary())
        freq_row.addWidget(self._spin_trim_end)
        freq_row.addStretch()
        algo_layout.addWidget(self._freq_widget)

        check_row = QHBoxLayout()
        self._check_extrap = QCheckBox("Theta 外推到 180°")
        self._check_extrap.toggled.connect(lambda: self._update_summary())
        check_row.addWidget(self._check_extrap)
        self._check_robust = QCheckBox("Robust peak detection（替代 np.max）")
        self._check_robust.toggled.connect(lambda: self._update_summary())
        check_row.addWidget(self._check_robust)
        self._cmb_ar_output = QComboBox()
        self._cmb_ar_output.addItem("AR 输出 dB", True)
        self._cmb_ar_output.addItem("AR 输出 线性", False)
        self._cmb_ar_output.setCurrentIndex(0)  # 默认 dB
        self._cmb_ar_output.setToolTip("AR 输出单位: dB (20·log₁₀) 或线性比值")
        self._cmb_ar_output.currentIndexChanged.connect(lambda: self._update_summary())
        check_row.addWidget(self._cmb_ar_output)
        check_row.addStretch()
        algo_layout.addLayout(check_row)

        param_layout.addWidget(algo_grp)
        param_layout.addStretch()

        # 将参数区域包裹进 QScrollArea
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)
        scroll.setWidget(param_widget)
        main_layout.addWidget(scroll, 1)

        # ── 已选参数概览 ──
        self._summary_grp = QGroupBox("📋 已选参数概览")
        summary_layout = QVBoxLayout(self._summary_grp)
        self._summary_label = QLabel()
        self._summary_label.setWordWrap(True)
        self._summary_label.setTextFormat(Qt.RichText)
        self._summary_label.setStyleSheet("padding: 4px; font-size: 12px;")
        summary_layout.addWidget(self._summary_label)
        main_layout.addWidget(self._summary_grp)

        # ── 按钮 ──
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        main_layout.addWidget(btns)

    # ═══════════════════════════════════════════════════════════
    # Tab 参数列表
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _get_params_for_tab(tab_index: int):
        """返回当前 Tab 的参数定义列表。

        Tab 0 (无源):    全部通用参数含 AR — Gain/Directivity/Efficiency/AR/波束/功率/XPI/总效率/相位中心。
        Tab 1 (有源 TRP): 通用参数(不含 AR) + TRP/NHPRP/PRP/比率 — 有源发射无相位数据, AR 无法计算。
        Tab 2 (有源 TIS): 仅灵敏度参数 — TIS/NHPIS/PIS/比率。TIS 测试不报告 Gain/Directivity/Efficiency 等无源参数。
        """
        if tab_index == 0:
            return list(CalcParamsDialog._COMMON_PARAMS)
        elif tab_index == 1:
            # TRP: 保留通用参数(不含 AR) + TRP 特有参数
            no_ar = [(g, plist) for g, plist in CalcParamsDialog._COMMON_PARAMS if g != "Axial Ratio"]
            return no_ar + list(CalcParamsDialog._TRP_PARAMS)
        else:
            # TIS: 仅灵敏度参数，不包含 Gain/Directivity/Efficiency 等无源参数
            return list(CalcParamsDialog._TIS_PARAMS)

    def _on_mode_changed(self, index: int):
        # 保存当前模式状态，加载新模式状态
        self._save_current_mode_state()
        self._load_mode_state(index)
        self._active_tab = index
        self._test_mode = index
        is_active = index in (1, 2)
        is_tis = index == 2
        self._grp_nh.setVisible(is_active)
        # 有源测试无 AR: 隐藏 AR 角度按钮; TIS 模式也隐藏 Gain 角度按钮
        self._btn_ar_angle.setVisible(not is_active)
        self._btn_gain_angle.setVisible(not is_tis)
        if is_active and self._btn_ar_angle.isChecked():
            self._btn_gain_angle.setChecked(True)
        # 有源测试(TRP/TIS)频点固定从数据源获取, 不需要来源选择和去除频点
        # 因为有源测试的频点由测量设备定义，模板中的频点列表不适用于 TRP/TIS
        self._freq_widget.setVisible(not is_active)
        self._rebuild_param_columns()
        self._update_selected_display()
        self._sync_angle_buttons()  # 模式切换后同步快捷角度按钮状态

    def _save_current_mode_state(self):
        """保存当前模式的状态。"""
        m = self._test_mode
        s = self._mode_states[m]
        s["singles"] = list(self._angle_singles)
        s["ranges"] = list(self._angle_ranges)
        s["ar_singles"] = list(self._ar_angle_singles)
        s["ar_ranges"] = list(self._ar_angle_ranges)
        s["nh_custom_angles"] = list(self._nh_custom_angles)
        s["extrapolate"] = self._check_extrap.isChecked()
        s["robust_peak"] = self._check_robust.isChecked()
        s["ar_output_db"] = self._cmb_ar_output.currentData()
        s["freq_source"] = self._cmb_freq_src.currentData()
        s["trim_start"] = self._spin_trim_start.value()
        s["trim_end"] = self._spin_trim_end.value()
        s["required"] = self._get_checked_keys(self._left_checkboxes)
        s["extra"] = self._get_checked_keys(self._right_checkboxes)

    def _load_mode_state(self, mode: int):
        """恢复指定模式的状态。"""
        self._test_mode = mode
        s = self._mode_states[mode]
        self._angle_singles = list(s["singles"])
        self._angle_ranges = list(s["ranges"])
        self._ar_angle_singles = list(s["ar_singles"])
        self._ar_angle_ranges = list(s["ar_ranges"])
        self._nh_custom_angles = list(s.get("nh_custom_angles", []))
        self._sync_nh_angle_display()
        self._check_extrap.setChecked(s["extrapolate"])
        self._check_robust.setChecked(s["robust_peak"])
        self._cmb_ar_output.setCurrentIndex(0 if s.get("ar_output_db", True) else 1)
        idx = self._cmb_freq_src.findData(s["freq_source"])
        if idx >= 0: self._cmb_freq_src.setCurrentIndex(idx)
        self._spin_trim_start.setValue(s["trim_start"])
        self._spin_trim_end.setValue(s["trim_end"])
        self._required_params = s["required"]
        self._extra_params = s["extra"]

    def _rebuild_param_columns(self):
        """重建参数列表 — 单列层级结构, Gain/AR 组内嵌角度设置按钮。"""
        params = self._get_params_for_tab(self._active_tab)

        # 单列内容
        content = QWidget()
        vbox = QVBoxLayout(content)
        vbox.setContentsMargins(4, 4, 4, 4)
        vbox.setSpacing(6)
        self._left_checkboxes.clear()
        self._right_checkboxes.clear()

        for grp_name, items in params:
            grp = QGroupBox(grp_name)
            gl = QVBoxLayout(grp); gl.setSpacing(2)
            for key, label in items:
                cb = QCheckBox(label)
                cb.setChecked(key in self._template_params)
                cb.toggled.connect(lambda checked, k=key: self._update_summary())
                gl.addWidget(cb)
                self._left_checkboxes[key] = cb
                self._right_checkboxes[key] = cb  # 统一管理
            # Gain / AR 组: 添加角度设置按钮
            if grp_name == "Gain":
                btn_angle = QPushButton("📡 Gain 角度设置...")
                btn_angle.clicked.connect(lambda: self._show_angle_popup("Gain"))
                gl.addWidget(btn_angle)
            elif grp_name == "Axial Ratio":
                btn_ar = QPushButton("🔄 AR 角度设置...")
                btn_ar.clicked.connect(lambda: self._show_angle_popup("AR"))
                gl.addWidget(btn_ar)
            vbox.addWidget(grp)
        vbox.addStretch()

        # 放入左列 scroll (右列 scroll 隐藏, 不再使用双列布局)
        self._left_scroll.setWidget(content)
        # 隐藏右列 scroll (改用单列)
        rw = self._right_scroll.parent()
        if rw and hasattr(rw, 'hide'):
            rw.hide()
        self._update_summary()

    def _show_angle_popup(self, target: str):
        """弹出角度选择窗口 (Gain 或 AR) — 流式标签 + 可拖动分隔条。

        所有增删操作实时更新显示，不关闭对话框。
        只有点击 OK 才提交修改，Cancel 放弃修改。
        """
        dlg = QDialog(self)
        dlg.setWindowTitle(f"{target} 角度配置")
        dlg.setMinimumSize(520, 460)

        is_ar = (target == "AR")
        # 深拷贝 — Cancel 时恢复原值
        import copy
        _src_singles = self._ar_angle_singles if is_ar else self._angle_singles
        _src_ranges = self._ar_angle_ranges if is_ar else self._angle_ranges
        _singles: List[float] = copy.deepcopy(_src_singles)
        _ranges: List[tuple] = copy.deepcopy(_src_ranges)

        # ── 帮助函数: 刷新已配置项显示 ──
        def _refresh_display():
            # 清空旧内容
            while _display_layout.count():
                item = _display_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
            if _singles or _ranges:
                scroll = QScrollArea()
                scroll.setWidgetResizable(True)
                dw = QWidget()
                fl = FlowLayout(dw, margin=4, h_spacing=6, v_spacing=4)
                for a in sorted(set(_singles)):
                    tag = QWidget()
                    tl = QHBoxLayout(tag); tl.setContentsMargins(2, 1, 2, 1); tl.setSpacing(2)
                    tl.addWidget(QLabel(f"{a}°"))
                    btn_del = QPushButton("✕")
                    btn_del.setFixedSize(20, 20); btn_del.setStyleSheet("padding:0;")
                    btn_del.clicked.connect(lambda checked, v=a: (_singles.remove(v), _refresh_display()))
                    tl.addWidget(btn_del)
                    fl.addWidget(tag)
                for lo, hi in sorted(set(_ranges), key=lambda x: (x[0], x[1])):
                    tag = QWidget()
                    tl = QHBoxLayout(tag); tl.setContentsMargins(2, 1, 2, 1); tl.setSpacing(2)
                    tl.addWidget(QLabel(f"{lo}°~{hi}°"))
                    btn_del = QPushButton("✕")
                    btn_del.setFixedSize(20, 20); btn_del.setStyleSheet("padding:0;")
                    btn_del.clicked.connect(lambda checked, lo=lo, hi=hi: (_ranges.remove((lo, hi)), _refresh_display()))
                    tl.addWidget(btn_del)
                    fl.addWidget(tag)
                scroll.setWidget(dw)
                _display_layout.addWidget(scroll)
                btn_clear = QPushButton("🗑 清空全部")
                btn_clear.clicked.connect(lambda: (_singles.clear(), _ranges.clear(), _refresh_display()))
                _display_layout.addWidget(btn_clear)
            else:
                _display_layout.addWidget(QLabel("  (暂无配置)"))
            _display_grp.setTitle(f"已配置: {len(_singles)} 个单角度, {len(_ranges)} 个范围")

        # ── 顶部: 已配置项显示 ──
        _display_grp = QGroupBox()
        _display_layout = QVBoxLayout(_display_grp)
        _refresh_display()

        # ── QSplitter: 上 (已配置) / 下 (操作控件) 可拖动 ──
        splitter = QSplitter(Qt.Vertical)
        bottom_ctls = QWidget()
        bottom_layout = QVBoxLayout(bottom_ctls)
        bottom_layout.setContentsMargins(0, 0, 0, 0)

        # 自定义
        cust_grp = QGroupBox("自定义")
        cust_layout = QHBoxLayout(cust_grp)
        spin_custom = QDoubleSpinBox(); spin_custom.setRange(0, 180); spin_custom.setValue(45)
        btn_add_custom = QPushButton("+ 添加")
        btn_add_custom.clicked.connect(lambda: (
            _singles.append(spin_custom.value()) if spin_custom.value() not in _singles else None,
            _refresh_display()
        ))
        cust_layout.addWidget(QLabel("角度:")); cust_layout.addWidget(spin_custom)
        cust_layout.addWidget(btn_add_custom); cust_layout.addStretch()
        bottom_layout.addWidget(cust_grp)

        # 步进生成
        step_grp = QGroupBox("步进批量生成")
        step_layout = QHBoxLayout(step_grp)
        spin_start = QDoubleSpinBox(); spin_start.setRange(0, 180); spin_start.setValue(0)
        spin_end = QDoubleSpinBox(); spin_end.setRange(0, 180); spin_end.setValue(90)
        spin_step = QDoubleSpinBox(); spin_step.setRange(1, 90); spin_step.setValue(10)
        btn_gen = QPushButton("生成")
        btn_gen.clicked.connect(lambda: (
            [_singles.append(round(float(a), 6)) for a in np.linspace(spin_start.value(), spin_end.value(), int((spin_end.value()-spin_start.value())/spin_step.value()+1))
             if round(float(a), 6) not in _singles],
            _refresh_display()
        ))
        step_layout.addWidget(QLabel("起:")); step_layout.addWidget(spin_start)
        step_layout.addWidget(QLabel("止:")); step_layout.addWidget(spin_end)
        step_layout.addWidget(QLabel("步:")); step_layout.addWidget(spin_step)
        step_layout.addWidget(btn_gen)
        bottom_layout.addWidget(step_grp)

        # 范围
        range_grp = QGroupBox("角度范围")
        range_layout = QHBoxLayout(range_grp)
        spin_rs = QDoubleSpinBox(); spin_rs.setRange(0, 180); spin_rs.setValue(0)
        spin_re = QDoubleSpinBox(); spin_re.setRange(0, 180); spin_re.setValue(90)
        btn_add_range = QPushButton("添加范围")
        def _add_range():
            lo, hi = spin_rs.value(), spin_re.value()
            key = (min(lo, hi), max(lo, hi))
            if key not in _ranges:
                _ranges.append(key)
                _refresh_display()
        btn_add_range.clicked.connect(_add_range)
        range_layout.addWidget(QLabel("起:")); range_layout.addWidget(spin_rs)
        range_layout.addWidget(QLabel("止:")); range_layout.addWidget(spin_re)
        range_layout.addWidget(btn_add_range); range_layout.addStretch()
        bottom_layout.addWidget(range_grp)

        # 确定 / 取消
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(lambda: (
            _src_singles.clear(), _src_singles.extend(sorted(set(_singles))),
            _src_ranges.clear(), _src_ranges.extend(sorted(set(_ranges), key=lambda x: (x[0], x[1]))),
            dlg.accept()
        ))
        btns.rejected.connect(dlg.reject)
        bottom_layout.addWidget(btns)

        splitter.addWidget(_display_grp)
        splitter.addWidget(bottom_ctls)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 0)

        layout = QVBoxLayout(dlg)
        layout.addWidget(splitter)
        dlg.exec()
        self._sync_angle_buttons()
        self._update_selected_display()

    def _sync_nh_angle_display(self):
        """更新 NH 自定义角度标签显示。"""
        if self._nh_custom_angles:
            angles_str = ", ".join(f"{a}°" for a in sorted(set(self._nh_custom_angles)))
            self._lbl_nh_angles.setText(angles_str)
        else:
            self._lbl_nh_angles.setText("（默认 45°）")

    def _show_nh_angle_popup(self):
        """弹出 NHPRP/NHPIS 自定义地平线边界角度选择窗口。

        复用 Gain/AR 角度弹窗模式: deep copy + OK 提交 / Cancel 放弃。
        仅支持单角度（NH 不需要范围），预置 Pi/N 快捷按钮。
        """
        dlg = QDialog(self)
        dlg.setWindowTitle("NHPRP / NHPIS 自定义角度")
        dlg.setMinimumSize(460, 380)

        import copy
        _src_angles = self._nh_custom_angles
        _angles: List[float] = copy.deepcopy(_src_angles)

        # ── 刷新已配置项显示 ──
        def _refresh_display():
            while _display_layout.count():
                item = _display_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
            if _angles:
                scroll = QScrollArea()
                scroll.setWidgetResizable(True)
                dw = QWidget()
                fl = FlowLayout(dw, margin=4, h_spacing=6, v_spacing=4)
                for a in sorted(set(_angles)):
                    tag = QWidget()
                    tl = QHBoxLayout(tag); tl.setContentsMargins(2, 1, 2, 1); tl.setSpacing(2)
                    tl.addWidget(QLabel(f"±{a}°"))
                    btn_del = QPushButton("✕")
                    btn_del.setFixedSize(20, 20); btn_del.setStyleSheet("padding:0;")
                    btn_del.clicked.connect(lambda checked, v=a: (_angles.remove(v), _refresh_display()))
                    tl.addWidget(btn_del)
                    fl.addWidget(tag)
                scroll.setWidget(dw)
                _display_layout.addWidget(scroll)
                btn_clear = QPushButton("🗑 清空全部")
                btn_clear.clicked.connect(lambda: (_angles.clear(), _refresh_display()))
                _display_layout.addWidget(btn_clear)
            else:
                _display_layout.addWidget(QLabel("  (默认 45°，添加自定义角度可覆盖默认值)"))
            _display_grp.setTitle(f"已配置: {len(_angles)} 个角度")

        # ── 顶部: 已配置项 ──
        _display_grp = QGroupBox()
        _display_layout = QVBoxLayout(_display_grp)
        _refresh_display()

        # ── 操作控件 ──
        bottom_ctls = QWidget()
        bottom_layout = QVBoxLayout(bottom_ctls)
        bottom_layout.setContentsMargins(0, 0, 0, 0)

        # 快捷预置
        quick_grp = QGroupBox("快捷预置")
        quick_layout = QHBoxLayout(quick_grp)
        presets = [
            ("22.5° (Pi/8)", 22.5), ("30° (Pi/6)", 30.0), ("45° (Pi/4)", 45.0),
            ("60° (Pi/3)", 60.0), ("75°", 75.0),
        ]
        for label, val in presets:
            btn = QPushButton(label)
            btn.clicked.connect(lambda checked, v=val: (_angles.append(v) if v not in _angles else None, _refresh_display()))
            quick_layout.addWidget(btn)
        bottom_layout.addWidget(quick_grp)

        # 自定义
        cust_grp = QGroupBox("自定义")
        cust_layout = QHBoxLayout(cust_grp)
        spin_custom = QDoubleSpinBox(); spin_custom.setRange(0, 90); spin_custom.setValue(45)
        spin_custom.setSuffix("°"); spin_custom.setDecimals(1)
        btn_add_custom = QPushButton("+ 添加")
        btn_add_custom.clicked.connect(lambda: (
            _angles.append(spin_custom.value()) if spin_custom.value() not in _angles else None,
            _refresh_display()
        ))
        cust_layout.addWidget(QLabel("角度:")); cust_layout.addWidget(spin_custom)
        cust_layout.addWidget(btn_add_custom); cust_layout.addStretch()
        bottom_layout.addWidget(cust_grp)

        # 步进生成
        step_grp = QGroupBox("步进批量生成")
        step_layout = QHBoxLayout(step_grp)
        spin_start = QDoubleSpinBox(); spin_start.setRange(0, 90); spin_start.setValue(0)
        spin_end = QDoubleSpinBox(); spin_end.setRange(0, 90); spin_end.setValue(90)
        spin_step = QDoubleSpinBox(); spin_step.setRange(1, 45); spin_step.setValue(15)
        btn_gen = QPushButton("生成")
        btn_gen.clicked.connect(lambda: (
            [_angles.append(round(float(a), 6)) for a in np.linspace(spin_start.value(), spin_end.value(), int((spin_end.value()-spin_start.value())/spin_step.value()+1))
             if round(float(a), 6) not in _angles],
            _refresh_display()
        ))
        step_layout.addWidget(QLabel("起:")); step_layout.addWidget(spin_start)
        step_layout.addWidget(QLabel("止:")); step_layout.addWidget(spin_end)
        step_layout.addWidget(QLabel("步:")); step_layout.addWidget(spin_step)
        step_layout.addWidget(btn_gen)
        bottom_layout.addWidget(step_grp)

        # 确定 / 取消
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(lambda: (
            _src_angles.clear(), _src_angles.extend(sorted(set(_angles))),
            dlg.accept()
        ))
        btns.rejected.connect(dlg.reject)
        bottom_layout.addWidget(btns)

        splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(_display_grp)
        splitter.addWidget(bottom_ctls)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 0)

        layout = QVBoxLayout(dlg)
        layout.addWidget(splitter)
        dlg.exec()
        self._sync_nh_angle_display()

    # ═══════════════════════════════════════════════════════════
    # 角度配置 — 共享 UI, 切换 Gain / AR 状态
    # ═══════════════════════════════════════════════════════════

    def _on_angle_target_changed(self, index: int):
        self._active_angle_tab = index
        self._btn_gain_angle.setChecked(index == 0)
        self._btn_ar_angle.setChecked(index == 1)
        self._sync_angle_buttons()
        self._update_selected_display()

    @property
    def _cur_singles(self) -> List[float]:
        return self._ar_angle_singles if self._active_angle_tab == 1 else self._angle_singles

    @property
    def _cur_ranges(self) -> List[tuple]:
        return self._ar_angle_ranges if self._active_angle_tab == 1 else self._angle_ranges

    def _build_angle_control_widget(self):
        """构建角度设置共享控件（Gain 和 AR 共用）。"""
        self._angle_ctrl_widget = QWidget()
        layout = QVBoxLayout(self._angle_ctrl_widget)
        layout.setSpacing(4)

        # 快捷按钮行
        quick_row = QHBoxLayout()
        quick_row.addWidget(QLabel("预设:"))
        self._angle_buttons: Dict[float, QPushButton] = {}
        for a in [0, 10, 20, 30, 40, 50, 60, 70, 80, 90]:
            btn = QPushButton(f"{a}°")
            btn.setCheckable(True)
            btn.setFixedWidth(48)
            btn.clicked.connect(lambda checked, angle=a: self._toggle_quick_angle(angle))
            quick_row.addWidget(btn)
            self._angle_buttons[a] = btn
        quick_row.addStretch()
        layout.addLayout(quick_row)

        # 自定义 + 步进行
        custom_row = QHBoxLayout()
        custom_row.addWidget(QLabel("自定义:"))
        self._spin_custom = QDoubleSpinBox()
        self._spin_custom.setRange(0, 180); self._spin_custom.setValue(0)
        self._spin_custom.setSuffix("°"); self._spin_custom.setFixedWidth(80)
        custom_row.addWidget(self._spin_custom)
        btn_custom_add = QPushButton("+")
        btn_custom_add.setFixedWidth(32)
        btn_custom_add.clicked.connect(self._add_custom_angle)
        custom_row.addWidget(btn_custom_add)
        custom_row.addSpacing(12)
        custom_row.addWidget(QLabel("步进:"))
        self._spin_step_start = QDoubleSpinBox()
        self._spin_step_start.setRange(0, 180); self._spin_step_start.setValue(0)
        self._spin_step_start.setSuffix("°"); self._spin_step_start.setFixedWidth(80)
        custom_row.addWidget(self._spin_step_start)
        custom_row.addWidget(QLabel("—"))
        self._spin_step_end = QDoubleSpinBox()
        self._spin_step_end.setRange(0, 180); self._spin_step_end.setValue(90)
        self._spin_step_end.setSuffix("°"); self._spin_step_end.setFixedWidth(80)
        custom_row.addWidget(self._spin_step_end)
        custom_row.addWidget(QLabel("步长:"))
        self._spin_step_by = QDoubleSpinBox()
        self._spin_step_by.setRange(1, 90); self._spin_step_by.setValue(10)
        self._spin_step_by.setSuffix("°"); self._spin_step_by.setFixedWidth(70)
        custom_row.addWidget(self._spin_step_by)
        btn_step_gen = QPushButton("生成")
        btn_step_gen.clicked.connect(self._on_step_generate)
        custom_row.addWidget(btn_step_gen)
        custom_row.addStretch()
        layout.addLayout(custom_row)

        # 范围行
        range_row = QHBoxLayout()
        range_row.addWidget(QLabel("角度范围:"))
        self._spin_range_start = QDoubleSpinBox()
        self._spin_range_start.setRange(0, 180); self._spin_range_start.setValue(0)
        self._spin_range_start.setSuffix("°"); self._spin_range_start.setFixedWidth(80)
        range_row.addWidget(QLabel("起始:")); range_row.addWidget(self._spin_range_start)
        self._spin_range_end = QDoubleSpinBox()
        self._spin_range_end.setRange(0, 180); self._spin_range_end.setValue(90)
        self._spin_range_end.setSuffix("°"); self._spin_range_end.setFixedWidth(80)
        range_row.addWidget(QLabel("结束:")); range_row.addWidget(self._spin_range_end)
        btn_add_range = QPushButton("添加范围")
        btn_add_range.clicked.connect(self._on_add_range)
        range_row.addWidget(btn_add_range)
        range_row.addStretch()
        layout.addLayout(range_row)
        layout.addStretch()

    def _toggle_quick_angle(self, angle: float):
        singles = self._cur_singles
        if angle in singles:
            singles.remove(angle)
        else:
            singles.append(angle)
        self._sync_angle_buttons()
        self._update_selected_display()

    def _add_custom_angle(self):
        a = self._spin_custom.value()
        singles = self._cur_singles
        if a not in singles:
            singles.append(a)
            self._sync_angle_buttons()
            self._update_selected_display()

    def _on_step_generate(self):
        start = self._spin_step_start.value()
        end = self._spin_step_end.value()
        step = self._spin_step_by.value()
        if step <= 0:
            return
        lo, hi = min(start, end), max(start, end)
        singles = self._cur_singles
        a = lo
        while a <= hi + 1e-9:
            rounded = round(a, 6)
            if rounded not in singles:
                singles.append(rounded)
            a += step
        self._sync_angle_buttons()
        self._update_selected_display()

    def _on_add_range(self):
        lo = self._spin_range_start.value()
        hi = self._spin_range_end.value()
        key = (min(lo, hi), max(lo, hi))
        ranges = self._cur_ranges
        if key not in ranges:
            ranges.append(key)
            self._update_selected_display()

    def _remove_single(self, angle: float):
        singles = self._cur_singles  # 由 _active_angle_tab 决定从 Gain 还是 AR 列表删除
        if angle in singles:
            singles.remove(angle)
        self._sync_angle_buttons()
        self._update_selected_display()

    def _remove_range(self, lo: float, hi: float):
        key = (lo, hi)
        ranges = self._cur_ranges  # 由 _active_angle_tab 决定从 Gain 还是 AR 列表删除
        if key in ranges:
            ranges.remove(key)
        self._update_selected_display()

    def _sync_angle_buttons(self):
        singles = self._cur_singles
        for a, btn in self._angle_buttons.items():
            btn.setChecked(a in singles)

    @staticmethod
    def _get_checked_keys(checkbox_dict: dict) -> set:
        return {k for k, cb in checkbox_dict.items() if cb.isChecked()}

    def _update_selected_display(self):
        """刷新「已选择」区域 — FlowLayout 一行多条, 自动换行。"""
        layout = self._selected_layout
        while layout.count():
            child = layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        gain_singles = sorted(set(self._angle_singles))
        gain_ranges = sorted(set(self._angle_ranges))
        ar_singles = sorted(set(self._ar_angle_singles))
        ar_ranges = sorted(set(self._ar_angle_ranges))

        has_gain = gain_singles or gain_ranges
        has_ar = ar_singles or ar_ranges

        if not has_gain and not has_ar:
            lbl = QLabel("（未选择角度）")
            lbl.setStyleSheet("color: #888; padding: 4px;")
            layout.addWidget(lbl)
            return

        def _add_flow_group(name, singles, ranges, color):
            lbl = QLabel(f"{name}:")
            layout.addWidget(lbl)
            fw = QWidget()
            fl = FlowLayout(fw, margin=0, h_spacing=4, v_spacing=2)
            for a in singles:
                tag = QWidget()
                tl = QHBoxLayout(tag); tl.setContentsMargins(4, 2, 4, 2); tl.setSpacing(2)
                tlabel = QLabel(f"{a}°")
                tlabel.setStyleSheet(f"background:{color};color:white;border-radius:3px;padding:1px 4px;")
                tl.addWidget(tlabel)
                btn_x = QPushButton("✕"); btn_x.setFixedSize(18, 18)
                btn_x.setStyleSheet("padding:0;")
                btn_x.clicked.connect(lambda checked, angle=a: self._remove_single(angle))
                tl.addWidget(btn_x)
                fl.addWidget(tag)
            for lo, hi in ranges:
                tag = QWidget()
                tl = QHBoxLayout(tag); tl.setContentsMargins(4, 2, 4, 2); tl.setSpacing(2)
                tlabel = QLabel(f"({lo}°–{hi}°)")
                tlabel.setStyleSheet(f"background:{color};color:white;border-radius:3px;padding:1px 4px;")
                tl.addWidget(tlabel)
                btn_x = QPushButton("✕"); btn_x.setFixedSize(18, 18)
                btn_x.setStyleSheet("padding:0;")
                btn_x.clicked.connect(lambda checked, l=lo, h=hi: self._remove_range(l, h))
                tl.addWidget(btn_x)
                fl.addWidget(tag)
            layout.addWidget(fw)

        if has_gain:
            _add_flow_group("Gain", gain_singles, gain_ranges, "#3a6fb5")
        if has_ar:
            _add_flow_group("AR", ar_singles, ar_ranges, "#b53a6f")
        self._update_summary()

    def _update_summary(self):
        """刷新底部「已选参数概览」— 显示所有已配置内容的 HTML 概览。"""
        mode_names = {0: "📡 无源天线", 1: "📶 有源发射 TRP", 2: "📻 有源接收 TIS"}
        mode_str = mode_names.get(self._test_mode, "未知")

        lines = [f"<b>测试模式:</b> {mode_str}"]

        # 已选参数
        checked = sorted(set(
            cb.text() for cb in self._left_checkboxes.values() if cb.isChecked()
        ))
        if checked:
            lines.append(f"<b>计算参数 ({len(checked)}):</b> {', '.join(checked)}")
        else:
            lines.append("<b>计算参数:</b> <span style='color:#888;'>(未选择)</span>")

        # Gain 角度
        gain_singles = sorted(set(self._angle_singles))
        gain_ranges = sorted(set(self._angle_ranges))
        if gain_singles or gain_ranges:
            parts = [f"{a}°" for a in gain_singles]
            parts += [f"({lo}°–{hi}°)" for lo, hi in gain_ranges]
            lines.append(f"<b>Gain 角度 ({len(parts)}):</b> {', '.join(parts)}")
        else:
            lines.append("<b>Gain 角度:</b> <span style='color:#888;'>(未设置)</span>")

        # AR 角度
        ar_singles = sorted(set(self._ar_angle_singles))
        ar_ranges = sorted(set(self._ar_angle_ranges))
        if ar_singles or ar_ranges:
            parts = [f"{a}°" for a in ar_singles]
            parts += [f"({lo}°–{hi}°)" for lo, hi in ar_ranges]
            lines.append(f"<b>AR 角度 ({len(parts)}):</b> {', '.join(parts)}")
        else:
            lines.append("<b>AR 角度:</b> <span style='color:#888;'>(未设置)</span>")

        # 算法选项
        algo_parts = []
        if self._check_extrap.isChecked():
            algo_parts.append("Theta 外推到 180°")
        if self._check_robust.isChecked():
            algo_parts.append("Robust peak detection")
        if not self._cmb_ar_output.currentData():
            algo_parts.append("AR 输出线性")
        algo_str = ", ".join(algo_parts) if algo_parts else "<span style='color:#888;'>(默认)</span>"
        lines.append(f"<b>算法选项:</b> {algo_str}")

        # 频点
        freq_src = self._cmb_freq_src.currentText()
        trim = f"去除: 前 {self._spin_trim_start.value()} / 后 {self._spin_trim_end.value()}"
        lines.append(f"<b>频点:</b> {freq_src} | {trim}")

        self._summary_label.setText("<br>".join(lines))

    # ═══════════════════════════════════════════════════════════
    # 加载 / 保存状态
    # ═══════════════════════════════════════════════════════════

    def _load_state(self):
        mw = self._mw
        # 恢复各模式的独立状态（跨对话框生命周期持久化）
        if hasattr(mw, '_mode_states') and mw._mode_states:
            self._mode_states = [dict(s) for s in mw._mode_states]
        # 测试模式
        if hasattr(mw, '_test_mode'):
            self._test_mode = mw._test_mode
            self._cmb_test_mode.blockSignals(True)
            self._cmb_test_mode.setCurrentIndex(mw._test_mode)
            self._cmb_test_mode.blockSignals(False)
        # 角度 — Gain
        if hasattr(mw, '_lag_config'):
            self._angle_singles = list(mw._lag_config.single_angles)
            self._angle_ranges = list(mw._lag_config.ranges)
        # 角度 — AR
        if hasattr(mw, '_ar_lag_config'):
            self._ar_angle_singles = list(mw._ar_lag_config.single_angles)
            self._ar_angle_ranges = list(mw._ar_lag_config.ranges)
        self._sync_angle_buttons()
        self._update_selected_display()
        # 频点
        if hasattr(mw, '_cmb_freq_source') and mw._cmb_freq_source:
            data = mw._cmb_freq_source.currentData()
            idx = self._cmb_freq_src.findData(data)
            if idx >= 0:
                self._cmb_freq_src.setCurrentIndex(idx)
        if hasattr(mw, '_spin_trim_start'):
            self._spin_trim_start.setValue(mw._spin_trim_start.value())
            self._spin_trim_end.setValue(mw._spin_trim_end.value())
        # 算法
        if hasattr(mw, '_cmb_extrapolate'):
            self._check_extrap.setChecked(mw._cmb_extrapolate.currentData() is not None)
        if hasattr(mw, '_check_robust_peak'):
            self._check_robust.setChecked(mw._check_robust_peak.isChecked())
        # AR dB
        self._cmb_ar_output.setCurrentIndex(0 if getattr(mw, '_ar_output_db', True) else 1)
        # NH 自定义角度
        if hasattr(mw, '_nh_custom_angles'):
            self._nh_custom_angles = list(mw._nh_custom_angles)
        self._sync_nh_angle_display()

    def _on_accept(self):
        # 保存当前模式状态
        self._save_current_mode_state()
        mw = self._mw
        # 保存所有三个模式的独立状态到 MainWindow
        mw._mode_states = [dict(s) for s in self._mode_states]
        # 测试模式
        mw._test_mode = self._cmb_test_mode.currentData() if hasattr(self, '_cmb_test_mode') else 0
        # 同步 Gain 角度
        if hasattr(mw, '_lag_config'):
            mw._lag_config.clear()
            for a in sorted(set(self._angle_singles)):
                mw._lag_config.add_single(a)
            for lo, hi in sorted(set(self._angle_ranges)):
                mw._lag_config.add_range(lo, hi)
            mw._sync_quick_buttons()
            mw._update_lag_display()
        # 同步 AR 角度
        if not hasattr(mw, '_ar_lag_config'):
            from src.lag_config import LagConfig
            mw._ar_lag_config = LagConfig()
        mw._ar_lag_config.clear()
        for a in sorted(set(self._ar_angle_singles)):
            mw._ar_lag_config.add_single(a)
        for lo, hi in sorted(set(self._ar_angle_ranges)):
            mw._ar_lag_config.add_range(lo, hi)
        # 保存额外参数 & 必需参数
        required = set(k for k, cb in self._left_checkboxes.items() if cb.isChecked())
        extra = set(k for k, cb in self._right_checkboxes.items() if cb.isChecked())
        mw._required_params = required
        mw._extra_params = extra
        # NH 自定义角度
        mw._nh_custom_angles = list(self._nh_custom_angles)
        # 频点
        if hasattr(mw, '_cmb_freq_source') and mw._cmb_freq_source:
            data = self._cmb_freq_src.currentData()
            idx = mw._cmb_freq_source.findData(data)
            if idx >= 0:
                mw._cmb_freq_source.setCurrentIndex(idx)
        if hasattr(mw, '_spin_trim_start'):
            mw._spin_trim_start.setValue(self._spin_trim_start.value())
            mw._spin_trim_end.setValue(self._spin_trim_end.value())
        # 算法
        if hasattr(mw, '_cmb_extrapolate'):
            if not self._check_extrap.isChecked():
                mw._cmb_extrapolate.setCurrentIndex(0)  # 不外推
        if hasattr(mw, '_check_robust_peak'):
            mw._check_robust_peak.setChecked(self._check_robust.isChecked())
        # AR 输出格式
        mw._ar_output_db = self._cmb_ar_output.currentData()
        self.accept()

    # ── 公共接口（外部调用） ──

    def set_template_params(self, params: set):
        """设置模板自动识别的参数集合 + 自动匹配测试模式和角度配置。"""
        self._template_params = set(params)
        trp_keys = {k for grp in self._TRP_PARAMS for k, _ in grp[1]}
        tis_keys = {k for grp in self._TIS_PARAMS for k, _ in grp[1]}
        if params & tis_keys:
            mode = 2
        elif params & trp_keys:
            mode = 1
        else:
            mode = 0
        if mode != self._active_tab:
            self._cmb_test_mode.blockSignals(True)
            self._cmb_test_mode.setCurrentIndex(mode)
            self._cmb_test_mode.blockSignals(False)
            self._on_mode_changed(mode)
        else:
            # 模式相同（如无源→无源），虽不切换tab但需要刷新参数列
            self._rebuild_param_columns()

    def set_angle_config(self, cfg: "LagConfig", is_ar: bool = False):
        """从模板自动配置角度 (Gain 或 AR)。"""
        if is_ar:
            self._ar_angle_singles = list(cfg.single_angles)
            self._ar_angle_ranges = list(cfg.ranges)
        else:
            self._angle_singles = list(cfg.single_angles)
            self._angle_ranges = list(cfg.ranges)
        self._sync_angle_buttons()
        self._update_selected_display()
        self._rebuild_param_columns()


# ═══════════════════════════════════════════════════════════════
# 图形配置对话框
# ═══════════════════════════════════════════════════════════════

class PlotConfigDialog(QDialog):
    """图形配置: 双列模式 — 报告需要 + 额外(full_report) + 视角参数"""

    def __init__(self, parent: "MainWindow"):
        super().__init__(parent)
        self._mw = parent
        self.setWindowTitle("图形配置")
        self.resize(780, 620)

        # ── 状态 ──
        self._chart_required: Dict[str, QCheckBox] = {}   # 左列: 报告需要
        self._chart_extra: Dict[str, QCheckBox] = {}      # 右列: 额外(full_report)
        self._collapse_map: Dict[str, dict] = {}          # 折叠状态

        self._setup_ui()
        self._load_state()
        auto_size_dialog(self, 720, 600)

    def _setup_ui(self):
        from src.chart_config import ChartConfig
        labels = ChartConfig.chart_labels()
        categories = ChartConfig.chart_categories()

        # ── 子角度选择状态 ──
        self._gain_angles: List[float] = []      # 左列 Gain 角度
        self._gain_ranges: List[tuple] = []      # 左列 Gain 范围
        self._ar_angles: List[float] = []        # 左列 AR 角度
        self._ar_ranges: List[tuple] = []        # 左列 AR 范围
        self._gain_angles_x: List[float] = []    # 右列 Gain 角度
        self._gain_ranges_x: List[tuple] = []    # 右列 Gain 范围
        self._ar_angles_x: List[float] = []      # 右列 AR 角度
        self._ar_ranges_x: List[tuple] = []      # 右列 AR 范围

        grp_list = []

        # ── 图形分类 + 双列 ──
        for cat_name, keys in categories.items():
            grp = QGroupBox(cat_name)
            grp.setCheckable(True)
            grp.setChecked(True)
            grp.setStyleSheet("""
                QGroupBox { font-weight: bold; padding-top: 16px; }
                QGroupBox::indicator { width: 14px; height: 14px; margin-right: 4px; }
                QGroupBox::indicator:unchecked { image: none; }
                QGroupBox::indicator:unchecked:hover { image: none; }
            """)
            grp.setCursor(Qt.PointingHandCursor)
            outer_layout = QVBoxLayout(grp)
            outer_layout.setSpacing(4)

            # 折叠内容容器 — 方便显示/隐藏
            self._collapse_map[cat_name] = {"grp": grp, "hidden": False}
            content_widget = QWidget()
            content_layout = QVBoxLayout(content_widget)
            content_layout.setContentsMargins(0, 0, 0, 0)
            content_layout.setSpacing(4)
            outer_layout.addWidget(content_widget)

            # 全选 / 取消全选 按钮行
            btn_row = QHBoxLayout()
            btn_row.setSpacing(6)
            btn_select_all = QPushButton(self.tr("全选"))
            btn_select_all.setFixedWidth(60)
            btn_deselect_all = QPushButton(self.tr("取消全选"))
            btn_deselect_all.setFixedWidth(72)
            # 连接信号：控制本分类左右两列的所有 checkbox
            btn_select_all.clicked.connect(
                lambda checked, ks=keys: [self._chart_required[k].setChecked(True) for k in ks if k in self._chart_required] or
                                         [self._chart_extra[k].setChecked(True) for k in ks if k in self._chart_extra])
            btn_deselect_all.clicked.connect(
                lambda checked, ks=keys: [self._chart_required[k].setChecked(False) for k in ks if k in self._chart_required] or
                                         [self._chart_extra[k].setChecked(False) for k in ks if k in self._chart_extra])
            btn_row.addWidget(btn_select_all)
            btn_row.addWidget(btn_deselect_all)
            btn_row.addStretch()
            content_layout.addLayout(btn_row)

            row_layout = QHBoxLayout()
            row_layout.setSpacing(8)

            # 左列: 报告需要
            left_box = QGroupBox("报告需要")
            left_layout = QVBoxLayout(left_box)
            left_layout.setSpacing(3)
            for key in keys:
                row = QHBoxLayout()
                cb = QCheckBox(labels.get(key, key))
                row.addWidget(cb)
                self._chart_required[key] = cb
                # Gain / AR 曲线 → 添加角度选择按钮
                if key in ("chart_gain_freq", "chart_ar_freq"):
                    btn = QPushButton("⚙ 角度...")
                    btn.setFixedWidth(80)
                    is_ar = (key == "chart_ar_freq")
                    btn.clicked.connect(lambda checked, k=key: self._show_chart_angle_popup(k, is_left=True))
                    row.addWidget(btn)
                elif key == "chart_lag_freq":
                    btn = QPushButton("⚙ 角度...")
                    btn.setFixedWidth(80)
                    btn.clicked.connect(lambda checked: self._show_chart_angle_popup("chart_lag_freq", is_left=True))
                    row.addWidget(btn)
                row.addStretch()
                left_layout.addLayout(row)
            left_layout.addStretch()
            row_layout.addWidget(left_box, 1)

            # 右列: 额外报告
            right_box = QGroupBox("额外报告")
            right_layout = QVBoxLayout(right_box)
            right_layout.setSpacing(3)
            for key in keys:
                row = QHBoxLayout()
                cb = QCheckBox(labels.get(key, key))
                row.addWidget(cb)
                self._chart_extra[key] = cb
                if key in ("chart_gain_freq", "chart_ar_freq"):
                    btn = QPushButton("⚙ 角度...")
                    btn.setFixedWidth(80)
                    btn.clicked.connect(lambda checked, k=key: self._show_chart_angle_popup(k, is_left=False))
                    row.addWidget(btn)
                elif key == "chart_lag_freq":
                    btn = QPushButton("⚙ 角度...")
                    btn.setFixedWidth(80)
                    btn.clicked.connect(lambda checked: self._show_chart_angle_popup("chart_lag_freq", is_left=False))
                    row.addWidget(btn)
                row.addStretch()
                right_layout.addLayout(row)
            right_layout.addStretch()
            row_layout.addWidget(right_box, 1)

            content_layout.addLayout(row_layout)
            grp_list.append(grp)

            # 折叠/展开切换
            def make_toggle(g=grp, cw=content_widget, name=cat_name):
                def toggle(checked):
                    cw.setVisible(checked)
                    self._collapse_map[name]["hidden"] = not checked
                return toggle
            grp.toggled.connect(make_toggle(grp, content_widget, cat_name))

        # ── 视角参数 ──
        view_grp = QGroupBox("视角参数")
        view_layout = QHBoxLayout(view_grp)
        view_layout.addWidget(QLabel("仰角:"))
        self._spin_elev = QDoubleSpinBox()
        self._spin_elev.setRange(-90, 90); self._spin_elev.setValue(30)
        self._spin_elev.setSuffix("°"); self._spin_elev.setFixedWidth(80)
        view_layout.addWidget(self._spin_elev)
        view_layout.addWidget(QLabel("方位角:"))
        self._spin_azim = QDoubleSpinBox()
        self._spin_azim.setRange(-180, 180); self._spin_azim.setValue(-60)
        self._spin_azim.setSuffix("°"); self._spin_azim.setFixedWidth(80)
        view_layout.addWidget(self._spin_azim)
        view_layout.addWidget(QLabel("DPI:"))
        self._spin_dpi = QSpinBox()
        self._spin_dpi.setRange(72, 300); self._spin_dpi.setValue(150)
        self._spin_dpi.setFixedWidth(70)
        view_layout.addWidget(self._spin_dpi)
        view_layout.addWidget(QLabel("采样精度:"))
        self._spin_step = QSpinBox()
        self._spin_step.setRange(1, 30)
        self._spin_step.setValue(5)
        self._spin_step.setSuffix("°")
        self._spin_step.setFixedWidth(70)
        self._spin_step.setToolTip(
            "3D 图形采样步进 (1°–30°):\n"
            "  1°=最精细(~40K点/频点,慢)\n"
            "  5°=标准(~1.7K点/频点)\n"
            "  30°=最快(~150点/频点)\n"
            "值越小图形越精细但计算越慢。\n"
            "此值将作为图形展示窗体的初始精度，\n"
            "可在展示窗体中通过 ⚙ 图形设置 独立调整。"
        )
        view_layout.addWidget(self._spin_step)
        view_layout.addStretch()
        grp_list.append(view_grp)

        # ── 输出方式 ──
        out_grp = QGroupBox("输出方式")
        out_layout = QHBoxLayout(out_grp)
        self._check_embed = QCheckBox("嵌入 Excel")
        self._check_embed.setChecked(True)
        self._check_png = QCheckBox("保存 PNG 文件夹")
        out_layout.addWidget(self._check_embed)
        out_layout.addWidget(self._check_png)
        out_layout.addStretch()
        grp_list.append(out_grp)

        # ── 按钮 ──
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)

        wrap_in_scroll(self, grp_list, btns)

    def _load_state(self):
        mw = self._mw
        # 视角
        if hasattr(mw, 'ui'):
            if hasattr(mw.ui, 'spinElev'):
                self._spin_elev.setValue(mw.ui.spinElev.value())
                self._spin_azim.setValue(mw.ui.spinAzim.value())
                self._spin_dpi.setValue(mw.ui.spinDpi.value())
                self._check_embed.setChecked(mw.ui.checkEmbedExcel.isChecked())
                self._check_png.setChecked(mw.ui.checkSavePng.isChecked())
        # 采样精度: 优先从 chart_config 读取
        step_deg = 5
        if hasattr(mw, '_chart_config_required') and mw._chart_config_required is not None:
            step_deg = int(getattr(mw._chart_config_required, 'step_deg', 5))
        self._spin_step.setValue(max(1, min(30, step_deg)))
        # 图表配置
        if hasattr(mw, '_chart_config_required') and mw._chart_config_required is not None:
            req = mw._chart_config_required
            for key, cb in self._chart_required.items():
                val = getattr(req, key, False)
                cb.setChecked(val)
            self._gain_angles = list(req.gain_chart_angles)
            self._gain_ranges = list(req.gain_chart_ranges)
            self._ar_angles = list(req.ar_chart_angles)
            self._ar_ranges = list(req.ar_chart_ranges)
        if hasattr(mw, '_chart_config_extra') and mw._chart_config_extra is not None:
            xtr = mw._chart_config_extra
            for key, cb in self._chart_extra.items():
                val = getattr(xtr, key, False)
                cb.setChecked(val)
            self._gain_angles_x = list(xtr.gain_chart_angles)
            self._gain_ranges_x = list(xtr.gain_chart_ranges)
            self._ar_angles_x = list(xtr.ar_chart_angles)
            self._ar_ranges_x = list(xtr.ar_chart_ranges)

    def _parse_step_deg(self) -> float:
        """从采样精度 spin box 获取步进值。"""
        return float(max(1, min(30, self._spin_step.value())))

    def _on_accept(self):
        from src.chart_config import ChartConfig
        mw = self._mw

        step_deg = self._parse_step_deg()

        # 构建 ChartConfig 对象
        required = ChartConfig()
        extra = ChartConfig()
        for key in ChartConfig.all_chart_keys():
            setattr(required, key, self._chart_required.get(key, QCheckBox()).isChecked())
            setattr(extra, key, self._chart_extra.get(key, QCheckBox()).isChecked())

        required.elev = self._spin_elev.value()
        required.azim = self._spin_azim.value()
        required.dpi = self._spin_dpi.value()
        required.step_deg = step_deg
        required.embed_in_excel = self._check_embed.isChecked()
        extra.elev = required.elev
        extra.azim = required.azim
        extra.dpi = required.dpi
        extra.step_deg = step_deg
        extra.embed_in_excel = False

        # 子角度选择
        required.gain_chart_angles = list(self._gain_angles)
        required.gain_chart_ranges = list(self._gain_ranges)
        required.ar_chart_angles = list(self._ar_angles)
        required.ar_chart_ranges = list(self._ar_ranges)
        extra.gain_chart_angles = list(self._gain_angles_x)
        extra.gain_chart_ranges = list(self._gain_ranges_x)
        extra.ar_chart_angles = list(self._ar_angles_x)
        extra.ar_chart_ranges = list(self._ar_ranges_x)

        mw._chart_config_required = required
        mw._chart_config_extra = extra

        # 同步视角到 UI 控件
        if hasattr(mw, 'ui'):
            mw.ui.spinElev.setValue(int(self._spin_elev.value()))
            mw.ui.spinAzim.setValue(int(self._spin_azim.value()))
            mw.ui.spinDpi.setValue(self._spin_dpi.value())
            mw.ui.checkEmbedExcel.setChecked(self._check_embed.isChecked())
            mw.ui.checkSavePng.setChecked(self._check_png.isChecked())

        self.accept()

    def _show_chart_angle_popup(self, chart_key: str, is_left: bool = True):
        """弹出角度选择窗口 — 与 CalcParamsDialog 的角度弹出窗口一致。"""
        dlg = QDialog(self)
        is_ar = (chart_key == "chart_ar_freq")
        if is_left:
            singles = self._ar_angles if is_ar else self._gain_angles
            ranges = self._ar_ranges if is_ar else self._gain_ranges
        else:
            singles = self._ar_angles_x if is_ar else self._gain_angles_x
            ranges = self._ar_ranges_x if is_ar else self._gain_ranges_x

        label_text = "AR" if is_ar else "Gain"
        dlg.setWindowTitle(f"选择 {label_text} 曲线角度 — 频点曲线")
        dlg.setMinimumSize(500, 420)

        # 深拷贝 — Cancel 时恢复原值
        import copy
        _singles = copy.deepcopy(singles)
        _ranges = copy.deepcopy(ranges)

        layout = QVBoxLayout(dlg)

        def _refresh_display():
            while _display_layout.count():
                item = _display_layout.takeAt(0)
                if item.widget(): item.widget().deleteLater()
            if _singles or _ranges:
                scroll = QScrollArea()
                scroll.setWidgetResizable(True)
                dw = QWidget()
                from ui.layout_utils import FlowLayout
                fl = FlowLayout(dw, margin=4, h_spacing=6, v_spacing=4)
                for a in sorted(set(_singles)):
                    tag = QWidget()
                    tl = QHBoxLayout(tag); tl.setContentsMargins(2, 1, 2, 1); tl.setSpacing(2)
                    tl.addWidget(QLabel(f"{a}°"))
                    btn_del = QPushButton("✕")
                    btn_del.setFixedSize(20, 20); btn_del.setStyleSheet("padding:0;")
                    btn_del.clicked.connect(lambda checked, v=a: (_singles.remove(v), _refresh_display()))
                    tl.addWidget(btn_del)
                    fl.addWidget(tag)
                for lo, hi in sorted(set(_ranges), key=lambda x: (x[0], x[1])):
                    tag = QWidget()
                    tl = QHBoxLayout(tag); tl.setContentsMargins(2, 1, 2, 1); tl.setSpacing(2)
                    tl.addWidget(QLabel(f"{lo}°~{hi}°"))
                    btn_del = QPushButton("✕")
                    btn_del.setFixedSize(20, 20); btn_del.setStyleSheet("padding:0;")
                    btn_del.clicked.connect(lambda checked, lo=lo, hi=hi: (_ranges.remove((lo, hi)), _refresh_display()))
                    tl.addWidget(btn_del)
                    fl.addWidget(tag)
                scroll.setWidget(dw)
                _display_layout.addWidget(scroll)
                btn_clear = QPushButton("🗑 清空全部")
                btn_clear.clicked.connect(lambda: (_singles.clear(), _ranges.clear(), _refresh_display()))
                _display_layout.addWidget(btn_clear)
            else:
                _display_layout.addWidget(QLabel("  (暂无选择 — 将自动使用默认值)"))

        _display_grp = QGroupBox(f"已选: {len(_singles)} 个单角度, {len(_ranges)} 个范围")
        _display_layout = QVBoxLayout(_display_grp)
        _refresh_display()

        splitter = QSplitter(Qt.Vertical)
        bottom_ctls = QWidget()
        bottom_layout = QVBoxLayout(bottom_ctls)
        bottom_layout.setContentsMargins(0, 0, 0, 0)

        # 快捷预设
        quick_grp = QGroupBox(f"快捷预设 — {label_text}")
        quick_layout = QHBoxLayout(quick_grp)
        for a in [0, 30, 45, 60, 90]:
            btn = QPushButton(f"{a}°")
            btn.clicked.connect(lambda checked, v=a: (_singles.append(v) if v not in _singles else None, _refresh_display()))
            quick_layout.addWidget(btn)
        quick_layout.addStretch()
        bottom_layout.addWidget(quick_grp)

        # 自定义
        cust_grp = QGroupBox("自定义")
        cust_layout = QHBoxLayout(cust_grp)
        spin_custom = QDoubleSpinBox(); spin_custom.setRange(0, 180); spin_custom.setValue(45)
        btn_add_custom = QPushButton("+ 添加")
        btn_add_custom.clicked.connect(lambda: (_singles.append(spin_custom.value()) if spin_custom.value() not in _singles else None, _refresh_display()))
        cust_layout.addWidget(QLabel("角度:")); cust_layout.addWidget(spin_custom)
        cust_layout.addWidget(btn_add_custom); cust_layout.addStretch()
        bottom_layout.addWidget(cust_grp)

        # 步进
        step_grp = QGroupBox("步进批量生成")
        step_layout = QHBoxLayout(step_grp)
        spin_start = QDoubleSpinBox(); spin_start.setRange(0, 180); spin_start.setValue(0)
        spin_end = QDoubleSpinBox(); spin_end.setRange(0, 180); spin_end.setValue(90)
        spin_step = QDoubleSpinBox(); spin_step.setRange(1, 90); spin_step.setValue(10)
        import numpy as np
        btn_gen = QPushButton("生成")
        btn_gen.clicked.connect(lambda: (
            [_singles.append(round(float(a), 6)) for a in np.linspace(spin_start.value(), spin_end.value(), int((spin_end.value()-spin_start.value())/spin_step.value()+1))
             if round(float(a), 6) not in _singles],
            _refresh_display()
        ))
        step_layout.addWidget(QLabel("起:")); step_layout.addWidget(spin_start)
        step_layout.addWidget(QLabel("止:")); step_layout.addWidget(spin_end)
        step_layout.addWidget(QLabel("步:")); step_layout.addWidget(spin_step)
        step_layout.addWidget(btn_gen)
        bottom_layout.addWidget(step_grp)

        # 范围
        range_grp = QGroupBox("角度范围")
        range_layout = QHBoxLayout(range_grp)
        spin_rs = QDoubleSpinBox(); spin_rs.setRange(0, 180); spin_rs.setValue(0)
        spin_re = QDoubleSpinBox(); spin_re.setRange(0, 180); spin_re.setValue(90)
        def _add_range():
            lo, hi = spin_rs.value(), spin_re.value()
            key = (min(lo, hi), max(lo, hi))
            if key not in _ranges:
                _ranges.append(key)
                _refresh_display()
        btn_add_range = QPushButton("添加范围")
        btn_add_range.clicked.connect(_add_range)
        range_layout.addWidget(QLabel("起:")); range_layout.addWidget(spin_rs)
        range_layout.addWidget(QLabel("止:")); range_layout.addWidget(spin_re)
        range_layout.addWidget(btn_add_range); range_layout.addStretch()
        bottom_layout.addWidget(range_grp)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(lambda: (
            singles.clear(), singles.extend(sorted(set(_singles))),
            ranges.clear(), ranges.extend(sorted(set(_ranges), key=lambda x: (x[0], x[1]))),
            dlg.accept()
        ))
        btns.rejected.connect(dlg.reject)
        bottom_layout.addWidget(btns)

        splitter.addWidget(_display_grp)
        splitter.addWidget(bottom_ctls)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 0)
        layout.addWidget(splitter)
        dlg.exec()


# ═══════════════════════════════════════════════════════════════
# 帮助搜索对话框
# ═══════════════════════════════════════════════════════════════

class HelpDialog(QDialog):
    """帮助搜索: BM25 全文搜索 + 可选语义搜索 + 可选 LLM RAG 问答"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("帮助 — 天线参数后处理工具")
        self.setMinimumSize(700, 550)
        self.resize(780, 620)

        from src.help_engine import HelpEngine, RAGSettings
        self._engine = HelpEngine()
        self._rag_settings = RAGSettings()
        self._latest_results = []

        self._load_rag_settings()
        self._setup_ui()
        auto_size_dialog(self, 700, 550)
        self._lbl_status.setText(f"帮助引擎已就绪 — {self._engine.chunk_count} 个章节")

    def _setup_ui(self):
        content_widgets = []

        # ── 搜索栏 ──
        search_widget = QWidget()
        search_row = QHBoxLayout(search_widget)
        self._edit_query = QLineEdit()
        self._edit_query.setPlaceholderText("输入问题或关键词，如: LAG怎么配置、模板列头格式...")
        self._edit_query.returnPressed.connect(self._on_search)
        search_row.addWidget(self._edit_query, 1)

        btn_search = QPushButton("🔍 搜索")
        btn_search.clicked.connect(self._on_search)
        search_row.addWidget(btn_search)
        content_widgets.append(search_widget)

        # ── 主体: 结果列表 (左) + RAG 回答 (右) ──
        main_widget = QWidget()
        main_split = QHBoxLayout(main_widget)
        main_split.setSpacing(8)

        # 左: 搜索结果列表
        left_panel = QVBoxLayout()
        left_panel.addWidget(QLabel("<b>搜索结果</b>"))

        self._result_list = QListWidget()
        self._result_list.setAlternatingRowColors(True)
        self._result_list.itemClicked.connect(self._on_result_clicked)
        left_panel.addWidget(self._result_list, 1)

        opt_row = QHBoxLayout()
        self._check_semantic = QCheckBox("语义搜索")
        self._check_semantic.setToolTip("启用语义搜索（需要安装 sentence-transformers + faiss）")
        self._check_semantic.setChecked(True)
        opt_row.addWidget(self._check_semantic)
        opt_row.addStretch()
        opt_row.addWidget(QLabel(f"共 {self._engine.chunk_count} 章节"))
        left_panel.addLayout(opt_row)

        # 把 left_panel 装进 QWidget 放入 main_split
        left_container = QWidget()
        left_container.setLayout(left_panel)
        main_split.addWidget(left_container, 1)

        # 右: RAG 回答
        right_panel = QVBoxLayout()
        right_header = QHBoxLayout()
        right_header.addWidget(QLabel("<b>AI 回答</b>"))
        self._btn_ask = QPushButton("🤖 提问 AI")
        self._btn_ask.clicked.connect(self._on_ask)
        self._btn_ask.setToolTip("使用 LLM 对检索到的文档生成回答")
        right_header.addWidget(self._btn_ask)
        self._btn_settings = QPushButton("⚙")
        self._btn_settings.setFixedWidth(32)
        self._btn_settings.setToolTip("LLM API 设置")
        self._btn_settings.clicked.connect(self._on_rag_settings)
        right_header.addWidget(self._btn_settings)
        right_panel.addLayout(right_header)

        self._rag_answer = QPlainTextEdit()
        self._rag_answer.setReadOnly(True)
        self._rag_answer.setPlaceholderText(
            "点击「🤖 提问 AI」使用 LLM 生成回答\n"
            "首次使用请在 ⚙ 中配置 API Key")
        right_panel.addWidget(self._rag_answer, 1)

        self._lbl_sources = QLabel("")
        self._lbl_sources.setStyleSheet("color: #888;")
        self._lbl_sources.setWordWrap(True)
        right_panel.addWidget(self._lbl_sources)

        right_container = QWidget()
        right_container.setLayout(right_panel)
        main_split.addWidget(right_container, 2)

        content_widgets.append(main_widget)

        # ── 底部 ──
        bottom_widget = QWidget()
        bottom_row = QHBoxLayout(bottom_widget)
        self._lbl_status = QLabel("")
        self._lbl_status.setStyleSheet("color: #666;")
        bottom_row.addWidget(self._lbl_status)
        bottom_row.addStretch()
        btn_open = QPushButton("📖 在浏览器中打开完整手册")
        btn_open.clicked.connect(self._on_open_browser)
        bottom_row.addWidget(btn_open)
        content_widgets.append(bottom_widget)

        # ── 按钮 ──
        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(self.reject)

        wrap_in_scroll(self, content_widgets, btns)

    def _on_search(self):
        query = self._edit_query.text().strip()
        if not query:
            return
        self._result_list.clear()
        self._lbl_status.setText("搜索中...")
        use_sem = self._check_semantic.isChecked()
        results = self._engine.search(query, top_k=8, use_semantic=use_sem)
        self._latest_results = results
        for r in results:
            label = f"[{r['source']}] {r['title']}  (score: {r['score']:.2f})"
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, r)
            self._result_list.addItem(item)
        n = len(results)
        srcs = set(r["source"] for r in results)
        self._lbl_status.setText(f"找到 {n} 个结果（来源: {', '.join(srcs)}）")

    def _on_result_clicked(self, item):
        r = item.data(Qt.UserRole)
        if r:
            detail = f"## {r['title']}\n\n{r['content']}"
            self._rag_answer.setPlainText(detail)

    def _on_ask(self):
        query = self._edit_query.text().strip()
        if not query:
            return
        self._rag_answer.setPlainText("正在请求 AI...")
        self._lbl_sources.setText("")
        QApplication.processEvents()
        result = self._engine.ask(query)
        if result.get("error"):
            self._rag_answer.setPlainText(f"❌ 错误: {result['error']}")
        else:
            self._rag_answer.setPlainText(result["answer"])
            if result["sources"]:
                self._lbl_sources.setText("📚 参考: " + " | ".join(result["sources"]))

    def _on_rag_settings(self):
        dlg = RAGSettingsDialog(self._rag_settings, self)
        if dlg.exec():
            self._rag_settings = dlg.settings
            self._engine.set_rag_settings(self._rag_settings)
            self._save_rag_settings()
            self._rag_answer.setPlaceholderText(
                f"✅ RAG 已配置 (Model: {self._rag_settings.model})")

    def _on_open_browser(self):
        import webbrowser, sys
        from pathlib import Path
        # 查找 USER_GUIDE.html（支持 PyInstaller 打包）
        if getattr(sys, 'frozen', False):
            base = sys._MEIPASS
            guide = os.path.join(base, 'USER_GUIDE.html')
        else:
            guide = str(Path(__file__).parent.parent / "USER_GUIDE.html")
        if os.path.exists(guide):
            webbrowser.open(f"file://{Path(guide).absolute()}")
        else:
            QMessageBox.warning(self, "提示", "帮助文件未找到。请确认 USER_GUIDE.html 存在。")

    def _load_rag_settings(self):
        try:
            from PySide6.QtCore import QSettings
            from src.license import decrypt_secret
            s = QSettings("AntennaPP", "AntennaPostProcessor")
            self._rag_settings.enabled = s.value("rag/enabled", False, type=bool)
            self._rag_settings.api_base = s.value("rag/api_base", self._rag_settings.api_base)
            self._rag_settings.api_key = decrypt_secret(s.value("rag/api_key", ""))
            self._rag_settings.model = s.value("rag/model", self._rag_settings.model)
            self._rag_settings.use_local = s.value("rag/use_local", False, type=bool)
            self._rag_settings.local_model = s.value("rag/local_model", "qwen2.5:7b")
            self._rag_settings.local_endpoint = s.value("rag/local_endpoint", "http://localhost:11434")
            self._engine.set_rag_settings(self._rag_settings)
        except Exception:
            pass

    def _save_rag_settings(self):
        try:
            from PySide6.QtCore import QSettings
            from src.license import encrypt_secret
            s = QSettings("AntennaPP", "AntennaPostProcessor")
            s.setValue("rag/enabled", self._rag_settings.enabled)
            s.setValue("rag/api_base", self._rag_settings.api_base)
            s.setValue("rag/api_key", encrypt_secret(self._rag_settings.api_key))
            s.setValue("rag/model", self._rag_settings.model)
            s.setValue("rag/use_local", self._rag_settings.use_local)
            s.setValue("rag/local_model", self._rag_settings.local_model)
            s.setValue("rag/local_endpoint", self._rag_settings.local_endpoint)
        except Exception:
            pass


class RAGSettingsDialog(QDialog):
    """LLM RAG API 设置对话框。"""

    def __init__(self, settings, parent=None):
        super().__init__(parent)
        self.setWindowTitle("LLM API 设置")
        self.setMinimumSize(450, 320)
        self.settings = settings
        self._setup_ui()

    def _setup_ui(self):
        from src.help_engine import RAGSettings
        layout = QVBoxLayout(self)

        fl = QFormLayout()
        fl.setSpacing(8)

        self._check_enable = QCheckBox("启用 RAG AI 问答")
        self._check_enable.setChecked(self.settings.enabled)
        fl.addRow("", self._check_enable)

        # 本地 Ollama 模式
        self._check_local = QCheckBox("使用本地 Ollama 模型（免费、离线、无需 API Key）")
        self._check_local.setChecked(self.settings.use_local)
        self._check_local.toggled.connect(self._on_local_toggled)
        fl.addRow("", self._check_local)

        self._edit_local_endpoint = QLineEdit(self.settings.local_endpoint)
        self._edit_local_endpoint.setPlaceholderText("http://localhost:11434")
        fl.addRow("Ollama 地址:", self._edit_local_endpoint)

        self._edit_local_model = QLineEdit(self.settings.local_model)
        self._edit_local_model.setPlaceholderText("qwen2.5:7b / llama3.2 / mistral")
        fl.addRow("Ollama 模型:", self._edit_local_model)

        # 云 API 模式
        self._cloud_widgets = []
        self._edit_base = QLineEdit(self.settings.api_base)
        self._edit_base.setPlaceholderText("https://api.anthropic.com/v1/messages")
        w1 = fl.addRow("API Base URL:", self._edit_base); self._cloud_widgets.append(w1)

        self._edit_key = QLineEdit(self.settings.api_key)
        self._edit_key.setEchoMode(QLineEdit.Password)
        self._edit_key.setPlaceholderText("sk-ant-... 或 sk-...")
        w2 = fl.addRow("API Key:", self._edit_key); self._cloud_widgets.append(w2)

        self._cmb_model = QComboBox()
        self._cmb_model.setEditable(True)
        self._cmb_model.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_model.lineEdit().setPlaceholderText("搜索...")
        self._cmb_model.setEditable(True)
        for m in ["claude-sonnet-4-6", "claude-opus-4-8", "gpt-4o", "gpt-4o-mini", "deepseek-chat"]:
            self._cmb_model.addItem(m)
        if self.settings.model:
            idx = self._cmb_model.findText(self.settings.model)
            if idx >= 0: self._cmb_model.setCurrentIndex(idx)
            else: self._cmb_model.setCurrentText(self.settings.model)
        w3 = fl.addRow("Model:", self._cmb_model); self._cloud_widgets.append(w3)

        self._on_local_toggled(self.settings.use_local)

        layout.addLayout(fl)

        info = QLabel(
            "本地 Ollama: 免费、离线、无需 API Key。需要先安装 Ollama 并拉取模型。\n"
            "云 API: 支持 Anthropic / OpenAI 兼容接口。API Key 存储在本地 QSettings。")
        info.setStyleSheet("color: #888;")
        layout.addWidget(info)

        layout.addStretch()

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _on_local_toggled(self, checked):
        for row in getattr(self, '_cloud_widgets', []):
            if row is not None:
                label = row.labelWidget
                field = row.fieldWidget
                if label: label.setVisible(not checked)
                if field: field.setVisible(not checked)
        if hasattr(self, '_edit_local_endpoint'):
            self._edit_local_endpoint.setVisible(checked)
            self._edit_local_model.setVisible(checked)

    def _on_accept(self):
        from src.help_engine import RAGSettings
        self.settings = RAGSettings(
            enabled=self._check_enable.isChecked(),
            api_base=self._edit_base.text().strip(),
            api_key=self._edit_key.text().strip(),
            model=self._cmb_model.currentText().strip(),
            use_local=self._check_local.isChecked(),
            local_model=self._edit_local_model.text().strip() or "qwen2.5:7b",
            local_endpoint=self._edit_local_endpoint.text().strip() or "http://localhost:11434",
        )
        self.accept()


# ═══════════════════════════════════════════════════════════════
# 系统设置对话框 — 整合主题/语言/字体/LLM
# ═══════════════════════════════════════════════════════════════

class SystemSettingsDialog(QDialog):
    """系统设置: 字体大小 + 主题 + 语言 + LLM API。"""

    def __init__(self, parent: "MainWindow"):
        super().__init__(parent)
        self._mw = parent
        self.setWindowTitle("系统设置")
        self._setup_ui()
        self._load_state()
        auto_size_dialog(self, 520, 620)
        # setStyleSheet 会重置 minimumWidth → 在 QSS 应用后恢复
        if hasattr(self, '_edit_api_base'):
            self._edit_api_base.setMinimumWidth(340)
        if hasattr(self, '_edit_api_key'):
            self._edit_api_key.setMinimumWidth(340)

    def _setup_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)

        # ── 字体大小 ──
        font_grp = QGroupBox("字体大小")
        font_layout = QHBoxLayout(font_grp)
        font_layout.addWidget(QLabel("A"))
        self._spin_font = QSpinBox()
        self._spin_font.setRange(8, 24)
        self._spin_font.setValue(13)
        self._spin_font.setSuffix(" px")
        font_layout.addWidget(self._spin_font)
        font_layout.addWidget(QLabel("A 大"))
        font_layout.addStretch()
        btn_font = QPushButton("应用字体")
        btn_font.clicked.connect(self._on_apply_font)
        font_layout.addWidget(btn_font)
        layout.addWidget(font_grp)

        # ── 主题 ──
        theme_grp = QGroupBox("主题")
        theme_layout = QHBoxLayout(theme_grp)
        self._cmb_theme = QComboBox()
        self._cmb_theme.setEditable(True)
        self._cmb_theme.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_theme.lineEdit().setPlaceholderText("搜索...")
        self._cmb_theme.setMinimumWidth(200)
        theme_layout.addWidget(QLabel("主题:"))
        theme_layout.addWidget(self._cmb_theme)
        theme_layout.addStretch()
        layout.addWidget(theme_grp)

        # ── 语言 ──
        lang_grp = QGroupBox("语言 / Language")
        lang_layout = QHBoxLayout(lang_grp)
        self._btn_lang = QPushButton("中文 / English")
        self._btn_lang.clicked.connect(self._on_toggle_lang)
        lang_layout.addWidget(self._btn_lang)
        lang_layout.addStretch()
        layout.addWidget(lang_grp)

        # ── 模板预设管理 ──
        tpl_grp = QGroupBox("模板预设管理")
        tpl_layout = QFormLayout(tpl_grp)
        tpl_layout.setSpacing(6)

        # 模板文件: 浏览选择
        tpl_path_row = QHBoxLayout()
        self._edit_tpl_path = QLineEdit()
        self._edit_tpl_path.setPlaceholderText("选择模板文件 (.xlsx .xls .csv .docx)")
        btn_browse_tpl = QPushButton("浏览...")
        btn_browse_tpl.clicked.connect(self._on_browse_template_file)
        tpl_path_row.addWidget(self._edit_tpl_path)
        tpl_path_row.addWidget(btn_browse_tpl)
        tpl_layout.addRow("模板文件:", tpl_path_row)

        # 厂商: 可搜索下拉
        self._cmb_mfr = QComboBox()
        self._cmb_mfr.setEditable(True)
        self._cmb_mfr.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_mfr.lineEdit().setPlaceholderText("搜索或输入新厂商...")
        self._cmb_mfr.addItem("", "")
        for mfr in self._mw._tm.manufacturers:
            self._cmb_mfr.addItem(mfr, mfr)
        self._cmb_mfr.currentIndexChanged.connect(self._on_tpl_mfr_changed)
        tpl_layout.addRow("厂商:", self._cmb_mfr)

        # 模板名: 可搜索下拉
        self._cmb_template = QComboBox()
        self._cmb_template.setEditable(True)
        self._cmb_template.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_template.lineEdit().setPlaceholderText("搜索模板...")
        self._cmb_template.setMinimumWidth(150)
        self._cmb_template.addItem("", "")
        self._cmb_template.currentIndexChanged.connect(self._on_tpl_selected)
        tpl_layout.addRow("模板名:", self._cmb_template)

        self._edit_tpl_output_dir = QLineEdit()
        self._edit_tpl_output_dir.setPlaceholderText("默认输出目录（可选）")
        tpl_layout.addRow("输出目录:", self._edit_tpl_output_dir)

        btn_row = QHBoxLayout()
        self._btn_save_preset = QPushButton("💾 保存为预设")
        self._btn_save_preset.clicked.connect(self._on_tpl_save)
        btn_row.addWidget(self._btn_save_preset)
        btn_row.addStretch()
        tpl_layout.addRow("", btn_row)

        layout.addWidget(tpl_grp)

        # ── RSP 校准预设管理 ──
        rsp_grp = QGroupBox("RSP 校准预设管理")
        rsp_layout = QFormLayout(rsp_grp)
        rsp_layout.setSpacing(6)

        self._cmb_rsp_name = QComboBox()
        self._cmb_rsp_name.setEditable(True)
        self._cmb_rsp_name.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_rsp_name.lineEdit().setPlaceholderText("输入新预设名称或选择已有...")
        self._cmb_rsp_name.currentIndexChanged.connect(self._on_rsp_preset_selected)
        rsp_layout.addRow("预设名称:", self._cmb_rsp_name)

        self._cmb_rsp_mode = QComboBox()
        self._cmb_rsp_mode.setEditable(True)
        self._cmb_rsp_mode.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_rsp_mode.lineEdit().setPlaceholderText("搜索...")
        self._cmb_rsp_mode.addItem("通用 (任意模式)", -1)
        self._cmb_rsp_mode.addItem("无源天线", 0)
        self._cmb_rsp_mode.addItem("有源发射 TRP", 1)
        self._cmb_rsp_mode.addItem("有源接收 TIS", 2)
        rsp_layout.addRow("关联测试模式:", self._cmb_rsp_mode)

        self._edit_rsp_h = QLineEdit()
        self._edit_rsp_h.setPlaceholderText("选择 H-pol RSP 校准文件 (Phi 分量)")
        btn_browse_rsp_h = QPushButton("浏览...")
        btn_browse_rsp_h.clicked.connect(self._on_browse_rsp_h)
        h_row = QHBoxLayout()
        h_row.addWidget(self._edit_rsp_h)
        h_row.addWidget(btn_browse_rsp_h)
        rsp_layout.addRow("H-pol RSP:", h_row)

        self._edit_rsp_v = QLineEdit()
        self._edit_rsp_v.setPlaceholderText("选择 V-pol RSP 校准文件 (Theta 分量)")
        btn_browse_rsp_v = QPushButton("浏览...")
        btn_browse_rsp_v.clicked.connect(self._on_browse_rsp_v)
        v_row = QHBoxLayout()
        v_row.addWidget(self._edit_rsp_v)
        v_row.addWidget(btn_browse_rsp_v)
        rsp_layout.addRow("V-pol RSP:", v_row)

        self._edit_rsp_desc = QLineEdit()
        self._edit_rsp_desc.setPlaceholderText("可选注释")
        rsp_layout.addRow("描述:", self._edit_rsp_desc)

        btn_rsp_row = QHBoxLayout()
        self._btn_save_rsp = QPushButton("💾 保存 RSP 预设")
        self._btn_save_rsp.clicked.connect(self._on_rsp_save)
        btn_rsp_row.addWidget(self._btn_save_rsp)
        self._btn_delete_rsp = QPushButton("🗑 删除预设")
        self._btn_delete_rsp.clicked.connect(self._on_rsp_delete)
        btn_rsp_row.addWidget(self._btn_delete_rsp)
        btn_rsp_row.addStretch()
        rsp_layout.addRow("", btn_rsp_row)

        # ── 当前默认值 ──
        defaults_grp = QGroupBox("当前默认值 (工具自动匹配)")
        defaults_layout = QFormLayout(defaults_grp)
        defaults_layout.setSpacing(4)

        self._cmb_default_passive = QComboBox()
        self._cmb_default_passive.setEditable(True)
        self._cmb_default_passive.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_default_passive.lineEdit().setPlaceholderText("搜索...")
        self._cmb_default_passive.addItem("(未设置)", "")
        defaults_layout.addRow("无源天线:", self._cmb_default_passive)

        self._cmb_default_trp = QComboBox()
        self._cmb_default_trp.setEditable(True)
        self._cmb_default_trp.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_default_trp.lineEdit().setPlaceholderText("搜索...")
        self._cmb_default_trp.addItem("(未设置)", "")
        defaults_layout.addRow("有源 TRP:", self._cmb_default_trp)

        self._cmb_default_tis = QComboBox()
        self._cmb_default_tis.setEditable(True)
        self._cmb_default_tis.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_default_tis.lineEdit().setPlaceholderText("搜索...")
        self._cmb_default_tis.addItem("(未设置)", "")
        defaults_layout.addRow("有源 TIS:", self._cmb_default_tis)

        btn_set_defaults = QPushButton("应用默认值")
        btn_set_defaults.clicked.connect(self._on_rsp_set_defaults)
        defaults_layout.addRow("", btn_set_defaults)

        rsp_layout.addRow(defaults_grp)
        layout.addWidget(rsp_grp)

        # ── 默认保存目录 ──
        dirs_grp = QGroupBox("默认保存目录")
        dirs_layout = QFormLayout(dirs_grp)
        dirs_layout.setSpacing(4)
        btn_default_dirs = QPushButton("设置默认保存目录...")
        btn_default_dirs.clicked.connect(self._show_default_dirs)
        dirs_layout.addRow(btn_default_dirs)
        layout.addWidget(dirs_grp)

        # ── LLM API ──
        llm_grp = QGroupBox("LLM API (RAG 问答)")
        llm_layout = QFormLayout(llm_grp)
        llm_layout.setSpacing(6)

        self._check_llm = QCheckBox("启用 RAG AI 问答")
        llm_layout.addRow("", self._check_llm)

        self._edit_api_base = QLineEdit()
        self._edit_api_base.setPlaceholderText("https://api.anthropic.com/v1/messages")
        self._edit_api_base.setMinimumWidth(340)
        llm_layout.addRow("API URL:", self._edit_api_base)

        self._edit_api_key = QLineEdit()
        self._edit_api_key.setEchoMode(QLineEdit.Password)
        self._edit_api_key.setPlaceholderText("sk-ant-... 或 sk-...")
        # setMinimumWidth 在 setStyleSheet 后会被重置，在 _on_accept 后
        # 由主窗口的 _apply_minimum_sizes 重新补充。此处保留一份调用。
        self._edit_api_key.setMinimumWidth(340)
        llm_layout.addRow("API Key:", self._edit_api_key)

        self._cmb_model = QComboBox()
        self._cmb_model.setEditable(True)
        self._cmb_model.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_model.lineEdit().setPlaceholderText("搜索...")
        self._cmb_model.setEditable(True)
        for m in ["claude-sonnet-4-6", "claude-opus-4-8", "gpt-4o", "gpt-4o-mini", "deepseek-chat"]:
            self._cmb_model.addItem(m)
        llm_layout.addRow("Model:", self._cmb_model)

        self._check_local = QCheckBox("使用本地 Ollama 模型（免费、离线、无需 API Key）")
        self._check_local.toggled.connect(self._on_local_toggled)
        llm_layout.addRow("", self._check_local)

        self._edit_local_model = QLineEdit()
        self._edit_local_model.setPlaceholderText("llama3:8b")
        llm_layout.addRow("本地模型名:", self._edit_local_model)

        self._edit_local_endpoint = QLineEdit()
        self._edit_local_endpoint.setPlaceholderText("http://localhost:11434/v1")
        llm_layout.addRow("本地端点:", self._edit_local_endpoint)

        layout.addWidget(llm_grp)

        # ── 智能识别 LLM (独立于 RAG 问答, 可配置不同模型) ──
        self._ai_grp = QGroupBox("智能识别 (AI 辅助) — 模板识别/数据源匹配/参数检测")
        ai_layout = QFormLayout(self._ai_grp)
        ai_layout.setSpacing(6)

        self._check_ai = QCheckBox("启用 AI 辅助识别（规则匹配失败时的兜底方案）")
        ai_layout.addRow("", self._check_ai)

        self._cmb_ai_mode = QComboBox()
        self._cmb_ai_mode.setEditable(True)
        self._cmb_ai_mode.setInsertPolicy(QComboBox.NoInsert)
        self._cmb_ai_mode.lineEdit().setPlaceholderText("搜索...")
        self._cmb_ai_mode.addItems(["cloud", "local"])
        self._cmb_ai_mode.currentIndexChanged.connect(self._on_ai_mode_changed)
        ai_layout.addRow("AI 模式:", self._cmb_ai_mode)

        self._edit_ai_api_base = QLineEdit()
        self._edit_ai_api_base.setPlaceholderText("https://api.anthropic.com/v1/messages")
        ai_layout.addRow("API URL:", self._edit_ai_api_base)

        self._edit_ai_api_key = QLineEdit()
        self._edit_ai_api_key.setEchoMode(QLineEdit.Password)
        self._edit_ai_api_key.setPlaceholderText("sk-ant-... 或 sk-...")
        ai_layout.addRow("API Key:", self._edit_ai_api_key)

        self._edit_ai_model = QLineEdit()
        self._edit_ai_model.setPlaceholderText("claude-sonnet-4-6 (云) / qwen2.5:7b (本地)")
        ai_layout.addRow("模型:", self._edit_ai_model)

        self._edit_ai_local_endpoint = QLineEdit()
        self._edit_ai_local_endpoint.setPlaceholderText("http://localhost:11434")
        ai_layout.addRow("本地地址:", self._edit_ai_local_endpoint)

        self._ai_cloud_fields = [self._edit_ai_api_base, self._edit_ai_api_key]
        self._ai_local_fields = [self._edit_ai_local_endpoint]
        self._on_ai_mode_changed(0)  # 初始化 cloud 模式

        layout.addWidget(self._ai_grp)

        layout.addStretch()

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        wrap_in_scroll(self, [font_grp, theme_grp, lang_grp, tpl_grp, dirs_grp, rsp_grp, llm_grp, self._ai_grp], btns)

        # 限制高度不超过屏幕 90%, 允许手动调整
        screen = QApplication.primaryScreen().availableGeometry()
        max_h = int(screen.height() * 0.9)
        self.setMaximumHeight(max_h)
        self.setMinimumHeight(400)
        self.resize(self.sizeHint().width(), min(self.sizeHint().height() + 40, max_h))

        # 存储 scroll area 引用供自动滚动使用
        self._ai_scroll_area = None
        scroll_item = self.layout().itemAt(0)
        if scroll_item is not None:
            w = scroll_item.widget()
            if isinstance(w, QScrollArea):
                self._ai_scroll_area = w

    def _load_state(self):
        mw = self._mw
        from src.config_manager import get_config_manager
        cfg = get_config_manager().config
        # 字体
        self._spin_font.setValue(cfg.font_size)
        # 主题
        from ui.theme_manager import ThemeManager
        for theme_id, name in ThemeManager.ALL_THEMES:
            self._cmb_theme.addItem(name, theme_id)
        cur = ThemeManager.current_theme()
        for i in range(self._cmb_theme.count()):
            if self._cmb_theme.itemData(i) == cur:
                self._cmb_theme.setCurrentIndex(i)
                break
        # 语言
        from i18n.i18n_manager import I18nManager
        lang = I18nManager.current_language()
        self._btn_lang.setText("English" if lang == "zh_CN" else "中文")
        # LLM
        self._check_llm.setChecked(cfg.llm.enabled)
        self._edit_api_base.setText(cfg.llm.api_base)
        self._edit_api_key.setText(get_config_manager().get_api_key("llm"))
        idx = self._cmb_model.findText(cfg.llm.model)
        if idx >= 0:
            self._cmb_model.setCurrentIndex(idx)
        else:
            self._cmb_model.setCurrentText(cfg.llm.model)
        self._check_local.setChecked(cfg.llm.use_local)
        self._edit_local_model.setText(cfg.llm.local_model)
        self._edit_local_endpoint.setText(cfg.llm.local_endpoint)
        self._on_local_toggled(cfg.llm.use_local)

        # AI 辅助设置
        self._check_ai.setChecked(cfg.ai.enabled)
        mode = cfg.ai.mode
        midx = self._cmb_ai_mode.findText(mode)
        if midx >= 0:
            self._cmb_ai_mode.setCurrentIndex(midx)
        self._edit_ai_api_base.setText(cfg.ai.api_base)
        self._edit_ai_api_key.setText(get_config_manager().get_api_key("ai"))
        self._edit_ai_model.setText(cfg.ai.model)
        self._edit_ai_local_endpoint.setText(cfg.ai.local_endpoint)
        self._on_ai_mode_changed(0)

        # RSP 预设
        self._refresh_rsp_presets()

    # ── 模板预设管理 ──

    def _on_browse_template_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择模板文件", "",
            "所有支持格式 (*.xlsx *.xls *.csv *.docx);;Excel (*.xlsx *.xls);;CSV (*.csv);;Word (*.docx);;所有文件 (*)")
        if path:
            self._edit_tpl_path.setText(path)

    def _on_tpl_mfr_changed(self, index: int):
        self._cmb_template.clear()
        self._cmb_template.addItem("", "")
        mfr = self._cmb_mfr.currentData()
        templates = self._mw._tm.get_templates(mfr) if mfr else self._mw._tm.get_all_templates()
        for tpl in templates:
            self._cmb_template.addItem(tpl.name, tpl)

    def _on_tpl_selected(self, index: int):
        from src.template_manager import TemplatePreset
        tpl = self._cmb_template.currentData()
        if isinstance(tpl, TemplatePreset):
            self._edit_tpl_path.setText(tpl.path)
            self._edit_tpl_output_dir.setText(tpl.default_output_dir)

    def _on_tpl_save(self):
        path = self._edit_tpl_path.text().strip()
        mfr = self._cmb_mfr.currentText().strip()
        tpl_name = self._cmb_template.currentText().strip()
        output_dir = self._edit_tpl_output_dir.text().strip()
        if not path:
            QMessageBox.warning(self, "保存预设", "请先选择模板文件。")
            return
        if not mfr:
            QMessageBox.warning(self, "保存预设", "请输入或选择厂商名称。")
            return
        if not tpl_name:
            QMessageBox.warning(self, "保存预设", "请输入模板名称。")
            return
        self._mw._tm.add_template(mfr, tpl_name, path, output_dir)
        self._refresh_tpl_lists()
        self._log_msg(f"✓ 模板预设已保存: {mfr} → {tpl_name}")

    def _log_msg(self, msg: str):
        if hasattr(self._mw, '_log'):
            self._mw._log(msg)

    def _refresh_tpl_lists(self):
        """刷新厂商和模板下拉列表。"""
        cur_mfr = self._cmb_mfr.currentData()
        self._cmb_mfr.blockSignals(True)
        self._cmb_mfr.clear()
        self._cmb_mfr.addItem("(所有厂商)", "")
        for mfr in self._mw._tm.manufacturers:
            self._cmb_mfr.addItem(mfr, mfr)
        idx = self._cmb_mfr.findData(cur_mfr)
        if idx >= 0: self._cmb_mfr.setCurrentIndex(idx)
        self._cmb_mfr.blockSignals(False)
        self._on_tpl_mfr_changed(0)

    # ── RSP 预设管理 ──

    def _on_browse_rsp_h(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 H-pol RSP 校准文件", "",
            "CSV/Excel 文件 (*.csv *.xlsx *.xls);;所有文件 (*)")
        if path:
            self._edit_rsp_h.setText(path)

    def _on_browse_rsp_v(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 V-pol RSP 校准文件", "",
            "CSV/Excel 文件 (*.csv *.xlsx *.xls);;所有文件 (*)")
        if path:
            self._edit_rsp_v.setText(path)

    def _on_rsp_preset_selected(self, _index: int):
        from src.rsp_preset_manager import RspPresetManager
        name = self._cmb_rsp_name.currentText().strip()
        if not name:
            return
        mgr = RspPresetManager()
        preset = mgr.get_by_name(name)
        if preset:
            self._cmb_rsp_name.blockSignals(True)
            mode_idx = self._cmb_rsp_mode.findData(preset.test_mode)
            if mode_idx >= 0:
                self._cmb_rsp_mode.setCurrentIndex(mode_idx)
            self._edit_rsp_h.setText(preset.rsp_h_path)
            self._edit_rsp_v.setText(preset.rsp_v_path)
            self._edit_rsp_desc.setText(preset.description)
            self._cmb_rsp_name.blockSignals(False)

    def _on_rsp_save(self):
        from src.rsp_preset_manager import RspPresetManager, RspPreset
        name = self._cmb_rsp_name.currentText().strip()
        if not name:
            QMessageBox.warning(self, "保存预设", "请输入预设名称。")
            return
        test_mode = self._cmb_rsp_mode.currentData()
        rsp_h = self._edit_rsp_h.text().strip()
        rsp_v = self._edit_rsp_v.text().strip()
        desc = self._edit_rsp_desc.text().strip()
        if not rsp_h and not rsp_v:
            QMessageBox.warning(self, "保存预设", "请至少选择一个 RSP 校准文件。")
            return
        mgr = RspPresetManager()
        preset = RspPreset(
            name=name, test_mode=test_mode,
            rsp_h_path=rsp_h, rsp_v_path=rsp_v, description=desc,
        )
        mgr.add_or_update(preset)
        self._refresh_rsp_presets()
        self._log_msg(f"RSP 预设已保存: {name}")

    def _on_rsp_delete(self):
        from src.rsp_preset_manager import RspPresetManager
        name = self._cmb_rsp_name.currentText().strip()
        if not name:
            return
        reply = QMessageBox.question(
            self, "删除预设",
            f"确定删除 RSP 预设「{name}」？\n（关联的默认值将同时清除）",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        mgr = RspPresetManager()
        mgr.delete(name)
        self._refresh_rsp_presets()
        self._log_msg(f"RSP 预设已删除: {name}")

    def _on_rsp_set_defaults(self):
        from src.rsp_preset_manager import RspPresetManager
        mgr = RspPresetManager()
        for mode, combo in [(0, self._cmb_default_passive),
                            (1, self._cmb_default_trp),
                            (2, self._cmb_default_tis)]:
            name = combo.currentData()
            mgr.set_default(mode, name if name else None)
        self._refresh_rsp_presets()
        self._log_msg("RSP 默认值已更新")

    def _refresh_rsp_presets(self):
        from src.rsp_preset_manager import RspPresetManager
        mgr = RspPresetManager()
        cur_name = self._cmb_rsp_name.currentText()
        self._cmb_rsp_name.blockSignals(True)
        self._cmb_rsp_name.clear()
        self._cmb_rsp_name.addItem("")  # 空选项
        for preset in mgr.presets:
            self._cmb_rsp_name.addItem(preset.name)
        idx = self._cmb_rsp_name.findText(cur_name)
        if idx >= 0:
            self._cmb_rsp_name.setCurrentIndex(idx)
        self._cmb_rsp_name.blockSignals(False)

        # 刷新默认值下拉框
        for mode, combo in [(0, self._cmb_default_passive),
                            (1, self._cmb_default_trp),
                            (2, self._cmb_default_tis)]:
            combo.blockSignals(True)
            combo.clear()
            combo.addItem("(未设置)", "")
            for preset in mgr.presets:
                if preset.test_mode in (mode, -1):
                    combo.addItem(preset.name, preset.name)
            default_name = mgr.get_default(mode)
            if default_name:
                idx2 = combo.findData(default_name)
                if idx2 >= 0:
                    combo.setCurrentIndex(idx2)
            combo.blockSignals(False)

    def _on_apply_font(self):
        size = self._spin_font.value()
        if self._mw is None:
            return
        from src.config_manager import get_config_manager
        get_config_manager().config.font_size = size
        get_config_manager()._dirty = True
        ScaleManager._font_scale = size / ScaleManager.BASE_FONT_SIZE
        ScaleManager.update(self._mw.width() if self._mw else 1920)
        ScaleManager.apply_full_qss(self._mw, self._mw._base_qss)
        self._mw.update()

    def _on_ai_mode_changed(self, idx):
        is_cloud = self._cmb_ai_mode.currentText() == "cloud"
        for w in self._ai_cloud_fields:
            w.setVisible(is_cloud)
        for w in self._ai_local_fields:
            w.setVisible(not is_cloud)

    def _on_toggle_lang(self):
        from i18n.i18n_manager import I18nManager
        new_lang = "en_US" if I18nManager.current_language() == "zh_CN" else "zh_CN"
        I18nManager.switch(QApplication.instance(), new_lang)
        self._btn_lang.setText("English" if new_lang == "zh_CN" else "中文")

    def _on_accept(self):
        # 模板预设 — 应用选中的模板
        from src.template_manager import TemplatePreset
        tpl = self._cmb_template.currentData()
        if isinstance(tpl, TemplatePreset):
            output_dir = self._edit_tpl_output_dir.text().strip() or tpl.default_output_dir
            self._mw.apply_template_preset(tpl.path, output_dir, tpl.name)
        # 字体 — 先应用字体设置
        self._on_apply_font()
        # 主题 — 重置后重建: theme_qss + custom_qss + ScaleManager
        theme_id = self._cmb_theme.currentData()
        if theme_id:
            from ui.theme_manager import ThemeManager
            ThemeManager.apply(theme_id)
            ThemeManager.save_theme(theme_id)
            if self._mw:
                self._mw._theme_qss = QApplication.instance().styleSheet()
                self._mw._base_qss = self._mw._theme_qss + self._mw._custom_qss
                ScaleManager.apply_full_qss(self._mw, self._mw._base_qss)
        # LLM 设置 → 统一配置文件
        from src.config_manager import get_config_manager
        mgr = get_config_manager()
        mgr.config.llm.enabled = self._check_llm.isChecked()
        mgr.config.llm.api_base = self._edit_api_base.text().strip()
        mgr.set_api_key("llm", self._edit_api_key.text().strip())
        mgr.config.llm.model = self._cmb_model.currentText().strip()
        mgr.config.llm.use_local = self._check_local.isChecked()
        mgr.config.llm.local_model = self._edit_local_model.text().strip()
        mgr.config.llm.local_endpoint = self._edit_local_endpoint.text().strip()
        # AI 辅助设置
        mgr.config.ai.enabled = self._check_ai.isChecked()
        mgr.config.ai.mode = self._cmb_ai_mode.currentText().strip()
        mgr.config.ai.api_base = self._edit_ai_api_base.text().strip()
        mgr.set_api_key("ai", self._edit_ai_api_key.text().strip())
        mgr.config.ai.model = self._edit_ai_model.text().strip()
        mgr.config.ai.local_endpoint = self._edit_ai_local_endpoint.text().strip()
        mgr.save()
        self.accept()

    def _on_local_toggled(self, checked):
        """本地模型开关 — 控制本地模型名和端点输入框的可见性。"""
        self._edit_local_model.setVisible(checked)
        self._edit_local_endpoint.setVisible(checked)

    def scroll_to_ai_settings(self):
        """滚动到「智能识别 (AI 辅助)」设置区域。"""
        if self._ai_scroll_area is not None and self._ai_grp is not None:
            self._ai_scroll_area.ensureWidgetVisible(self._ai_grp)

    def _show_default_dirs(self):
        """打开默认保存目录设置对话框。"""
        dlg = DefaultDirsDialog(self)
        dlg.exec()


# ═══════════════════════════════════════════════════════════════
# 默认保存目录对话框
# ═══════════════════════════════════════════════════════════════

class DefaultDirsDialog(QDialog):
    """为四类输出文件设置默认保存目录。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(self.tr("默认保存目录"))
        self.setMinimumWidth(600)
        from src.config_manager import get_config_manager
        self._cfg = get_config_manager()
        self._setup_ui()
        self._load()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        info = QLabel(self.tr("以下目录将作为各类文件的默认保存位置。\n"
                               "新建任务时自动使用此处设置的目录，可随时在界面中临时修改。"))
        info.setWordWrap(True)
        layout.addWidget(info)

        form = QFormLayout()
        form.setSpacing(8)

        # 1) 天线参数测试报告 (.xlsx)
        self._edit_excel = QLineEdit()
        self._edit_excel.setPlaceholderText(self.tr("默认: 源文件目录"))
        form.addRow(self.tr("天线参数报告 (.xlsx):"), self._make_dir_row(self._edit_excel))

        # 2) 图表报告 (.docx)
        self._edit_word = QLineEdit()
        self._edit_word.setPlaceholderText(self.tr("默认: 源文件目录"))
        form.addRow(self.tr("Word 图表报告 (.docx):"), self._make_dir_row(self._edit_word))

        # 3) 中间数据 (.xlsx)
        self._edit_data = QLineEdit()
        self._edit_data.setPlaceholderText(self.tr("默认: 源文件目录"))
        form.addRow(self.tr("中间数据 (.xlsx):"), self._make_dir_row(self._edit_data))

        # 4) 任务包 (.ant)
        self._edit_ant = QLineEdit()
        self._edit_ant.setPlaceholderText(self.tr("默认: 源文件目录"))
        form.addRow(self.tr("任务包 (.ant):"), self._make_dir_row(self._edit_ant))

        layout.addLayout(form)
        layout.addStretch()

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._on_save)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _make_dir_row(self, edit):
        row = QHBoxLayout()
        row.addWidget(edit, 1)
        btn = QPushButton(self.tr("浏览..."))
        btn.clicked.connect(lambda: self._browse_dir(edit))
        row.addWidget(btn)
        return row

    def _browse_dir(self, edit):
        start = edit.text() or str(Path.cwd())
        path = QFileDialog.getExistingDirectory(self, self.tr("选择目录"), start)
        if path:
            edit.setText(path)

    def _load(self):
        dd = getattr(self._cfg.config, 'default_dirs', None) or {}
        self._edit_excel.setText(dd.get('excel', ''))
        self._edit_word.setText(dd.get('word', ''))
        self._edit_data.setText(dd.get('data', ''))
        self._edit_ant.setText(dd.get('ant', ''))

    def _on_save(self):
        dd = {
            'excel': self._edit_excel.text().strip(),
            'word': self._edit_word.text().strip(),
            'data': self._edit_data.text().strip(),
            'ant': self._edit_ant.text().strip(),
        }
        self._cfg.config.default_dirs = dd
        self._cfg._dirty = True
        self._cfg._save()
        self.accept()


# ═══════════════════════════════════════════════════════════════
# 报告元数据编辑对话框
# ═══════════════════════════════════════════════════════════════

class ReportMetadataDialog(QDialog):
    """编辑测试报告元数据: 客户信息、项目信息、测试配置等。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(self.tr("报告元数据"))
        self.setMinimumSize(600, 500)
        from src.config_manager import get_config_manager
        self._cfg = get_config_manager()
        self._setup_ui()
        self._load()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        info = QLabel(self.tr("以下信息将填入 Word 测试报告模板的 SDT Tag 中。"))
        info.setWordWrap(True)
        layout.addWidget(info)

        tabs = QTabWidget()

        # Tab 1: 项目信息
        tab1 = QWidget()
        form1 = QFormLayout(tab1)
        form1.setSpacing(6)
        self._fields = {}
        fields_tab1 = [
            ('customer',       '客户名称'),
            ('project',        '项目名称'),
            ('contract_no',    '合同号'),
            ('antenna_model',  '天线型号'),
            ('report_no',      '报告编号'),
            ('test_standard',  '测试标准'),
        ]
        for key, label in fields_tab1:
            edit = QLineEdit()
            edit.setPlaceholderText(label)
            form1.addRow(label + ":", edit)
            self._fields[key] = edit
        tabs.addTab(tab1, self.tr("项目信息"))

        # Tab 2: 测试信息
        tab2 = QWidget()
        form2 = QFormLayout(tab2)
        form2.setSpacing(6)
        fields_tab2 = [
            ('test_lab',        '测试实验室'),
            ('test_lab_addr',   '实验室地址'),
            ('test_engineer',   '测试工程师'),
            ('reviewer',        '审核人'),
            ('test_start_date', '测试开始日期'),
            ('test_end_date',   '测试结束日期'),
            ('test_plan_no',    '测试计划编号'),
            ('test_plan_ver',   '测试计划版本'),
        ]
        for key, label in fields_tab2:
            edit = QLineEdit()
            edit.setPlaceholderText(label)
            form2.addRow(label + ":", edit)
            self._fields[key] = edit
        tabs.addTab(tab2, self.tr("测试信息"))

        # Tab 3: 备注
        tab3 = QWidget()
        form3 = QVBoxLayout(tab3)
        self._edit_notes = QPlainTextEdit()
        self._edit_notes.setPlaceholderText(self.tr("备注信息（可选）"))
        form3.addWidget(self._edit_notes)
        tabs.addTab(tab3, self.tr("备注"))

        layout.addWidget(tabs)

        # 按钮栏
        btn_row = QHBoxLayout()
        btn_extract = QPushButton(self.tr("🔍 从数据源提取"))
        btn_extract.setToolTip(self.tr("从当前加载的 JSON/CSV 数据源中自动提取元数据"))
        btn_extract.clicked.connect(self._extract_from_datasource)
        btn_row.addWidget(btn_extract)
        btn_import = QPushButton(self.tr("📥 从 Excel 导入..."))
        btn_import.clicked.connect(self._import_excel)
        btn_row.addWidget(btn_import)
        btn_row.addStretch()
        self._lbl_source = QLabel("")
        self._lbl_source.setStyleSheet("color: gray; font-style: italic;")
        btn_row.addWidget(self._lbl_source)
        layout.addLayout(btn_row)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._on_save)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _load(self):
        md = getattr(self._cfg.config, 'metadata', None) or {}
        for key, edit in self._fields.items():
            edit.setText(md.get(key, ''))
        self._edit_notes.setPlainText(md.get('notes', ''))

    def _on_save(self):
        md = {key: edit.text().strip() for key, edit in self._fields.items()}
        md['notes'] = self._edit_notes.toPlainText().strip()
        self._cfg.config.metadata = md
        self._cfg._dirty = True
        self._cfg._save()
        self.accept()

    def _extract_from_datasource(self):
        """从当前数据源 (JSON/CSV) 自动提取元数据。"""
        parent_widget = self.parent()
        mw = None
        # 向上查找 MainWindow
        while parent_widget:
            if hasattr(parent_widget, '_data_file_paths'):
                mw = parent_widget
                break
            parent_widget = parent_widget.parent()

        if not mw or not getattr(mw, '_data_file_paths', None):
            QMessageBox.warning(self, self.tr("无数据源"), self.tr("请先在系统设置中添加数据文件。"))
            return

        # 查找 JSON 文件
        json_paths = [p for p in mw._data_file_paths if p.lower().endswith('.json')]
        if not json_paths:
            QMessageBox.information(self, self.tr("无 JSON 数据源"),
                self.tr("当前数据源中无 JSON 文件。\n"
                        "JSON 文件由 EMQuest 导出，包含被测件型号、操作员、测试时间等元数据。"))
            return

        extracted = {}
        for jp in json_paths:
            try:
                from src.json_reader import JsonDataSource
                ds = JsonDataSource(jp)
                md = ds.get_metadata()
                extracted.update(md)
                self._lbl_source.setText(self.tr("已从 {n} 个 JSON 提取").format(n=len(json_paths)))
                break  # 取第一个成功的
            except Exception as e:
                continue

        if not extracted:
            QMessageBox.warning(self, self.tr("提取失败"), self.tr("无法从 JSON 文件中提取元数据。"))
            return

        # 映射 JSON 字段 → 表单字段
        field_map = {
            'model': 'antenna_model',
            'manufacturer': 'customer',
            'serialno': 'contract_no',
            'operator': 'test_engineer',
            'test_method': 'test_standard',
            'test_time': 'test_start_date',
            'test_end_time': 'test_end_date',
        }
        filled = 0
        for src_key, field_key in field_map.items():
            if src_key in extracted and field_key in self._fields:
                self._fields[field_key].setText(str(extracted[src_key]))
                filled += 1

        # 备注: 合并额外信息
        notes_parts = []
        for k in ('opcomments', 'parm_file', 'app_version', 'freq_range', 'data_format'):
            if k in extracted:
                notes_parts.append(f"{k}: {extracted[k]}")
        if notes_parts:
            self._edit_notes.setPlainText('\n'.join(notes_parts))

        # User Defined 字段追加到备注
        for i in range(1, 13):
            uk = f'user_{i}'
            if uk in extracted:
                label = extracted.get(f'user_{i}_label', '')
                existing = self._edit_notes.toPlainText()
                self._edit_notes.setPlainText(
                    existing + f"\n{label}: {extracted[uk]}")

        self._lbl_source.setText(
            self.tr("已提取 {f} 项").format(f=filled))

    def _import_excel(self):
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getOpenFileName(
            self, self.tr("选择元数据 Excel"), "",
            self.tr("Excel 文件 (*.xlsx *.xls)"))
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            if '项目信息' in wb.sheetnames:
                ws = wb['项目信息']
                # 读取键值对 (A=key, B=value, C=key, D=value)
                mapping = {
                    '合同号': 'contract_no', '客户名称': 'customer',
                    '项目名称': 'project', '天线型号': 'antenna_model',
                    '测试标准': 'test_standard', '报告编号': 'report_no',
                    '测试实验室': 'test_lab', '实验室地址': 'test_lab_addr',
                    '测试工程师': 'test_engineer', '审核人': 'reviewer',
                    '测试开始日期': 'test_start_date', '测试结束日期': 'test_end_date',
                    '测试计划编号': 'test_plan_no', '测试计划版本': 'test_plan_ver',
                }
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for col_pair in [(0, 1), (2, 3)]:
                        key_cell = row[col_pair[0]].value
                        val_cell = row[col_pair[1]].value
                        if key_cell and key_cell in mapping:
                            field_key = mapping[key_cell]
                            if field_key in self._fields:
                                self._fields[field_key].setText(str(val_cell) if val_cell else '')
            wb.close()
            QMessageBox.information(self, self.tr("导入完成"), self.tr("元数据已从 Excel 导入。"))
        except Exception as e:
            QMessageBox.warning(self, self.tr("导入失败"), str(e))


# ═══════════════════════════════════════════════════════════════
# 步进重采样对话框
# ═══════════════════════════════════════════════════════════════

class ResampleDialog(QDialog):
    """多步进数据提取: 从源文件按多组步进值批量导出重采样 CSV。"""

    # 常用步进值
    COMMON_STEPS = [2, 5, 10, 15, 20, 30, 45]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(self.tr("多步进数据提取"))
        self.setMinimumSize(660, 520)
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # ── 1. 源文件 ──
        src_grp = QGroupBox(self.tr("源文件"))
        src_row = QHBoxLayout(src_grp)
        self._edit_src = QLineEdit()
        self._edit_src.setPlaceholderText(self.tr("选择 merged CSV 文件..."))
        src_row.addWidget(self._edit_src, 1)
        btn_src = QPushButton(self.tr("浏览..."))
        btn_src.clicked.connect(self._on_browse_src)
        src_row.addWidget(btn_src)
        layout.addWidget(src_grp)

        # ── 源文件信息 ──
        self._lbl_info = QLabel("")
        self._lbl_info.setStyleSheet("color: #666;")
        layout.addWidget(self._lbl_info)

        # ── 2. 目标步进 (checkbox 多选 + 自定义) ──
        step_grp = QGroupBox(self.tr("目标步进（度）— 可多选"))
        step_layout = QVBoxLayout(step_grp)

        # Checkbox 网格
        self._step_checks: Dict[int, QCheckBox] = {}
        cb_grid = QHBoxLayout()
        cb_grid.setSpacing(10)
        for s in self.COMMON_STEPS:
            cb = QCheckBox(f"{s}°")
            cb.setChecked(s in [5, 10, 15])  # 默认选中常用项
            cb.toggled.connect(lambda checked, val=s: self._on_step_toggled(val, checked))
            self._step_checks[s] = cb
            cb_grid.addWidget(cb)
        cb_grid.addStretch()
        step_layout.addLayout(cb_grid)

        # 自定义输入行
        custom_row = QHBoxLayout()
        custom_row.addWidget(QLabel(self.tr("自定义:")))
        self._edit_custom = QLineEdit()
        self._edit_custom.setPlaceholderText(self.tr("如: 3, 8, 25 (逗号分隔)"))
        self._edit_custom.textChanged.connect(self._on_custom_steps_changed)
        custom_row.addWidget(self._edit_custom, 1)
        step_layout.addLayout(custom_row)

        layout.addWidget(step_grp)

        # ── 3. 输出目录 + 文件命名 ──
        out_grp = QGroupBox(self.tr("输出设置"))
        out_layout = QVBoxLayout(out_grp)

        # 输出目录
        dir_row = QHBoxLayout()
        dir_row.addWidget(QLabel(self.tr("输出目录:")))
        self._edit_dir = QLineEdit()
        self._edit_dir.setPlaceholderText(self.tr("默认: 源文件所在目录"))
        dir_row.addWidget(self._edit_dir, 1)
        btn_dir = QPushButton(self.tr("浏览..."))
        btn_dir.clicked.connect(self._on_browse_dir)
        dir_row.addWidget(btn_dir)
        out_layout.addLayout(dir_row)

        # 文件命名预览
        name_row = QHBoxLayout()
        name_row.addWidget(QLabel(self.tr("命名规则:")))
        self._lbl_naming = QLabel(self.tr("源文件名_step{步进}deg.csv"))
        self._lbl_naming.setStyleSheet("color: #888; font-size:0.9em;")
        name_row.addWidget(self._lbl_naming)
        name_row.addStretch()
        out_layout.addLayout(name_row)

        layout.addWidget(out_grp)

        # ── 4. 预览 ──
        grp_preview = QGroupBox(self.tr("输出预览"))
        preview_layout = QVBoxLayout(grp_preview)
        self._preview_table = QTableWidget()
        self._preview_table.setColumnCount(4)
        self._preview_table.setHorizontalHeaderLabels(
            [self.tr("步进"), self.tr("θ点数"), self.tr("φ点数"), self.tr("输出文件名")])
        self._preview_table.horizontalHeader().setStretchLastSection(True)
        self._preview_table.setMaximumHeight(200)
        self._preview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        preview_layout.addWidget(self._preview_table)
        layout.addWidget(grp_preview)

        # ── 5. 按钮 ──
        btn_row = QHBoxLayout()
        self._btn_run = QPushButton(self.tr("▶ 开始批量导出"))
        self._btn_run.clicked.connect(self._on_run)
        self._btn_run.setMinimumHeight(36)
        self._btn_run.setEnabled(False)
        btn_row.addWidget(self._btn_run)
        btn_row.addStretch()
        self._btn_close = QPushButton(self.tr("关闭"))
        self._btn_close.clicked.connect(self.reject)
        btn_row.addWidget(self._btn_close)
        layout.addLayout(btn_row)

        # 信号连接
        self._edit_src.textChanged.connect(self._on_src_changed)

    # ── 步进选择 ──

    def _on_step_toggled(self, val: int, checked: bool):
        """checkbox 切换时更新预览。"""
        self._update_preview()

    def _on_custom_steps_changed(self, _text):
        """自定义步进文本变化时更新预览。"""
        self._update_preview()

    def _get_selected_steps(self) -> List[float]:
        """收集所有选中的步进值（checkbox + 自定义）。"""
        steps = []
        for s, cb in self._step_checks.items():
            if cb.isChecked():
                steps.append(float(s))
        # 自定义输入
        custom = self._edit_custom.text().strip()
        if custom:
            for part in custom.split(","):
                part = part.strip()
                if part:
                    try:
                        v = float(part)
                        if v > 0 and v not in steps:
                            steps.append(v)
                    except ValueError:
                        pass
        return sorted(steps)

    # ── 浏览 ──

    def _on_browse_src(self):
        path, _ = QFileDialog.getOpenFileName(
            self, self.tr("选择源 CSV 文件"), "",
            self.tr("CSV 文件 (*.csv);;所有文件 (*)"))
        if path:
            self._edit_src.setText(path)

    def _on_browse_dir(self):
        d = QFileDialog.getExistingDirectory(self, self.tr("选择输出目录"))
        if d:
            self._edit_dir.setText(d)

    # ── 源文件变化 ──

    def _on_src_changed(self):
        path = self._edit_src.text().strip()
        if path and os.path.exists(path):
            try:
                from src.step_resampler import _read_all
                _, theta, phi, sfreqs = _read_all(path)
                freqs = list(sfreqs.values())[0] if sfreqs else []
                t_step = theta[1] - theta[0] if len(theta) > 1 else "?"
                p_step = phi[1] - phi[0] if len(phi) > 1 else "?"
                self._src_theta = theta
                self._src_phi = phi
                self._src_step = t_step if isinstance(t_step, (int, float)) else 1.0
                self._lbl_info.setText(
                    f"θ={theta[0]:.0f}~{theta[-1]:.0f}° (步进{t_step}°), "
                    f"φ={phi[0]:.0f}~{phi[-1]:.0f}° (步进{p_step}°), {len(freqs)} 频点")
                self._btn_run.setEnabled(True)
            except Exception as e:
                self._lbl_info.setText(f"{self.tr('读取失败')}: {e}")
                self._btn_run.setEnabled(False)
        else:
            self._lbl_info.setText("")
            self._btn_run.setEnabled(False)
        self._update_preview()

    # ── 预览 ──

    def _update_preview(self):
        path = self._edit_src.text().strip()
        steps = self._get_selected_steps()

        stem = Path(path).stem if path else ""
        self._preview_table.setRowCount(0)

        if not steps:
            return

        has_src = path and os.path.exists(path)
        theta_count = len(self._src_theta) if has_src and hasattr(self, '_src_theta') else 0
        phi_count = len(self._src_phi) if has_src and hasattr(self, '_src_phi') else 0

        self._preview_table.setRowCount(len(steps))
        for ri, s in enumerate(steps):
            s_str = str(int(s)) if s == int(s) else str(s).replace(".", "p")
            fname = f"{stem}_step{s_str}deg.csv" if stem else ""

            self._preview_table.setItem(ri, 0, QTableWidgetItem(f"{s}°"))
            if has_src:
                t_n = len(self._src_theta[::max(1, int(round(s / self._src_step)))])
                p_n = 360  # phi always 360 points for 0-359
                self._preview_table.setItem(ri, 1, QTableWidgetItem(str(t_n)))
                self._preview_table.setItem(ri, 2, QTableWidgetItem(str(p_n)))
            else:
                self._preview_table.setItem(ri, 1, QTableWidgetItem("—"))
                self._preview_table.setItem(ri, 2, QTableWidgetItem("—"))
            self._preview_table.setItem(ri, 3, QTableWidgetItem(fname))

        if self._preview_table.rowCount() > 0:
            self._preview_table.resizeColumnsToContents()
            self._btn_run.setEnabled(has_src)
        else:
            self._btn_run.setEnabled(False)

    # ── 执行 ──

    def _on_run(self):
        path = self._edit_src.text().strip()
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, self.tr("提示"), self.tr("请选择有效的源 CSV 文件。"))
            return

        steps = self._get_selected_steps()
        if not steps:
            QMessageBox.warning(self, self.tr("提示"), self.tr("请选择目标步进值。"))
            return

        out_dir = self._edit_dir.text().strip()
        if not out_dir:
            out_dir = str(Path(path).parent)
        os.makedirs(out_dir, exist_ok=True)

        try:
            from src.step_resampler import batch_resample
            self._btn_run.setEnabled(False)
            self._btn_run.setText(self.tr("处理中..."))
            QApplication.processEvents()

            outputs = batch_resample(path, out_dir, steps)

            self._btn_run.setText(self.tr("▶ 开始批量导出"))
            self._btn_run.setEnabled(True)
            QMessageBox.information(self, self.tr("完成"),
                f"{self.tr('成功导出')} {len(outputs)} {self.tr('个文件')}:\n" +
                "\n".join(f"  • {Path(o).name}" for o in outputs))
        except Exception as e:
            self._btn_run.setText(self.tr("▶ 开始批量导出"))
            self._btn_run.setEnabled(True)
            QMessageBox.critical(self, self.tr("错误"), f"{self.tr('重采样失败')}: {e}")


def _parse_steps(text: str) -> List[float]:
    """解析步进字符串: '5, 10, 15' → [5.0, 10.0, 15.0]"""
    steps = []
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            val = float(part)
            if val > 0:
                steps.append(val)
        except ValueError:
            pass
    return steps


# ═══════════════════════════════════════════════════════════════
# 批量数据检查与转换对话框
# ═══════════════════════════════════════════════════════════════

class BatchCalibrateDialog(QDialog):
    """批量检查CSV格式 → 实部/虚部格式自动发现 → 可选RSP校准 → 一键转换。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(self.tr("数据检查与转换"))
        self.setMinimumSize(700, 550)
        self._paths: List[str] = []
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # ── 文件列表 ──
        grp_files = QGroupBox(self.tr("待检查的 CSV 文件"))
        fl = QVBoxLayout(grp_files)
        br = QHBoxLayout()
        btn_add = QPushButton(self.tr("📂 添加文件..."))
        btn_add.clicked.connect(self._on_add_files)
        br.addWidget(btn_add)
        btn_clr = QPushButton(self.tr("清除"))
        btn_clr.clicked.connect(self._on_clear)
        br.addWidget(btn_clr)
        br.addStretch()
        self._lbl_count = QLabel("")
        br.addWidget(self._lbl_count)
        fl.addLayout(br)

        self._file_table = QTableWidget()
        self._file_table.setColumnCount(3)
        self._file_table.setHorizontalHeaderLabels(
            [self.tr("文件名"), self.tr("格式"), self.tr("状态")])
        self._file_table.horizontalHeader().setStretchLastSection(True)
        self._file_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._file_table.setMaximumHeight(180)
        fl.addWidget(self._file_table)
        layout.addWidget(grp_files)

        # ── RSP 校准 ──
        grp_rsp = QGroupBox(self.tr("RSP 路径损耗校准 (可选 — 仅对实部/虚部文件生效)"))
        rl = QFormLayout(grp_rsp)
        rl.setSpacing(6)

        hr = QHBoxLayout()
        self._edit_rsp_h = QLineEdit()
        self._edit_rsp_h.setPlaceholderText(self.tr("H-pol RSP CSV..."))
        hr.addWidget(self._edit_rsp_h, 1)
        bh = QPushButton(self.tr("浏览..."))
        bh.clicked.connect(lambda: self._browse_rsp("h"))
        hr.addWidget(bh)
        rl.addRow("H-pol:", hr)

        vr = QHBoxLayout()
        self._edit_rsp_v = QLineEdit()
        self._edit_rsp_v.setPlaceholderText(self.tr("V-pol RSP CSV..."))
        vr.addWidget(self._edit_rsp_v, 1)
        bv = QPushButton(self.tr("浏览..."))
        bv.clicked.connect(lambda: self._browse_rsp("v"))
        vr.addWidget(bv)
        rl.addRow("V-pol:", vr)
        layout.addWidget(grp_rsp)

        # ── 输出 ──
        grp_out = QGroupBox(self.tr("输出目录"))
        orow = QHBoxLayout(grp_out)
        self._edit_out = QLineEdit()
        self._edit_out.setPlaceholderText(self.tr("默认: 源文件所在目录"))
        orow.addWidget(self._edit_out, 1)
        bo = QPushButton(self.tr("浏览..."))
        bo.clicked.connect(self._on_browse_out)
        orow.addWidget(bo)
        layout.addWidget(grp_out)

        layout.addStretch()

        # ── 按钮 ──
        brow = QHBoxLayout()
        self._btn_check = QPushButton(self.tr("🔍 扫描格式"))
        self._btn_check.clicked.connect(self._on_scan)
        brow.addWidget(self._btn_check)
        self._btn_run = QPushButton(self.tr("▶ 开始转换"))
        self._btn_run.clicked.connect(self._on_run)
        self._btn_run.setMinimumHeight(36)
        self._btn_run.setEnabled(False)
        brow.addWidget(self._btn_run)
        brow.addStretch()
        bc = QPushButton(self.tr("关闭"))
        bc.clicked.connect(self.reject)
        brow.addWidget(bc)
        layout.addLayout(brow)

    # ── 文件 ──

    def _on_add_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, self.tr("选择要检查的 CSV 文件 (可多选)"), "",
            self.tr("CSV 文件 (*.csv);;所有文件 (*)"))
        if paths:
            for p in paths:
                if p not in self._paths:
                    self._paths.append(p)
            self._refresh_table()
            if not self._edit_out.text().strip() and self._paths:
                self._edit_out.setText(str(Path(self._paths[0]).parent))

    def _on_clear(self):
        self._paths.clear()
        self._file_table.setRowCount(0)
        self._lbl_count.setText("")
        self._btn_run.setEnabled(False)

    def _refresh_table(self):
        from src.raw_converter import _detect_format
        self._file_table.setRowCount(len(self._paths))
        self._lbl_count.setText(f"{len(self._paths)} {self.tr('个文件')}")
        has_aborted = False
        for i, p in enumerate(self._paths):
            self._file_table.setItem(i, 0, QTableWidgetItem(Path(p).name))
            try:
                fmt = _detect_format(p)
                if fmt == 'standard':
                    self._file_table.setItem(i, 1, QTableWidgetItem(
                        self.tr("对数域 (LogMag/Phase)")))
                    self._file_table.setItem(i, 2, QTableWidgetItem("✅ OK"))
                elif fmt == 'aborted':
                    has_aborted = True
                    self._file_table.setItem(i, 1, QTableWidgetItem(
                        self.tr("实部/虚部 (Real/Imag)")))
                    self._file_table.setItem(i, 2, QTableWidgetItem(
                        self.tr("⚠ 需转换")))
                else:
                    self._file_table.setItem(i, 1, QTableWidgetItem(
                        self.tr("未知")))
                    self._file_table.setItem(i, 2, QTableWidgetItem("❓"))
            except Exception:
                self._file_table.setItem(i, 1, QTableWidgetItem("?"))
                self._file_table.setItem(i, 2, QTableWidgetItem(
                    self.tr("读取失败")))
        self._file_table.resizeColumnsToContents()
        self._btn_run.setEnabled(has_aborted)

    # ── 扫描 ──

    def _on_scan(self):
        if not self._paths:
            self._on_add_files()
        else:
            self._refresh_table()

    # ── RSP ──

    def _browse_rsp(self, pol: str):
        path, _ = QFileDialog.getOpenFileName(
            self, self.tr("选择 RSP 校准文件"), "",
            self.tr("CSV 文件 (*.csv);;所有文件 (*)"))
        if path:
            if pol == "h":
                self._edit_rsp_h.setText(path)
            else:
                self._edit_rsp_v.setText(path)

    def _on_browse_out(self):
        d = QFileDialog.getExistingDirectory(self, self.tr("选择输出目录"))
        if d:
            self._edit_out.setText(d)

    # ── 执行 ──

    def _on_run(self):
        aborted = [p for p in self._paths if self._check_fmt(p) == 'aborted']
        if not aborted:
            QMessageBox.information(self, self.tr("提示"),
                self.tr("没有需要转换的文件 (全部已是标准格式)。"))
            return

        out_dir = self._edit_out.text().strip()
        if not out_dir and aborted:
            out_dir = str(Path(aborted[0]).parent)
        os.makedirs(out_dir, exist_ok=True)

        rsp_h = self._edit_rsp_h.text().strip() or None
        rsp_v = self._edit_rsp_v.text().strip() or None

        # RSP 频率覆盖检查
        if rsp_h or rsp_v:
            from src.raw_converter import parse_rsp_csv, batch_check_rsp_coverage
            rh = parse_rsp_csv(rsp_h) if rsp_h else {}
            rv = parse_rsp_csv(rsp_v) if rsp_v else {}
            cov = batch_check_rsp_coverage(aborted, rh, rv, only_fmt='aborted')
            if not cov.ok:
                warn = (self.tr("RSP 频率范围不足:\n") +
                    f"H-pol: {cov.rsp_h_bounds}\nV-pol: {cov.rsp_v_bounds}\n\n" +
                    "\n".join(cov.warnings[:10]) +
                    self.tr("\n\n继续使用边界值外推？"))
                if QMessageBox.question(self, self.tr("⚠ RSP 频率范围不足"), warn,
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
                    return

        self._btn_run.setEnabled(False)
        self._btn_run.setText(self.tr("⏳ 转换中..."))
        QApplication.processEvents()

        try:
            from src.raw_converter import batch_check_and_convert
            result = batch_check_and_convert(
                self._paths, out_dir, rsp_h_path=rsp_h, rsp_v_path=rsp_v)
            ok = len(result['converted'])
            fail = len(result['failed'])
            summary = f"{self.tr('转换完成')}:\n\n✅ {self.tr('成功')}: {ok}\n❌ {self.tr('失败')}: {fail}"
            if ok > 0:
                summary += f"\n\n{self.tr('输出目录')}:\n{out_dir}"
            if fail > 0:
                summary += "\n\n{self.tr('失败详情')}:\n"
                summary += "\n".join(
                    f"  • {Path(f['source']).name}: {f['error']}" for f in result['failed'])
            QMessageBox.information(self, self.tr("完成"), summary)
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, self.tr("错误"), f"{self.tr('转换失败')}: {e}")
        finally:
            self._btn_run.setText(self.tr("▶ 开始转换"))
            self._btn_run.setEnabled(True)

    @staticmethod
    def _check_fmt(path: str) -> str:
        try:
            from src.raw_converter import _detect_format
            return _detect_format(path)
        except Exception:
            return 'unknown'


# ═══════════════════════════════════════════════════════════════
# 数据合并对话框
# ═══════════════════════════════════════════════════════════════

class MergeDialog(QDialog):
    """多段数据拼接: 选文件 → 设RSP校准 → 输出路径 → 一键合并。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(self.tr("数据合并 (多段拼接)"))
        self.setMinimumSize(660, 520)
        self._rsp_h_path: str = ""
        self._rsp_v_path: str = ""
        self._paths: List[str] = []
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # ── 1. 源文件列表 ──
        grp_files = QGroupBox(self.tr("要合并的 CSV 文件 (至少2个)"))
        files_layout = QVBoxLayout(grp_files)
        btn_row = QHBoxLayout()
        btn_add = QPushButton(self.tr("📂 添加文件..."))
        btn_add.clicked.connect(self._on_add_files)
        btn_row.addWidget(btn_add)
        btn_clear = QPushButton(self.tr("清除"))
        btn_clear.clicked.connect(self._on_clear_files)
        btn_row.addWidget(btn_clear)
        btn_row.addStretch()
        self._lbl_count = QLabel("")
        btn_row.addWidget(self._lbl_count)
        files_layout.addLayout(btn_row)

        self._file_list = QListWidget()
        files_layout.addWidget(self._file_list)
        layout.addWidget(grp_files)

        # ── 2. RSP 路径损耗校准 ──
        grp_rsp = QGroupBox(self.tr("RSP 路径损耗校准 (可选)"))
        rsp_layout = QFormLayout(grp_rsp)
        rsp_layout.setSpacing(6)

        h_row = QHBoxLayout()
        self._edit_rsp_h = QLineEdit()
        self._edit_rsp_h.setPlaceholderText(self.tr("H-pol RSP CSV..."))
        h_row.addWidget(self._edit_rsp_h, 1)
        btn_h = QPushButton(self.tr("浏览..."))
        btn_h.clicked.connect(lambda: self._browse_rsp("h"))
        h_row.addWidget(btn_h)
        btn_h_pre = QPushButton(self.tr("从预设选择..."))
        btn_h_pre.clicked.connect(lambda: self._pick_rsp_preset("h"))
        h_row.addWidget(btn_h_pre)
        rsp_layout.addRow("H-pol:", h_row)

        v_row = QHBoxLayout()
        self._edit_rsp_v = QLineEdit()
        self._edit_rsp_v.setPlaceholderText(self.tr("V-pol RSP CSV..."))
        v_row.addWidget(self._edit_rsp_v, 1)
        btn_v = QPushButton(self.tr("浏览..."))
        btn_v.clicked.connect(lambda: self._browse_rsp("v"))
        v_row.addWidget(btn_v)
        btn_v_pre = QPushButton(self.tr("从预设选择..."))
        btn_v_pre.clicked.connect(lambda: self._pick_rsp_preset("v"))
        v_row.addWidget(btn_v_pre)
        rsp_layout.addRow("V-pol:", v_row)

        self._lbl_rsp_info = QLabel(
            self.tr("校准仅对实部/虚部格式文件生效。对数域文件不需要。"))
        self._lbl_rsp_info.setStyleSheet("color: #888; font-size:0.9em;")
        rsp_layout.addRow(self._lbl_rsp_info)
        layout.addWidget(grp_rsp)

        # ── 3. 输出 ──
        grp_out = QGroupBox(self.tr("输出文件"))
        out_row = QHBoxLayout(grp_out)
        self._edit_out = QLineEdit()
        self._edit_out.setPlaceholderText(self.tr("默认: 首个文件所在目录/merged.csv"))
        out_row.addWidget(self._edit_out, 1)
        btn_out = QPushButton(self.tr("浏览..."))
        btn_out.clicked.connect(self._on_browse_out)
        out_row.addWidget(btn_out)
        layout.addWidget(grp_out)

        # ── 4. 信息 ──
        self._lbl_info = QLabel("")
        self._lbl_info.setStyleSheet("color: #666;")
        layout.addWidget(self._lbl_info)

        layout.addStretch()

        # ── 5. 按钮 ──
        btn_row2 = QHBoxLayout()
        self._btn_run = QPushButton(self.tr("▶ 开始合并"))
        self._btn_run.clicked.connect(self._on_run)
        self._btn_run.setMinimumHeight(36)
        self._btn_run.setEnabled(False)
        btn_row2.addWidget(self._btn_run)
        btn_row2.addStretch()
        btn_close = QPushButton(self.tr("关闭"))
        btn_close.clicked.connect(self.reject)
        btn_row2.addWidget(btn_close)
        layout.addLayout(btn_row2)

    # ── 文件操作 ──

    def _on_add_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, self.tr("选择要合并的 CSV 文件 (可多选)"), "",
            self.tr("CSV 文件 (*.csv);;所有文件 (*)"))
        if paths:
            for p in paths:
                if p not in self._paths:
                    self._paths.append(p)
                    self._file_list.addItem(str(Path(p).name))
            self._update_info()
            self._set_default_output()

    def _on_clear_files(self):
        self._paths.clear()
        self._file_list.clear()
        self._update_info()

    def _set_default_output(self):
        if self._paths and not self._edit_out.text().strip():
            out = str(Path(self._paths[0]).parent / "merged.csv")
            self._edit_out.setText(out)

    # ── RSP ──

    def _browse_rsp(self, pol: str):
        path, _ = QFileDialog.getOpenFileName(
            self, self.tr("选择 RSP 校准文件"), "",
            self.tr("CSV 文件 (*.csv);;所有文件 (*)"))
        if path:
            if pol == "h":
                self._edit_rsp_h.setText(path)
                self._rsp_h_path = path
            else:
                self._edit_rsp_v.setText(path)
                self._rsp_v_path = path

    def _pick_rsp_preset(self, pol: str):
        """从 RSP 预设中选择。"""
        from ui.rsp_picker_dialog import RspPickerDialog
        h_path, v_path = RspPickerDialog.pick(self)
        if pol == "h" and h_path:
            self._edit_rsp_h.setText(h_path)
            self._rsp_h_path = h_path
        elif pol == "v" and v_path:
            self._edit_rsp_v.setText(v_path)
            self._rsp_v_path = v_path

    # ── 输出 ──

    def _on_browse_out(self):
        path, _ = QFileDialog.getSaveFileName(
            self, self.tr("保存合并结果"), "",
            self.tr("CSV 文件 (*.csv)"))
        if path:
            self._edit_out.setText(path)

    # ── 信息更新 ──

    def _update_info(self):
        n = len(self._paths)
        self._lbl_count.setText(f"{n} {self.tr('个文件')}")
        self._btn_run.setEnabled(n >= 2)

        # 检测是否有实部/虚部格式文件
        if self._paths:
            try:
                from src.raw_converter import _detect_format
                has_ri = any(_detect_format(p) == 'aborted' for p in self._paths)
                if has_ri:
                    self._lbl_info.setText(
                        self.tr("⚠ 检测到实部/虚部格式文件。建议加载 RSP 路径损耗校准文件。"))
                    self._lbl_info.setStyleSheet("color: #e67e22;")
                else:
                    self._lbl_info.setText(self.tr("✓ 全部为对数域格式文件，无需 RSP 校准。"))
                    self._lbl_info.setStyleSheet("color: #27ae60;")
            except Exception:
                self._lbl_info.setText("")

    # ── 执行 ──

    def _on_run(self):
        if len(self._paths) < 2:
            QMessageBox.warning(self, self.tr("提示"), self.tr("请至少选择2个文件。"))
            return

        out = self._edit_out.text().strip()
        if not out:
            out = str(Path(self._paths[0]).parent / "merged.csv")
        os.makedirs(str(Path(out).parent), exist_ok=True)

        rsp_h = self._edit_rsp_h.text().strip() or None
        rsp_v = self._edit_rsp_v.text().strip() or None

        # 如果加载了 RSP，检查频率覆盖
        if rsp_h or rsp_v:
            from src.raw_converter import parse_rsp_csv, batch_check_rsp_coverage
            rsp_h_data = parse_rsp_csv(rsp_h) if rsp_h else {}
            rsp_v_data = parse_rsp_csv(rsp_v) if rsp_v else {}
            cov = batch_check_rsp_coverage(self._paths, rsp_h_data, rsp_v_data, only_fmt='aborted')
            if not cov.ok:
                warn_text = (
                    self.tr("RSP 频率范围不足:\n") +
                    f"H-pol: {cov.rsp_h_bounds}\nV-pol: {cov.rsp_v_bounds}\n\n" +
                    "\n".join(cov.warnings[:10]) +
                    ("\n..." if len(cov.warnings) > 10 else "") +
                    self.tr("\n\n继续使用边界值外推？")
                )
                reply = QMessageBox.question(
                    self, self.tr("⚠ RSP 频率范围不足"), warn_text,
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply != QMessageBox.Yes:
                    return

        self._btn_run.setEnabled(False)
        self._btn_run.setText(self.tr("⏳ 合并中..."))
        QApplication.processEvents()

        try:
            from src.raw_converter import merge_csv_files
            result = merge_csv_files(self._paths, out,
                rsp_h_path=rsp_h, rsp_v_path=rsp_v)
            QMessageBox.information(self, self.tr("完成"),
                f"{self.tr('合并完成')}:\n{result}")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, self.tr("错误"), f"{self.tr('合并失败')}: {e}")
        finally:
            self._btn_run.setText(self.tr("▶ 开始合并"))
            self._btn_run.setEnabled(True)


# ═══════════════════════════════════════════════════════════════
# 数据修复对话框
# ═══════════════════════════════════════════════════════════════

class RepairDialog(QDialog):
    """数据修复 — 扫描→预览→选方法→执行。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(self.tr("数据修复"))
        self.setMinimumSize(700, 550)
        self._paths: List[str] = []
        self._scan_results: List[dict] = []
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        # ── 文件 ──
        grp = QGroupBox(self.tr("待修复的 CSV 文件"))
        fl = QVBoxLayout(grp)
        br = QHBoxLayout()
        btn_add = QPushButton(self.tr("📂 添加文件..."))
        btn_add.clicked.connect(self._on_add_files)
        br.addWidget(btn_add)
        btn_clr = QPushButton(self.tr("清除"))
        btn_clr.clicked.connect(self._on_clear)
        br.addWidget(btn_clr)
        self._lbl_count = QLabel("")
        br.addWidget(self._lbl_count)
        br.addStretch()
        btn_scan = QPushButton(self.tr("🔍 扫描数据质量"))
        btn_scan.clicked.connect(self._on_scan)
        br.addWidget(btn_scan)
        fl.addLayout(br)
        layout.addWidget(grp)

        # ── 扫描结果表 ──
        self._result_table = QTableWidget()
        self._result_table.setColumnCount(5)
        self._result_table.setHorizontalHeaderLabels([
            self.tr("文件"), self.tr("格式"), self.tr("坏点数"),
            self.tr("坏点位置"), self.tr("建议方法")
        ])
        self._result_table.horizontalHeader().setStretchLastSection(True)
        self._result_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._result_table.setMaximumHeight(200)
        self._result_table.setAlternatingRowColors(True)
        layout.addWidget(self._result_table)

        # ── 修复方法 ──
        grp_method = QGroupBox(self.tr("修复方法"))
        ml = QVBoxLayout(grp_method)
        self._rb_mad = QRadioButton(
            self.tr("MAD 异常检测 — 中位数绝对偏差，适合标准格式"))
        self._rb_mad.setChecked(True)
        ml.addWidget(self._rb_mad)
        self._rb_q25 = QRadioButton(
            self.tr("Q25 比率检测 — 四分位数比率，适合异常终止格式"))
        ml.addWidget(self._rb_q25)
        self._rb_knn = QRadioButton(
            self.tr("KNN 插值修复 — 逆距离加权 K 近邻插值，通用"))
        ml.addWidget(self._rb_knn)
        self._rb_manual = QRadioButton(
            self.tr("手动指定 phi — 直接输入需修复的 phi 索引"))
        ml.addWidget(self._rb_manual)
        self._edit_manual_phi = QLineEdit()
        self._edit_manual_phi.setPlaceholderText(self.tr("如: 5, 7, 9 (逗号分隔, 0-based)"))
        self._edit_manual_phi.setEnabled(False)
        self._rb_manual.toggled.connect(lambda c: self._edit_manual_phi.setEnabled(c))
        ml.addWidget(self._edit_manual_phi)
        layout.addWidget(grp_method)

        # ── 输出目录 ──
        out_row = QHBoxLayout()
        out_row.addWidget(QLabel(self.tr("输出目录:")))
        self._edit_out = QLineEdit()
        self._edit_out.setPlaceholderText(self.tr("默认: 源文件目录 (文件名 _repaired.csv)"))
        out_row.addWidget(self._edit_out)
        btn_out = QPushButton(self.tr("浏览..."))
        btn_out.clicked.connect(self._on_browse_out)
        out_row.addWidget(btn_out)
        layout.addLayout(out_row)

        # ── 状态信息（standalone 模式反馈） ──
        self._lbl_status = QLabel("")
        self._lbl_status.setStyleSheet("color: #666; padding: 2px 0;")
        layout.addWidget(self._lbl_status)

        # ── 按钮 ──
        brow = QHBoxLayout()
        brow.addStretch()
        self._btn_run = QPushButton(self.tr("▶ 执行修复"))
        self._btn_run.clicked.connect(self._on_run)
        self._btn_run.setMinimumHeight(36)
        self._btn_run.setEnabled(False)
        brow.addWidget(self._btn_run)
        bc = QPushButton(self.tr("关闭"))
        bc.clicked.connect(self.reject)
        brow.addWidget(bc)
        layout.addLayout(brow)

    def _on_add_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, self.tr("选择要修复的 CSV 文件"), "",
            self.tr("CSV 文件 (*.csv);;所有文件 (*)"))
        if paths:
            for p in paths:
                if p not in self._paths:
                    self._paths.append(p)
            self._lbl_count.setText(f"{len(self._paths)} {self.tr('个文件')}")
            self._btn_run.setEnabled(True)

    def _on_clear(self):
        self._paths.clear()
        self._scan_results.clear()
        self._result_table.setRowCount(0)
        self._lbl_count.setText("")
        self._btn_run.setEnabled(False)

    def _on_scan(self):
        """扫描文件，检测坏点。"""
        if not self._paths:
            QMessageBox.warning(self, self.tr("提示"), self.tr("请先添加文件。"))
            return
        from src.data_quality import detect_phi_anomalies
        from src.raw_converter import _detect_format
        self._scan_results.clear()
        for p in self._paths:
            try:
                fmt = _detect_format(p)
                anomalies = detect_phi_anomalies(p)
                bad_phis = sorted(anomalies.get("phi_indices", []))
                suggested = anomalies.get("suggested_method", "MAD")
                self._scan_results.append({
                    "path": p, "format": fmt,
                    "bad_count": len(bad_phis),
                    "bad_phis": bad_phis,
                    "suggested": suggested,
                })
            except Exception as e:
                self._scan_results.append({
                    "path": p, "format": "error",
                    "bad_count": 0, "bad_phis": [], "suggested": "",
                    "error": str(e),
                })
        self._refresh_scan_table()

    def _refresh_scan_table(self):
        self._result_table.setRowCount(len(self._scan_results))
        for i, r in enumerate(self._scan_results):
            self._result_table.setItem(i, 0, QTableWidgetItem(Path(r["path"]).name))
            fmt_text = r.get("format", "?")
            if "error" in r:
                fmt_text = f"❌ {r['error']}"
            self._result_table.setItem(i, 1, QTableWidgetItem(fmt_text))
            self._result_table.setItem(i, 2, QTableWidgetItem(str(r["bad_count"])))
            bad_str = ", ".join(str(b) for b in r["bad_phis"][:10])
            if len(r["bad_phis"]) > 10:
                bad_str += "..."
            self._result_table.setItem(i, 3, QTableWidgetItem(bad_str or "—"))
            self._result_table.setItem(i, 4, QTableWidgetItem(r.get("suggested", "—")))
        self._result_table.resizeColumnsToContents()
        self._btn_run.setEnabled(True)

    def _on_browse_out(self):
        d = QFileDialog.getExistingDirectory(self, self.tr("选择输出目录"))
        if d:
            self._edit_out.setText(d)

    def _on_run(self):
        if not self._paths:
            return
        output_dir = self._edit_out.text().strip()
        # 确定修复方法
        method = "mad"
        force_phis = None
        if self._rb_q25.isChecked():
            method = "q25"
        elif self._rb_knn.isChecked():
            method = "knn"
        elif self._rb_manual.isChecked():
            method = "manual"
            try:
                force_phis = [int(x.strip()) for x in self._edit_manual_phi.text().split(",") if x.strip()]
            except ValueError:
                QMessageBox.warning(self, self.tr("输入错误"),
                    self.tr("手动 phi 格式无效，请用逗号分隔数字。"))
                return

        self._btn_run.setEnabled(False)
        self._btn_run.setText(self.tr("⏳ 修复中..."))
        QApplication.processEvents()
        ok, fail = 0, 0
        try:
            from src.data_quality import auto_detect_and_repair
            for p in self._paths:
                pobj = Path(p)
                out = str(Path(output_dir or pobj.parent) / f"{pobj.stem}_repaired.csv")
                try:
                    r = auto_detect_and_repair(
                        p, out, method=method, force_phis=force_phis)
                    ok += 1
                except Exception as e:
                    fail += 1
                    self._log(f"✗ {pobj.name}: {e}")
            msg = self.tr("修复完成: {} 成功, {} 失败").format(ok, fail)
            QMessageBox.information(self, self.tr("完成"), msg)
            if fail == 0:
                self.accept()
        except Exception as e:
            QMessageBox.critical(self, self.tr("错误"),
                self.tr(f"修复失败: {e}"))
        finally:
            self._btn_run.setText(self.tr("▶ 执行修复"))

    def _log(self, msg: str):
        parent_mw = self.parent()
        if hasattr(parent_mw, '_log'):
            parent_mw._log(msg)
        if hasattr(self, '_lbl_status') and self._lbl_status:
            self._lbl_status.setText(msg)
            from PySide6.QtWidgets import QApplication
            QApplication.processEvents()



class ActivationDialog(QDialog):
    """在线激活对话框。

    用户输入激活码 → 调用激活服务器 → 获取许可文件 → 保存到本地。
    """

    ACTIVATION_CODE_PLACEHOLDER = "XXXX-XXXX-XXXX-XXXX"  # noqa: S105 — placeholder, not a real key

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("软件激活")
        self.setMinimumWidth(480)
        self._activated = False
        self._setup_ui()

    # ── 属性 ──

    @property
    def is_activated(self) -> bool:
        return self._activated

    # ── UI ──

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # 说明文字
        info_label = QLabel(
            "<h3>🔑 软件激活</h3>"
            "<p>请输入从供应商获取的激活码。<br>"
            "激活需要网络连接以验证激活码并获取许可文件。</p>"
        )
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        # 激活码输入
        code_layout = QFormLayout()
        self._edit_code = QLineEdit()
        self._edit_code.setPlaceholderText(self.tr(self.ACTIVATION_CODE_PLACEHOLDER))
        self._edit_code.setMaxLength(30)
        self._edit_code.textChanged.connect(self._on_code_changed)
        code_layout.addRow(self.tr("激活码:"), self._edit_code)
        layout.addLayout(code_layout)

        # 服务器配置（可折叠）
        server_group = QGroupBox(self.tr("激活服务器设置（高级）"))
        server_layout = QFormLayout(server_group)
        self._edit_server = QLineEdit()
        self._edit_server.setPlaceholderText("http://activation.antenna-pp.local:8899")
        try:
            from src.activation import get_server_url
            current = get_server_url()
            if current:
                self._edit_server.setText(current)
        except Exception:
            pass
        server_layout.addRow(self.tr("服务器 URL:"), self._edit_server)
        server_group.setVisible(False)
        layout.addWidget(server_group)
        self._server_group = server_group

        # 展开/折叠服务器设置
        toggle_btn = QPushButton(self.tr("⚙ 服务器设置..."))
        toggle_btn.setFlat(True)
        toggle_btn.clicked.connect(lambda: server_group.setVisible(not server_group.isVisible()))
        layout.addWidget(toggle_btn)

        # 进度条（激活过程）
        self._progress = QProgressBar()
        self._progress.setRange(0, 0)  # indeterminate
        self._progress.setVisible(False)
        layout.addWidget(self._progress)

        # 状态标签
        self._lbl_status = QLabel("")
        self._lbl_status.setWordWrap(True)
        self._lbl_status.setStyleSheet("color: #666;")
        layout.addWidget(self._lbl_status)

        layout.addStretch()

        # 按钮
        btn_layout = QHBoxLayout()
        self._btn_activate = QPushButton(self.tr("🔓 激活"))
        self._btn_activate.setEnabled(False)
        self._btn_activate.clicked.connect(self._on_activate)
        btn_layout.addStretch()
        btn_layout.addWidget(self._btn_activate)
        layout.addLayout(btn_layout)

    # ── 逻辑 ──

    def _on_code_changed(self, text: str):
        self._btn_activate.setEnabled(len(text.strip()) >= 8)

    def _on_activate(self):
        code = self._edit_code.text().strip()
        if not code:
            return

        # 保存服务器 URL
        from src.activation import set_server_url, get_machine_id
        server = self._edit_server.text().strip()
        if server:
            set_server_url(server)

        # 开始激活
        self._btn_activate.setEnabled(False)
        self._edit_code.setEnabled(False)
        self._progress.setVisible(True)
        self._lbl_status.setText(self.tr("正在连接激活服务器..."))
        QApplication.processEvents()

        from src.activation import activate
        machine_id = get_machine_id()
        ok, result = activate(code, machine_id, server if server else None)

        self._progress.setVisible(False)
        if ok:
            self._lbl_status.setStyleSheet("color: #2e7d32; font-weight: bold;")
            self._lbl_status.setText(
                self.tr(f"✅ 激活成功！\n许可已保存到: {result}")
            )
            self._activated = True
            QMessageBox.information(
                self, self.tr("激活成功"),
                self.tr("许可已安装。程序将在重新启动后生效。")
            )
            self.accept()
        else:
            self._lbl_status.setStyleSheet("color: #c62828; font-weight: bold;")
            self._lbl_status.setText(self.tr(f"❌ 激活失败: {result}"))
            self._btn_activate.setEnabled(True)
            self._edit_code.setEnabled(True)

# ═══════════════════════════════════════════════════════════════
# PathLossDialog — 路径损耗补偿（独立于数据检查与转换）
# ═══════════════════════════════════════════════════════════════

class PathLossDialog(QDialog):
    """路径损耗补偿: 检查文件格式 + 选RSP校准 + 转换+补偿。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(self.tr("路径损耗补偿"))
        self.setMinimumSize(650, 500)
        self._paths: List[str] = []
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        # ── 文件列表 ──
        grp_files = QGroupBox(self.tr("CSV 数据文件"))
        fl = QVBoxLayout(grp_files)
        br = QHBoxLayout()
        btn_add = QPushButton(self.tr("📂 添加文件..."))
        btn_add.clicked.connect(self._on_add_files)
        br.addWidget(btn_add)
        btn_clr = QPushButton(self.tr("清除"))
        btn_clr.clicked.connect(self._on_clear)
        br.addWidget(btn_clr)
        br.addStretch()
        self._lbl_count = QLabel("")
        br.addWidget(self._lbl_count)
        fl.addLayout(br)
        self._file_table = QTableWidget()
        self._file_table.setColumnCount(3)
        self._file_table.setHorizontalHeaderLabels([
            self.tr("文件名"), self.tr("格式"), self.tr("状态")])
        self._file_table.horizontalHeader().setStretchLastSection(True)
        self._file_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._file_table.setMaximumHeight(160)
        fl.addWidget(self._file_table)
        layout.addWidget(grp_files)

        # ── RSP 校准文件 ──
        grp_rsp = QGroupBox(self.tr("RSP 路径损耗校准"))
        rl = QFormLayout(grp_rsp)
        rl.setSpacing(6)

        h_row = QHBoxLayout()
        self._edit_rsp_h = QLineEdit()
        self._edit_rsp_h.setPlaceholderText(self.tr("选择 H-pol RSP 校准文件..."))
        h_row.addWidget(self._edit_rsp_h)
        btn_h = QPushButton(self.tr("浏览..."))
        btn_h.clicked.connect(lambda: self._browse_rsp("h"))
        h_row.addWidget(btn_h)
        btn_h_pre = QPushButton(self.tr("从预设选择..."))
        btn_h_pre.clicked.connect(lambda: self._pick_rsp_preset("h"))
        h_row.addWidget(btn_h_pre)
        rl.addRow("H-pol:", h_row)

        v_row = QHBoxLayout()
        self._edit_rsp_v = QLineEdit()
        self._edit_rsp_v.setPlaceholderText(self.tr("选择 V-pol RSP 校准文件..."))
        v_row.addWidget(self._edit_rsp_v)
        btn_v = QPushButton(self.tr("浏览..."))
        btn_v.clicked.connect(lambda: self._browse_rsp("v"))
        v_row.addWidget(btn_v)
        btn_v_pre = QPushButton(self.tr("从预设选择..."))
        btn_v_pre.clicked.connect(lambda: self._pick_rsp_preset("v"))
        v_row.addWidget(btn_v_pre)
        rl.addRow("V-pol:", v_row)

        layout.addWidget(grp_rsp)

        # ── 处理选项 ──
        grp_opt = QGroupBox(self.tr("处理选项"))
        ol = QVBoxLayout(grp_opt)
        self._chk_convert = QCheckBox(
            self.tr("自动转换实部/虚部文件（必须先转换为对数域）"))
        self._chk_convert.setChecked(True)
        ol.addWidget(self._chk_convert)
        self._chk_apply = QCheckBox(self.tr("应用路径损耗补偿"))
        self._chk_apply.setChecked(True)
        ol.addWidget(self._chk_apply)
        layout.addWidget(grp_opt)

        # ── 输出目录 ──
        out_row = QHBoxLayout()
        out_row.addWidget(QLabel(self.tr("输出目录:")))
        self._edit_out = QLineEdit()
        self._edit_out.setPlaceholderText(self.tr("默认: 源文件目录"))
        out_row.addWidget(self._edit_out)
        btn_out = QPushButton(self.tr("浏览..."))
        btn_out.clicked.connect(self._on_browse_out)
        out_row.addWidget(btn_out)
        layout.addLayout(out_row)

        # ── 按钮 ──
        btn_row = QHBoxLayout()
        self._btn_scan = QPushButton(self.tr("🔍 检查兼容性"))
        self._btn_scan.clicked.connect(self._on_scan)
        btn_row.addWidget(self._btn_scan)
        btn_row.addStretch()
        self._btn_run = QPushButton(self.tr("▶ 执行"))
        self._btn_run.clicked.connect(self._on_run)
        self._btn_run.setEnabled(False)
        self._btn_run.setMinimumHeight(36)
        btn_row.addWidget(self._btn_run)
        btn_close = QPushButton(self.tr("关闭"))
        btn_close.clicked.connect(self.reject)
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)

    def _on_add_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, self.tr("选择 CSV 文件"), "",
            self.tr("CSV 文件 (*.csv);;所有文件 (*)"))
        if paths:
            for p in paths:
                if p not in self._paths:
                    self._paths.append(p)
            self._refresh_table()
            if not self._edit_out.text().strip() and self._paths:
                self._edit_out.setText(str(Path(self._paths[0]).parent))

    def _on_clear(self):
        self._paths.clear()
        self._file_table.setRowCount(0)
        self._lbl_count.setText("")
        self._btn_run.setEnabled(False)

    def _refresh_table(self):
        from src.raw_converter import _detect_format
        self._file_table.setRowCount(len(self._paths))
        self._lbl_count.setText(f"{len(self._paths)} {self.tr('个文件')}")
        has_aborted = False
        for i, p in enumerate(self._paths):
            self._file_table.setItem(i, 0, QTableWidgetItem(Path(p).name))
            try:
                fmt = _detect_format(p)
                if fmt == 'standard':
                    self._file_table.setItem(i, 1, QTableWidgetItem(
                        self.tr("对数域 (LogMag/Phase)")))
                    self._file_table.setItem(i, 2, QTableWidgetItem("✅"))
                elif fmt == 'aborted':
                    has_aborted = True
                    self._file_table.setItem(i, 1, QTableWidgetItem(
                        self.tr("实部/虚部 (Real/Imag)")))
                    self._file_table.setItem(i, 2, QTableWidgetItem(
                        self.tr("⚠ 需转换")))
                else:
                    self._file_table.setItem(i, 1, QTableWidgetItem(self.tr("未知")))
                    self._file_table.setItem(i, 2, QTableWidgetItem("❓"))
            except Exception:
                self._file_table.setItem(i, 1, QTableWidgetItem("?"))
                self._file_table.setItem(i, 2, QTableWidgetItem(self.tr("读取失败")))
        self._file_table.resizeColumnsToContents()
        self._btn_run.setEnabled(True)

    def _on_scan(self):
        if not self._paths:
            self._on_add_files()
        else:
            self._refresh_table()

    def _browse_rsp(self, pol: str):
        path, _ = QFileDialog.getOpenFileName(
            self, self.tr("选择 RSP 校准文件"), "",
            self.tr("CSV 文件 (*.csv);;所有文件 (*)"))
        if path:
            if pol == "h":
                self._edit_rsp_h.setText(path)
            else:
                self._edit_rsp_v.setText(path)

    def _pick_rsp_preset(self, pol: str):
        """从 RSP 预设中选择。"""
        from ui.rsp_picker_dialog import RspPickerDialog
        h_path, v_path = RspPickerDialog.pick(self)
        if pol == "h" and h_path:
            self._edit_rsp_h.setText(h_path)
        elif pol == "v" and v_path:
            self._edit_rsp_v.setText(v_path)

    def _on_browse_out(self):
        d = QFileDialog.getExistingDirectory(self, self.tr("选择输出目录"))
        if d:
            self._edit_out.setText(d)

    def _on_run(self):
        """执行转换 + RSP 校准。"""
        if not self._paths:
            QMessageBox.warning(self, self.tr("提示"), self.tr("请先添加文件。"))
            return
        from src.raw_converter import _detect_format, convert_file
        rsp_h = self._edit_rsp_h.text().strip() or None
        rsp_v = self._edit_rsp_v.text().strip() or None
        output_dir = self._edit_out.text().strip() or str(Path(self._paths[0]).parent)
        os.makedirs(output_dir, exist_ok=True)
        convert_needed = self._chk_convert.isChecked()
        apply_rsp = self._chk_apply.isChecked()

        results = {"ok": 0, "fail": 0, "errors": []}
        for p in self._paths:
            try:
                fmt = _detect_format(p)
                fname = Path(p).name
                # Step 1: 如果是不完整格式且勾选了自动转换
                if fmt == 'aborted' and convert_needed:
                    out_name = f"converted_{fname}"
                    out_path = str(Path(output_dir) / out_name)
                    convert_file(p, out_path)
                    self._log(f"✓ {fname} → {out_name}")
                    if apply_rsp:
                        p = out_path  # 后续用转换后的文件做 RSP
                    else:
                        results["ok"] += 1
                        continue
                elif fmt == 'aborted' and not convert_needed:
                    results["fail"] += 1
                    results["errors"].append(f"{fname}: {self.tr('需先转换')}")
                    continue

                # Step 2: 应用 RSP 校准
                if apply_rsp and (rsp_h or rsp_v):
                    from src.rsp_calibration import parse_rsp_csv, parse_rsp_phase, _apply_rsp_calibration
                    import re
                    is_hpol = bool(re.search(r'[Hh].*[Pp]ol|[Pp]ol.*[Hh]|[Hh]_', fname))
                    rsp_path = rsp_h if is_hpol else (rsp_v or rsp_h)
                    if rsp_path:
                        rsp_amp = parse_rsp_csv(rsp_path)
                        rsp_phase_data = parse_rsp_phase(rsp_path) if rsp_path else {}
                        out_name = f"calibrated_{fname}"
                        out_path = str(Path(output_dir) / out_name)
                        _apply_rsp_calibration(p, rsp_amp, rsp_phase_data, out_path)
                        self._log(f"✓ {fname} → {out_name} (RSP 已应用)")
                    else:
                        self._log(f"  {fname}: {self.tr('无对应 RSP 文件，跳过')}")
                results["ok"] += 1
            except Exception as e:
                results["fail"] += 1
                results["errors"].append(f"{Path(p).name}: {e}")

        msg = self.tr("处理完成: {} 成功, {} 失败").format(results["ok"], results["fail"])
        if results["errors"]:
            msg += "\n\n" + self.tr("错误详情:\n") + "\n".join(results["errors"][:5])
        QMessageBox.information(self, self.tr("完成"), msg)
        self.accept()

    def _log(self, msg: str):
        """简易日志（打印到父窗口日志或 stderr）。"""
        parent_mw = self.parent()
        if hasattr(parent_mw, '_log'):
            parent_mw._log(msg)
        else:
            print(msg)
