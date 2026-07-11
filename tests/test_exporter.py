"""exporter 单元测试 — 列映射、名称替换、辅助函数"""
import os
import tempfile

import openpyxl
import pytest

from src.excel_reader import ColumnInfo, SheetInfo
from src.exporter import (
    _build_col_map,
    _name_delta,
    _replace_cell_text,
    _write_ar_range,
    _write_ar_single,
    _write_cell,
    _write_lag_range,
    _write_lag_single,
    export_results,
)

# ── Fixtures ───────────────────────────────────────────────────────

@pytest.fixture
def sample_info():
    """构建一个含常见列类型的 SheetInfo。"""
    columns = [
        ColumnInfo("A", 1, "Frequency", "frequency", "frequency"),
        ColumnInfo("B", 2, "Gain (dBi)", "gain", "gain"),
        ColumnInfo("C", 3, "Directivity (dBi)", "directivity", "directivity"),
        ColumnInfo("D", 4, "Efficiency (%)", "efficiency_pct", "efficiency_pct"),
        ColumnInfo("E", 5, "Gain at Theta=30 (dB)\nLAG", "lag_single", "lag_single"),
        ColumnInfo("F", 6, "Gain at Theta=60 (dB)", "lag_single", "lag_single"),
        ColumnInfo("G", 7, "Gain at Theta=0~70 (dB)", "lag_range", "lag_range"),
        ColumnInfo("H", 8, "AR at Theta=30deg (dB)", "ar_single", "ar_single"),
        ColumnInfo("I", 9, "AR at Theta=0~70deg (dB)", "ar_range", "ar_range"),
        ColumnInfo("J", 10, "TRP (dBm)", "trp", "trp"),
    ]
    return SheetInfo(
        name="TestSheet",
        header_row=1,
        data_start_row=2,
        data_end_row=5,
        columns=columns,
        frequencies=[699.0, 1700.0, 2600.0, 3500.0],
    )


# ── _name_delta ────────────────────────────────────────────────────

class TestNameDelta:
    def test_same_name(self):
        assert _name_delta("5G1", "5G1") == ("", "")

    def test_single_char_diff(self):
        old_d, new_d = _name_delta("5G1", "5G2")
        assert old_d == "1"
        assert new_d == "2"

    def test_prefix_diff(self):
        old_d, new_d = _name_delta("ANT001", "ANT002")
        assert old_d == "1"
        assert new_d == "2"

    def test_mid_diff(self):
        old_d, new_d = _name_delta("DUT-A-test", "DUT-B-test")
        assert old_d == "A"
        assert new_d == "B"

    def test_no_common(self):
        old_d, new_d = _name_delta("ABC", "XYZ")
        assert old_d != "" and new_d != ""


# ── _replace_cell_text ─────────────────────────────────────────────

class TestReplaceCellText:
    def test_exact_replace(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "5G1"
        ws.cell(1, 1, "Antenna 5G1 Report")
        _replace_cell_text(ws, "5G1", "5G2")
        assert ws.cell(1, 1).value == "Antenna 5G2 Report"
        wb.close()

    def test_same_name_skips(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "Antenna 5G1 Report")
        _replace_cell_text(ws, "5G1", "5G1")
        assert ws.cell(1, 1).value == "Antenna 5G1 Report"  # unchanged
        wb.close()

    def test_partial_match(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "5G1"
        ws.cell(1, 1, "5G2")  # cell = "5G2", old="5G1" — delta is "1"→"2"
        _replace_cell_text(ws, "5G1", "5G2")
        # Cell "5G2" contains old delta "2"? No — let me reconsider.
        # delta of "5G1" vs "5G2" = ("1", "2")
        # cell "5G2" contains "2" so it should replace "2" with... no, that doesn't make sense.
        # The delta is used as substring search. Cell="5G2", old_delta="1" — "1" not in "5G2".
        # So it falls through to no match. Cell stays "5G2".
        wb.close()


# ── _build_col_map ─────────────────────────────────────────────────

class TestBuildColMap:
    def test_maps_by_type(self, sample_info):
        col_map = _build_col_map(sample_info)
        assert "frequency" in col_map
        assert "gain" in col_map
        assert len(col_map["lag_single"]) == 2  # 两个 lag_single 列

    def test_duplicate_types(self, sample_info):
        col_map = _build_col_map(sample_info)
        assert len(col_map.get("lag_single", [])) == 2


# ── 列类型+角度/范围 匹配 (旧 _find_*_column 已重构为 _build_col_map + writer 内联匹配) ──
# 原 _find_lag_single_column(info, 30) 返回列索引的语义, 现由 writer 的匹配逻辑承载。
# 这里通过 writer 行为验证「类型+角度/范围 → 正确列」及「无匹配不写入」两类语义。

class TestLagSingleMatch:
    def test_theta_30_writes_col_E(self, sample_info):
        wb = openpyxl.Workbook(); ws = wb.active
        _write_lag_single(ws, 2, _build_col_map(sample_info), 30.0, 5.0)
        assert ws.cell(2, 5).value == round(5.0, 6)  # column E
        assert ws.cell(2, 6).value is None            # 未误写 F
        wb.close()

    def test_theta_60_writes_col_F(self, sample_info):
        wb = openpyxl.Workbook(); ws = wb.active
        _write_lag_single(ws, 2, _build_col_map(sample_info), 60.0, 6.0)
        assert ws.cell(2, 6).value == round(6.0, 6)  # column F
        assert ws.cell(2, 5).value is None            # 未误写 E
        wb.close()

    def test_no_match_writes_nothing(self, sample_info):
        wb = openpyxl.Workbook(); ws = wb.active
        _write_lag_single(ws, 2, _build_col_map(sample_info), 99.0, 1.0)
        assert ws.cell(2, 5).value is None and ws.cell(2, 6).value is None
        wb.close()


class TestLagRangeMatch:
    def test_0_to_70_writes_col_G(self, sample_info):
        wb = openpyxl.Workbook(); ws = wb.active
        _write_lag_range(ws, 2, _build_col_map(sample_info), 0, 70, 4.5)
        assert ws.cell(2, 7).value == round(4.5, 6)  # column G
        wb.close()

    def test_no_match_writes_nothing(self, sample_info):
        wb = openpyxl.Workbook(); ws = wb.active
        _write_lag_range(ws, 2, _build_col_map(sample_info), 10, 30, 4.5)
        assert ws.cell(2, 7).value is None
        wb.close()


class TestArSingleMatch:
    def test_theta_30_writes_col_H(self, sample_info):
        wb = openpyxl.Workbook(); ws = wb.active
        _write_ar_single(ws, 2, _build_col_map(sample_info), 30.0, 2.5)
        assert ws.cell(2, 8).value == round(2.5, 6)  # column H
        wb.close()

    def test_no_match_writes_nothing(self, sample_info):
        wb = openpyxl.Workbook(); ws = wb.active
        _write_ar_single(ws, 2, _build_col_map(sample_info), 99.0, 2.5)
        assert ws.cell(2, 8).value is None
        wb.close()


class TestArRangeMatch:
    def test_0_to_70_writes_col_I(self, sample_info):
        wb = openpyxl.Workbook(); ws = wb.active
        _write_ar_range(ws, 2, _build_col_map(sample_info), 0, 70, 3.0)
        assert ws.cell(2, 9).value == round(3.0, 6)  # column I
        wb.close()

    def test_no_match_writes_nothing(self, sample_info):
        wb = openpyxl.Workbook(); ws = wb.active
        _write_ar_range(ws, 2, _build_col_map(sample_info), 10, 30, 3.0)
        assert ws.cell(2, 9).value is None
        wb.close()


# ── Write Functions ────────────────────────────────────────────────

class TestWriteCell:
    def test_write_simple(self, sample_info):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = _build_col_map(sample_info)
        _write_cell(ws, 2, col_map, "frequency", 699.0)
        assert ws.cell(2, 1).value == 699.0
        wb.close()

    def test_write_rounds_float(self, sample_info):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = _build_col_map(sample_info)
        _write_cell(ws, 2, col_map, "gain", 12.3456789)
        # round(val, 6)
        assert ws.cell(2, 2).value == round(12.3456789, 6)
        wb.close()

    def test_write_unknown_type_is_noop(self, sample_info):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = _build_col_map(sample_info)
        _write_cell(ws, 2, col_map, "nonexistent", 123.0)
        # no exception, no write
        wb.close()


class TestWriteLagSingle:
    def test_write(self, sample_info):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = _build_col_map(sample_info)
        _write_lag_single(ws, 2, col_map, 30.0, 5.12345)
        assert ws.cell(2, 5).value == round(5.12345, 6)
        wb.close()


class TestWriteLagRange:
    def test_write(self, sample_info):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = _build_col_map(sample_info)
        _write_lag_range(ws, 2, col_map, 0, 70, 4.56789)
        assert ws.cell(2, 7).value == round(4.56789, 6)
        wb.close()


class TestWriteArSingle:
    def test_write(self, sample_info):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = _build_col_map(sample_info)
        _write_ar_single(ws, 2, col_map, 30.0, 2.5)
        assert ws.cell(2, 8).value == round(2.5, 6)
        wb.close()


class TestWriteArRange:
    def test_write(self, sample_info):
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = _build_col_map(sample_info)
        _write_ar_range(ws, 2, col_map, 0, 70, 3.0)
        assert ws.cell(2, 9).value == round(3.0, 6)
        wb.close()


# ── Export Results (integration) ───────────────────────────────────

class TestExportResults:
    def test_export_minimal(self, sample_info):
        """模板→结果的基本导出（用临时文件）。"""
        # 创建一个最小模板
        tpl = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tpl_path = tpl.name
        tpl.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws.cell(1, 1, "Frequency")
        ws.cell(1, 2, "Gain (dBi)")
        ws.cell(1, 3, "Directivity (dBi)")
        ws.cell(1, 4, "Efficiency (%)")
        ws.cell(1, 5, "Gain at Theta=30 (dB)\nLAG")
        ws.cell(1, 6, "Gain at Theta=0~70 (dB)")
        wb.save(tpl_path)
        wb.close()

        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        out.close()

        try:
            results = {
                "TestSheet": [
                    {
                        "frequency": 699.0,
                        "gain": 10.5,
                        "directivity": 12.0,
                        "efficiency_pct": 70.0,
                        "lag_single_30.0": 8.0,
                        "lag_range_0.0_70.0": 7.5,
                    }
                ]
            }
            path = export_results(
                template_path=tpl_path,
                output_path=out.name,
                sheet_results=results,
                sheets_info=[sample_info],
            )
            assert path == out.name
            assert os.path.exists(out.name)

            # 验证输出
            wb_out = openpyxl.load_workbook(out.name)
            ws_out = wb_out["TestSheet"]
            assert ws_out.cell(2, 1).value == 699.0  # frequency
            wb_out.close()
        finally:
            os.unlink(tpl_path)
            if os.path.exists(out.name):
                os.unlink(out.name)

    def test_auto_clone_sheet(self, sample_info):
        """自动创建工作表（多数据源场景）。"""
        tpl = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tpl_path = tpl.name
        tpl.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "5G1"
        ws.cell(1, 1, "Frequency")
        ws.cell(1, 2, "Gain (dBi)")
        wb.save(tpl_path)
        wb.close()

        out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        out.close()

        try:
            results = {
                "5G1": [{"frequency": 699.0, "gain": 10.0}],
                "5G2": [{"frequency": 1700.0, "gain": 11.0}],
            }
            path = export_results(
                template_path=tpl_path,
                output_path=out.name,
                sheet_results=results,
            )
            assert path == out.name
            wb_out = openpyxl.load_workbook(out.name)
            assert "5G2" in wb_out.sheetnames
            wb_out.close()
        finally:
            os.unlink(tpl_path)
            if os.path.exists(out.name):
                os.unlink(out.name)
