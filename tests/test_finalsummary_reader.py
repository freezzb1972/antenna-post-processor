"""FinalSummarySource 单元测试 — 自适应结构探测 + 频点读取"""
import os
import tempfile
from pathlib import Path

import numpy as np
import openpyxl
import pytest

from src.finalsummary_reader import (
    FinalSummarySource,
    _freq_sheet_name,
    _is_numeric,
    _read_matrix,
    _to_float,
)

# ── Type Helpers ───────────────────────────────────────────────────

class TestIsNumeric:
    def test_int(self):
        assert _is_numeric(42)

    def test_float(self):
        assert _is_numeric(3.14)

    def test_numeric_string(self):
        assert _is_numeric("123.45")

    def test_non_numeric_string(self):
        assert not _is_numeric("hello")

    def test_none(self):
        assert not _is_numeric(None)

    def test_empty_string(self):
        assert not _is_numeric("")


class TestToFloat:
    def test_int(self):
        assert _to_float(42) == 42.0

    def test_float(self):
        assert _to_float(3.14) == 3.14

    def test_string(self):
        assert _to_float("99.5") == 99.5

    def test_none(self):
        assert _to_float(None) is None

    def test_non_numeric_string(self):
        assert _to_float("hello") is None


class TestFreqSheetName:
    def test_integer_freq(self):
        assert _freq_sheet_name(1154.0) == "1154"

    def test_fractional_freq(self):
        assert _freq_sheet_name(699.5) == "699.5"


# ── Matrix Reader ──────────────────────────────────────────────────

class TestReadMatrix:
    def test_read_small_matrix(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        # _read_matrix reads from min_col=2, so col B=1st value, col C=2nd, etc.
        # Row 1: B=1.0, C=2.0, D=3.0
        # Row 2: B=4.0, C=5.0, D=6.0
        # Row 3: B=7.0, C=8.0, D=9.0
        for r in range(1, 4):
            for c in range(1, 4):
                ws.cell(r, c + 1, float((r - 1) * 3 + c))

        result = _read_matrix(ws, 1, 3, 3)
        assert result.shape == (3, 3)
        assert result[0, 0] == 1.0
        assert result[2, 2] == 9.0
        wb.close()

    def test_read_with_none_cells(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 2, 10.0)
        ws.cell(1, 3, None)  # empty
        ws.cell(2, 2, 20.0)

        result = _read_matrix(ws, 1, 2, 2)
        assert result[0, 0] == 10.0
        assert np.isnan(result[0, 1])  # None → NaN
        assert result[1, 0] == 20.0
        wb.close()


# ── FinalSummarySource with synthetic data ─────────────────────────

def _make_finalsummary_xlsx(freqs=None, with_phase=True, with_phi=True):
    """构造一个最小 FinalSummary Excel 用于测试。"""
    if freqs is None:
        freqs = [699.0, 1700.0]

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for freq in freqs:
        sn = _freq_sheet_name(freq)
        ws = wb.create_sheet(title=sn)

        # Row 1-2: description
        ws.cell(1, 1, "Antenna Test Report")
        ws.cell(2, 1, f"Frequency: {freq} MHz")

        # Row 3: Theta header
        ws.cell(3, 1, "Theta/Phi")
        for ti, theta in enumerate([0, 30, 60, 90]):
            ws.cell(3, 2 + ti, float(theta))

        # Rows 4-5: Theta data (2 phi points)
        ws.cell(4, 1, 0.0)
        ws.cell(4, 2, -5.0)
        ws.cell(4, 3, -4.0)
        ws.cell(4, 4, -3.0)
        ws.cell(4, 5, -2.0)
        ws.cell(5, 1, 180.0)
        ws.cell(5, 2, -8.0)
        ws.cell(5, 3, -7.0)
        ws.cell(5, 4, -6.0)
        ws.cell(5, 5, -5.0)

        row = 6

        if with_phase:
            ws.cell(row, 1, "Phase")
            row += 1
            # Phase header row
            ws.cell(row, 1, "Theta/Phi")
            for ti in range(4):
                ws.cell(row, 2 + ti, float([0, 30, 60, 90][ti]))
            row += 1
            # Phase data
            ws.cell(row, 1, 0.0)
            ws.cell(row, 2, 10.0)
            ws.cell(row, 3, 20.0)
            ws.cell(row, 4, 30.0)
            ws.cell(row, 5, 40.0)
            row += 1
            ws.cell(row, 1, 180.0)
            ws.cell(row, 2, 50.0)
            ws.cell(row, 3, 60.0)
            ws.cell(row, 4, 70.0)
            ws.cell(row, 5, 80.0)
            row += 1

        if with_phi:
            ws.cell(row, 1, "Phi Polarization")
            row += 1
            ws.cell(row, 1, "Power")
            row += 1
            ws.cell(row, 1, "Theta/Phi")
            for ti in range(4):
                ws.cell(row, 2 + ti, float([0, 30, 60, 90][ti]))
            row += 1
            ws.cell(row, 1, 0.0)
            ws.cell(row, 2, -6.0)
            ws.cell(row, 3, -5.0)
            ws.cell(row, 4, -4.0)
            ws.cell(row, 5, -3.0)
            row += 1
            ws.cell(row, 1, 180.0)
            ws.cell(row, 2, -9.0)
            ws.cell(row, 3, -8.0)
            ws.cell(row, 4, -7.0)
            ws.cell(row, 5, -6.0)
            row += 1

            if with_phase:
                ws.cell(row, 1, "Phase")
                row += 1
                ws.cell(row, 1, "Theta/Phi")
                for ti in range(4):
                    ws.cell(row, 2 + ti, float([0, 30, 60, 90][ti]))
                row += 1
                ws.cell(row, 1, 0.0)
                ws.cell(row, 2, 100.0)
                ws.cell(row, 3, 110.0)
                ws.cell(row, 4, 120.0)
                ws.cell(row, 5, 130.0)
                row += 1
                ws.cell(row, 1, 180.0)
                ws.cell(row, 2, 140.0)
                ws.cell(row, 3, 150.0)
                ws.cell(row, 4, 160.0)
                ws.cell(row, 5, 170.0)

    wb.save(tmp.name)
    wb.close()
    return tmp.name


class TestFinalSummarySourceSynthetic:
    def test_frequencies(self):
        path = _make_finalsummary_xlsx()
        try:
            ds = FinalSummarySource(path)
            assert ds.frequencies == [699.0, 1700.0]
            ds.close()
        finally:
            os.unlink(path)

    def test_theta_angles(self):
        path = _make_finalsummary_xlsx()
        try:
            ds = FinalSummarySource(path)
            assert ds.theta_angles == [0.0, 30.0, 60.0, 90.0]
            ds.close()
        finally:
            os.unlink(path)

    def test_phi_angles(self):
        path = _make_finalsummary_xlsx()
        try:
            ds = FinalSummarySource(path)
            assert ds.phi_angles == [0.0, 180.0]
            ds.close()
        finally:
            os.unlink(path)

    def test_read_sections(self):
        path = _make_finalsummary_xlsx()
        try:
            ds = FinalSummarySource(path)
            result = ds.read_sections(0)  # first freq
            assert "theta_logmag" in result
            assert result["theta_logmag"].shape == (2, 4)
            assert result["theta_logmag"][0, 0] == -5.0
            assert result["phi_logmag"][0, 0] == -6.0
            ds.close()
        finally:
            os.unlink(path)

    def test_read_sections_has_phase(self):
        path = _make_finalsummary_xlsx(with_phase=True)
        try:
            ds = FinalSummarySource(path)
            result = ds.read_sections(0)
            assert result["theta_phase"] is not None
            assert result["phi_phase"] is not None
            ds.close()
        finally:
            os.unlink(path)

    def test_read_sections_no_phase(self):
        path = _make_finalsummary_xlsx(with_phase=False)
        try:
            ds = FinalSummarySource(path)
            result = ds.read_sections(0)
            assert result["theta_phase"] is None
            ds.close()
        finally:
            os.unlink(path)

    def test_read_batch(self):
        path = _make_finalsummary_xlsx(freqs=[699.0, 700.0, 701.0])
        try:
            ds = FinalSummarySource(path)
            result = ds.read_batch([0, 1, 2])
            assert len(result) == 3
            assert 699.0 in result
            assert 700.0 in result
            assert 701.0 in result
            ds.close()
        finally:
            os.unlink(path)

    def test_cache_reuse(self):
        path = _make_finalsummary_xlsx()
        try:
            ds = FinalSummarySource(path)
            r1 = ds.read_sections(0)
            r2 = ds.read_sections(0)  # should hit cache
            assert np.array_equal(r1["theta_logmag"], r2["theta_logmag"])
            ds.close()
        finally:
            os.unlink(path)

    def test_no_digit_sheets_raises(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"  # not a number
        wb.save(tmp.name)
        wb.close()
        try:
            with pytest.raises(ValueError, match="数字命名"):
                FinalSummarySource(tmp.name)
        finally:
            os.unlink(tmp.name)

    def test_close(self):
        path = _make_finalsummary_xlsx()
        ds = FinalSummarySource(path)
        ds.close()
        ds.close()  # should be safe to call twice


# ── Real Data ──────────────────────────────────────────────────────

class TestFinalSummarySourceRealData:
    @pytest.fixture
    def ds_5g1(self):
        p = Path(__file__).parent.parent / "data" / "5G1FinalSummary.xlsx"
        if not p.exists():
            pytest.skip("5G1FinalSummary.xlsx not found")
        return FinalSummarySource(str(p))

    def test_has_frequencies(self, ds_5g1):
        assert len(ds_5g1.frequencies) > 0

    def test_read_sections_no_crash(self, ds_5g1):
        result = ds_5g1.read_sections(0)
        assert isinstance(result["theta_logmag"], np.ndarray)
        assert result["theta_logmag"].ndim == 2
