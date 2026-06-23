#!/usr/bin/env python3
"""
AFN 天线参数汇总处理
====================
使用 antenna-post-processor 的 run_pipeline() 处理 L5+L1 原始数据,
计算 Summary.xlsx 中列出的全部参数,结果写入 Summary.xlsx。

依赖 antenna-post-processor 管线:
  - FinalSummarySource 加载 RawData
  - run_pipeline() 执行全部计算
  - sheet_results 填入 Summary.xlsx
"""

import sys
import time
import openpyxl
import numpy as np

from src.finalsummary_reader import FinalSummarySource
from src.pipeline import run_pipeline
from src.lag_config import LagConfig

# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------
L5_DATA  = "data/AFN/NO2_L5_RawData.xlsx"    # L5: 1154-1224 MHz (71 频点)
L1_DATA  = "data/AFN/NO2_L1_RawData.xlsx"    # L1: 1549-1616 MHz (68 频点)
TEMPLATE = "data/template_AFN_L1.xlsx"        # 管线模板
SUMMARY  = "data/AFN/Summary.xlsx"            # 输出目标

# Summary.xlsx 中需要计算的参数 (Row → pipeline key)
ROW_MAP = {
    2:  "trp",                # Tot. Rad. Pwr. (dBm)
    3:  "peak_eirp",          # Peak EIRP (dBm)
    4:  "directivity",        # Directivity (dBi)
    5:  "efficiency_db",      # Efficiency (dB)
    6:  "efficiency_pct",     # Efficiency (%)
    7:  "gain",               # Gain (dBi)
    8:  "nhprp_45",           # NHPRP +-Pi/4 (dBm)
    9:  "nhprp_30",           # NHPRP +-Pi/6 (dBm)
    10: "nhprp_225",          # NHPRP +-Pi/8 (dBm)
    11: "uh_prp",             # Upper Hem. PRP (dBm)
    12: "lh_prp",             # Lower Hem. PRP (dBm)
    13: "nhprp45_ratio_db",   # NHPRP4 / TRP Ratio (dB)
    14: "nhprp45_ratio_pct",  # NHPRP4 / TRP Ratio (%)
    15: "nhprp_45",           # Near Horz. TRP +-Pi/4 (dBm)
    16: "nhprp30_ratio_db",   # NHPRP6 / TRP Ratio (dB)
    17: "nhprp30_ratio_pct",  # NHPRP6 / TRP Ratio (%)
    18: "nhprp_30",           # Near Horz. TRP +-Pi/6 (dBm)
    19: "nhprp225_ratio_db",  # NHPRP8 / TRP Ratio (dB)
    20: "nhprp225_ratio_pct", # NHPRP8 / TRP Ratio (%)
    21: "nhprp_225",          # Near Horz. TRP +-Pi/8 (dBm)
    22: "uh_ratio_db",        # UHPRP / TRP Ratio (dB)
    23: "uh_ratio_pct",       # UHPRP / TRP Ratio (%)
    24: "uh_prp",             # Upper Hem.Total Radiated Pwr (dBm)
    25: "lh_ratio_db",        # LHPRP / TRP Ratio (dB)
    26: "lh_ratio_pct",       # LHPRP / TRP Ratio (%)
    27: "lh_prp",             # Lower Hem.Total Radiated Pwr (dBm)
    28: "prp_120",            # PRP (dBm) (theta = 0 to 120)
    29: "boresight_phi",      # Boresight Phi (Degrees)
    30: "boresight_theta",    # Boresight Th. (Degrees)
    31: "max_power",          # Maximum Power (dBm)
    32: "min_power",          # Minimum Power (dBm)
    33: "avg_gain",           # Average Gain (dB)
    34: "avg_power",          # Average Power (dBm)
}

ALL_PARAMS = set(ROW_MAP.values())
# 比率参数的基键名（_process_one_frequency 中用基键触发比率计算）
# 注意: 管线内部检查 ratio_need 用的是无后缀的基键
ALL_PARAMS_BASE = ALL_PARAMS | {
    "nhprp45_ratio", "nhprp30_ratio", "nhprp225_ratio",
    "uh_ratio", "lh_ratio",
    "nhpis45_ratio", "nhpis30_ratio", "nhpis225_ratio",
}


def parse_summary_freqs(path: str):
    """从 Summary.xlsx 解析 列号 → 频率 映射"""
    wb = openpyxl.load_workbook(path)
    ws = wb["Attributes"]
    col_freqs = {}
    for col in range(2, ws.max_column + 1):
        hdr = ws.cell(row=1, column=col).value
        if hdr is None:
            continue
        try:
            col_freqs[col] = float(hdr.split()[0])
        except (ValueError, IndexError):
            pass
    wb.close()
    return col_freqs


def main():
    print("=" * 60)
    print("AFN 天线参数汇总 — 使用 antenna-post-processor 管线")
    print("=" * 60)

    # ---- 1. 逐数据源调用 run_pipeline() ----
    all_results = {}  # {freq: {param_key: value}}

    for label, data_path in [("L5", L5_DATA), ("L1", L1_DATA)]:
        print(f"\n--- 处理 {label}: {data_path} ---")
        t0 = time.time()

        ds = FinalSummarySource(data_path)
        freqs = ds.frequencies
        print(f"  频点: {len(freqs)} ({freqs[0]:.0f}-{freqs[-1]:.0f} MHz)")

        # 调用管线 (空 LagConfig: 不计算 LAG/AR, Summary.xlsx 不含这些列)
        try:
            sheet_results = run_pipeline(
                datasource=ds,
                template_path=TEMPLATE,
                output_path="/tmp/_afn_temp_output.xlsx",
                extrapolate_theta=False,
                freq_source="datasource",
                extra_params=ALL_PARAMS_BASE,
                lag_config_override=LagConfig(single_angles=[], ranges=[]),
            )
        except Exception as e:
            print(f"  错误: {type(e).__name__}: {e}")
            ds.close()
            continue

        # 收集所有频点的计算结果
        row_count = 0
        for sheet_name, rows in sheet_results.items():
            for row in rows:
                freq = row.get("frequency")
                if freq is None:
                    continue
                all_results[freq] = {k: v for k, v in row.items()}
                row_count += 1

        ds.close()
        elapsed = time.time() - t0
        print(f"  完成: {row_count} 行, {elapsed:.1f}s")

    print(f"\n总计: {len(all_results)} 个频点有计算结果")

    # ---- 2. 填入 Summary.xlsx ----
    print(f"\n--- 填入 {SUMMARY} ---")
    col_freqs = parse_summary_freqs(SUMMARY)
    wb = openpyxl.load_workbook(SUMMARY)
    ws = wb["Attributes"]

    filled = 0
    missing = []
    for col, target_freq in sorted(col_freqs.items()):
        # 最近邻匹配 (容差 ±1 MHz, 因为 Summary.xlsx 含整数 MHz)
        if target_freq in all_results:
            result = all_results[target_freq]
        else:
            best = min(all_results.keys(), key=lambda f: abs(f - target_freq))
            if abs(best - target_freq) <= 1.0:
                result = all_results[best]
            else:
                missing.append(f"  {target_freq} MHz (Col {col})")
                continue

        for row_num, key in ROW_MAP.items():
            if key in result:
                ws.cell(row=row_num, column=col).value = result[key]

        filled += 1

    wb.save(SUMMARY)
    wb.close()
    print(f"  填入 {filled}/{len(col_freqs)} 个频率列")

    if missing:
        print(f"  无匹配数据 ({len(missing)}):")
        for m in missing:
            print(m)

    print(f"\n{'=' * 60}")
    print(f"处理完成: {filled} 频点 → {SUMMARY}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
