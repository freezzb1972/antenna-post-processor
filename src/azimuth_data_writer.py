"""
方位面切面 — 中间数据 Excel 导出
================================
将每频点选定 Theta 角度的 Gain/AR 值导出到 Excel，
便于用户验证或二次绘图。

每频点一个 worksheet（名 = 频率整数如 "1154"），
列 = Theta 角度，行 = Phi 坐标 (0-360°)。
"""

from __future__ import annotations

import os

import numpy as np
import openpyxl


def write_azimuth_data(
    freq_data: list[tuple[float, dict[float, np.ndarray]]],
    output_path: str,
    value_label: str = "Gain (dBi)",
) -> None:
    """将方位面切面中间数据写入 Excel。

    Args:
        freq_data: [(freq_mhz, {theta_deg: values_over_phi}), ...]
                    values_over_phi 是形状 (n_phi,) 的 1D 数组
        output_path: 输出 .xlsx 路径
        value_label: 值标签 (如 "Gain (dBi)", "AR (dB)")
    """
    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    for freq_mhz, theta_data in freq_data:
        if not theta_data:
            continue

        # worksheet 名 = 频率整数
        ws_name = str(int(round(freq_mhz)))

        # 确保 worksheet 名唯一（多 batch 可能重名）
        base_name = ws_name
        counter = 1
        while ws_name in wb.sheetnames:
            ws_name = f"{base_name}_{counter}"
            counter += 1

        ws = wb.create_sheet(title=ws_name)

        # 排序 theta 角度
        sorted_thetas = sorted(theta_data.keys())
        n_phi = len(next(iter(theta_data.values())))

        # Row 1: 列头
        ws.cell(row=1, column=1, value="Phi (°)")
        for ti, theta_deg in enumerate(sorted_thetas):
            ws.cell(row=1, column=ti + 2, value=f"{theta_deg:.0f}°")

        # Data rows (phi 0..360)
        for phi_idx in range(n_phi):
            row_num = phi_idx + 2
            ws.cell(row=row_num, column=1, value=phi_idx)
            for ti, theta_deg in enumerate(sorted_thetas):
                val = theta_data[theta_deg][phi_idx]
                if np.isfinite(val):
                    ws.cell(row=row_num, column=ti + 2,
                            value=round(float(val), 6))
                else:
                    ws.cell(row=row_num, column=ti + 2, value="")

        # 列宽
        ws.column_dimensions['A'].width = 8
        for ti in range(len(sorted_thetas)):
            col_letter = openpyxl.utils.get_column_letter(ti + 2)
            ws.column_dimensions[col_letter].width = 12

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    wb.close()
