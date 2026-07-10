"""
中间数据导出 (泛化, 注册表驱动)
================================
把每频点的计算中间量写入 Excel, 供人工核对"哪一步算错了"。

设计: 中间数据只有 3 种形态, 每种一个写入器, 靠 row 字段命名约定 + 注册表驱动。
新增参数类型只需向对应注册表加一行, 无需改写入代码。

- KIND 1 完整矩阵 (φ×θ): Gain/AR/RHCP/LHCP/CP-XPI → `<Label>_<源sheet>`
- KIND 2 峰值 vs φ (θ≤N 归约): 自动发现 `_gain_pk_{N}_db` → `PkGain_<源sheet>`
- KIND 3 角度切片+结果值: LAG/AR@θ/RHCP@θ/CP-XPI@θ → `<Label>_<源sheet>`
  (从源矩阵切固定 θ 的 φ 切片, 与 row 里已算好的结果值并排, 不重算聚合)

驱动源是 row 字段(计算参数), 与图表配置无关: 只勾"中间数据"也能出。
"""

from __future__ import annotations

from typing import Any, Callable

import numpy as np

# ── KIND 1: 完整矩阵 (φ×θ) 注册表 ── (label, row_field, unit)
MATRIX_SPECS: list[tuple[str, str, str]] = [
    ("Gain",   "_chart_gain_dbi", "dBi"),
    ("AR",     "_chart_ar_db",    "dB"),
    ("RHCP",   "_rhcp_gain",      "dBi"),
    ("LHCP",   "_lhcp_gain",      "dBi"),
    ("CP-XPI", "_cp_xpi",         "dB"),
]

# ── KIND 3: 角度切片+结果值 注册表 ── (label, matrix_field, result_key_prefix, unit)
# 从 matrix_field 切固定 θ 的 φ 切片, 结果值取 row["<prefix>_<angle>"]。
ANGLE_SLICE_SPECS: list[tuple[str, str, str, str]] = [
    ("LAG",      "_chart_gain_dbi", "lag_single",    "dB"),
    ("AR@θ",     "_chart_ar_db",    "ar_single",     "dB"),
    ("RHCP@θ",   "_rhcp_gain",      "rhcp_single",   "dBi"),
    ("CP-XPI@θ", "_cp_xpi",         "cp_xpi_single", "dB"),
]


def _safe_sheet(name: str) -> str:
    """Excel sheet 名 ≤31 字符, 去非法字符。"""
    for ch in r'[]:*?/\\':
        name = name.replace(ch, "_")
    return name[:31]


def _get_or_create(wb, name: str):
    name = _safe_sheet(name)
    return wb[name] if name in wb.sheetnames else wb.create_sheet(name)


def _next_row(ws) -> int:
    """在已有数据后追加(空行分隔)。"""
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        return 1
    return ws.max_row + 2


# ── KIND 1 写入器: 一个频点的 φ×θ 矩阵块 ──
def _write_matrix(ws, freq_label: str, unit: str,
                  matrix: np.ndarray, phi: np.ndarray, theta: np.ndarray) -> None:
    r = _next_row(ws)
    ws.cell(r, 1, f"Frequency: {freq_label}  ({unit})"); r += 1
    ws.cell(r, 1, "Phi \\ Theta (°)")
    for ci, tv in enumerate(theta):
        ws.cell(r, ci + 2, round(float(tv), 1))
    r += 1
    for pi in range(matrix.shape[0]):
        ws.cell(r, 1, round(float(phi[pi]), 1) if pi < len(phi) else pi)
        for ci in range(matrix.shape[1]):
            ws.cell(r, ci + 2, round(float(matrix[pi, ci]), 4))
        r += 1


# ── KIND 2 写入器: 峰值 vs φ (θ≤N) ──
def _write_peak(ws, freq_label: str, ranges: dict[str, np.ndarray], phi: np.ndarray) -> None:
    """ranges: {N: peak_vs_phi}; 每范围一列, φ 一行。"""
    r = _next_row(ws)
    ws.cell(r, 1, f"Frequency: {freq_label}  (峰值 Gain over θ≤N°, dBi)"); r += 1
    ws.cell(r, 1, "Phi (°)")
    keys = sorted(ranges.keys(), key=lambda x: float(x))
    for ci, k in enumerate(keys):
        ws.cell(r, ci + 2, f"θ≤{k}°")
    r += 1
    n_phi = max((len(v) for v in ranges.values()), default=0)
    for pi in range(n_phi):
        ws.cell(r, 1, round(float(phi[pi]), 1) if pi < len(phi) else pi)
        for ci, k in enumerate(keys):
            v = ranges[k]
            if pi < len(v):
                ws.cell(r, ci + 2, round(float(v[pi]), 4))
        r += 1


# ── KIND 3 写入器: 角度切片 + 结果值 ──
def _write_angle_slice(ws, freq_label: str, label: str, unit: str,
                       angle: float, result_val: Any,
                       slice_vals: np.ndarray, phi: np.ndarray) -> None:
    r = _next_row(ws)
    ws.cell(r, 1, f"Frequency: {freq_label}  |  {label} θ={angle:g}° = {result_val} {unit}"); r += 1
    ws.cell(r, 1, "Phi (°)"); ws.cell(r, 2, f"值 @ θ={angle:g}° ({unit})"); r += 1
    for pi in range(len(slice_vals)):
        ws.cell(r, 1, round(float(phi[pi]), 1) if pi < len(phi) else pi)
        ws.cell(r, 2, round(float(slice_vals[pi]), 4))
        r += 1


def _theta_index(theta: np.ndarray, angle: float) -> int:
    return int(np.argmin(np.abs(theta - angle)))


def _parse_angle(key: str, prefix: str) -> float | None:
    """从 row key 提取角度, 如 'lag_single_30.0' + 'lag_single' → 30.0。"""
    suffix = key[len(prefix) + 1:]  # 去掉 'prefix_'
    try:
        return float(suffix)
    except ValueError:
        return None


def write_intermediate_data(
    sheet_results: dict[str, list[dict[str, Any]]],
    data_path: str,
    log_callback: Callable[[str], None] | None = None,
) -> bool:
    """把 sheet_results 里各频点的中间量写入 Excel。返回是否成功。

    自动裁剪: 只对 row 里实际存在的字段出 sheet。
    """
    import openpyxl as _xl

    def _log(m: str):
        if log_callback:
            log_callback(m)

    wb = _xl.Workbook()
    wb.remove(wb.active)
    index_rows: list[tuple[str, str]] = []  # (sheet 名, 说明)
    produced: set[str] = set()

    try:
        for sn, rows in sheet_results.items():
            for row in rows:
                if row.get("_error"):
                    continue
                theta = row.get("_chart_theta_deg")
                phi = row.get("_chart_phi_deg")
                freq = row.get("frequency")
                if theta is None or phi is None or freq is None:
                    continue
                theta = np.asarray(theta); phi = np.asarray(phi)
                flabel = f"{freq:.0f}MHz"

                # KIND 1: 完整矩阵
                for label, field, unit in MATRIX_SPECS:
                    m = row.get(field)
                    if m is None:
                        continue
                    ws_name = f"{label}_{sn}"
                    _write_matrix(_get_or_create(wb, ws_name), flabel, unit,
                                  np.asarray(m), phi, theta)
                    if ws_name not in produced:
                        produced.add(ws_name)
                        index_rows.append((_safe_sheet(ws_name), f"{label} 完整矩阵 φ×θ ({unit})"))

                # KIND 2: 峰值 vs φ (自动发现 _gain_pk_{N}_db)
                pk = {k[len("_gain_pk_"):-len("_db")]: np.asarray(v)
                      for k, v in row.items()
                      if k.startswith("_gain_pk_") and k.endswith("_db")}
                if pk:
                    ws_name = f"PkGain_{sn}"
                    _write_peak(_get_or_create(wb, ws_name), flabel, pk, phi)
                    if ws_name not in produced:
                        produced.add(ws_name)
                        index_rows.append((_safe_sheet(ws_name), "θ≤N° 峰值 Gain vs φ (dBi)"))

                # KIND 3: 角度切片 + 结果值
                for label, mfield, prefix, unit in ANGLE_SLICE_SPECS:
                    matrix = row.get(mfield)
                    if matrix is None:
                        continue
                    matrix = np.asarray(matrix)
                    angle_keys = [k for k in row
                                  if k.startswith(prefix + "_")
                                  and _parse_angle(k, prefix) is not None]
                    if not angle_keys:
                        continue
                    ws_name = f"{label}_{sn}"
                    ws = _get_or_create(wb, ws_name)
                    for k in sorted(angle_keys, key=lambda x: _parse_angle(x, prefix)):
                        angle = _parse_angle(k, prefix)
                        ti = _theta_index(theta, angle)
                        if ti >= matrix.shape[1]:
                            continue
                        _write_angle_slice(ws, flabel, label, unit, angle,
                                           row.get(k), matrix[:, ti], phi)
                    if ws_name not in produced:
                        produced.add(ws_name)
                        index_rows.append((_safe_sheet(ws_name),
                                          f"{label} 各角度 φ 切片 + 结果值 ({unit})"))

        if not produced:
            _log("  ⚠ 中间数据为空 — 无可导出的计算矩阵")
            return False

        # 索引 sheet (置首)
        idx = wb.create_sheet("索引", 0)
        idx.cell(1, 1, "Worksheet"); idx.cell(1, 2, "内容说明")
        for i, (name, desc) in enumerate(sorted(index_rows), start=2):
            idx.cell(i, 1, name); idx.cell(i, 2, desc)

        wb.save(data_path); wb.close()
        _log(f"  ✓ 中间数据已保存 ({len(produced)} 个 worksheet)")
        return True
    except Exception as e:
        _log(f"  ✗ 中间数据导出失败: {e}")
        return False
