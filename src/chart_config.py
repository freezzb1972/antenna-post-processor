"""
图形配置数据模型
================
ChartConfig 定义需要生成哪些图形、视角参数、输出方式。
支持从模板自动检测图形需求。
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

# ═══════════════════════════════════════════════════════════════
# 标题匹配正则 — 英文 + 中文
# ═══════════════════════════════════════════════════════════════

_CHART_PATTERNS: dict[str, list[str]] = {
    # A 类: 3D 方向图
    "pattern_3d_gain": [
        r"3D.*Gain.*(Pattern|Radiation|方向图)",
        r"3D.*增益.*方向图",
        r"球面.*(增益|Gain).*(方向图|Pattern)",
    ],
    "pattern_3d_eirp": [
        r"3D.*EIRP.*(Pattern|方向图)",
        r"3D.*EIRP.*方向图",
    ],
    "pattern_3d_ar": [
        r"3D.*(AR|Axial|Polarization).*(Pattern|方向图)",
        r"3D.*(轴比|AR|极化).*方向图",
    ],
    # B 类: 频点曲线
    "chart_eff_freq": [
        r"(Eff|Efficiency).*(vs|over|Frequency|频率)",
        r"效率.*(vs|频率)",
    ],
    "chart_gain_freq": [
        r"(Peak\s?)?Gain.*(vs|over|Frequency|频率)",
        r"(峰值\s?)?增益.*(vs|频率)",
    ],
    "chart_dir_freq": [
        r"Directivity.*(vs|over|Frequency|频率)",
        r"方向性.*(vs|频率)",
    ],
    "chart_lag_freq": [
        r"Gain.*Theta.*(vs|over|频率)",
        r"(增益|LAG).*(角度|Theta).*(vs|频率)",
    ],
    "chart_trp_freq": [
        r"TRP((?!NHPRP).)*(vs|over|Frequency|频率)",
        r"TRP.*(vs|频率)",
    ],
    "chart_trp_nhprp": [
        r"TRP.*NHPRP",
        r"TRP.*NHPRP",
    ],
    "chart_ar_freq": [
        r"(AR|Axial)((?!3D).)*(vs|over|Frequency|频率)",
        r"(轴比|AR)((?!3D).)*(vs|频率)",
    ],
    # C 类: 俯仰面切面
    "cut_2d_polar": [
        r"Polar.*(Cut|Plot|切面|图|Elevation)",
        r"极坐标.*(切面|图|俯仰)",
    ],
    "cut_2d_rect": [
        r"(Rect|Cartesian).*(Cut|Plot|切面|图|Elevation)",
        r"直角坐标.*(切面|图|俯仰)",
    ],
}

# 从列类型推导图表 — B 类(频点曲线) + A 类(3D) + C 类(切面)
_COLTYPE_TO_CHART = {
    # B 类: 频点曲线
    "efficiency_pct":  "chart_eff_freq",
    "efficiency_db":   "chart_eff_freq",
    "gain":            "chart_gain_freq|pattern_3d_gain|cut_azimuth_polar",
    "directivity":     "chart_dir_freq",
    "lag_range":       "chart_lag_freq",
    "lag_single":      "chart_lag_freq",
    "trp":             "chart_trp_freq",
    "ar_single":       "chart_ar_freq|pattern_3d_ar|cut_azimuth_polar_ar",
    "ar_range":        "chart_ar_freq|pattern_3d_ar|cut_azimuth_polar_ar",
    "avg_gain":        "chart_gain_freq|pattern_3d_gain",
    "peak_eirp":       "chart_trp_freq|pattern_3d_eirp",
    "nhprp_45":        "chart_trp_nhprp",
    "nhprp_30":        "chart_trp_nhprp",
    # A 类: 3D 方向图
    "rhcp_single":     "pattern_3d_gain|pattern_3d_rhcp|cut_azimuth_polar_rhcp",
    "cp_xpi_single":   "pattern_3d_gain",
    # C 类: 切面图 (由 azimuth config 独立控制, 不在此映射)
}

# 图表自动检测: 从模板参数推断哪些图表应启用
def auto_detect_charts(template_params: set) -> dict[str, bool]:
    """根据模板检测到的参数, 返回应启用的图表 key → True。"""
    result: dict[str, bool] = {}
    for col_type, chart_keys in _COLTYPE_TO_CHART.items():
        if col_type in template_params:
            for ck in chart_keys.split("|"):
                result[ck.strip()] = True
    # C 类切面图: 只要有 Gain 或 AR 就启用 (默认只 Gain 参数)
    if "gain" in template_params or "ar_single" in template_params or "ar_range" in template_params:
        result["cut_2d_polar"] = True
        result["cut_2d_rect"] = True
    return result


@dataclass
class ChartConfig:
    """图形生成配置。

    所有图表分为三类:
      - A 类: 3D 球面方向图（每频点 1 张 PNG）
      - B 类: 频点-参数曲线（openpyxl 原生图表嵌入 Excel）
      - C 类: 俯仰面切面图（每频点 PNG）
    """

    # A 类: 3D 方向图
    pattern_3d_gain: bool = False
    pattern_3d_eirp: bool = False
    pattern_3d_ar: bool = False
    pattern_3d_etheta: bool = False    # E_θ 分量 3D 方向图
    pattern_3d_ephi: bool = False      # E_φ 分量 3D 方向图

    # B 类: 频点曲线
    chart_eff_freq: bool = False
    chart_gain_freq: bool = False
    chart_dir_freq: bool = False
    chart_lag_freq: bool = False
    chart_trp_freq: bool = False
    chart_trp_nhprp: bool = False
    chart_ar_freq: bool = False

    # B 类子选择: 具体角度/范围 (为空时自动使用模板默认值)
    gain_chart_angles: list[float] = field(default_factory=list)   # PK Gain + 指定 θ 单角度
    gain_chart_ranges: list[tuple] = field(default_factory=list)   # 指定 θ 范围
    ar_chart_angles: list[float] = field(default_factory=list)     # AR 指定 θ 单角度
    ar_chart_ranges: list[tuple] = field(default_factory=list)     # AR 指定 θ 范围

    # ── 图表渲染总开关 ──
    render_charts: bool = True  # False=跳过所有图表渲染, 仅计算参数

    # C 类: 2D 切面 (俯仰面 + 方位面)
    cut_2d_polar: bool = False
    cut_2d_rect: bool = False
    cut_azimuth_polar: bool = False      # 方位面极坐标
    cut_azimuth_rect: bool = False       # 方位面直角坐标
    cut_2d_phi_angles: list[float] = field(default_factory=list)    # 向后兼容
    cut_2d_theta_angles: list[float] = field(default_factory=list)  # 向后兼容
    cut_2d_params: set = field(default_factory=lambda: {"gain"})    # 向后兼容
    # 4 组独立图表条目 (每组 = (param, [angles]) 的 list)
    cut_2d_polar_entries: list = field(default_factory=list)
    cut_2d_rect_entries: list = field(default_factory=list)
    cut_azimuth_polar_entries: list = field(default_factory=list)
    cut_azimuth_rect_entries: list = field(default_factory=list)

    # 视角参数
    elev: float = 30.0
    azim: float = -60.0
    view_angle_pairs: list[tuple[float, float]] = field(default_factory=list)  # [(elev, azim), ...]
    dpi: int = 100
    step_deg: float = 5.0          # 3D 图形采样精度 (°)

    # 频点过滤 (空列表 = 所有频点)
    selected_frequencies_a: list[float] = field(default_factory=list)  # A类
    selected_frequencies_b: list[float] = field(default_factory=list)  # B类
    selected_frequencies_c: list[float] = field(default_factory=list)  # C类

    # 输出方式
    embed_in_excel: bool = True
    save_png_folder: str | None = None

    # ── 属性 ──

    @property
    def has_any_a_class(self) -> bool:
        return self.pattern_3d_gain or self.pattern_3d_eirp or self.pattern_3d_ar

    @property
    def has_any_b_class(self) -> bool:
        return (self.chart_eff_freq or self.chart_gain_freq or
                self.chart_dir_freq or self.chart_lag_freq or
                self.chart_trp_freq or self.chart_trp_nhprp or
                self.chart_ar_freq)

    @property
    def has_any_c_class(self) -> bool:
        return (self.cut_2d_polar or self.cut_2d_rect or
                self.cut_azimuth_polar or self.cut_azimuth_rect)

    @property
    def has_any_pattern_or_cut(self) -> bool:
        """是否有需要逐频点生成的图形（A 或 C 类）。"""
        return self.has_any_a_class or self.has_any_c_class

    # ── 合并 ──

    def merge(self, other: ChartConfig) -> ChartConfig:
        """合并两个配置（OR 逻辑），视角参数取 self 的值。"""
        fields = [
            "pattern_3d_gain", "pattern_3d_eirp", "pattern_3d_ar",
            "pattern_3d_etheta", "pattern_3d_ephi",
            "chart_eff_freq", "chart_gain_freq", "chart_dir_freq",
            "chart_lag_freq", "chart_trp_freq", "chart_trp_nhprp",
            "chart_ar_freq", "cut_2d_polar", "cut_2d_rect",
            "cut_azimuth_polar", "cut_azimuth_rect",
        ]
        merged = ChartConfig(
            elev=self.elev, azim=self.azim, dpi=self.dpi,
            step_deg=self.step_deg,
            embed_in_excel=self.embed_in_excel,
            save_png_folder=self.save_png_folder,
            gain_chart_angles=list(set(self.gain_chart_angles + other.gain_chart_angles)),
            gain_chart_ranges=list(set(self.gain_chart_ranges + other.gain_chart_ranges)),
            ar_chart_angles=list(set(self.ar_chart_angles + other.ar_chart_angles)),
            ar_chart_ranges=list(set(self.ar_chart_ranges + other.ar_chart_ranges)),
            cut_2d_phi_angles=list(set(self.cut_2d_phi_angles + other.cut_2d_phi_angles)),
        )
        for f in fields:
            setattr(merged, f, getattr(self, f) or getattr(other, f))
        # 非 bool 字段合并
        merged.render_charts = self.render_charts and other.render_charts
        merged.view_angle_pairs = self.view_angle_pairs or other.view_angle_pairs
        merged.selected_frequencies_a = self.selected_frequencies_a or other.selected_frequencies_a
        merged.selected_frequencies_b = self.selected_frequencies_b or other.selected_frequencies_b
        merged.selected_frequencies_c = self.selected_frequencies_c or other.selected_frequencies_c
        merged.cut_2d_phi_angles = list(set(self.cut_2d_phi_angles + other.cut_2d_phi_angles))
        merged.cut_2d_theta_angles = list(set(self.cut_2d_theta_angles + other.cut_2d_theta_angles))
        merged.cut_2d_params = self.cut_2d_params | other.cut_2d_params
        # 合并 4 组图表条目
        for attr in ("cut_2d_polar_entries", "cut_2d_rect_entries",
                      "cut_azimuth_polar_entries", "cut_azimuth_rect_entries"):
            sl = list(getattr(self, attr, []))
            ol = list(getattr(other, attr, []))
            setattr(merged, attr, sl + ol)
        return merged

    # ── 工厂方法 ──

    @classmethod
    def from_template(cls, template_path: str) -> ChartConfig:
        """扫描模板文件：元数据行 + 列头 → 自动检测图形需求。

        扫描策略:
          1. 在数据列头行之前的所有行中搜索匹配标题正则
          2. 从模板列类型推导 B 类图表
          3. 合并为默认启用的 ChartConfig
        """
        import openpyxl
        wb = openpyxl.load_workbook(template_path, data_only=True)
        config = ChartConfig()
        col_types: set[str] = set()

        try:
            for ws in wb.worksheets:
                max_row = ws.max_row or 100
                max_col = ws.max_column or 20

                # 扫描元数据行（在 Frequency 列头之前）
                header_row = None
                for row_idx in range(1, min(max_row + 1, 200)):
                    row_texts = []
                    for c_idx in range(1, max_col + 1):
                        v = ws.cell(row_idx, c_idx).value
                        if v is not None:
                            row_texts.append(str(v).strip())

                    # 检查是否到了 Frequency 列头行
                    for t in row_texts:
                        if _match_text(t, [r"(?i)freq", r"(?i)频率"]):
                            header_row = row_idx
                            break

                    if header_row is not None:
                        # 此行为列头行 —— 解析列类型
                        for c_idx in range(1, max_col + 1):
                            v = ws.cell(header_row, c_idx).value
                            if v is not None:
                                # 简化: 直接通过文本判断
                                col_type = _classify_column_text(str(v).strip())
                                if col_type:
                                    col_types.add(col_type)
                        break

                    # 元数据行：搜索图表标题
                    for t in row_texts:
                        for chart_key, patterns in _CHART_PATTERNS.items():
                            if _match_text(t, patterns):
                                setattr(config, chart_key, True)
                                break
        finally:
            wb.close()

        # 从列类型推导 B 类图表
        for ct in col_types:
            chart_key = _COLTYPE_TO_CHART.get(ct)
            if chart_key:
                setattr(config, chart_key, True)

        # NHPRP 存在 → 多线图
        if any(c in col_types for c in ("nhprp_45", "nhprp_30", "nhprp_225")):
            config.chart_trp_nhprp = True

        return config

    @classmethod
    def from_template_headers(cls, headers: list[str], col_types: set[str]) -> ChartConfig:
        """从列头文本列表和 col_type 集合推导图形需求。

        用于 UI 对话框在已有 SheetInfo 的情况下快速推导。
        """
        config = ChartConfig()

        # 扫描列头文本
        for h in headers:
            for chart_key, patterns in _CHART_PATTERNS.items():
                if _match_text(h, patterns):
                    setattr(config, chart_key, True)

        # 从 col_type 推导
        for ct in col_types:
            chart_key = _COLTYPE_TO_CHART.get(ct)
            if chart_key:
                setattr(config, chart_key, True)

        if any(c in col_types for c in ("nhprp_45", "nhprp_30", "nhprp_225")):
            config.chart_trp_nhprp = True

        return config

    @classmethod
    def all_chart_keys(cls) -> list[str]:
        """返回所有图形 flag 的 key 列表（不含视角参数和角度列表）。"""
        return [
            "pattern_3d_gain", "pattern_3d_eirp", "pattern_3d_ar",
            "pattern_3d_etheta", "pattern_3d_ephi",
            "chart_eff_freq", "chart_gain_freq", "chart_dir_freq",
            "chart_lag_freq", "chart_trp_freq", "chart_trp_nhprp",
            "chart_ar_freq", "cut_2d_polar", "cut_2d_rect",
            "cut_azimuth_polar", "cut_azimuth_rect",
        ]

    @classmethod
    def all_sub_angle_keys(cls) -> list[str]:
        """返回所有子角度列表的 key。"""
        return ["gain_chart_angles", "gain_chart_ranges",
                "ar_chart_angles", "ar_chart_ranges",
                "cut_2d_phi_angles"]

    @classmethod
    def get_angles_from_data_config(cls, mw, chart_key: str) -> list[float]:
        """从天线参数角度配置读取图表角度 (统一数据源)。

        避免用户在数据参数和图表参数中重复配置角度。
        """
        if chart_key in ("chart_gain_freq", "chart_lag_freq", "cut_2d_polar", "cut_2d_rect",
                         "pattern_3d_gain", "pattern_3d_etheta", "pattern_3d_ephi"):
            cfg = getattr(mw, '_lag_config', None)
            return cfg.singles_sorted if cfg else []
        elif chart_key in ("chart_ar_freq", "pattern_3d_ar"):
            cfg = getattr(mw, '_ar_lag_config', None)
            return cfg.singles_sorted if cfg else []
        return []

    @classmethod
    def get_chart_ranges_from_data_config(cls, mw, chart_key: str) -> list[tuple]:
        """从天线参数角度配置读取图表角度范围。"""
        if chart_key in ("chart_gain_freq", "chart_lag_freq"):
            cfg = getattr(mw, '_lag_config', None)
            return cfg.ranges_sorted if cfg else []
        elif chart_key == "chart_ar_freq":
            cfg = getattr(mw, '_ar_lag_config', None)
            return cfg.ranges_sorted if cfg else []
        return []

    @classmethod
    def chart_labels(cls) -> dict[str, str]:
        """返回图形 key → 统一命名映射 (英文技术缩写 + 中文类别词)。"""
        return {
            "pattern_3d_gain": "3D Gain 方向图",
            "pattern_3d_eirp": "3D EIRP 方向图",
            "pattern_3d_ar": "3D AR 方向图",
            "pattern_3d_etheta": "3D Eθ 方向图",
            "pattern_3d_ephi": "3D Eφ 方向图",
            "chart_eff_freq": "Efficiency vs 频率",
            "chart_gain_freq": "Gain vs 频率",
            "chart_dir_freq": "Directivity vs 频率",
            "chart_trp_freq": "TRP vs 频率",
            "chart_ar_freq": "AR vs 频率",
            "cut_2d_polar": "极坐标俯仰面切面图",
            "cut_2d_rect": "直角坐标俯仰面切面图",
            "cut_azimuth_polar": "极坐标方位面切面图",
            "cut_azimuth_rect": "直角坐标方位面切面图",
        }

    @classmethod
    def chart_categories(cls, mode: int = 0) -> dict[str, list[str]]:
        """返回图形分类: 类别名 → [chart_key, ...]。按测试模式过滤。

        Args:
            mode: 0=无源天线, 1=有源发射(TRP), 2=有源接收(TIS)
        """
        # 共用: 3D 方向图 (所有模式都有 Gain)
        pattern_3d = ["pattern_3d_gain", "pattern_3d_etheta", "pattern_3d_ephi"]
        if mode == 1:
            pattern_3d.append("pattern_3d_eirp")
        else:
            pattern_3d.append("pattern_3d_ar")

        # vs 频率曲线: 按模式分组
        vs_freq = ["chart_gain_freq", "chart_eff_freq", "chart_dir_freq"]
        if mode == 0:
            vs_freq.append("chart_ar_freq")
        elif mode == 1:
            vs_freq.append("chart_trp_freq")

        # 2D 切面图: 俯仰面 + 方位面
        cuts = ["cut_2d_polar", "cut_azimuth_polar", "cut_2d_rect", "cut_azimuth_rect"]

        return {
            "A 类: 3D 方向图": pattern_3d,
            "B 类: 频率曲线": vs_freq,
            "C 类: 2D 切面图": cuts,
        }


# ═══════════════════════════════════════════════════════════════
# 内部辅助
# ═══════════════════════════════════════════════════════════════

def _match_text(text: str, patterns: list[str]) -> bool:
    """检查 text 是否匹配任一正则 pattern。"""
    for pat in patterns:
        if re.search(pat, text, re.IGNORECASE):
            return True
    return False


def _classify_column_text(text: str) -> str | None:
    """从列头文本推导 col_type（简化版，JSON 模式优先，fallback 到内置正则）。"""
    from .excel_reader import (
        _classify_by_json_patterns,
        detect_ratio_column_type,
        is_ar_range_column,
        is_ar_single_column,
        is_avg_gain_column,
        is_avg_power_column,
        is_boresight_phi_column,
        is_boresight_theta_column,
        is_directivity_column,
        is_efficiency_column,
        is_frequency_column,
        is_gain_column,
        is_lh_prp_column,
        is_max_power_column,
        is_min_power_column,
        is_nhprp_30_column,
        is_nhprp_45_column,
        is_nhprp_225_column,
        is_peak_eirp_column,
        is_trp_column,
        is_uh_prp_column,
    )
    # JSON 用户模式优先
    json_type = _classify_by_json_patterns(text)
    if json_type is not None:
        return json_type
    if is_frequency_column(text): return "frequency"
    if is_directivity_column(text): return "directivity"
    if is_efficiency_column(text):
        return "efficiency_pct"
    if is_gain_column(text): return "gain"
    if is_trp_column(text): return "trp"
    if is_nhprp_45_column(text): return "nhprp_45"
    if is_nhprp_30_column(text): return "nhprp_30"
    if is_peak_eirp_column(text): return "peak_eirp"
    if is_ar_single_column(text): return "ar_single"
    if is_ar_range_column(text): return "ar_range"
    if is_nhprp_225_column(text): return "nhprp_225"
    if is_uh_prp_column(text): return "uh_prp"
    if is_lh_prp_column(text): return "lh_prp"
    ratio = detect_ratio_column_type(text)
    if ratio: return ratio
    if is_boresight_phi_column(text): return "boresight_phi"
    if is_boresight_theta_column(text): return "boresight_theta"
    if is_max_power_column(text): return "max_power"
    if is_min_power_column(text): return "min_power"
    if is_avg_gain_column(text): return "avg_gain"
    if is_avg_power_column(text): return "avg_power"
    return None
