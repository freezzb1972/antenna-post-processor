"""
Excel 输出模块
==============
基于模板填充天线参数数据 + 嵌入 3D 辐射方向图。

使用 openpyxl 复制模板并写入结果，保留原始格式。
"""

from __future__ import annotations

import io
import math
import os
import re
from collections.abc import Callable
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from .chart_config import ChartConfig

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from .excel_reader import ColumnInfo, SheetInfo, read_template
from .lag_config import (
    _RE_LAG_RANGE,
    _RE_LAG_RANGE_NO_PREFIX,
    _RE_LAG_SINGLE,
    _RE_LAG_SINGLE_NO_PREFIX,
    normalize_header,
)


def _replace_cell_text(ws, old_sheet_name: str, new_sheet_name: str, max_scan_rows: int = 15):
    """在工作表的前 N 行中查找并替换天线名称。

    智能匹配策略（按优先级）：
    1. 查找包含「旧全名」的单元格 → 替换为「新全名」
    2. 查找包含「差异部分」（两名的变化字符段）的单元格 → 精准替换
    3. 两策略都匹配多个单元格时，选「得分最高」的唯一匹配；否则跳过不替换
    """
    if old_sheet_name == new_sheet_name:
        return

    # 找出两个名字的差异段
    old_delta, new_delta = _name_delta(old_sheet_name, new_sheet_name)

    # 策略 1: 全名精确匹配（得分最高）
    candidates_full = _find_matching_cells(ws, old_sheet_name, max_scan_rows)

    # 策略 2: 差异段匹配（更灵活，如 "5G1"→"5G2" 中 "G1"→"G2"）
    candidates_delta = []
    if old_delta and len(old_delta) >= 2:  # 差异段至少 2 字符，避免 "1"→"2" 过度匹配
        candidates_delta = _find_matching_cells(ws, old_delta, max_scan_rows)

    # 选最佳匹配：优先全名唯一匹配 → 差异段唯一匹配 → 全名最高分 → 跳过
    best = None
    if len(candidates_full) == 1:
        best = ("full", candidates_full[0])
    elif len(candidates_delta) == 1:
        best = ("delta", candidates_delta[0])
    elif candidates_full:
        best = ("full", candidates_full[0])  # 全名匹配多个时选最高分
    elif candidates_delta:
        best = ("delta", candidates_delta[0])  # 差异段匹配多个时选最高分

    if best is None:
        return  # 找不到可靠匹配，安全跳过

    strategy, (row, col, _score) = best
    cell = ws.cell(row, col)
    old_text = str(cell.value) if cell.value else ""

    if strategy == "full":
        cell.value = old_text.replace(old_sheet_name, new_sheet_name)
    else:
        cell.value = old_text.replace(old_delta, new_delta)


def _find_matching_cells(ws, search: str, max_rows: int):
    """扫描前 N 行，返回匹配单元格列表 [(row, col, score), ...]。

    分值规则:
      - 单元格文本 == search        → 10 分（精确匹配）
      - 单元格文本 以 search 开头    → 8 分
      - 单元格文本 包含 search       → 5 分
    """
    matches = []
    for row in range(1, max_rows + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if not val or not isinstance(val, str):
                continue
            if search not in val:
                continue
            if val == search:
                matches.append((row, col, 10))
            elif val.startswith(search):
                matches.append((row, col, 8))
            else:
                matches.append((row, col, 5))
    matches.sort(key=lambda x: -x[2])  # 按分值降序
    return matches


def _name_delta(old: str, new: str):
    """提取两个名字的「差异段」。

    "5G1" vs "5G2" → ("5G1", "5G2") — 无共同字符，返回全名
    "ANT001" vs "ANT002" → ("1", "2") — 共同前缀 ANT00，差异在末尾
    "DUT-A" vs "DUT-B" → ("A", "B")
    """
    if old == new:
        return "", ""

    # 找共同前缀
    i = 0
    while i < min(len(old), len(new)) and old[i] == new[i]:
        i += 1

    # 找共同后缀（从差异点之后开始）
    old_rest = old[i:]
    new_rest = new[i:]
    j = 0
    while j < min(len(old_rest), len(new_rest)) and old_rest[-(j+1)] == new_rest[-(j+1)]:
        j += 1

    old_delta = old_rest[:len(old_rest)-j] if j > 0 else old_rest
    new_delta = new_rest[:len(new_rest)-j] if j > 0 else new_rest

    return old_delta, new_delta


# ---------------------------------------------------------------------------
# 公开 API
# ---------------------------------------------------------------------------

def export_results(
    template_path: str,
    output_path: str,
    sheet_results: dict[str, list[dict[str, Any]]],
    *,
    pattern_images: dict[str, dict[float, io.BytesIO]] | None = None,
    sheets_info: list[SheetInfo] | None = None,
    chart_config: ChartConfig | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
    log_callback: Callable[[str], None] | None = None,
    remove_template_sheets: list[str] | None = None,
    **kwargs,
) -> str:
    """基于模板填充数据 + 嵌入图片。

    Args:
        template_path:  模板 Excel 路径。
        output_path:    输出 Excel 路径。
        sheet_results:  {sheet_name: [row_dict, ...]}，
                        row_dict key = column type ("directivity", "lag_single_60.0", ...)。
        pattern_images: {sheet_name: {freq_mhz: PNG_buffer}}（可选）。
        sheets_info:    预解析的 Sheet 信息（可选，避免重复读模板）。
        remove_template_sheets: 要删除的模板原始工作表名列表（None=不删除）。
        progress_callback: (current, total, message)。
        log_callback:   (message)。

    Returns:
        输出文件路径。
    """
    # 读取模板（仅当调用方未提供 sheets_info 时）
    if sheets_info is None:
        sheets_info = read_template(template_path)
    info_map = {s.name: s for s in sheets_info}

    # 复制模板
    wb = openpyxl.load_workbook(template_path)

    total_ops = sum(len(rows) for rows in sheet_results.values()) + len(sheets_info)
    current = 0

    # 找到参考 worksheet（用于克隆）
    ref_ws = wb.worksheets[0] if wb.worksheets else None

    for sheet_name, rows in sheet_results.items():
        if sheet_name not in wb.sheetnames:
            # 自动克隆：用第一个 worksheet 为模板创建新 sheet
            if ref_ws is not None:
                if log_callback:
                    log_callback(f"  ↗ 自动创建工作表: {sheet_name}")
                ws = wb.copy_worksheet(ref_ws)
                ws.title = sheet_name
                # 替换单元格中旧的 sheet 名 → 新的 sheet 名（如 "5G1"→"5G2"）
                _replace_cell_text(ws, ref_ws.title, sheet_name)
            else:
                continue
        else:
            ws = wb[sheet_name]

        if sheet_name not in info_map:
            # 自动扩增的 sheet 不在原始 info_map 中 — 从 sheets_info 查找
            if sheets_info:
                for si in sheets_info:
                    if si.name == sheet_name:
                        info_map[sheet_name] = si
                        break
            if sheet_name not in info_map:
                continue

        info = info_map[sheet_name]

        # 构建列映射: col_type → ColumnInfo
        col_map = _build_col_map(info)

        for row_idx_offset, row_data in enumerate(rows):
            excel_row = info.data_start_row + row_idx_offset

            # 匹配频点
            freq = row_data.get("frequency")
            if freq is None:
                continue

            # 填入各列
            for key, value in row_data.items():
                if key == "frequency":
                    _write_cell(ws, excel_row, col_map, "frequency", freq)
                elif key == "directivity":
                    _write_cell(ws, excel_row, col_map, "directivity", value)
                elif key == "efficiency_pct":
                    _write_cell(ws, excel_row, col_map, "efficiency_pct", value)
                elif key == "efficiency_db":
                    _write_cell(ws, excel_row, col_map, "efficiency_db", value)
                elif key == "gain":
                    # 可能有多个 Gain 列 (5G4)
                    _write_cell(ws, excel_row, col_map, "gain", value)
                elif key == "trp":
                    _write_cell(ws, excel_row, col_map, "trp", value)
                elif key == "nhprp_45":
                    _write_cell(ws, excel_row, col_map, "nhprp_45", value)
                elif key == "nhprp_30":
                    _write_cell(ws, excel_row, col_map, "nhprp_30", value)
                elif key == "peak_eirp":
                    _write_cell(ws, excel_row, col_map, "peak_eirp", value)
                elif key.startswith("lag_single_"):
                    angle_str = key[len("lag_single_"):]
                    _write_lag_single(ws, excel_row, col_map, float(angle_str), value)
                elif key.startswith("lag_range_"):
                    parts = key[len("lag_range_"):].split("_")
                    if len(parts) == 2:
                        lo, hi = float(parts[0]), float(parts[1])
                        _write_lag_range(ws, excel_row, col_map, lo, hi, value)
                elif key.startswith("ar_single_"):
                    angle_str = key[len("ar_single_"):]
                    _write_ar_single(ws, excel_row, col_map, float(angle_str), value)
                elif key.startswith("ar_range_"):
                    parts = key[len("ar_range_"):].split("_")
                    if len(parts) == 2:
                        lo, hi = float(parts[0]), float(parts[1])
                        _write_ar_range(ws, excel_row, col_map, lo, hi, value)
                elif key.startswith("rhcp_single_"):
                    angle_str = key[len("rhcp_single_"):]
                    _write_rhcp_single(ws, excel_row, col_map, float(angle_str), value)
                elif key.startswith("rhcp_range_"):
                    parts = key[len("rhcp_range_"):].split("_")
                    if len(parts) == 2:
                        _write_rhcp_range(ws, excel_row, col_map, float(parts[0]), float(parts[1]), value)
                elif key.startswith("cp_xpi_single_"):
                    angle_str = key[len("cp_xpi_single_"):]
                    _write_cp_xpi_single(ws, excel_row, col_map, float(angle_str), value)
                elif key.startswith("cp_xpi_range_"):
                    parts = key[len("cp_xpi_range_"):].split("_")
                    if len(parts) == 2:
                        _write_cp_xpi_range(ws, excel_row, col_map, float(parts[0]), float(parts[1]), value)
                elif key in ("nhprp_225", "uh_prp", "lh_prp", "prp_120",
                             "max_power", "min_power", "avg_gain", "avg_power",
                             "boresight_theta", "boresight_phi",
                             "xpi_boresight", "xpi_mean", "xpi_min",
                             "total_efficiency_pct", "mismatch_loss_db",
                             "pc_theta_mm", "pc_phi_mm") or key.endswith("_ratio_db") or key.endswith("_ratio_pct"):
                    _write_cell(ws, excel_row, col_map, key, value)

            current += 1
            if progress_callback:
                progress_callback(current, total_ops, f"写入 {sheet_name} Row {excel_row}")

        # ---- 嵌入 3D 方向图 ----
        if pattern_images and sheet_name in pattern_images:
            _embed_images(ws, info, pattern_images[sheet_name], log_callback)

        current += 1
        if progress_callback:
            progress_callback(current, total_ops, f"完成 {sheet_name}")

    # ---- 删除模板原始工作表 (数据源文件名模式) ----
    if remove_template_sheets:
        desired = set(sheet_results.keys())
        sheets_to_remove = [s for s in remove_template_sheets if s in wb.sheetnames and s not in desired]
        if sheets_to_remove:
            if log_callback:
                log_callback(f"  ✕ 删除模板原始工作表: {', '.join(sheets_to_remove)}")
            for name in sheets_to_remove:
                del wb[name]

    wb.save(output_path)
    wb.close()
    return output_path


def _build_col_map(info: SheetInfo) -> dict[str, list[ColumnInfo]]:
    """构建 col_type → [ColumnInfo, ...] 映射（可能有重复类型列，如 5G4 的两个 Gain 列）。"""
    m: dict[str, list[ColumnInfo]] = {}
    for cinfo in info.columns:
        m.setdefault(cinfo.col_type, []).append(cinfo)
    return m


def _write_ar_single(ws, row, col_map, angle, value):
    """写入 AR 单角度到匹配的列。"""
    for cinfo in col_map.get("ar_single", []):
        norm = normalize_header(cinfo.raw_header)
        m = re.search(r"(\d+)\s*deg", norm) or re.search(r"theta[= ]*(\d+)", norm, re.IGNORECASE)
        if m and abs(float(m.group(1)) - angle) < 0.01:
            cell = ws.cell(row, cinfo.col_index)
            cell.value = round(value, 6) if isinstance(value, float) else value
            return


def _write_ar_range(ws, row, col_map, lo, hi, value):
    """写入 AR 范围到匹配的列。"""
    for cinfo in col_map.get("ar_range", []):
        norm = normalize_header(cinfo.raw_header)
        m = re.search(r"(\d+)\s*[~\-–—]\s*(\d+)", norm)
        if m:
            clo, chi = float(m.group(1)), float(m.group(2))
            if abs(clo - lo) < 0.01 and abs(chi - hi) < 0.01:
                cell = ws.cell(row, cinfo.col_index)
                cell.value = round(value, 6) if isinstance(value, float) else value
                return


def _write_cell(
    ws,
    row: int,
    col_map: dict[str, list[ColumnInfo]],
    ctype: str,
    value: Any,
):
    """写入一个单元格。"""
    cols = col_map.get(ctype, [])
    if not cols:
        return
    for cinfo in cols:
        cell = ws.cell(row, cinfo.col_index)
        if value is not None:
            # 保留 2 位小数（科学计算常用）
            if isinstance(value, float):
                cell.value = round(value, 6)
            else:
                cell.value = value


def _write_lag_single(ws, row, col_map, angle, value):
    """写入单角度参数到匹配的列 (LAG/RHCP/CP-XPI 共用)。"""
    _write_typed_single(ws, row, col_map, "lag_single", angle, value)

def _write_rhcp_single(ws, row, col_map, angle, value):
    _write_typed_single(ws, row, col_map, "rhcp_single", angle, value)

def _write_cp_xpi_single(ws, row, col_map, angle, value):
    _write_typed_single(ws, row, col_map, "cp_xpi_single", angle, value)

def _write_typed_single(
    ws, row: int, col_map: dict[str, list[ColumnInfo]],
    col_type: str, angle: float, value: Any,
):
    """写入单角度参数到匹配的列 (角度匹配, 不跨类型覆盖)。"""
    for cinfo in col_map.get(col_type, []):
        norm = normalize_header(cinfo.raw_header)
        m = _RE_LAG_SINGLE.search(norm) or _RE_LAG_SINGLE_NO_PREFIX.search(norm)
        if not m:
            m = re.search(r"theta[= ]*(\d+)", norm, re.IGNORECASE)
        if not m:
            m = re.search(r"(\d+\.?\d*)", norm)
        if m and abs(float(m.group(1)) - angle) < 0.01:
            cell = ws.cell(row, cinfo.col_index)
            cell.value = round(value, 6) if isinstance(value, float) else value
            return


def _write_lag_range(ws, row, col_map, lo, hi, value):
    """写入范围参数到匹配的列。"""
    _write_typed_range(ws, row, col_map, "lag_range", lo, hi, value)

def _write_rhcp_range(ws, row, col_map, lo, hi, value):
    _write_typed_range(ws, row, col_map, "rhcp_range", lo, hi, value)

def _write_cp_xpi_range(ws, row, col_map, lo, hi, value):
    _write_typed_range(ws, row, col_map, "cp_xpi_range", lo, hi, value)

def _write_typed_range(
    ws, row: int, col_map: dict[str, list[ColumnInfo]],
    col_type: str, lo: float, hi: float, value: Any,
):
    """写入范围参数到匹配的列 (角度匹配, 不跨类型覆盖)。"""
    for cinfo in col_map.get(col_type, []):
        norm = normalize_header(cinfo.raw_header)
        m = _RE_LAG_RANGE.search(norm) or _RE_LAG_RANGE_NO_PREFIX.search(norm)
        if not m:
            m = re.search(r"theta[= ]*(\d+)\s*[~\-–—]\s*(\d+)", norm, re.IGNORECASE)
        if not m:
            m = re.search(r"(\d+)\s*[~\-–—]\s*(\d+)", norm)
        if m:
            clo, chi = float(m.group(1)), float(m.group(2))
            if abs(clo - lo) < 0.01 and abs(chi - hi) < 0.01:
                cell = ws.cell(row, cinfo.col_index)
                cell.value = round(value, 6) if isinstance(value, float) else value
                return


def _embed_images(
    ws,
    info: SheetInfo,
    images: dict[float, io.BytesIO],
    log_callback=None,
    *,
    col_offset: int = 3,
    row_step: int = 22,
    img_width: int = 280,
    img_height: int = 210,
):
    """在数据区域右侧嵌入 3D 方向图。

    图片布局：从数据区域右侧 col_offset 列开始，每 row_step 行放一张图。

    Args:
        ws:         工作表对象。
        info:       工作表结构信息。
        images:     {freq_mhz: PNG_buffer}。
        log_callback: 日志回调。
        col_offset: 图片列相对数据区域右侧的偏移（列数）。
        row_step:   每张图占用的行高。
        img_width:  图片宽度 (px)。
        img_height: 图片高度 (px)。
    """
    if not images:
        return

    max_col = max((c.col_index for c in info.columns), default=10)
    image_col = max_col + col_offset
    image_row = info.data_start_row

    for idx, (freq, buf) in enumerate(sorted(images.items())):
        try:
            buf.seek(0)
            img = XLImage(buf)
            img.width = img_width
            img.height = img_height
            cell = f"{get_column_letter(image_col)}{image_row}"
            ws.add_image(img, cell)
            image_row += row_step
        except Exception as e:
            # 图片嵌入失败不阻塞数据输出，但记录日志
            if log_callback:
                log_callback(f"  ⚠ {freq} MHz: 图片嵌入失败 — {e}")
