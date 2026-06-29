"""
后台处理 Worker
================
QThread 封装的异步处理任务，通过 Signal 与 GUI 通信。
支持多步进同时计算模式。
"""

from __future__ import annotations
from typing import Any, Dict, List, Optional
from PySide6.QtCore import QObject, Signal
from .lag_config import LagConfig
from .plot_config import PlotConfig
from .chart_config import ChartConfig
from .pipeline import run_pipeline, run_batch_pipeline
import math
import traceback
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import NumericAxis
from openpyxl.utils import get_column_letter
from .datasource import DataSource, ResampledDataSource


# ── 工具 ──────────────────────────────────────────────────────

def _safe_create_sheet(wb, base_name, suffix="", max_len=31):
    """创建工作表: 自动截断到 Excel 31字符限制，冲突时加 _01/_02 编号。"""
    max_base = max_len - len(suffix)
    safe = base_name[:max_base] if len(base_name) > max_base else base_name
    name = f"{safe}{suffix}"
    existing = {ws.title for ws in wb.worksheets}
    if name not in existing:
        return wb.create_sheet(title=name)
    for i in range(1, 100):
        alt = f"{safe[:max_base - 3]}_{i:02d}{suffix}"
        if alt not in existing:
            return wb.create_sheet(title=alt)
    raise RuntimeError(f"无法为 '{base_name}' 生成唯一工作表名")


class ProcessingWorker(QObject):
    """后台天线参数处理 Worker。"""

    progress = Signal(int, int, str)
    log = Signal(str)
    finished = Signal(object, object)
    error = Signal(str)

    def __init__(
        self,
        csv_path: str = "",
        template_path: str = "",
        output_path: str = "",
        *,
        datasource: Optional[DataSource] = None,
        datasource_map: Optional[Dict[str, DataSource]] = None,
        sheet_mode_map: Optional[Dict[str, int]] = None,
        lag_config: Optional[LagConfig] = None,
        plot_config: Optional[PlotConfig] = None,
        full_report_path: Optional[str] = None,
        extrapolate_theta: bool = False,
        freq_source: str = "datasource",
        trim_start: int = 0,
        trim_end: int = 0,
        robust_peak: bool = False,
        extra_params: Optional[set] = None,
        chart_config_obj: Optional[ChartConfig] = None,
        ar_lag_config: Optional[LagConfig] = None,
        nh_custom_angles: Optional[List[float]] = None,
        ar_output_db: bool = True,
        worksheet_naming_mode: int = 0,
        # 多步进参数
        step_values: Optional[List[float]] = None,
        skip_original: bool = False,
        gen_diff: bool = False,
        gen_diff_chart: bool = False,
    ):
        super().__init__()
        self.csv_path = csv_path
        self.template_path = template_path
        self.output_path = output_path
        self.datasource = datasource
        self.datasource_map = datasource_map
        self.sheet_mode_map = sheet_mode_map or {}
        self.lag_config = lag_config
        self.ar_lag_config = ar_lag_config
        self.plot_config = plot_config or PlotConfig()
        self.full_report_path = full_report_path
        self.extrapolate_theta = extrapolate_theta
        self.freq_source = freq_source
        self.trim_start = trim_start
        self.trim_end = trim_end
        self.robust_peak = robust_peak
        self.extra_params = extra_params
        self.chart_config_obj = chart_config_obj
        self.nh_custom_angles = nh_custom_angles
        self.ar_output_db = ar_output_db
        self.worksheet_naming_mode = worksheet_naming_mode
        self.step_values = step_values or []
        self.skip_original = skip_original
        self.gen_diff = gen_diff
        self.gen_diff_chart = gen_diff_chart
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    def _is_cancelled(self) -> bool:
        return self._cancelled

    def run(self):
        try:
            if self._cancelled:
                self.log.emit("处理已取消")
                return

            kwargs = dict(
                template_path=self.template_path,
                output_path=self.output_path,
                lag_config_override=self.lag_config,
                ar_lag_config_override=self.ar_lag_config,
                sheet_mode_map=self.sheet_mode_map,
                plot_config=self.plot_config,
                full_report_path=self.full_report_path,
                extrapolate_theta=self.extrapolate_theta,
                freq_source=self.freq_source,
                trim_start=self.trim_start,
                trim_end=self.trim_end,
                chart_config_obj=self.chart_config_obj,
                robust_peak=self.robust_peak,
                extra_params=self.extra_params,
                nh_custom_angles=self.nh_custom_angles,
                ar_output_db=self.ar_output_db,
                worksheet_naming_mode=self.worksheet_naming_mode,
                cancel_callback=self._is_cancelled,
                progress_callback=self._on_progress,
                log_callback=self._on_log,
            )

            if self.gen_diff_chart and not self.gen_diff:
                raise ValueError("gen_diff_chart requires gen_diff")

            # ── 多步进模式 ──
            if self.step_values:
                # 获取基础 DataSource（单文件或 map 首项）
                base = self.datasource
                if base is None and self.datasource_map:
                    base = next(iter(self.datasource_map.values()))
                if base is not None:
                    results = self._run_multi_step(base, **kwargs)
                else:
                    results = run_batch_pipeline(csv_path=self.csv_path, **kwargs)
            elif self.datasource_map:
                results = run_pipeline(datasource_map=self.datasource_map, **kwargs)
            elif self.datasource:
                results = run_pipeline(datasource=self.datasource, **kwargs)
            else:
                results = run_batch_pipeline(csv_path=self.csv_path, **kwargs)

            if not self._cancelled:
                self.finished.emit(results, {})

        except Exception as e:
            import traceback
            self.error.emit(f"{e}\n{traceback.format_exc()}")

    def _run_multi_step(self, base_ds: DataSource, **kwargs) -> Dict[str, list]:
        """多步进并行计算：源文件一次读取，各步进在独立线程中同时计算。

        自动适配硬件：
          - 最大线程数 = min(步进数, CPU核心数-1, 6)
          - 4核以下: 串行或2线程; 8核+: 最多6线程并行
        """
        import os
        import openpyxl
        from concurrent.futures import ThreadPoolExecutor, as_completed
        from tempfile import NamedTemporaryFile

        # 原始步进
        orig_theta_step = round(
            base_ds.theta_angles[1] - base_ds.theta_angles[0], 6
        ) if len(base_ds.theta_angles) > 1 else 1.0
        orig_phi_step = round(
            base_ds.phi_angles[1] - base_ds.phi_angles[0], 6
        ) if len(base_ds.phi_angles) > 1 else 1.0

        # 构建任务列表: (suffix, display_label, DataSource)
        tasks = []
        if not self.skip_original:
            tasks.append(("", "原始步进", base_ds))

        for step in sorted(self.step_values):
            theta_stride = max(1, int(round(step / orig_theta_step)))
            phi_stride = max(1, int(round(step / orig_phi_step)))
            suffix = f"_step{int(step)}"
            resampled = ResampledDataSource(base_ds, theta_stride, phi_stride)
            tasks.append((suffix, f"步进 {step}°", resampled))

        output_path = kwargs.pop("output_path", self.output_path)
        temp_dir = os.path.dirname(output_path) or "."
        os.makedirs(temp_dir, exist_ok=True)

        # ── 自动适配并行度 ──
        cpu_count = os.cpu_count() or 2
        max_workers = max(1, min(len(tasks), cpu_count - 1, 6))
        self.log.emit(
            f"📏 多步进并行计算: {len(tasks)} 个步进, "
            f"{cpu_count} 核 CPU, 使用 {max_workers} 个并行线程")

        # ── 并行计算 ──
        def _compute_one(suffix, label, ds):
            """单个步进的计算任务（在线程池中运行）。"""
            if self._cancelled:
                return None
            self.log.emit(f"  ⏳ 计算: {label}...")

            with NamedTemporaryFile(suffix=".xlsx", dir=temp_dir, delete=False) as tf:
                tmp_out = tf.name

            try:
                kw = dict(kwargs)
                kw["output_path"] = tmp_out
                kw["progress_callback"] = lambda c, t, m: None  # 不跨线程发信号
                kw["log_callback"] = lambda m: None

                run_pipeline(datasource=ds, **kw)
                if self._cancelled:
                    return None
                self.log.emit(f"  ✓ 完成: {label}")
                return (suffix, tmp_out)
            except Exception as e:
                self.log.emit(f"  ✗ 失败: {label} — {e}")
                try:
                    os.unlink(tmp_out)
                except OSError:
                    pass
                return None

        results = []
        if max_workers == 1:
            # 串行模式（单核或只有1个任务）
            for suffix, label, ds in tasks:
                if self._cancelled:
                    break
                r = _compute_one(suffix, label, ds)
                if r:
                    results.append(r)
        else:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {
                    executor.submit(_compute_one, suffix, label, ds): (suffix, label)
                    for suffix, label, ds in tasks
                }
                for future in as_completed(futures):
                    if self._cancelled:
                        executor.shutdown(wait=False, cancel_futures=True)
                        break
                    r = future.result()
                    if r:
                        results.append(r)

        if self._cancelled:
            return {}

        # ── 合并工作簿（串行，按步进排序） ──
        results.sort(key=lambda x: (x[0] == "", float(x[0].replace("_step", "") or "0")))

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for suffix, tmp_out in results:
            try:
                twb = openpyxl.load_workbook(tmp_out, data_only=True)
                for ws in twb.worksheets:
                    new_name = (ws.title + suffix)[:31]
                    nws = _safe_create_sheet(wb, ws.title, suffix)
                    for row in ws.iter_rows():
                        for cell in row:
                            ncell = nws.cell(row=cell.row, column=cell.column)
                            ncell.value = cell.value
                            if cell.has_style:
                                ncell.font = cell.font
                                ncell.fill = cell.fill
                                ncell.alignment = cell.alignment
                                ncell.border = cell.border
                                ncell.number_format = cell.number_format
                    for col_letter, dim in ws.column_dimensions.items():
                        nws.column_dimensions[col_letter].width = dim.width
                    if ws.freeze_panes:
                        nws.freeze_panes = ws.freeze_panes
                twb.close()
            finally:
                try:
                    os.unlink(tmp_out)
                except OSError:
                    pass

        # 步进差值比较表
        if self.gen_diff and len(results) > 1:
            try:
                self._add_diff_sheet(wb, results)
            except Exception as e:
                self.log.emit(f"\u26a0 步进差值表生成失败: {e}")
                self.log.emit(traceback.format_exc())

        wb.save(output_path)
        wb.close()

        return {}

    def _on_progress(self, current, total, message):
        if not self._cancelled:
            self.progress.emit(current, total, message)

    def _on_log(self, message):
        if not self._cancelled:
            self.log.emit(message)

    def _add_diff_sheet(self, wb, results):
        """生成步进差值比较表 + 差值图表。"""
        # 按 suffix 分组: original(suffix="") + steps
        sheets_by_base = {}
        for suffix, _ in results:
            for ws in wb.worksheets:
                if ws.title.endswith(suffix):
                    base = ws.title[:-len(suffix)] if suffix else ws.title
                    sheets_by_base.setdefault(base, {})[suffix] = ws
        # 对每个基础 sheet 生成差值表; 同时收集 flat data 用于 PivotChart
        all_flat_rows = []  # (freq_val, param_name, step_label, diff_val)
        for base, suffixed in sheets_by_base.items():
            orig_ws = suffixed.get("")
            if orig_ws is None or len(suffixed) < 2:
                continue
            # 读取原始数据
            rows = list(orig_ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(c or "") for c in rows[0]]
            data_rows = rows[1:]
            # 创建差值 sheet
            dws = _safe_create_sheet(wb, base, "_diff")
            # 表头: 频率 | 参数_步进N | 差值_N (与数据写入顺序一致: 外层=步进, 内层=参数)
            col = 1
            dws.cell(1, col, "Frequency (MHz)")
            col += 1
            for suffix_key in sorted(suffixed.keys()):
                if suffix_key == "":
                    continue  # 跳过原始, 只写步进列 (与数据写入对齐)
                label = f"步进{int(float(suffix_key.replace('_step','')))}°"
                for hi, h in enumerate(headers):
                    if h and h.lower() not in ("frequency", "frequency (mhz)", "", "freq"):
                        dws.cell(1, col, f"{h} ({label})")
                        dws.cell(1, col + 1, f"{h} 差值 ({label})")
                        col += 2
            # 写入数据: 逐行计算差值
            freq_col = 0
            for fi, h in enumerate(headers):
                if h and h.lower() in ("frequency", "frequency (mhz)", "freq"):
                    freq_col = fi
                    break
            # 预加载所有步进工作表数据 (避免每行重读 O(N*M) -> O(M))
            all_step_rows = {}
            for suffix_key in sorted(suffixed.keys()):
                sws = suffixed[suffix_key]
                all_step_rows[suffix_key] = list(sws.iter_rows(values_only=True))

            for ri, data_row in enumerate(data_rows):
                er = ri + 2
                freq_val = data_row[freq_col] if freq_col < len(data_row) else None
                dws.cell(er, 1, freq_val)
                col = 2
                step_data = {}
                for suffix_key in sorted(suffixed.keys()):
                    srows = all_step_rows[suffix_key]
                    if ri + 1 < len(srows):
                        step_data[suffix_key] = srows[ri + 1]
                    else:
                        step_data[suffix_key] = []
                orig_row = step_data.get("", [])
                for suffix_key in sorted(suffixed.keys()):
                    if suffix_key == "":
                        continue
                    step_label = f"步进{int(float(suffix_key.replace('_step','')))}°"
                    srow = step_data.get(suffix_key, [])
                    for hi, h in enumerate(headers):
                        if h and h.lower() not in ("frequency", "frequency (mhz)", "", "freq"):
                            oval = orig_row[hi] if hi < len(orig_row) else None
                            sval = srow[hi] if hi < len(srow) else None
                            try:
                                oval_f = float(oval) if oval is not None else 0
                                sval_f = float(sval) if sval is not None else 0
                                diff = sval_f - oval_f
                            except (ValueError, TypeError):
                                oval_f = 0; sval_f = 0; diff = 0
                            dws.cell(er, col, round(sval_f, 4))
                            dws.cell(er, col + 1, round(diff, 4))
                                            # 收集 flat data 用于交互式 PivotChart (跳过空频率)
                            if freq_val is not None:
                                all_flat_rows.append((freq_val, h, step_label, diff))
                            col += 2
            self.log.emit(f"  📊 差值表已生成: {dws.title}")
        # ── 交互式 PivotChart + 切片器（替代每参数静态图表）──
        if all_flat_rows and self.gen_diff_chart:
            # 内存保护: 超大数据集跳过 PivotChart，直接回退静态图表
            if len(all_flat_rows) > 50000:
                self.log.emit(f"  ⚠ 差值数据量过大({len(all_flat_rows)}行), 跳过高开销交互式图表")
                self._add_static_diff_charts_fallback(wb, sheets_by_base)
            else:
                try:
                    self._add_pivot_diff(wb, all_flat_rows)
                    self.log.emit("  📈 交互式差值图表（PivotChart + 切片器）已生成")
                except Exception as e:
                    self.log.emit(f"  ⚠ PivotChart 生成失败，回退到静态图表: {e}")
                    self.log.emit(traceback.format_exc())
                    # Fallback: 为每个 diff sheet 生成静态图表
                    self._add_static_diff_charts_fallback(wb, sheets_by_base)

    def _add_pivot_diff(self, wb, all_flat_rows):
        """创建交互式 PivotTable + 散点图（参数/步进角度切片切换）。"""
        flat_ws, n_data = self._write_flat_data(wb, all_flat_rows)
        pt_ws = self._build_pivot_table(wb, flat_ws, n_data)
        self._add_diff_scatter_chart(pt_ws, flat_ws, n_data)

    # ── _add_pivot_diff 子方法 ──────────────────────────────────

    @staticmethod
    def _write_flat_data(wb, all_flat_rows):
        """写入 long-format 数据到 _diff_flat sheet, 返回 (ws, n_data)。"""
        flat_ws = _safe_create_sheet(wb, "_diff_flat")
        flat_ws.append(["Frequency", "Parameter", "StepAngle", "DiffValue"])
        for freq_val, param_name, step_label, diff_val in all_flat_rows:
            try:
                f = float(freq_val) if freq_val is not None else 0.0
            except (ValueError, TypeError):
                f = 0.0
            try:
                d = float(diff_val) if diff_val is not None else 0.0
            except (ValueError, TypeError):
                d = 0.0
            flat_ws.append([f, str(param_name), str(step_label), round(d, 6)])
        return flat_ws, len(all_flat_rows)

    @staticmethod
    def _build_pivot_table(wb, flat_ws, n_data):
        """在 DiffChart sheet 创建含 PageField 下拉筛选的 PivotTable。"""
        from openpyxl.pivot.cache import CacheDefinition, CacheSource, WorksheetSource
        from openpyxl.pivot.table import (
            TableDefinition, Location, PageField, DataField, RowColField, PivotField
        )
        from openpyxl.styles import Font, Alignment

        last_col_letter = get_column_letter(4)
        data_ref = f"A1:{last_col_letter}{n_data + 1}"

        cache_src = CacheSource(
            type="worksheet",
            worksheetSource=WorksheetSource(ref=data_ref, sheet="_diff_flat")
        )
        cache_def = CacheDefinition(cacheSource=cache_src)
        existing_ids = {getattr(p, "cacheId", None) for ws in wb.worksheets
                        for p in getattr(ws, "_pivots", [])}
        existing_ids.discard(None)
        cache_def.cacheId = max(existing_ids, default=-1) + 1

        pt_ws = _safe_create_sheet(wb, "DiffChart")
        pivot_fields = [
            PivotField(name="Frequency"),
            PivotField(name="Parameter", axis="axisPage"),
            PivotField(name="StepAngle", axis="axisPage"),
            PivotField(name="DiffValue"),
        ]
        pt = TableDefinition(
            name="DiffPivot", cacheId=cache_def.cacheId,
            dataOnRows=True, dataCaption="Values",
            grandTotalCaption="Grand Total", errorCaption="#VALUE!", showError=False,
            missingCaption="", showMissing=True, updatedVersion=3, minRefreshableVersion=3,
            asteriskTotals=False, showItems=True, editData=False, disableFieldList=False,
            showCalcMbrs=True, visualTotals=True, showMultipleLabel=True,
            showDataDropDown=True, showDrill=True, printDrill=False,
            showMemberPropertyTips=True, showDataTips=True,
            location=Location(ref="A3:D20", firstHeaderRow=3, firstDataRow=4, firstDataCol=1),
            pivotFields=pivot_fields,
            rowFields=[RowColField(x=0)],
            pageFields=[PageField(fld=1, name="参数"), PageField(fld=2, name="步进")],
            dataFields=[DataField(name="差值", fld=3, numFmtId=2)],
        )
        pt.cache = cache_def
        pt_ws.add_pivot(pt)

        pt_ws.cell(1, 1, "📊 交互式差值分析 — PivotTable 支持 参数/步进角度 下拉筛选")
        pt_ws.cell(2, 1, '💡 右键 PivotTable → "插入切片器" 添加视觉切片按钮 | 下方图表为全量数据概览')
        for r in (1, 2):
            cell = pt_ws.cell(r, 1)
            cell.font = Font(bold=True, color="4472C4")
            cell.alignment = Alignment(horizontal="left")
        return pt_ws

    @staticmethod
    def _add_diff_scatter_chart(pt_ws, flat_ws, n_data):
        """在 PivotTable sheet 添加全量差值概览散点图。"""
        chart = ScatterChart()
        chart.title = "步进差值 vs Frequency（全量概览 — 使用 PivotTable 筛选分析）"
        chart.style = 2; chart.width = 18; chart.height = 11
        chart.x_axis.title = "Frequency (MHz)"
        chart.y_axis.title = "差值"
        chart.y_axis.numFmt = "0.000"
        chart.legend.position = "b"

        x_vals = Reference(flat_ws, min_col=1, min_row=2, max_row=n_data + 1)
        y_vals = Reference(flat_ws, min_col=4, min_row=2, max_row=n_data + 1)
        series = Series(y_vals, x_vals, title="差值")
        series.marker.symbol = "circle"; series.marker.size = 5
        series.graphicalProperties.line.width = 14000
        series.smooth = False
        chart.series.append(series)

        pt_ws.add_chart(chart, f"A{4 + n_data + 3}")

    def _add_static_diff_charts_fallback(self, wb, sheets_by_base):
        """回退方案: 为每个参数生成独立静态图表 (旧逻辑)。"""
        for base, suffixed in sheets_by_base.items():
            orig_ws = suffixed.get("")
            if orig_ws is None or len(suffixed) < 2:
                continue
            # 用 _safe_create_sheet 计算名称，但 fallback 只需要读取已存在的 sheet
            max_base = 31 - len("_diff")
            safe_base = base[:max_base] if len(base) > max_base else base
            diff_name = f"{safe_base}_diff"
            if diff_name not in {ws.title for ws in wb.worksheets}:
                continue
            dws = wb[diff_name]
            rows = list(dws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(c or "") for c in rows[0]]
            data_rows = rows[1:]
            if not data_rows:
                continue
            n_rows = len(data_rows)

            # 解析列头, 按参数分组找到对应的差值列
            # 列头格式: "Gain (步进3°)", "Gain 差值 (步进3°)", "AR (步进3°)", ...
            # 布局: 每个步进块内交替出现 value 列和 diff 列 (diff 列在步进块中的奇数位置)
            # 差值列特征: 列头包含 " 差值 (" 而非参数名本身含 "差值"
            param_diff_map = {}  # param_name -> [(step_label, diff_col)]
            for ci, h in enumerate(headers):
                marker = " 差值 ("
                if not h or marker not in h:
                    continue
                # 安全解析: 用 " 差值 (" 分隔，确保参数名不含此标记
                idx = h.index(marker)
                param_name = h[:idx].strip()
                step_label = h[idx + len(marker):].rstrip(")")
                param_diff_map.setdefault(param_name, []).append((step_label, ci + 1))

            if not param_diff_map:
                continue

            chart_row = n_rows + 5
            param_idx = 0
            x_vals = Reference(dws, min_col=1, min_row=2, max_row=n_rows + 1)

            for param_name, step_diffs in param_diff_map.items():
                chart = ScatterChart()
                chart.title = f"{param_name} 步进差值 vs Frequency"
                chart.style = 2
                chart.width = 16; chart.height = 10
                for step_label, diff_col in step_diffs:
                    y_vals = Reference(dws, min_col=diff_col, min_row=2, max_row=n_rows + 1)
                    series = Series(y_vals, x_vals, title=step_label)
                    series.marker.symbol = 'circle'
                    series.marker.size = 5
                    series.graphicalProperties.line.width = 14000
                    series.smooth = True
                    chart.series.append(series)
                chart.x_axis.title = "Frequency (MHz)"
                chart.y_axis.title = f"{param_name} 差值"
                chart.y_axis.numFmt = '0.000'
                chart.legend.position = 'b'
                chart_row_offset = chart_row + param_idx * 18
                dws.add_chart(chart, f"E{chart_row_offset}")
                param_idx += 1
