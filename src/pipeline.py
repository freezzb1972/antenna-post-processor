"""
批处理管线 (v2)
================
协调 datasource → extrapolate → calculator → plotter → exporter。

支持:
  - 任意 DataSource (MergedCSV / FinalSummary)
  - 内存中 Theta 外推 (无中间文件)
  - 逐频点处理 (峰值内存 ~2MB/频点)
  - 可选多进程并行
"""

from __future__ import annotations

import copy
import io
import math
import os
import re
import time
from collections.abc import Callable
from concurrent.futures import ProcessPoolExecutor
from typing import Any

import numpy as np

from .azimuth_config import AzimuthReportConfig
from .calculator import (
    compute_ar_at_angles,
    compute_ar_range,
    compute_average_gain_db,
    compute_average_power_dbm,
    compute_axial_ratio,
    compute_beamwidth,
    compute_boresight,
    compute_directivity,
    compute_efficiency,
    compute_lag_at_angles,
    compute_lag_ranges,
    compute_lower_hemisphere_prp,
    compute_min_power_dbm,
    compute_nhprp,
    compute_nhprp_flex,
    compute_partial_prp,
    compute_peak_eirp,
    compute_phase_center,
    compute_power_ratios,
    compute_prp_trp_ratio,
    compute_total_efficiency,
    compute_total_gain_linear,
    compute_trp,
    compute_upper_hemisphere_prp,
    compute_xpi,
)
from .chart_config import ChartConfig
from .datasource import DataSource
from .excel_reader import ColumnInfo, SheetInfo, read_template
from .exporter import export_results
from .lag_config import LagConfig
from .parser import MergedCSVParser
from .plot_config import PlotConfig
from .report_exporter import export_full_report

# ---------------------------------------------------------------------------
# Theta 外推 (pipeline 层)
# ---------------------------------------------------------------------------

def extrapolate_theta(
    theta_deg: np.ndarray,
    data: np.ndarray,  # (n_phi, n_theta)
    method: str = "linear",
) -> tuple[np.ndarray, np.ndarray]:
    """将 Theta 范围外推到 0-180°。

    Args:
        theta_deg: 原始 Theta 角度 (°)。
        data:      数据矩阵 (n_phi, n_theta)。
        method:    'linear' | 'constant' | 'mirror'。

    Returns:
        (new_theta_deg, new_data)，new_data 形状 (n_phi, n_new)。
    """
    max_t = theta_deg[-1]
    if max_t >= 179:
        return theta_deg.copy(), data.copy()

    n_phi, n_theta = data.shape
    dtheta = theta_deg[1] - theta_deg[0] if len(theta_deg) > 1 else 1.0

    new_theta = list(theta_deg)
    t = max_t + dtheta
    while t <= 180.01:
        new_theta.append(round(t, 6))
        t += dtheta

    n_new = len(new_theta)
    new_data = np.zeros((n_phi, n_new), dtype=np.float64)
    new_data[:, :n_theta] = data

    if method == "constant":
        tail_avg = np.mean(data[:, -10:], axis=1)
        new_data[:, n_theta:] = tail_avg[:, np.newaxis]

    elif method == "mirror":
        for i in range(n_new - n_theta):
            mirror_idx = n_theta - 2 - i
            idx = mirror_idx if mirror_idx >= 0 else 0
            new_data[:, n_theta + i] = data[:, idx]

    elif method == "linear":
        tail_n = min(10, n_theta)
        xv = theta_deg[-tail_n:]
        n = len(xv)
        sx = float(np.sum(xv))
        sxx = float(np.sum(xv * xv))
        denom_ok = n > 1 and sxx * n - sx * sx != 0
        for pi in range(n_phi):
            yv = data[pi, -tail_n:]
            sy = float(np.sum(yv))
            sxy = float(np.sum(xv * yv))
            if denom_ok:
                slope = (n * sxy - sx * sy) / (n * sxx - sx * sx)
                intercept = (sy - slope * sx) / n
                if slope > 0:
                    slope = 0; intercept = sy / n
                peak = float(np.max(yv))
                floor = peak - 40
                for i in range(n_theta, n_new):
                    val = slope * new_theta[i] + intercept
                    val = min(val, yv[-1])
                    new_data[pi, i] = max(val, floor)

    return np.array(new_theta), new_data


# ---------------------------------------------------------------------------
# 单频点处理
# ---------------------------------------------------------------------------

def _process_one_frequency(
    raw: dict[str, np.ndarray | None],
    freq: float,
    theta_deg: np.ndarray,
    lag_config: LagConfig,
    *,
    do_extrapolate: bool = False,
    robust_peak: bool = False,
    needed_params: set = None,
    extra_params: set = None,
    chart_config: ChartConfig = None,
    ar_lag_config: LagConfig = None,
    azimuth_config: AzimuthReportConfig = None,
    nh_custom_angles: list[float] | None = None,
    ar_output_db: bool = True,
    dir_extrap_method: str = "linear",
    compute_only: bool = False,
    log_cb=None,
) -> dict[str, Any]:
    """处理单个频点。按 needed_params（模板列）+ extra_params（用户额外）计算。"""
    theta_lm = raw["theta_logmag"]
    phi_lm = raw["phi_logmag"]
    need = needed_params or set()
    extra = extra_params or set()
    compute_set = need | extra  # 实际计算: 模板需求 + 用户额外

    need_extrap = do_extrapolate and theta_deg[-1] < 175
    if need_extrap:
        theta_orig = theta_deg.copy()
        new_theta, theta_lm = extrapolate_theta(theta_deg, theta_lm, "linear")
        _, phi_lm = extrapolate_theta(theta_deg, phi_lm, "linear")
        theta_deg = new_theta

    theta_rad = np.deg2rad(theta_deg)
    gain_linear, peak_dbi = compute_total_gain_linear(theta_lm, phi_lm, robust=robust_peak)
    row: dict[str, Any] = {"frequency": freq}

    # Gain (always include if template has it)
    if "gain" in need or "peak_eirp" in compute_set or not need:
        row["gain"] = round(peak_dbi, 6)

    # Directivity (需全球面积分; theta < 175° 时临时外推补全)
    directivity_dbi = None
    need_dir_or_eff = ("directivity" in compute_set or "efficiency_pct" in need
                       or "efficiency_db" in compute_set or not need)
    if need_dir_or_eff:
        if theta_deg[-1] < 175 and not do_extrapolate:
            # 外推 LogMag 数据到 180° 仅用于 Directivity, 不覆盖原始数据
            ext_th, ext_tl = extrapolate_theta(theta_deg, theta_lm, dir_extrap_method)
            _, ext_pl = extrapolate_theta(theta_deg, phi_lm, dir_extrap_method)
            ext_gl, _ = compute_total_gain_linear(ext_tl, ext_pl)
            dir_gl = ext_gl
            dir_theta = np.deg2rad(ext_th)
        else:
            dir_gl = gain_linear
            dir_theta = theta_rad
        directivity_dbi = compute_directivity(dir_gl, dir_theta)
        if "directivity" in compute_set or not need:
            row["directivity"] = round(directivity_dbi, 6)

    # Efficiency
    if "efficiency_pct" in need or "efficiency_db" in compute_set or not need:
        if directivity_dbi is None:
            directivity_dbi = compute_directivity(dir_gl, dir_theta)
        eff_pct, eff_db = compute_efficiency(peak_dbi, directivity_dbi)
        if "efficiency_pct" in compute_set or not need:
            row["efficiency_pct"] = round(eff_pct, 6)
        if "efficiency_db" in compute_set or not need:
            row["efficiency_db"] = round(eff_db, 6)

    # TRP / NHPRP / Peak EIRP
    for ct, fn in [("trp", lambda: compute_trp(gain_linear, theta_rad)),
                   ("nhprp_45", lambda: compute_nhprp(gain_linear, theta_rad, 45.0)),
                   ("nhprp_30", lambda: compute_nhprp(gain_linear, theta_rad, 30.0)),
                   ("peak_eirp", lambda: compute_peak_eirp(gain_linear))]:
        if ct in compute_set or not need:
            row[ct] = round(fn(), 2)

    # ---- Extended antenna parameters (NHPRP flex, PRP, ratios, boresight, averages) ----
    n_phi = phi_lm.shape[0]
    phi_angles_deg = np.linspace(0, 360, n_phi, endpoint=False)

    nhprp_225 = compute_nhprp_flex(gain_linear, theta_rad, 22.5)
    nhprp_45_flex = compute_nhprp_flex(gain_linear, theta_rad, 45.0)
    nhprp_30_flex = compute_nhprp_flex(gain_linear, theta_rad, 30.0)
    if "nhprp_225" in compute_set or not need: row["nhprp_225"] = round(nhprp_225, 2)

    # TIS = TRP (同公式, 不同测量模式)
    if "tis" in compute_set or not need: row["tis"] = round(compute_trp(gain_linear, theta_rad), 2)
    # NHPIS = NHPRP (同公式, 不同测量模式)
    for label, edge in [("nhpis_45", 45.0), ("nhpis_30", 30.0), ("nhpis_225", 22.5)]:
        if label in compute_set or not need: row[label] = round(compute_nhprp_flex(gain_linear, theta_rad, edge), 2)
    # Custom NHPRP/NHPIS angle
    _nh_angles = nh_custom_angles if nh_custom_angles else [45.0]  # default: 45°
    if "nhprp_custom" in compute_set or not need:
        for edge in _nh_angles:
            row[f"nhprp_custom_{int(edge)}"] = round(compute_nhprp_flex(gain_linear, theta_rad, edge), 2)
        # 向后兼容: 保留第一个角度键
        row["nhprp_custom"] = round(compute_nhprp_flex(gain_linear, theta_rad, _nh_angles[0]), 2)
    if "nhpis_custom" in compute_set or not need:
        for edge in _nh_angles:
            row[f"nhpis_custom_{int(edge)}"] = round(compute_nhprp_flex(gain_linear, theta_rad, edge), 2)
        row["nhpis_custom"] = round(compute_nhprp_flex(gain_linear, theta_rad, _nh_angles[0]), 2)

    uh_prp = compute_upper_hemisphere_prp(gain_linear, theta_rad)
    lh_prp = compute_lower_hemisphere_prp(gain_linear, theta_rad)
    if "uh_prp" in compute_set or not need: row["uh_prp"] = round(uh_prp, 2)
    if "lh_prp" in compute_set or not need: row["lh_prp"] = round(lh_prp, 2)
    if "uh_pis" in compute_set or not need: row["uh_pis"] = round(uh_prp, 2)
    if "lh_pis" in compute_set or not need: row["lh_pis"] = round(lh_prp, 2)

    prp_120 = compute_partial_prp(gain_linear, theta_rad, 0, 120)
    if "prp_120" in compute_set or not need: row["prp_120"] = round(prp_120, 2)
    if "pis_120" in compute_set or not need: row["pis_120"] = round(prp_120, 2)

    ratio_need = compute_set & {"nhprp45_ratio", "nhprp30_ratio", "nhprp225_ratio", "nhpis45_ratio",
                                  "nhpis30_ratio", "nhpis225_ratio", "uh_ratio", "lh_ratio"}
    if ratio_need or not need:
        trp_val = compute_trp(gain_linear, theta_rad)
        tis_val = trp_val  # same formula
        for ratio_key, prp_val, ref_val in [
            ("nhprp45", nhprp_45_flex, trp_val), ("nhprp30", nhprp_30_flex, trp_val),
            ("nhprp225", nhprp_225, trp_val), ("nhpis45", nhprp_45_flex, tis_val),
            ("nhpis30", nhprp_30_flex, tis_val), ("nhpis225", nhprp_225, tis_val),
            ("uh", uh_prp, trp_val), ("lh", lh_prp, trp_val)]:
            rdb, rpct = compute_prp_trp_ratio(prp_val, ref_val)
            if f"{ratio_key}_ratio_db" in compute_set or not need: row[f"{ratio_key}_ratio_db"] = round(rdb, 2)
            if f"{ratio_key}_ratio_pct" in compute_set or not need: row[f"{ratio_key}_ratio_pct"] = round(rpct, 2)

    bs_need = compute_set & {"boresight_phi", "boresight_theta"}
    if bs_need or not need:
        bs_theta, bs_phi = compute_boresight(gain_linear, theta_deg, phi_angles_deg)
        if "boresight_theta" in compute_set or not need: row["boresight_theta"] = round(bs_theta, 1)
        if "boresight_phi" in compute_set or not need: row["boresight_phi"] = round(bs_phi, 1)

    if "max_power" in compute_set or not need: row["max_power"] = round(compute_peak_eirp(gain_linear), 2)
    if "min_power" in compute_set or not need: row["min_power"] = round(compute_min_power_dbm(gain_linear), 2)
    if "avg_gain" in compute_set or not need: row["avg_gain"] = round(compute_average_gain_db(gain_linear), 2)
    if "avg_power" in compute_set or not need: row["avg_power"] = round(compute_average_power_dbm(gain_linear), 2)

    # Power ratios + Beamwidth
    if any(c in compute_set for c in ("max_min_ratio", "max_avg_ratio", "min_avg_ratio",
                                 "theta_bw", "phi_bw", "front_back_ratio")) or not need:
        max_p = float(np.max(gain_linear)); min_p = float(np.min(gain_linear[gain_linear>1e-15]) if np.any(gain_linear>1e-15) else 1e-15)
        avg_p = float(np.mean(gain_linear))
        ratios = compute_power_ratios(10*np.log10(max(max_p,1e-15)), 10*np.log10(max(min_p,1e-15)), 10*np.log10(max(avg_p,1e-15)))
        for k, v in ratios.items(): row[k] = v
        bw = compute_beamwidth(gain_linear, theta_deg, phi_angles_deg)
        for k, v in bw.items(): row[k] = v if v is not None else 0

    # Axial Ratio (仅当有 Phase 数据且需要 AR 列)
    # 若 azimuth_config 要求 AR 方位面图，强制计算 AR
    ar_need = compute_set & {"axial_ratio", "ar_single", "ar_range"}
    az_force_ar = (azimuth_config is not None and azimuth_config.cut_azimuth_polar_ar)
    if ar_need or az_force_ar or not need:
        tp = raw.get("theta_phase"); pp = raw.get("phi_phase")
        if tp is not None and pp is not None:
            try:
                if need_extrap:
                    _, tp = extrapolate_theta(theta_orig, tp, "constant")
                    _, pp = extrapolate_theta(theta_orig, pp, "constant")
                ar_result = compute_axial_ratio(theta_lm, tp, phi_lm, pp)
                if ar_result is not None and ar_result[0].size > 0:
                    ar, _, _ = ar_result
                    # AR 使用独立的 ar_lag_config 或 fallback 到 lag_config
                    ar_cfg = ar_lag_config if ar_lag_config is not None and not ar_lag_config.is_empty() else lag_config
                    ar_singles = ar_cfg.singles_sorted
                    if ar_singles:
                        for angle, val in compute_ar_at_angles(ar, theta_deg, ar_singles).items():
                            if ar_output_db:
                                val = 20.0 * math.log10(max(val, 1e-15))
                            row[f"ar_single_{angle}"] = round(val, 6)
                    # AR 范围
                    ar_ranges = ar_cfg.ranges_sorted
                    if ar_ranges:
                        for (lo, hi), val in [(r, compute_ar_range(ar, theta_deg, r[0], r[1])) for r in ar_ranges]:
                            if ar_output_db:
                                val = 20.0 * math.log10(max(val, 1e-15))
                            row[f"ar_range_{lo}_{hi}"] = round(val, 6)
                    # 向后兼容的 axial_ratio 字段
                    if "axial_ratio" in compute_set or not need:
                        legacy_ar = float(np.mean(ar[0, :5]))  # ar is set above from ar_result
                        if ar_output_db:
                            legacy_ar = 10.0 * math.log10(max(legacy_ar, 1e-15))
                        row["axial_ratio"] = round(legacy_ar, 6)
            except Exception as e:
                row["axial_ratio_error"] = str(e)

    # Cross-Polarization Isolation (XPI)
    xpi_need = compute_set & {"xpi_boresight", "xpi_mean", "xpi_min"}
    if xpi_need or not need:
        xpi_result = compute_xpi(theta_lm, phi_lm)
        if "xpi_boresight" in compute_set or not need:
            row["xpi_boresight"] = round(xpi_result["xpi_boresight"], 6)
        if "xpi_mean" in compute_set or not need:
            row["xpi_mean"] = round(xpi_result["xpi_mean"], 6)
        if "xpi_min" in compute_set or not need:
            row["xpi_min"] = round(xpi_result["xpi_min"], 6)

    # Total Efficiency (含 S11 反射损耗 — S11 当前不可用, 标记为 None)
    te_need = compute_set & {"total_efficiency_pct", "mismatch_loss_db"}
    if te_need or not need:
        if directivity_dbi is None:
            directivity_dbi = compute_directivity(gain_linear, theta_rad)
        te_eff_pct, _ = compute_efficiency(peak_dbi, directivity_dbi)
        te_result = compute_total_efficiency(te_eff_pct)
        if te_result and ("total_efficiency_pct" in compute_set or not need):
            if te_result.get("total_efficiency_pct") is not None:
                row["total_efficiency_pct"] = round(te_result["total_efficiency_pct"], 6)
        if te_result and ("mismatch_loss_db" in compute_set or not need):
            if te_result.get("mismatch_loss_db") is not None:
                row["mismatch_loss_db"] = round(te_result["mismatch_loss_db"], 6)
        # total_efficiency_pct 为 None 说明 S11 数据未提供
        if te_result["total_efficiency_pct"] is None and log_cb:
            _log(log_cb,
                 f"  ℹ {freq} MHz: Total Efficiency 需 S11 (回波损耗) 数据，当前标记为 None")

    # Phase Center (仅当有 Phase 数据)
    pc_need = compute_set & {"pc_theta_mm", "pc_phi_mm"}
    if pc_need or not need:
        tp_r = raw.get("theta_phase"); pp_r = raw.get("phi_phase")
        if tp_r is not None and pp_r is not None:
            pc_result = compute_phase_center(tp_r, pp_r, theta_deg, freq)
            if "pc_theta_mm" in compute_set or not need:
                row["pc_theta_mm"] = round(pc_result["pc_theta_mm"], 6)
            if "pc_phi_mm" in compute_set or not need:
                row["pc_phi_mm"] = round(pc_result["pc_phi_mm"], 6)

    # LAG
    singles = lag_config.singles_sorted
    if singles:
        for angle, val in compute_lag_at_angles(gain_linear, theta_deg, singles).items():
            row[f"lag_single_{angle}"] = round(val, 6)
    ranges = lag_config.ranges_sorted
    if ranges:
        for (lo, hi), val in compute_lag_ranges(gain_linear, theta_deg, ranges).items():
            row[f"lag_range_{lo}_{hi}"] = round(val, 6)

    # ── RHCP/LHCP Gain + CP-XPI (始终计算, 查看器需要 + 轻量公式) ──
    always_compute_rhcp = True
    if always_compute_rhcp:
        tp = raw.get("theta_phase"); pp = raw.get("phi_phase")
        if tp is not None and pp is not None:
            try:
                from .calculator import compute_cp_xpi, compute_rhcp_lhcp_gain
                if need_extrap:
                    _, tp = extrapolate_theta(theta_orig, tp, "constant")
                    _, pp = extrapolate_theta(theta_orig, pp, "constant")
                rhcp_g, lhcp_g = compute_rhcp_lhcp_gain(theta_lm, tp, phi_lm, pp)
                cp_xpi = compute_cp_xpi(rhcp_g, lhcp_g)

                if True:  # RHCP single always computed
                    for angle, val in compute_lag_at_angles(
                        rhcp_g, theta_deg, lag_config.singles_sorted
                    ).items():
                        row[f"rhcp_single_{angle}"] = round(val, 6)
                if lag_config.ranges_sorted:
                    for (lo, hi), val in compute_lag_ranges(
                        rhcp_g, theta_deg, lag_config.ranges_sorted
                    ).items():
                        row[f"rhcp_range_{lo}_{hi}"] = round(val, 6)

                if True:
                    for angle, val in compute_lag_at_angles(
                        cp_xpi, theta_deg, lag_config.singles_sorted
                    ).items():
                        row[f"cp_xpi_single_{angle}"] = round(val, 6)
                if True:
                    for (lo, hi), val in compute_lag_ranges(
                        cp_xpi, theta_deg, lag_config.ranges_sorted
                    ).items():
                        row[f"cp_xpi_range_{lo}_{hi}"] = round(val, 6)
            except Exception as e:
                row["rhcp_error"] = str(e)

        # RHCP/LHCP 矩阵存储供查看器使用
        if tp is not None and pp is not None and "_rhcp_gain" not in row:
            try:
                from .calculator import compute_rhcp_lhcp_gain
                rhcp_g, lhcp_g = compute_rhcp_lhcp_gain(theta_lm, tp, phi_lm, pp)
                row["_rhcp_gain"] = rhcp_g
                row["_lhcp_gain"] = lhcp_g
                row["_cp_xpi"] = rhcp_g - lhcp_g
            except Exception:
                pass

    # 计算总增益 dB 矩阵 (供图形和 2D Cuts 使用)
    gain_dbi = 10.0 * np.log10(np.maximum(gain_linear, 1e-15))

    # ── 图形生成 (A/C 类: 每频点 PNG + 方位面) ──
    az_need = (azimuth_config is not None and azimuth_config.has_any_azimuth)
    if (chart_config is not None and chart_config.has_any_pattern_or_cut) or az_need:
        try:
            from .plotter import generate_all_for_frequency
            # 确保 chart_config 不为 None（纯 azimuth 模式时 ChartConfig 可能为 None）
            ccfg = chart_config if chart_config is not None else ChartConfig()
            n_phi = phi_lm.shape[0]
            phi_angles = np.linspace(0, 360, n_phi, endpoint=False)
            # AR 线性值（如果需要 3D AR 或 方位面 AR）
            ar_lin = None
            need_ar_for_graphics = (
                ccfg.pattern_3d_ar or
                (azimuth_config is not None and azimuth_config.cut_azimuth_polar_ar)
            )
            if need_ar_for_graphics and "axial_ratio" not in str(row.get("axial_ratio_error", "")):
                tp = raw.get("theta_phase"); pp = raw.get("phi_phase")
                if tp is not None and pp is not None:
                    ar_result = compute_axial_ratio(theta_lm, tp, phi_lm, pp)
                    if ar_result is not None:
                        ar_lin = ar_result[0]
            # compute_only 模式跳过 Matplotlib 渲染
            if not compute_only:
                # 构建 E_θ/E_φ 分量额外数据
                extra_patterns = {}
                if ccfg.pattern_3d_etheta:
                    extra_patterns["3d_etheta"] = theta_lm
                if ccfg.pattern_3d_ephi:
                    extra_patterns["3d_ephi"] = phi_lm
                extra_patterns = extra_patterns if extra_patterns else None
                # RHCP/LHCP 增益矩阵 (供方位图)
                rhcp_db = row.get("_rhcp_gain")
                lhcp_db = row.get("_lhcp_gain")
                images = generate_all_for_frequency(
                    theta_deg, phi_angles, gain_dbi,
                    freq, ccfg, ar_linear=ar_lin,
                    rhcp_db=rhcp_db, lhcp_db=lhcp_db,
                    antenna_name="",
                    azimuth_config=azimuth_config,
                    extra_patterns=extra_patterns,
                )
                if images:
                    row["_images"] = images
            # 存储中间数据供方位面导出使用
            if azimuth_config.cut_azimuth_polar and azimuth_config.azimuth_cut_angles:
                row["_azimuth_gain_dbi"] = gain_dbi
                row["_azimuth_theta_deg"] = theta_deg.copy()
            if azimuth_config.cut_azimuth_polar_ar and ar_lin is not None and azimuth_config.azimuth_cut_angles_ar:
                row["_azimuth_ar_db"] = 20.0 * np.log10(np.maximum(ar_lin, 1e-15))
                if "_azimuth_theta_deg" not in row:
                    row["_azimuth_theta_deg"] = theta_deg.copy()
            if azimuth_config.cut_azimuth_polar_rhcp and rhcp_db is not None and azimuth_config.azimuth_cut_angles_rhcp:
                row["_azimuth_rhcp_db"] = rhcp_db
                if "_azimuth_theta_deg" not in row:
                    row["_azimuth_theta_deg"] = theta_deg.copy()
            if azimuth_config.cut_azimuth_polar_lhcp and lhcp_db is not None and azimuth_config.azimuth_cut_angles_lhcp:
                row["_azimuth_lhcp_db"] = lhcp_db
                if "_azimuth_theta_deg" not in row:
                    row["_azimuth_theta_deg"] = theta_deg.copy()
            # Gain 0-70° Pk 中间数据
            if azimuth_config and azimuth_config.cut_azimuth_polar_pk070:
                mask = theta_deg <= 70.1
                row["_gain_pk070_deg"] = theta_deg[mask].copy()
                row["_gain_pk070_db"] = np.max(gain_dbi[:, mask], axis=1)
        except Exception as e:
            row["_graph_error"] = str(e)  # 图形生成失败不阻塞数据处理

    # 存储原始数据供图形展示使用
    # NOTE: 每频点存储 _raw_data 会大幅增加内存开销。
    # 若有 N 个频点，每个频点的数据为 (n_phi × n_theta) float64 矩阵，
    # 总内存占用 = N × n_phi × n_theta × 8 字节。
    # 对高分辨率扫描 (如 361×181) 和大量频点，可能达到数百 MB。
    # 建议通过 chart_config 限制图形生成频率数，或在 pipeline 层面
    # 仅存储必要的 raw_data。
    row["_raw_data"] = {k: v for k, v in raw.items() if v is not None}
    row["_theta_angles"] = list(theta_deg)
    row["_phi_angles"] = [float(i) for i in np.linspace(0, 360, phi_lm.shape[0], endpoint=False)]
    # 2D Cuts 模式数据源: E_θ/E_φ 分量 + 总增益
    row["theta_db"] = theta_lm
    row["phi_db"] = phi_lm
    row["gain_db"] = gain_dbi

    return row


# ---------------------------------------------------------------------------
# 查找最近频点
# ---------------------------------------------------------------------------

def _find_closest_freq(csv_freqs: list[float], target: float, tol=5.0) -> int | None:
    if not csv_freqs:
        return None
    best_idx = int(np.argmin([abs(f - target) for f in csv_freqs]))
    return best_idx if abs(csv_freqs[best_idx] - target) <= tol else None


# ---------------------------------------------------------------------------
# 模板工作表自动扩增
# ---------------------------------------------------------------------------

def _derive_sheet_name(reference_name: str, target_key: str) -> str:
    """从参考工作表名推导新工作表名。

    "5G1" + key="G2" → "5G2"
    "Antenna_G1" + key="G3" → "Antenna_G3"
    """
    m = re.search(r'G\d+', reference_name, re.IGNORECASE)
    tk = re.search(r'G\d+', target_key, re.IGNORECASE)
    if m and tk:
        return reference_name[:m.start()] + tk.group(0).upper() + reference_name[m.end():]
    return target_key


def _expand_template_sheets(
    sheets_info: list[SheetInfo],
    datasource_map: dict[str, DataSource],
    freq_source: str = "datasource",
    use_raw_name: bool = False,
) -> list[SheetInfo]:
    """当模板工作表数少于数据源数时，用第一个 sheet 为模板克隆其余 sheet。

    Args:
        sheets_info:    read_template() 返回的原始列表。
        datasource_map: {sheet_name: DataSource}。
        freq_source:    "datasource" → 新 sheet 用数据源频点；
                        "template" → 新 sheet 用模板最近邻匹配。
        use_raw_name:   True → 直接用 datasource_map 的键作工作表名，
                        丢弃旧工作表名，为每个数据源创建新工作表；
                        False → 从键名推导工作表名。

    Returns:
        扩展后的 SheetInfo 列表。
    """
    ref = sheets_info[0]

    if use_raw_name:
        # 文件名模式：丢弃旧工作表名，为每个数据源创建新工作表
        expanded: list[SheetInfo] = []
        matched_names: set = set()
    else:
        if len(sheets_info) >= len(datasource_map):
            return list(sheets_info)
        expanded = list(sheets_info)
        matched_names = {si.name for si in sheets_info}

    existing_ds_names = set(datasource_map.keys())

    # 找出有 datasource 但没对应 sheet 的名称
    unmatched = existing_ds_names - matched_names
    if not unmatched:
        return list(sheets_info)

    from .sheet_file_matcher import sanitize_sheet_name

    for ds_name in sorted(unmatched):
        if use_raw_name:
            new_name = sanitize_sheet_name(ds_name)
        else:
            from .sheet_file_matcher import extract_key
            key = extract_key(ds_name)
            new_name = sanitize_sheet_name(_derive_sheet_name(ref.name, key))

        # 深拷贝列头结构
        new_columns = [
            ColumnInfo(
                col_letter=c.col_letter,
                col_index=c.col_index,
                raw_header=c.raw_header,
                normalized_header=c.normalized_header,
                col_type=c.col_type,
            )
            for c in ref.columns
        ]

        ds = datasource_map[ds_name]
        if freq_source == "template":
            frequencies = list(ref.frequencies)
        else:
            frequencies = list(ds.frequencies)

        new_si = SheetInfo(
            name=new_name,
            header_row=ref.header_row,
            data_start_row=ref.data_start_row,
            data_end_row=ref.data_start_row + len(frequencies) - 1,
            columns=new_columns,
            frequencies=frequencies,
            lag_config=copy.deepcopy(ref.lag_config),
            theta_range=ref.theta_range,
        )
        expanded.append(new_si)

    return expanded


# ---------------------------------------------------------------------------
# 管线助手 — 任务收集 / 数据加载+计算
# ---------------------------------------------------------------------------

def _collect_tasks(
    sheets_info: list[Any],
    datasource: DataSource | None,
    datasource_map: dict[str, DataSource] | None,
    freq_source: str = "datasource",
    trim_start: int = 0,
    trim_end: int = 0,
    sheet_mode_map: dict[str, int] | None = None,
    log_cb=None,
) -> list[tuple[str, float, int, Any, DataSource]]:
    """收集所有 (sheet_name, freq, csv_idx, lag_cfg, ds) 任务。"""
    use_multi = datasource_map is not None
    tasks: list[tuple[str, float, int, Any, DataSource]] = []

    if use_multi:
        original_sheets = {si.name for si in sheets_info}
        all_ds_names = set(datasource_map.keys())
        expanded_names = all_ds_names - original_sheets
    else:
        expanded_names = set()

    for si in sheets_info:
        ds: DataSource | None = datasource_map.get(si.name) if use_multi else datasource
        if ds is None:
            if use_multi:
                _log(log_cb, f"  ⚠ {si.name}: 无匹配数据源 — 跳过")
            continue

        # 提取模板需要的参数类型
        needed_params = {c.col_type for c in si.columns}

        # 混合批处理: 记录 sheet 对应的测试模式
        smap = sheet_mode_map or {}
        sheet_mode = smap.get(si.name, 0)
        mode_label = {0: "无源", 1: "TRP", 2: "TIS"}.get(sheet_mode, "?")

        is_expanded = si.name in expanded_names if use_multi else False
        use_ds_freqs = (is_expanded and freq_source == "datasource")

        # 调试: 记录本 sheet 的频点来源信息
        ds_freq_count = len(ds.frequencies) if ds else 0
        tmpl_freq_count = len(si.frequencies) if si.frequencies else 0
        _log(log_cb, f"  [{si.name}] 数据源={ds_freq_count}频点, 模板={tmpl_freq_count}频点, expanded={is_expanded}")

        if not use_ds_freqs:
            dsfreqs = ds.frequencies
            match_count = 0
            for freq in si.frequencies:
                idx = _find_closest_freq(dsfreqs, freq)
                if idx is not None:
                    tasks.append((si.name, freq, idx, si.lag_config, ds, needed_params))
                    match_count += 1
            if match_count == 0 and dsfreqs:
                _log(log_cb, f"  ↻ {si.name}: 模板频点无匹配 → 使用数据源全部 {len(dsfreqs)} 个频点")
                use_ds_freqs = True
            elif match_count > 0:
                _log(log_cb, f"  ✓ {si.name}: 模板匹配 {match_count}/{tmpl_freq_count} 个频点")

        if use_ds_freqs:
            dsfreqs = ds.frequencies
            for idx, freq in enumerate(dsfreqs):
                tasks.append((si.name, freq, idx, si.lag_config, ds, needed_params))

    # ---- 频点裁剪 (trim_start/trim_end) ----
    if trim_start > 0 or trim_end > 0:
        # 按 sheet 分组、按 freq 排序后裁剪首尾
        from collections import OrderedDict
        grouped: dict[str, list] = OrderedDict()
        for t in tasks:
            grouped.setdefault(t[0], []).append(t)
        tasks = []
        for sn, group in grouped.items():
            group.sort(key=lambda x: x[1])  # 按 freq 排序
            end = len(group) - trim_end if trim_end > 0 else len(group)
            trimmed = group[trim_start:end]
            if trimmed:
                tasks.extend(trimmed)
            if trim_start > 0 or trim_end > 0:
                removed = len(group) - len(trimmed)
                if removed > 0:
                    _log(log_cb, f"  ✂ {sn}: 去除 {removed} 个频点 (前{trim_start}后{trim_end})")

    _log(log_cb, f"共 {len(tasks)} 个待处理频点")
    return tasks


def _load_and_compute(
    tasks: list[tuple[str, float, int, Any, DataSource]],
    sheets_info: list[Any],
    extrapolate_theta: bool,
    robust_peak: bool,
    parallel: int,
    extra_params: set = None,
    chart_config: ChartConfig = None,
    ar_lag_config: LagConfig = None,
    sheet_ar_configs: dict[str, LagConfig] = None,
    azimuth_config: AzimuthReportConfig = None,
    nh_custom_angles: list[float] | None = None,
    ar_output_db: bool = True,
    dir_extrap_method: str = "linear",
    compute_only: bool = False,
    cancel_callback=None,
    progress_callback=None,
    log_callback=None,
) -> dict[str, list[dict[str, Any]]]:
    """加载原始数据并执行计算，返回 sheet_results。"""
    total = len(tasks)
    sheet_results: dict[str, list[dict[str, Any]]] = {si.name: [] for si in sheets_info}
    if total == 0:
        return sheet_results

    if sheet_ar_configs is None:
        sheet_ar_configs = {}

    progress_max = total * 2 + 50  # 加载+计算=2×total, 导出留 50 步

    # 阶段 A: 加载数据
    _log(log_callback, f"读取 {total} 个频点数据...")
    _report(progress_callback, 0, progress_max, "📂 加载数据 0%")
    compute_tasks = []
    for i, (sheet_name, freq, csv_idx, lag_cfg, task_ds, needed_params) in enumerate(tasks):
        if cancel_callback and cancel_callback():
            break
        raw = task_ds.read_sections(csv_idx)
        theta_list = list(task_ds.theta_angles)
        ar_cfg = ar_lag_config if ar_lag_config is not None and not ar_lag_config.is_empty() else sheet_ar_configs.get(sheet_name, LagConfig())
        compute_tasks.append((sheet_name, freq, raw, lag_cfg, theta_list, extrapolate_theta, robust_peak, needed_params, extra_params, chart_config, ar_cfg, nh_custom_angles, ar_output_db, azimuth_config, compute_only))
        _report(progress_callback, i + 1, progress_max, f"📂 加载数据 ({i+1}/{total})")

    data_done = len(compute_tasks)
    _report(progress_callback, data_done, progress_max, "🧮 计算中...")

    # 阶段 B: 计算（支持并行）
    if parallel > 1 and data_done > 1:
        _log(log_callback, f"并行计算: {parallel} 进程 × {data_done} 频点")
        chunk_size = max(1, len(compute_tasks) // parallel)
        chunks = [compute_tasks[i:i + chunk_size] for i in range(0, len(compute_tasks), chunk_size)]
        with ProcessPoolExecutor(max_workers=parallel) as executor:
            futures = [executor.submit(_compute_chunk, chunk) for chunk in chunks]
            completed = 0
            for fut in futures:
                if cancel_callback and cancel_callback():
                    for f in futures:
                        f.cancel()
                    break
                for sheet_name, row in fut.result():
                    sheet_results[sheet_name].append(row)
                    completed += 1
                _report(progress_callback, data_done + completed, progress_max,
                        f"🧮 计算中 {int((data_done+completed)/progress_max*100)}%")
    else:
        _run_compute_serial(compute_tasks, sheet_results, data_done, progress_max,
                            cancel_callback, progress_callback, log_cb=log_callback,
                            azimuth_config=azimuth_config,
                            dir_extrap_method=dir_extrap_method)

    return sheet_results


def _run_compute_serial(
    compute_tasks, sheet_results, data_done, progress_max,
    cancel_callback, progress_callback, log_cb=None, azimuth_config=None,
    dir_extrap_method="linear",
):
    """串行逐频点计算（单进程或 parallel=1）。"""
    for i, (sheet_name, freq, raw, lag_cfg, theta_list, do_extrap, rpk, nparams, xparams, ccfg, ar_cfg, nh_angles, ar_out_db, az_cfg, co) in enumerate(compute_tasks):
        if cancel_callback and cancel_callback():
            break
        try:
            theta_arr = np.array(theta_list)
            row = _process_one_frequency(raw, freq, theta_arr, lag_cfg, do_extrapolate=do_extrap, robust_peak=rpk, needed_params=nparams, extra_params=xparams, chart_config=ccfg, ar_lag_config=ar_cfg, azimuth_config=az_cfg, nh_custom_angles=nh_angles, ar_output_db=ar_out_db, dir_extrap_method=dir_extrap_method, compute_only=co, log_cb=log_cb)
            sheet_results[sheet_name].append(row)
        except Exception as e:
            sheet_results[sheet_name].append({"frequency": freq, "_error": str(e)})
        if (i + 1) % 10 == 0 or (i + 1) == data_done:
            _report(progress_callback, data_done + i + 1, progress_max,
                    f"🧮 计算中 {int((data_done+i+1)/progress_max*100)}%")


def _close_datasources(
    use_multi: bool,
    datasource: DataSource | None,
    datasource_map: dict[str, DataSource] | None,
):
    """安全关闭所有数据源。"""
    if use_multi and datasource_map:
        for ds in datasource_map.values():
            try:
                ds.close()
            except Exception:
                pass
    elif datasource:
        datasource.close()


# ---------------------------------------------------------------------------
# 主管线
# ---------------------------------------------------------------------------

def run_pipeline(
    datasource: DataSource | None = None,
    template_path: str = "",
    output_path: str = "",
    *,
    datasource_map: dict[str, DataSource] | None = None,
    sheet_mode_map: dict[str, int] | None = None,
    lag_config_override: LagConfig | None = None,
    ar_lag_config_override: LagConfig | None = None,
    plot_config: PlotConfig | None = None,
    full_report_path: str | None = None,
    compute_only: bool = False,  # True → 只计算不导出 (预览模式)
    extrapolate_theta: bool = False,
    freq_source: str = "datasource",
    trim_start: int = 0,
    trim_end: int = 0,
    chart_config_obj: ChartConfig | None = None,
    azimuth_config: AzimuthReportConfig | None = None,
    out_excel: bool = True,
    out_word: bool = False,
    out_data: bool = False,
    word_template_path: str | None = None,  # Word 模板路径 (为空则自动生成)
    dir_extrap_method: str = "linear",  # Directivity 外推方法: linear|constant|mirror
    robust_peak: bool = False,
    extra_params: set | None = None,
    nh_custom_angles: list[float] | None = None,
    ar_output_db: bool = True,
    worksheet_naming_mode: int = 0,  # 0=保留模板工作表名, 1=用数据源名
    parallel: int = 1,
    cancel_callback: Callable[[], bool] | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
    log_callback: Callable[[str], None] | None = None,
) -> dict[str, list[dict[str, Any]]]:
    """执行完整处理管线。

    Args:
        datasource:          数据源 (单源模式，向后兼容)。
        template_path:       模板 Excel 路径。
        output_path:         输出 Excel 路径。
        datasource_map:      工作表名→DataSource 映射 (多源模式)。
                             为 None 时使用 datasource 参数 (单源模式)。
        lag_config_override: LAG 配置覆盖。
        plot_config:         3D 图配置 (默认: 不生成图)。
        full_report_path:    完整报告路径。
        extrapolate_theta:   Theta 外推开关。
        freq_source:         "datasource" 或 "template"。
                             当模板 sheet 数<数据源数时，新 sheet 的频点来源。
        parallel:            (保留参数，当前仅串行)。
        cancel_callback / progress_callback / log_callback: 同旧版。

    Returns:
        {sheet_name: [row_dict, ...]}
    """
    if plot_config is None:
        plot_config = PlotConfig(embed_in_excel=False)

    t0 = time.time()

    # ---- 0. 参数校验 ----
    _log(log_callback, f"AR 输出: {'dB (20·log₁₀)' if ar_output_db else '线性比值'}")
    if datasource_map is not None and datasource is not None:
        raise ValueError("datasource 和 datasource_map 互斥，只能提供一个")
    if datasource_map is None and datasource is None:
        raise ValueError("必须提供 datasource 或 datasource_map")
    use_multi_ds = datasource_map is not None

    # ---- 1. 读取模板 + LAG ----
    _log(log_callback, f"读取模板: {template_path}")
    sheets_info = read_template(template_path)
    for si in sheets_info:
        _log(log_callback, f"  {si.name}: {len(si.frequencies)} 频点")

    # ---- 1.5: 自动扩增工作表 (模板 sheet 数 < 数据源数, 或文件名模式) ----
    template_sheet_names_to_remove: list[str] | None = None
    if use_multi_ds:
        template_names = {si.name for si in sheets_info}
        ds_names = set(datasource_map.keys())
        matched_count = len(template_names & ds_names)
        if worksheet_naming_mode == 1:
            # 文件名模式: 用第一个 sheet 做模板, datasource_map 键直接作工作表名
            _log(log_callback, f"文件名模式: {len(datasource_map)} 个数据源 → 创建 {len(datasource_map)} 个工作表...")
            template_sheet_names_to_remove = [si.name for si in sheets_info]
            ref_sheet = sheets_info[0:1]
            sheets_info = _expand_template_sheets(ref_sheet, datasource_map, freq_source, use_raw_name=True)
            for si in sheets_info:
                _log(log_callback, f"  {si.name}: {len(si.frequencies)} 频点")
        elif len(sheets_info) < len(datasource_map):
            _log(log_callback, f"模板 {len(sheets_info)} 个工作表 → {len(datasource_map)} 个数据源，自动扩增...")
            sheets_info = _expand_template_sheets(sheets_info, datasource_map, freq_source)
            for si in sheets_info:
                _log(log_callback, f"  {si.name}: {len(si.frequencies)} 频点 (来源: {'数据源' if freq_source == 'datasource' else '模板'})")

    if lag_config_override is not None and not lag_config_override.is_empty():
        for si in sheets_info:
            si.lag_config = lag_config_override
        _log(log_callback, "使用用户指定的 LAG 配置")

    # ---- 构建 sheet→AR 配置映射 (自动检测或使用覆盖) ----
    sheet_ar_configs: dict[str, LagConfig] = {}
    if ar_lag_config_override is not None and not ar_lag_config_override.is_empty():
        # 用户覆盖: 所有 sheet 用同一个 AR 配置
        for si in sheets_info:
            sheet_ar_configs[si.name] = ar_lag_config_override
        _log(log_callback, "使用用户指定的 AR 配置")
    else:
        # 自动检测: 从模板列头解析 AR 角度
        for si in sheets_info:
            if si.ar_config is not None and not si.ar_config.is_empty():
                sheet_ar_configs[si.name] = si.ar_config
        if sheet_ar_configs:
            total_angles = sum(len(c.singles_sorted) + len(c.ranges_sorted) for c in sheet_ar_configs.values())
            _log(log_callback, f"自动检测到 {total_angles} 个 AR 角度 (来自 {len(sheet_ar_configs)} 个 sheet)")

    # ---- 2. 收集任务 + 加载数据 + 计算 ----
    tasks = _collect_tasks(sheets_info, datasource, datasource_map, freq_source, trim_start, trim_end, sheet_mode_map, log_callback)
    try:
        sheet_results = _load_and_compute(
            tasks, sheets_info, extrapolate_theta, robust_peak, parallel,
            extra_params=extra_params, chart_config=chart_config_obj,
            ar_lag_config=ar_lag_config_override,
            sheet_ar_configs=sheet_ar_configs,
            azimuth_config=azimuth_config,
            nh_custom_angles=nh_custom_angles,
            ar_output_db=ar_output_db,
            dir_extrap_method=dir_extrap_method,
            compute_only=compute_only,
            cancel_callback=cancel_callback, progress_callback=progress_callback, log_callback=log_callback,
        )
    finally:
        _close_datasources(use_multi_ds, datasource, datasource_map)

    # ── compute_only 模式下跳过所有导出步骤 ──
    if compute_only:
        _log(log_callback, "⏭ 预览模式 — 跳过 Excel/Word/报告导出")
        elapsed = time.time() - t0
        total_rows = sum(len(v) for v in sheet_results.values())
        _log(log_callback, f"✓ 计算完成: {total_rows} 行, {elapsed:.1f}s")
        _report(progress_callback, 1, 1, "✅ 预览就绪")
        return sheet_results

    # ---- 3. 写入 Excel (可选) ----
    total = len(tasks)
    progress_max = total * 2 + 50  # 加载+计算=2×total, 导出留 50 步
    if out_excel and output_path:
        _log(log_callback, f"写入输出: {output_path}")
        _report(progress_callback, progress_max - 3, progress_max, "💾 写入 Excel...")
        export_results(
            template_path=template_path,
            output_path=output_path,
            sheet_results=sheet_results,
            pattern_images=None,
            sheets_info=sheets_info,
            chart_config=chart_config_obj,
            log_callback=log_callback,
            remove_template_sheets=template_sheet_names_to_remove,
        )
        _report(progress_callback, progress_max - 1, progress_max, "✅ Excel 写入完成")
    elif not out_excel:
        _log(log_callback, "⏭ 跳过天线参数 Excel 输出")

    # ---- 4. 完整报告 (可选) ----
    if full_report_path:
        _log(log_callback, f"生成完整报告: {full_report_path}")
        imgs_3d, imgs_2d_polar, imgs_2d_rect = _collect_report_images(sheet_results)
        export_full_report(
            output_path=full_report_path,
            sheet_results=sheet_results,
            pattern_images_3d=imgs_3d if imgs_3d else None,
            pattern_images_2d_polar=imgs_2d_polar if imgs_2d_polar else None,
            pattern_images_2d_rect=imgs_2d_rect if imgs_2d_rect else None,
        )

    # ---- 5. 方位面报告 (可选) ----
    if out_word or out_data:
        _export_azimuth(sheet_results, azimuth_config, log_callback,
                        out_word=out_word, out_data=out_data,
                        word_template_path=word_template_path)

    elapsed = time.time() - t0
    total_rows = sum(len(v) for v in sheet_results.values())
    _log(log_callback, f"✓ 完成: {total_rows} 行, {elapsed:.1f}s")
    _report(progress_callback, progress_max, progress_max, "✅ 完成")

    return sheet_results

# ---------------------------------------------------------------------------
# 报告图片收集
# ---------------------------------------------------------------------------

def _collect_report_images(
    sheet_results: dict[str, list[dict[str, Any]]],
) -> tuple:
    """从处理结果收集图片，按 3D/2D Polar/2D Rect 分类。

    图片在 _process_one_frequency 中已渲染为 PNG BytesIO，
    存入 row["_images"]。此函数按 report_exporter 期望的格式重组。

    Returns:
        (images_3d, images_2d_polar, images_2d_rect)
        - images_3d:        {sheet_name: {freq: BytesIO}}
        - images_2d_polar:  {sheet_name: {freq: {cut_label: BytesIO}}}
        - images_2d_rect:   {sheet_name: {freq: {cut_label: BytesIO}}}
    """
    images_3d: dict[str, dict[float, io.BytesIO]] = {}
    images_2d_polar: dict[str, dict[float, dict[str, io.BytesIO]]] = {}
    images_2d_rect: dict[str, dict[float, dict[str, io.BytesIO]]] = {}

    for sheet_name, rows in sheet_results.items():
        for row in rows:
            freq = row.get("frequency")
            if freq is None:
                continue
            imgs = row.get("_images", {})
            if not imgs:
                continue
            for img_key, buf in imgs.items():
                if buf is None:
                    continue
                if img_key.startswith("3d_"):
                    images_3d.setdefault(sheet_name, {})[freq] = buf
                elif img_key.startswith("2d_polar_"):
                    cut_label = img_key[len("2d_polar_"):]
                    images_2d_polar.setdefault(sheet_name, {}).setdefault(freq, {})[cut_label] = buf
                elif img_key.startswith("2d_rect_"):
                    cut_label = img_key[len("2d_rect_"):]
                    images_2d_rect.setdefault(sheet_name, {}).setdefault(freq, {})[cut_label] = buf

    return images_3d, images_2d_polar, images_2d_rect


# ---------------------------------------------------------------------------
# 方位面报告导出
# ---------------------------------------------------------------------------

def _export_azimuth(
    sheet_results: dict[str, list[dict[str, Any]]],
    azimuth_config: AzimuthReportConfig,
    log_callback=None,
    out_word: bool = True,
    out_data: bool = True,
    word_template_path: str | None = None,
):
    """从处理结果中收集方位面图片和中间数据，写入 Word 和 Excel。

    当 word_template_path 不为空时，使用 WordReporter 填充模板；
    否则使用 chart_word_writer 自动生成 Word 报告。
    """
    from pathlib import Path

    from .azimuth_data_writer import write_azimuth_data
    from .chart_word_writer import write_chart_word_report

    # ── 收集所有图片和中间数据 ──
    image_groups: dict[str, dict[float, io.BytesIO]] = {}
    freq_gain_data: list[tuple[float, dict[float, np.ndarray]]] = []
    freq_ar_data: list[tuple[float, dict[float, np.ndarray]]] = []
    freq_rhcp_data: list[tuple[float, dict[float, np.ndarray]]] = []
    freq_lhcp_data: list[tuple[float, dict[float, np.ndarray]]] = []
    freq_gain_vs_theta: list[tuple[float, np.ndarray]] = []  # [(freq, pk070_values_over_phi)]


    # 图片类型 → 用户可读组名
    def _label_for_image_key(img_key: str) -> str:
        """将 image key 映射为用户可读组名。"""
        if img_key.startswith("2d_polar_phi"):
            phi = img_key[len("2d_polar_phi"):]
            return f"2D Polar Cut (φ={phi}°)"
        if img_key.startswith("2d_rect_phi"):
            phi = img_key[len("2d_rect_phi"):]
            return f"2D Rectangular Cut (φ={phi}°)"
        # 3D multi-view keys: 3d_gain_v0, 3d_gain_v1, ...
        if "_v" in img_key and any(img_key.startswith(p) for p in ("3d_gain", "3d_eirp", "3d_ar")):
            base = img_key.rsplit("_v", 1)[0]
            known = {"3d_gain": "3D Gain Pattern", "3d_eirp": "3D EIRP Pattern", "3d_ar": "3D Axial Ratio Pattern"}
            return known.get(base, img_key)
        known = {
            "3d_gain": "3D Gain Pattern",
            "3d_eirp": "3D EIRP Pattern",
            "3d_ar": "3D Axial Ratio Pattern",
            "azimuth_polar": "Gain Azimuth Cut",
            "azimuth_polar_pk070": "Gain Azimuth (θ=0°-70°)",
            "azimuth_polar_ar": "AR Azimuth Cut",
            "azimuth_polar_rhcp": "RHCP Azimuth Cut",
            "azimuth_polar_lhcp": "LHCP Azimuth Cut",
        }
        return known.get(img_key, img_key)

    for sheet_name, rows in sheet_results.items():
        for row in rows:
            freq = row.get("frequency")
            if freq is None:
                continue
            images = row.get("_images", {})
            for img_key, buf in images.items():
                if buf is None:
                    continue
                label = _label_for_image_key(img_key)
                if label not in image_groups:
                    image_groups[label] = {}
                image_groups[label][freq] = buf

    # ── B 类: 频点曲线 PNG (Word 报告) ──
    _B_PARAM_MAP = {
        "efficiency_pct": "Efficiency (%)", "gain": "Peak Gain (dBi)",
        "directivity": "Directivity (dBi)", "trp": "TRP (dBm)",
        "peak_eirp": "Peak EIRP (dBm)", "avg_gain": "Average Gain (dB)",
        "nhprp_45": "NHPRP ±45°", "nhprp_30": "NHPRP ±30°",
    }
    # 当模板无 Efficiency 列但有 Gain+Directivity 时, 推导效率
    _flat = [r for rows in sheet_results.values() for r in rows]
    if _flat and "efficiency_pct" not in _flat[0] and "gain" in _flat[0] and "directivity" in _flat[0]:
        for r in _flat:
            g = r.get("gain"); d = r.get("directivity")
            if g is not None and d is not None:
                r["efficiency_pct"] = 10 ** ((g - d) / 10) * 100
    from .plotter import _renderer as _freq_renderer
    for sheet_name, rows in sheet_results.items():
        for param_key, param_label in _B_PARAM_MAP.items():
            freqs = []; values = []
            for row in rows:
                v = row.get(param_key)
                if v is not None and row.get("frequency") is not None:
                    freqs.append(row["frequency"]); values.append(v)
            if len(freqs) > 1:
                try:
                    gap = getattr(azimuth_config, 'freq_gap_mhz', 10) if azimuth_config else 10
                    png = _freq_renderer.render_freq_curve(
                        freqs, values, ylabel=param_label,
                        title=f"{param_label} vs Frequency",
                        gap_mhz=gap)
                    group = f"B: {param_label} vs Freq"
                    # B 类图表不按频点分组，用 0.0 做占位 key
                    if group not in image_groups:
                        image_groups[group] = {}
                    image_groups[group][0.0] = png
                except Exception:
                    pass

    # ── 双Y轴配对 (B 类) ──
    _dual_y = getattr(azimuth_config, 'dual_y_enabled', False) if azimuth_config else False
    if _dual_y and _flat:
        # 预定义配对: (%类, dB类) → 双Y轴
        _DUAL_PAIRS = [
            (("efficiency_pct", "Efficiency (%)"), ("gain", "Peak Gain (dBi)")),
            (("directivity", "Directivity (dBi)"), ("trp", "TRP (dBm)")),
        ]
        gap = getattr(azimuth_config, 'freq_gap_mhz', 10) if azimuth_config else 10
        for (k1, l1), (k2, l2) in _DUAL_PAIRS:
            if k1 not in _flat[0] or k2 not in _flat[0]:
                continue
            freqs = []; v1 = []; v2 = []
            for row in _flat:
                f = row.get("frequency")
                a = row.get(k1); b = row.get(k2)
                if f is not None and a is not None and b is not None:
                    freqs.append(f); v1.append(a); v2.append(b)
            if len(freqs) > 1:
                try:
                    png = _freq_renderer.render_freq_curve_dual(
                        freqs, v1, l1, v2, l2, gap_mhz=gap,
                        title=f"{l1} + {l2} vs Frequency")
                    group = f"B: {l1} + {l2} vs Freq"
                    if group not in image_groups:
                        image_groups[group] = {}
                    image_groups[group][0.0] = png
                except Exception:
                    pass

    # ── 中间数据收集 ──
    for sheet_name, rows in sheet_results.items():
        for row in rows:
            freq = row.get("frequency")
            if freq is None: continue
            gain_dbi = row.get("_azimuth_gain_dbi")
            ar_db_v = row.get("_azimuth_ar_db")
            theta_deg_arr = row.get("_azimuth_theta_deg")
            if gain_dbi is not None and theta_deg_arr is not None and azimuth_config and azimuth_config.azimuth_cut_angles:
                gd = {}
                for angle in azimuth_config.angles_sorted:
                    idx = int(np.argmin(np.abs(theta_deg_arr - angle)))
                    nearest = float(theta_deg_arr[idx])
                    gd[nearest] = gain_dbi[:, idx].copy()
                freq_gain_data.append((freq, gd))
            if ar_db_v is not None and theta_deg_arr is not None and azimuth_config and azimuth_config.azimuth_cut_angles_ar:
                ad = {}
                for angle in azimuth_config.angles_ar_sorted:
                    idx = int(np.argmin(np.abs(theta_deg_arr - angle)))
                    nearest = float(theta_deg_arr[idx])
                    ad[nearest] = ar_db_v[:, idx].copy()
                freq_ar_data.append((freq, ad))
            rhcp_db_v = row.get("_azimuth_rhcp_db")
            if rhcp_db_v is not None and theta_deg_arr is not None and azimuth_config and azimuth_config.azimuth_cut_angles_rhcp:
                rd = {}
                for angle in azimuth_config.angles_rhcp_sorted:
                    idx = int(np.argmin(np.abs(theta_deg_arr - angle)))
                    nearest = float(theta_deg_arr[idx])
                    rd[nearest] = rhcp_db_v[:, idx].copy()
                freq_rhcp_data.append((freq, rd))
            lhcp_db_v = row.get("_azimuth_lhcp_db")
            if lhcp_db_v is not None and theta_deg_arr is not None and azimuth_config and azimuth_config.azimuth_cut_angles_lhcp:
                ld = {}
                for angle in azimuth_config.angles_lhcp_sorted:
                    idx = int(np.argmin(np.abs(theta_deg_arr - angle)))
                    nearest = float(theta_deg_arr[idx])
                    ld[nearest] = lhcp_db_v[:, idx].copy()
                freq_lhcp_data.append((freq, ld))
            # Gain 0-70° Pk 中间数据 (每 phi 的 Theta 范围峰值)
            pk070_db = row.get("_gain_pk070_db")
            if pk070_db is not None and freq is not None:
                freq_gain_vs_theta.append((freq, pk070_db.copy()))


    # ── 按频点配对: azimuth_polar + azimuth_polar_pk070 并排 ──
    freq_pairs: dict[float, dict[str, io.BytesIO]] = {}
    if out_word and azimuth_config and azimuth_config.cut_azimuth_polar \
            and azimuth_config.cut_azimuth_polar_pk070:
        for row in (r for rows in sheet_results.values() for r in rows):
            f = row.get("frequency")
            if f is None: continue
            imgs = row.get("_images", {})
            if "azimuth_polar" in imgs and "azimuth_polar_pk070" in imgs:
                freq_pairs[f] = {"azimuth_polar": imgs["azimuth_polar"],
                                 "azimuth_polar_pk070": imgs["azimuth_polar_pk070"]}

    # 分离 B 类 (非 azimuth) 图片
    extra_groups = {k: v for k, v in image_groups.items()
                    if not k.startswith("Gain Azimuth") and not k.startswith("Gain 0-70")} \
        if image_groups else {}

    # Write Word
    if out_word and (freq_pairs or extra_groups):
        az = azimuth_config or None
        word_path = az.chart_output_path if az else ""
        if not word_path:
            _log(log_callback, "  ⚠ 未设置 Word 输出路径, 跳过图表报告")
        elif word_template_path and os.path.exists(word_template_path):
            # 模板模式略 (暂不处理 freq_pairs)
            pass
        elif freq_pairs:
            from .chart_word_writer import write_chart_word_report_by_freq
            _log(log_callback, f"生成图表报告 (按频点): {word_path}")
            try:
                write_chart_word_report_by_freq(
                    freq_pairs,
                    pair_order=["azimuth_polar", "azimuth_polar_pk070"],
                    pair_labels={
                        "azimuth_polar": "Gain Azimuth Cut",
                        "azimuth_polar_pk070": "Gain Azimuth (θ=0°-70°)",
                    },
                    output_path=word_path,
                    antenna_name=az.antenna_name if az else "",
                    image_width_cm=getattr(az, 'image_width_cm', 7.5) if az else 7.5,
                    show_caption=getattr(az, 'show_caption', True) if az else True,
                    extra_groups=extra_groups if extra_groups else None,
                )
                total = len(freq_pairs) * 2 + sum(len(v) for v in extra_groups.values())
                _log(log_callback, f"  ✓ Word 报告已保存 ({len(freq_pairs)} 频点, {total} 张图)")
            except Exception as e:
                _log(log_callback, f"  ✗ Word 报告生成失败: {e}")
        else:
            # 无 azimuth 对, 回退到旧 writer
            _log(log_callback, f"生成图表报告: {word_path}")
            try:
                angles_str = ", ".join(
                    f"{a:.0f}°" for a in (az.azimuth_cut_angles if az else [])
                ) if (az and az.azimuth_cut_angles) else ""
                write_chart_word_report(
                    extra_groups, word_path,
                    antenna_name=az.antenna_name if az else "",
                    angles_str=angles_str,
                    layout_columns=az.word_columns if az else 2,
                    image_width_pct=az.word_image_width_pct if az else 90,
                )
                total_imgs = sum(len(v) for v in extra_groups.values())
                _log(log_callback, f"  ✓ Word 报告已保存 ({len(extra_groups)} 组, {total_imgs} 张图)")
            except Exception as e:
                _log(log_callback, f"  ✗ Word 报告生成失败: {e}")


    # ── 中间数据: 单文件多 sheet (按启用的图表类型) ──
    if out_data:
        data_path = azimuth_config.data_output_path if azimuth_config else ""
        if not data_path:
            data_path = ""
        # 收集启用的图表类型 → 数据映射
        data_sheets = {}
        if azimuth_config:
            if azimuth_config.cut_azimuth_polar and freq_gain_data:
                data_sheets["Gain Azimuth"] = freq_gain_data
            if azimuth_config.cut_azimuth_polar_ar and freq_ar_data:
                data_sheets["AR Azimuth"] = freq_ar_data
            if azimuth_config.cut_azimuth_polar_rhcp and freq_rhcp_data:
                data_sheets["RHCP Azimuth"] = freq_rhcp_data
            if azimuth_config.cut_azimuth_polar_lhcp and freq_lhcp_data:
                data_sheets["LHCP Azimuth"] = freq_lhcp_data
            if azimuth_config.cut_azimuth_polar_pk070 and freq_gain_vs_theta:
                data_sheets["Gain 0-70 Pk"] = [("phi_matrix", freq_gain_vs_theta)]
        if data_sheets:
            if not data_path:
                gdir = getattr(azimuth_config, 'data_output_dir', '') if azimuth_config else ''
                gfn = getattr(azimuth_config, 'data_output_filename', '') if azimuth_config else ''
                if gdir and gfn:
                    data_path = str(Path(gdir) / gfn)
            if data_path:
                _log(log_callback, f"中间数据: {data_path}")
                try:
                    import openpyxl as _xl
                    wb = _xl.Workbook(); wb.remove(wb.active)
                    for sheet_name, fd in data_sheets.items():
                        if not fd: continue
                        if sheet_name == "Gain 0-70 Pk":
                            ws = wb.create_sheet(sheet_name)
                            _, pk_data = fd[0]
                            ws.cell(1, 1, "Phi (°)")
                            for ci, (f, _) in enumerate(pk_data):
                                ws.cell(1, ci + 2, f"{f:.1f} MHz")
                            n_phi = len(pk_data[0][1])
                            for pi in range(n_phi):
                                ws.cell(pi + 2, 1, pi)
                                for ci, (_, vals) in enumerate(pk_data):
                                    ws.cell(pi + 2, ci + 2, round(float(vals[pi]), 6))
                        else:
                            _write_data_sheet(wb, sheet_name, fd)
                    wb.save(data_path); wb.close()
                    _log(log_callback, f"  ✓ 中间数据已保存 ({len(data_sheets)} sheets)")
                except Exception as e:
                    _log(log_callback, f"  ✗ 中间数据导出失败: {e}")

    def _write_data_sheet(wb, sheet_name, freq_data):
        """在 workbook 中添加一个数据 sheet (每频点一个 block)。"""
        ws = wb.create_sheet(sheet_name[:31])
        for freq_mhz, theta_data in freq_data:
            sorted_thetas = sorted(theta_data.keys())
            n_phi = len(next(iter(theta_data.values())))
            ws.cell(row=ws.max_row + 1 if ws.max_row else 1, column=1,
                    value=f"Frequency: {freq_mhz:.1f} MHz")
            r0 = ws.max_row + 1
            ws.cell(r0, 1, "Phi (°)")
            for ti, theta in enumerate(sorted_thetas):
                ws.cell(r0, ti + 2, f"{theta:.0f}°")
            for pi in range(n_phi):
                r = r0 + 1 + pi
                ws.cell(r, 1, pi)
                for ti, theta in enumerate(sorted_thetas):
                    v = theta_data[theta][pi]
                    if np.isfinite(v):
                        ws.cell(r, ti + 2, round(float(v), 6))
            ws.max_row  # force update

            if azimuth_config.cut_azimuth_polar and freq_gain_data:
                data_sheets["Gain Azimuth"] = freq_gain_data
            if azimuth_config.cut_azimuth_polar_ar and freq_ar_data:
                data_sheets["AR Azimuth"] = freq_ar_data
            if azimuth_config.cut_azimuth_polar_rhcp and freq_rhcp_data:
                data_sheets["RHCP Azimuth"] = freq_rhcp_data
            if azimuth_config.cut_azimuth_polar_lhcp and freq_lhcp_data:
                data_sheets["LHCP Azimuth"] = freq_lhcp_data
            if azimuth_config.cut_azimuth_polar_pk070 and freq_gain_vs_theta:
                data_sheets["Gain 0-70 Pk"] = [("phi_matrix", freq_gain_vs_theta)]
        if data_sheets:
            # 生成路径
            if not data_path:
                gdir = getattr(azimuth_config, 'data_output_dir', '') if azimuth_config else ''
                gfn = getattr(azimuth_config, 'data_output_filename', '') if azimuth_config else ''
                if gdir and gfn:
                    data_path = str(Path(gdir) / gfn)
            if data_path:
                _log(log_callback, f"中间数据: {data_path}")
                try:
                    import openpyxl as _xl
                    wb = _xl.Workbook(); wb.remove(wb.active)
                    for sheet_name, fd in data_sheets.items():
                        if not fd: continue
                        if sheet_name == "Gain 0-70 Pk":
                            ws = wb.create_sheet(sheet_name)
                            _, pk_data = fd[0]
                            ws.cell(1, 1, "Phi (°)")
                            for ci, (f, _) in enumerate(pk_data):
                                ws.cell(1, ci + 2, f"{f:.1f} MHz")
                            n_phi = len(pk_data[0][1])
                            for pi in range(n_phi):
                                ws.cell(pi + 2, 1, pi)
                                for ci, (_, vals) in enumerate(pk_data):
                                    ws.cell(pi + 2, ci + 2, round(float(vals[pi]), 6))
                        else:
                            _write_data_sheet(wb, sheet_name, fd)
                    wb.save(data_path); wb.close()
                    _log(log_callback, f"  ✓ 中间数据已保存 ({len(data_sheets)} sheets)")
                except Exception as e:
                    _log(log_callback, f"  ✗ 中间数据导出失败: {e}")


# ---------------------------------------------------------------------------
# 向后兼容: 保留旧版 run_batch_pipeline
# ---------------------------------------------------------------------------

def run_batch_pipeline(
    csv_path: str,
    template_path: str,
    output_path: str,
    *,
    lag_config_override=None,
    plot_config=None,
    full_report_path=None,
    extrapolate_theta: bool = False,
    cancel_callback=None,
    progress_callback=None,
    log_callback=None,
):
    """旧版 pipeline: 接受 CSV 路径，内部创建 MergedCSVParser。"""
    ds = MergedCSVParser(csv_path)
    return run_pipeline(
        datasource=ds,
        template_path=template_path,
        output_path=output_path,
        lag_config_override=lag_config_override,
        plot_config=plot_config,
        full_report_path=full_report_path,
        extrapolate_theta=extrapolate_theta,
        cancel_callback=cancel_callback,
        progress_callback=progress_callback,
        log_callback=log_callback,
    )


# ---------------------------------------------------------------------------
# 辅助
# ---------------------------------------------------------------------------


def _log(cb, msg):
    if cb: cb(msg)


def _report(cb, cur, tot, msg):
    if cb: cb(cur, tot, msg)


# ---------------------------------------------------------------------------
# 并行计算 worker（模块级，可 pickle，纯 numpy 数学）
# ---------------------------------------------------------------------------

def _compute_chunk(
    compute_tasks: list[tuple[str, float, dict[str, Any], Any, list[float], bool]],
) -> list[tuple[str, dict[str, Any]]]:
    """子进程中处理一批频点的纯计算任务。不读文件。

    每个任务: (sheet_name, freq, raw_data, lag_cfg, theta_list, extrapolate_theta)
    """
    import numpy as np
    results = []
    for sheet_name, freq, raw, lag_cfg, theta_list, do_extrap, rpk, nparams, xparams, ccfg, ar_cfg, nh_angles, ar_out_db, az_cfg, co in compute_tasks:
        try:
            theta_raw = np.array(theta_list)
            row = _process_one_frequency(raw, freq, theta_raw, lag_cfg,
                                         do_extrapolate=do_extrap, robust_peak=rpk, needed_params=nparams, extra_params=xparams, chart_config=ccfg, ar_lag_config=ar_cfg, azimuth_config=az_cfg, nh_custom_angles=nh_angles, ar_output_db=ar_out_db, compute_only=co,
                                         log_cb=None)
            results.append((sheet_name, row))
        except Exception as e:
            results.append((sheet_name, {"frequency": freq, "_error": str(e)}))
    return results
