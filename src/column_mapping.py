"""
列映射公共模块
==============
统一模板列头识别、映射配置、加载/保存逻辑。
被 excel_reader, template_recognizer, report_exporter, SystemSettings 共用。
"""
from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from pathlib import Path

# ═══════════════════════════════════════════════════════════════
# 数据结构
# ═══════════════════════════════════════════════════════════════

@dataclass
class ColumnMapping:
    """单个列的检测与映射信息。"""
    col_letter: str          # "A", "B", ...
    col_index: int           # 1-based
    raw_header: str          # 原始列头文本
    detected_type: str       # 自动检测的 col_type
    confirmed_type: str = ""  # 用户确认的 (空=使用detected)

    @property
    def effective_type(self) -> str:
        return self.confirmed_type or self.detected_type

    def to_dict(self) -> dict:
        return {
            "col_letter": self.col_letter,
            "col_index": self.col_index,
            "raw_header": self.raw_header,
            "detected_type": self.detected_type,
            "confirmed_type": self.confirmed_type,
        }

    @classmethod
    def from_dict(cls, d: dict) -> ColumnMapping:
        return cls(
            col_letter=d.get("col_letter", ""),
            col_index=d.get("col_index", 0),
            raw_header=d.get("raw_header", ""),
            detected_type=d.get("detected_type", "unknown"),
            confirmed_type=d.get("confirmed_type", ""),
        )


@dataclass
class TemplatePreset:
    """模板预设 — 包含列映射和计算参数。"""
    name: str
    path: str
    manufacturer: str = ""
    default_output_dir: str = ""
    file_type: str = "xlsx"  # xlsx, xls, docx, doc
    column_mappings: list[ColumnMapping] = field(default_factory=list)
    calc_params: dict = field(default_factory=dict)  # extrapolate_theta, etc.
    graph_config: dict = field(default_factory=dict)  # elevation, DPI, etc.

    def to_dict(self) -> dict:
        return {
            "name": self.name,
            "path": self.path,
            "manufacturer": self.manufacturer,
            "default_output_dir": self.default_output_dir,
            "file_type": self.file_type,
            "column_mappings": [m.to_dict() for m in self.column_mappings],
            "calc_params": self.calc_params,
            "graph_config": self.graph_config,
        }

    @classmethod
    def from_dict(cls, d: dict) -> TemplatePreset:
        mappings = [ColumnMapping.from_dict(m)
                    for m in d.get("column_mappings", [])]
        return cls(
            name=d.get("name", ""),
            path=d.get("path", ""),
            manufacturer=d.get("manufacturer", ""),
            default_output_dir=d.get("default_output_dir", ""),
            file_type=d.get("file_type", "xlsx"),
            column_mappings=mappings,
            calc_params=d.get("calc_params", {}),
            graph_config=d.get("graph_config", {}),
        )


# ═══════════════════════════════════════════════════════════════
# 列检测 — 统一入口
# ═══════════════════════════════════════════════════════════════

def detect_columns_from_template(template_path: str,
                                  header_row: int = None) -> list[ColumnMapping]:
    """从模板文件检测所有列的 col_type。

    支持 .xlsx, .xls, .docx, .doc 格式。
    复用 excel_reader 和 column_patterns.json。
    """
    ext = Path(template_path).suffix.lower()
    if ext in ('.xlsx', '.xls'):
        return _detect_excel_columns(template_path, header_row)
    elif ext in ('.docx', '.doc'):
        return _detect_word_columns(template_path, header_row)
    else:
        raise ValueError(f"不支持的模板格式: {ext}")


def _detect_excel_columns(path: str, header_row: int = None) -> list[ColumnMapping]:
    """Excel 模板列检测。"""
    import openpyxl

    from src.excel_reader import is_frequency_column

    wb = openpyxl.load_workbook(path, data_only=True)

    # 自动找标题行
    if header_row is None:
        header_row = 1
        ws0 = wb.worksheets[0]
        best_score = 0
        for r in range(1, min(ws0.max_row or 100, 30)):
            score = 0
            for c in range(1, (ws0.max_column or 20) + 1):
                v = ws0.cell(r, c).value
                if v and str(v).strip():
                    score += 1
                    if is_frequency_column(str(v).strip()):
                        score += 10
            if score > best_score:
                best_score = score
                header_row = r

    mappings = []
    ws = wb.worksheets[0]
    max_col = ws.max_column or 30

    for c in range(1, max_col + 1):
        raw = str(ws.cell(header_row, c).value or "").strip()
        if not raw:
            continue

        col_letter = openpyxl.utils.get_column_letter(c)
        ctype = classify_header(raw)
        mappings.append(ColumnMapping(
            col_letter=col_letter,
            col_index=c,
            raw_header=raw,
            detected_type=ctype,
        ))

    wb.close()
    return mappings


def _detect_word_columns(path: str, header_row: int = None) -> list[ColumnMapping]:
    """Word 模板列检测 (.docx/.doc)。"""
    # 对于 .doc，先尝试转换为 .docx
    actual_path = path
    tmp_docx = None
    ext = Path(path).suffix.lower()
    if ext == '.doc':
        actual_path = _convert_doc_to_docx(path)
        if actual_path is None:
            return []  # 无法转换

    try:
        return _detect_word_columns_docx(actual_path, header_row)
    finally:
        if tmp_docx and os.path.exists(tmp_docx):
            try:
                os.unlink(tmp_docx)
            except OSError:
                pass


def _detect_word_columns_docx(path: str, header_row: int = None) -> list[ColumnMapping]:
    """读取 .docx 文件表格中的列头。"""
    try:
        from docx import Document
    except ImportError:
        return []  # python-docx 未安装

    try:
        doc = Document(path)
    except Exception:
        return []

    mappings = []
    # 找第一个包含表格的内容
    for table in doc.tables:
        if not table.rows:
            continue
        # 第一行作为表头
        header_cells = table.rows[0].cells
        for ci, cell in enumerate(header_cells):
            raw = cell.text.strip()
            if not raw:
                continue
            col_letter = _col_index_to_letter(ci + 1)
            ctype = _classify_full_static(raw)

            mappings.append(ColumnMapping(
                col_letter=col_letter,
                col_index=ci + 1,
                raw_header=raw,
                detected_type=ctype,
            ))
        break  # 只处理第一个表格

    return mappings


# ═══════════════════════════════════════════════════════════════
# 分类函数
# ═══════════════════════════════════════════════════════════════

def classify_header(raw_header: str) -> str:
    """统一列头分类入口。JSON 模式优先 → 内置函数 fallback → regex fallback。

    单一入口，供 excel_reader, template_recognizer, chart_config 共用。
    """
    from src.excel_reader import (
        _classify_by_json_patterns,
        detect_ratio_column_type,
        is_ar_range_column,
        is_ar_single_column,
        is_avg_gain_column,
        is_avg_power_column,
        is_boresight_phi_column,
        is_boresight_theta_column,
        is_directivity_column,
        is_efficiency_column,
        is_frequency_column,
        is_gain_column,
        is_lh_prp_column,
        is_max_power_column,
        is_min_power_column,
        is_mismatch_loss_column,
        is_nhprp_30_column,
        is_nhprp_45_column,
        is_nhprp_225_column,
        is_pc_phi_column,
        is_pc_theta_column,
        is_peak_eirp_column,
        is_total_efficiency_column,
        is_trp_column,
        is_uh_prp_column,
        is_xpi_boresight_column,
        is_xpi_mean_column,
        is_xpi_min_column,
        normalize_header,
    )
    from src.lag_config import (
        _RE_LAG_RANGE,
        _RE_LAG_RANGE_NO_PREFIX,
        _RE_LAG_SINGLE,
        _RE_LAG_SINGLE_NO_PREFIX,
    )

    # JSON 模式优先
    json_type = _classify_by_json_patterns(raw_header)
    if json_type is not None:
        return json_type

    # 内置函数 fallback
    if is_frequency_column(raw_header):          return "frequency"
    if is_directivity_column(raw_header):        return "directivity"

    if is_total_efficiency_column(raw_header):
        return "total_efficiency_db" if "db" in normalize_header(raw_header).lower() else "total_efficiency_pct"
    if is_efficiency_column(raw_header):
        return "efficiency_db" if "db" in normalize_header(raw_header).lower() else "efficiency_pct"

    if is_gain_column(raw_header):               return "gain"
    if is_trp_column(raw_header):                return "trp"
    if is_nhprp_45_column(raw_header):           return "nhprp_45"
    if is_nhprp_30_column(raw_header):           return "nhprp_30"
    if is_peak_eirp_column(raw_header):          return "peak_eirp"
    if is_ar_single_column(raw_header):          return "ar_single"
    if is_ar_range_column(raw_header):           return "ar_range"
    if is_nhprp_225_column(raw_header):          return "nhprp_225"
    if is_uh_prp_column(raw_header):             return "uh_prp"
    if is_lh_prp_column(raw_header):             return "lh_prp"

    ratio_type = detect_ratio_column_type(raw_header)
    if ratio_type:                               return ratio_type

    if is_boresight_phi_column(raw_header):      return "boresight_phi"
    if is_boresight_theta_column(raw_header):    return "boresight_theta"
    if is_max_power_column(raw_header):          return "max_power"
    if is_min_power_column(raw_header):          return "min_power"
    if is_avg_gain_column(raw_header):           return "avg_gain"
    if is_avg_power_column(raw_header):          return "avg_power"
    if is_xpi_boresight_column(raw_header):      return "xpi_boresight"
    if is_xpi_mean_column(raw_header):           return "xpi_mean"
    if is_xpi_min_column(raw_header):            return "xpi_min"
    if is_mismatch_loss_column(raw_header):      return "mismatch_loss_db"
    if is_pc_theta_column(raw_header):           return "pc_theta_mm"
    if is_pc_phi_column(raw_header):             return "pc_phi_mm"

    # RHCP/LHCP/CP-XPI 检测 (在 LAG 之前, 避免 "RHCP Gain at Theta=" 被 LAG 误匹配)
    _RE_RHCP = re.compile(r"RHCP\s*(?:Gain)?", re.IGNORECASE)
    _RE_CP_XPI = re.compile(r"CP[\s-]*XPI", re.IGNORECASE)
    if _RE_RHCP.search(raw_header) or _RE_CP_XPI.search(raw_header):
        normalized = normalize_header(raw_header)
        if "range" in normalized.lower() or re.search(r"\d+.*[-–].*\d+", raw_header):
            return "rhcp_range" if _RE_RHCP.search(raw_header) else "cp_xpi_range"
        return "rhcp_single" if _RE_RHCP.search(raw_header) else "cp_xpi_single"
    # Regex fallback (LAG) — 用原始列头而非 _norm，因 _normalize_key 会去掉 = 号
    if _RE_LAG_RANGE.search(raw_header) or _RE_LAG_RANGE_NO_PREFIX.search(raw_header):
        return "lag_range"
    if _RE_LAG_SINGLE.search(raw_header) or _RE_LAG_SINGLE_NO_PREFIX.search(raw_header):
        return "lag_single"
    return "unknown"


def _classify_full(raw_header, classify_json, *classifiers) -> str:
    """(已弃用) 保留兼容接口，直接调用 classify_header。"""
    return classify_header(raw_header)


def _classify_full_static(raw_header: str) -> str:
    """(已弃用) 保留兼容接口，直接调用 classify_header。"""
    return classify_header(raw_header)


# ═══════════════════════════════════════════════════════════════
# 预设管理 (templates.json)
# ═══════════════════════════════════════════════════════════════

_TEMPLATES_PATH = Path(__file__).resolve().parent.parent / "config" / "templates.json"


def load_presets() -> list[TemplatePreset]:
    """从 templates.json 加载所有模板预设。"""
    if not _TEMPLATES_PATH.exists():
        return []
    try:
        with open(_TEMPLATES_PATH, encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return []

    presets = []
    for mfr in data.get("manufacturers", []):
        for tpl in mfr.get("templates", []):
            presets.append(TemplatePreset.from_dict({
                **tpl, "manufacturer": mfr.get("name", ""),
            }))
    return presets


def save_preset(preset: TemplatePreset) -> None:
    """保存一个模板预设到 templates.json（合并写入）。"""
    if _TEMPLATES_PATH.exists():
        try:
            with open(_TEMPLATES_PATH, encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            data = {"manufacturers": []}
    else:
        data = {"manufacturers": []}

    # 查找或创建厂商
    mfr_name = preset.manufacturer or "默认"
    mfr_entry = None
    for m in data.get("manufacturers", []):
        if m.get("name") == mfr_name:
            mfr_entry = m
            break
    if mfr_entry is None:
        mfr_entry = {"name": mfr_name, "templates": []}
        data.setdefault("manufacturers", []).append(mfr_entry)

    # 更新或添加模板
    tpl_dict = preset.to_dict()
    tpl_dict.pop("manufacturer", None)
    found = False
    for i, t in enumerate(mfr_entry.get("templates", [])):
        if t.get("name") == preset.name:
            mfr_entry["templates"][i] = tpl_dict
            found = True
            break
    if not found:
        mfr_entry.setdefault("templates", []).append(tpl_dict)

    with open(_TEMPLATES_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ═══════════════════════════════════════════════════════════════
# Word 格式转换
# ═══════════════════════════════════════════════════════════════

def _convert_doc_to_docx(doc_path: str) -> str | None:
    """将 .doc 转换为 .docx（使用 libreoffice）。"""
    import subprocess
    import tempfile
    tmp_dir = tempfile.mkdtemp()
    try:
        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "docx",
             "--outdir", tmp_dir, doc_path],
            capture_output=True, timeout=30,
        )
        if result.returncode == 0:
            base = os.path.splitext(os.path.basename(doc_path))[0]
            docx_path = os.path.join(tmp_dir, base + ".docx")
            if os.path.exists(docx_path):
                return docx_path
    except (subprocess.TimeoutExpired, FileNotFoundError, OSError):
        pass
    return None


def _col_index_to_letter(idx: int) -> str:
    """1 → 'A', 2 → 'B', ..., 27 → 'AA'。"""
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


# ═══════════════════════════════════════════════════════════════
# 列类型显示标签
# ═══════════════════════════════════════════════════════════════

# ── 合并键 → 列头级子键展开 ──
_MERGED_EXPANSIONS = {
    "gain": [("lag_single", "LAG 单角度"), ("lag_range", "LAG 范围")],
    "ar":   [("ar_single", "AR 单角度"), ("ar_range", "AR 范围")],
    "total_efficiency_pct": [("total_efficiency_pct", "总效率(%)")],
}


def get_col_type_labels(mode: int = 0) -> list[tuple[str, str]]:
    """返回指定测试模式的列类型下拉列表。

    数据源: AntennaParamsPage 参数定义 (单一数据源)。

    Args:
        mode: 0=无源天线, 1=有源发射(TRP), 2=有源接收(TIS)
    """
    from ui.pages import AntennaParamsPage

    result = [("frequency", "频率"), ("unknown", "未知")]

    if mode == 0:
        params = list(AntennaParamsPage._COMMON_PARAMS)
    elif mode == 1:
        no_ar = [(g, plist) for g, plist in AntennaParamsPage._COMMON_PARAMS
                 if g not in ("Axial Ratio",)]
        params = no_ar + list(AntennaParamsPage._TRP_PARAMS)
    else:
        params = list(AntennaParamsPage._TIS_PARAMS)

    for group_name, items in params:
        for key, label in items:
            # 合并键展开为模板列头级子键 (gain → lag_single + lag_range)
            if key in _MERGED_EXPANSIONS:
                result.extend(_MERGED_EXPANSIONS[key])
            else:
                result.append((key, label))
    return result


# 向后兼容 (无源模式)
ALL_COL_TYPE_LABELS = get_col_type_labels(0)
