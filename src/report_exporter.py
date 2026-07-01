"""
完整报告导出
=============
独立的 Excel 报告 — 不依赖客户模板，包含：
  - 指标数据表（冻结首行）
  - 3D 球面方向图
  - 2D 极坐标切面图
  - 2D 直角坐标切面图
  - 图标题/描述
"""

from __future__ import annotations

import io
import json
import os
import sys
from typing import Any, Callable, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 样式常量
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

PLOT_DESC_FONT = Font(name="Calibri", size=9, italic=True, color="808080")

# 图像尺寸
IMG_3D_SIZE = (300, 225)    # 3D 球面图
IMG_2D_SIZE = (280, 210)    # 2D 切面图

# ── JSON 列配置缓存 ──
_REPORT_COLUMNS: Optional[List[dict]] = None
_REPORT_VALIDATION: Optional[dict] = None


def _load_report_columns() -> List[dict]:
    """加载 config/full_report_columns.json 中的列定义。"""
    global _REPORT_COLUMNS, _REPORT_VALIDATION
    if _REPORT_COLUMNS is not None:
        return _REPORT_COLUMNS

    candidates = []
    # 打包模式: EXE 同目录 config/ 优先（用户外部编辑）
    if getattr(sys, 'frozen', False):
        candidates.append(os.path.join(os.path.dirname(sys.executable), "config", "full_report_columns.json"))
    # 内嵌默认值 (MEIPASS) 或开发模式项目根目录
    candidates.append(os.path.join(os.path.dirname(__file__), "..", "config", "full_report_columns.json"))
    # 当前工作目录 fallback
    candidates.append(os.path.join(os.getcwd(), "config", "full_report_columns.json"))
    for candidate in candidates:
        path = os.path.normpath(candidate)
        if os.path.isfile(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                _REPORT_COLUMNS = data.get("columns", [])
                _REPORT_VALIDATION = data.get("validation", {})
                return _REPORT_COLUMNS
            except (json.JSONDecodeError, OSError):
                pass

    _REPORT_COLUMNS = []
    _REPORT_VALIDATION = {}
    return _REPORT_COLUMNS


def _get_column_order(all_keys: List[str]) -> List[str]:
    """根据 JSON 配置对列排序，未知列追加到末尾。"""
    config = _load_report_columns()
    # JSON 中声明的 key（非动态）
    ordered = []
    for col in config:
        if col.get("_dynamic"):
            # 动态列 — 匹配 all_keys 中匹配的前缀
            base = col["key"].rstrip("*")
            matched = sorted(k for k in all_keys if k.startswith(base))
            ordered.extend(matched)
        elif col["key"] in all_keys:
            ordered.append(col["key"])
    # 追加 JSON 中未声明的列
    declared = {c["key"].rstrip("*") for c in config}
    remaining = sorted(k for k in all_keys if k not in ordered
                       and not any(k.startswith(d) for d in declared if d.endswith("_*")))
    ordered += [k for k in remaining if not k.startswith("_")]
    return ordered


def _get_column_header(key: str, lang: str = "cn") -> str:
    """从 JSON 配置获取列头显示文本。"""
    config = _load_report_columns()
    header_field = "header_cn" if lang == "cn" else "header_en"

    for col in config:
        if col.get("_dynamic"):
            base = col["key"].rstrip("*")
            if key.startswith(base):
                tmpl = col.get(header_field, key)
                # 解析模板变量
                if "lag_single_" in base or "ar_single_" in base:
                    angle = key.replace(base, "")
                    return tmpl.replace("{angle}", angle)
                elif "lag_range_" in base:
                    parts = key.replace(base, "").split("_")
                    if len(parts) >= 2:
                        return tmpl.replace("{start}", parts[0]).replace("{end}", parts[1])
                return tmpl
        elif col["key"] == key:
            return col.get(header_field, key)

    # Fallback: 用旧逻辑
    if key.startswith("lag_single_"):
        angle = key.replace("lag_single_", "")
        return f"LAG θ={angle}° (dB)"
    if key.startswith("lag_range_"):
        parts = key.replace("lag_range_", "").split("_")
        return f"LAG ({parts[0]}°-{parts[1]}°) (dB)"
    if key.startswith("ar_single_"):
        angle = key.replace("ar_single_", "")
        return f"AR θ={angle}° (dB)"
    return key


# ---------------------------------------------------------------------------
# 工具
# ---------------------------------------------------------------------------


def _auto_width(ws, min_width: int = 10, max_width: int = 30):
    """自动调整列宽（基于内容）。"""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def export_full_report(
    output_path: str,
    sheet_results: Dict[str, List[Dict[str, Any]]],
    *,
    pattern_images_3d: Optional[Dict[str, Dict[float, io.BytesIO]]] = None,
    pattern_images_2d_polar: Optional[Dict[str, Dict[float, Dict[str, io.BytesIO]]]] = None,
    pattern_images_2d_rect: Optional[Dict[str, Dict[float, Dict[str, io.BytesIO]]]] = None,
    bands: Optional[Dict[str, Tuple[float, float]]] = None,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
) -> None:
    """生成完整天线指标报告。

    Args:
        output_path:          输出 .xlsx 路径。
        sheet_results:        {sheet_name: [row_dict, ...]}
        pattern_images_3d:    {sheet_name: {freq: 3D_PNG_BytesIO}}
        pattern_images_2d_polar: {sheet_name: {freq: {cut_label: PNG}}}
        pattern_images_2d_rect:  {sheet_name: {freq: {cut_label: PNG}}}
        bands:                {band_name: (lo_freq, hi_freq)} 频段配置。
        progress_callback:    (current, total, message)
    """
    wb = openpyxl.Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    total_sheets = len(sheet_results)
    sheet_count = 0

    for sheet_name, rows in sheet_results.items():
        sheet_count += 1
        if progress_callback:
            progress_callback(sheet_count, total_sheets, f"写入报告: {sheet_name}")

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars

        # 标题行
        _write_sheet_title(ws, sheet_name)

        # 数据表
        if rows:
            _write_data_table(ws, rows, start_row=3)

        # 图片
        _embed_report_images(
            ws,
            sheet_name,
            rows,
            pattern_images_3d,
            pattern_images_2d_polar,
            pattern_images_2d_rect,
        )

    # 摘要 Sheet
    _write_summary_sheet(wb, sheet_results, bands)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# 内部函数
# ---------------------------------------------------------------------------

def _write_sheet_title(ws, sheet_name: str):
    """写入 Sheet 标题和频段说明。"""
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"天线指标报告 — {sheet_name}"
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30


def _write_data_table(ws, rows: List[Dict[str, Any]], start_row: int):
    """写入完整指标数据表。

    动态检测所有 row key，按固定顺序排列已知列，其余 LAG 列追加。
    """
    if not rows:
        return

    # 收集所有列名
    seen = set()
    all_keys = []
    for row in rows:
        for k in row:
            if k not in seen:
                seen.add(k)
                all_keys.append(k)

    # 排序：JSON 配置优先，剩余 LAG 列追加
    ordered = _get_column_order(all_keys)

    # 写表头
    header_row = start_row
    for ci, key in enumerate(ordered, 1):
        cell = ws.cell(row=header_row, column=ci)
        cell.value = _get_column_header(key)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 写数据行
    for ri, row in enumerate(rows):
        data_row = header_row + 1 + ri
        for ci, key in enumerate(ordered, 1):
            cell = ws.cell(row=data_row, column=ci)
            val = row.get(key, "")
            # 跳过内部数据键和非标量值 (numpy array / dict / BytesIO)
            if key.startswith("_") or isinstance(val, (list, dict, io.BytesIO)):
                val = ""
            # 只跳过多维 array (ndim>0), 保留标量 np.float64 等 (openpyxl 可接受)
            if hasattr(val, 'ndim') and val.ndim > 0:
                val = ""
            cell.value = val
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

    # 冻结表头
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # 自动列宽
    _auto_width(ws)

    # 行高
    ws.row_dimensions[header_row].height = 28
    for ri in range(len(rows)):
        ws.row_dimensions[header_row + 1 + ri].height = 20


def _embed_report_images(
    ws,
    sheet_name: str,
    rows: List[Dict[str, Any]],
    images_3d: Optional[Dict[str, Dict[float, io.BytesIO]]],
    images_2d_polar: Optional[Dict[str, Dict[float, Dict[str, io.BytesIO]]]],
    images_2d_rect: Optional[Dict[str, Dict[float, Dict[str, io.BytesIO]]]],
):
    """在数据表右侧嵌入 3D + 2D 图像及说明。

    布局（每频点一行）：
      Col: [3D 图]   [图描述]   [2D 极坐标]   [2D 直角坐标]
    """
    if not rows:
        return

    # 找到数据表最后一列
    data_last_col = 0
    for row in rows:
        data_last_col = max(data_last_col, len(row))

    img_start_col = data_last_col + 2  # 留一列空白
    header_row = 3  # 数据表标题行

    sheet_3d = images_3d.get(sheet_name, {}) if images_3d else {}
    sheet_polar = images_2d_polar.get(sheet_name, {}) if images_2d_polar else {}
    sheet_rect = images_2d_rect.get(sheet_name, {}) if images_2d_rect else {}

    for ri, row in enumerate(rows):
        freq = row.get("frequency")
        if freq is None:
            continue

        excel_row = header_row + 1 + ri
        col = img_start_col

        # ---- 3D 球面方向图 ----
        if freq in sheet_3d:
            buf = sheet_3d[freq]
            buf.seek(0)  # 重置 buffer 位置（可能已被模板导出消耗）
            img = XLImage(buf)
            img.width = IMG_3D_SIZE[0]
            img.height = IMG_3D_SIZE[1]
            anchor_cell = ws.cell(row=excel_row, column=col).coordinate
            ws.add_image(img, anchor_cell)
            # 图描述
            desc_cell = ws.cell(row=excel_row, column=col + 1)
            desc_cell.value = f"3D Radiation Pattern\n{freq} MHz"
            desc_cell.font = PLOT_DESC_FONT
            desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            col += 2
        else:
            col += 2

        # ---- 2D 极坐标切面 ----
        polar_cuts = sheet_polar.get(freq, {})
        if polar_cuts:
            for cut_label, buf in polar_cuts.items():
                buf.seek(0)
                img = XLImage(buf)
                img.width = IMG_2D_SIZE[0]
                img.height = IMG_2D_SIZE[1]
                anchor_cell = ws.cell(row=excel_row, column=col).coordinate
                ws.add_image(img, anchor_cell)
                desc_cell = ws.cell(row=excel_row, column=col + 1)
                desc_cell.value = f"Polar Cut\n{cut_label}\n{freq} MHz"
                desc_cell.font = PLOT_DESC_FONT
                desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                col += 2
                break  # 每个频点只放一个极坐标图

        # ---- 2D 直角坐标切面 ----
        rect_cuts = sheet_rect.get(freq, {})
        if rect_cuts:
            for cut_label, buf in rect_cuts.items():
                buf.seek(0)
                img = XLImage(buf)
                img.width = IMG_2D_SIZE[0]
                img.height = IMG_2D_SIZE[1]
                anchor_cell = ws.cell(row=excel_row, column=col).coordinate
                ws.add_image(img, anchor_cell)
                desc_cell = ws.cell(row=excel_row, column=col + 1)
                desc_cell.value = f"Rectangular Cut\n{cut_label}\n{freq} MHz"
                desc_cell.font = PLOT_DESC_FONT
                desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                col += 2
                break

        # 调整图表行高度
        ws.row_dimensions[excel_row].height = max(
            ws.row_dimensions[excel_row].height or 20,
            IMG_3D_SIZE[1] * 0.75
        )

    # 图表区域列宽
    for c in range(img_start_col, img_start_col + 8):
        ws.column_dimensions[get_column_letter(c)].width = 42


def _write_summary_sheet(
    wb,
    sheet_results: Dict[str, List[Dict[str, Any]]],
    bands: Optional[Dict[str, Tuple[float, float]]] = None,
):
    """写入摘要 Sheet（概览统计）。"""
    ws = wb.create_sheet(title="Summary", index=0)

    ws.merge_cells("A1:G1")
    ws["A1"].value = "天线指标报告 — 摘要"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    start_row = 3
    for ci, header in enumerate(["Sheet", "频点数", "Gain Min(dBi)", "Gain Max(dBi)",
                                  "Gain Avg(dBi)", "Eff Min(%)", "Eff Max(%)", "Eff Avg(%)"], 1):
        cell = ws.cell(row=start_row, column=ci)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for ri, (name, rows) in enumerate(sheet_results.items()):
        data_row = start_row + 1 + ri
        valid = [r for r in rows if r.get("gain") is not None and r.get("efficiency_pct") is not None]
        if not valid:
            ws.cell(row=data_row, column=1).value = name
            ws.cell(row=data_row, column=2).value = f"{len(rows)} (无有效数据)"
            continue

        gains = [r["gain"] for r in valid]
        effs = [r["efficiency_pct"] for r in valid]

        ws.cell(row=data_row, column=1).value = name
        ws.cell(row=data_row, column=2).value = len(rows)
        ws.cell(row=data_row, column=3).value = round(min(gains), 2)
        ws.cell(row=data_row, column=4).value = round(max(gains), 2)
        ws.cell(row=data_row, column=5).value = round(sum(gains) / len(gains), 2)
        ws.cell(row=data_row, column=6).value = round(min(effs), 2)
        ws.cell(row=data_row, column=7).value = round(max(effs), 2)
        ws.cell(row=data_row, column=8).value = round(sum(effs) / len(effs), 2)

        for ci in range(1, 9):
            cell = ws.cell(row=data_row, column=ci)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

    _auto_width(ws)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# ---------------------------------------------------------------------------
# 数据验证
# ---------------------------------------------------------------------------

def validate_report_data(
    rows: List[Dict[str, Any]],
    sheet_name: str = "",
) -> Dict[str, Any]:
    """验证报告数据是否符合 JSON 中定义的规则。

    Returns:
        {"valid": bool, "errors": [str], "warnings": [str], "stats": {...}}
    """
    _load_report_columns()
    validation = _REPORT_VALIDATION or {}
    rules = validation.get("rules", []) if validation.get("enabled", True) else []

    errors = []
    warnings = []
    stats = {"total_rows": len(rows)}

    # 检查 required 列
    required_cols = [c["key"] for c in _REPORT_COLUMNS if c.get("required")
                     and not c.get("_dynamic")]
    for req_key in required_cols:
        missing = sum(1 for r in rows if req_key not in r or r.get(req_key) is None)
        if missing > 0:
            errors.append(f"[{sheet_name}] 缺少必需列 '{req_key}': {missing}/{len(rows)} 行无数据")

    for col_def in rules:
        key = col_def.get("key", "")
        vmin = col_def.get("min")
        vmax = col_def.get("max")
        col_type = col_def.get("type", "float")

        # 检查值范围
        for ri, row in enumerate(rows):
            val = row.get(key)
            if val is None or val == "":
                continue
            try:
                val = float(val)
                if vmin is not None and val < vmin:
                    warnings.append(
                        f"[{sheet_name}] 行{ri+1}: {key}={val} < min({vmin})")
                if vmax is not None and val > vmax:
                    warnings.append(
                        f"[{sheet_name}] 行{ri+1}: {key}={val} > max({vmax})")
            except (ValueError, TypeError):
                if col_type == "float":
                    warnings.append(
                        f"[{sheet_name}] 行{ri+1}: {key}={val} 不是数字")

        # 统计
        values = []
        for r in rows:
            v = r.get(key)
            if v is not None and v != "":
                try:
                    values.append(float(v))
                except (ValueError, TypeError):
                    pass
        if values:
            stats[key] = {"min": min(values), "max": max(values),
                          "avg": sum(values) / len(values), "count": len(values)}

    valid = len(errors) == 0
    return {"valid": valid, "errors": errors, "warnings": warnings, "stats": stats}


def export_full_report_with_validation(
    output_path: str,
    sheet_results: Dict[str, List[Dict[str, Any]]],
    **kwargs,
) -> Tuple[bool, Dict[str, Any]]:
    """生成报告 + 数据验证，返回 (success, validation_result)。"""
    # 先验证
    all_errors = []
    all_warnings = []
    all_stats = {}
    for sheet_name, rows in sheet_results.items():
        result = validate_report_data(rows, sheet_name)
        all_errors.extend(result["errors"])
        all_warnings.extend(result["warnings"])
        all_stats[sheet_name] = result["stats"]

    # 始终导出（验证警告不阻止）
    export_full_report(output_path, sheet_results, **kwargs)

    validation_result = {
        "valid": len(all_errors) == 0,
        "errors": all_errors,
        "warnings": all_warnings,
        "stats": all_stats,
    }
    return len(all_errors) == 0, validation_result
