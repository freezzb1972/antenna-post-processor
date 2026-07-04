"""
Excel 模板读取器
================
读取输出模板 Excel，解析所有 Sheet 的结构信息：
  - 列头定义（Frequency / Directivity / LAG 角度等）
  - 频点列表
  - 自动检测 LAG 需求
"""

from __future__ import annotations

import json
import os
import re
import sys
from dataclasses import dataclass, field

import openpyxl

from .lag_config import (
    _RE_LAG_RANGE,
    _RE_LAG_RANGE_NO_PREFIX,
    _RE_LAG_SINGLE,
    _RE_LAG_SINGLE_NO_PREFIX,
    LagConfig,
    normalize_header,
)

# ---------------------------------------------------------------------------
# 列头规范化 & 分类 (从 lag_config.py 移入 — 它们只被模板解析使用)
# ---------------------------------------------------------------------------

def _normalize_key(name: str) -> str:
    """将列头转成小写无空格键，用于固定列匹配。"""
    return re.sub(r"[^a-z%％()db]+", "", name.lower())


# ═══════════════════════════════════════════════════════════════
# JSON 列头模式加载器 — 用户可编辑的外部配置
# ═══════════════════════════════════════════════════════════════

_COLUMN_PATTERNS: list[dict] | None = None


def _load_column_patterns() -> list[dict]:
    """加载 config/column_patterns.json，若文件不存在则返回空列表。"""
    global _COLUMN_PATTERNS
    if _COLUMN_PATTERNS is not None:
        return _COLUMN_PATTERNS

    candidates = []
    # 打包模式: EXE 同目录 config/ 优先（用户外部编辑，可写入）
    if getattr(sys, 'frozen', False):
        candidates.append(os.path.join(os.path.dirname(sys.executable), "config", "column_patterns.json"))
    # 内嵌默认值 (MEIPASS) 或开发模式项目根目录
    candidates.append(os.path.join(os.path.dirname(__file__), "..", "config", "column_patterns.json"))
    # 当前工作目录 fallback
    candidates.append(os.path.join(os.getcwd(), "config", "column_patterns.json"))
    for candidate in candidates:
        path = os.path.normpath(candidate)
        if os.path.isfile(path):
            try:
                with open(path, encoding="utf-8") as f:
                    data = json.load(f)
                _COLUMN_PATTERNS = data.get("patterns", [])
                return _COLUMN_PATTERNS
            except (json.JSONDecodeError, OSError):
                pass

    _COLUMN_PATTERNS = []
    return _COLUMN_PATTERNS


def _classify_by_json_patterns(raw_header: str) -> str | None:
    """用 config/column_patterns.json 中的规则匹配列头（唯一分类器）。

    匹配优先级（按 JSON 顺序，先匹配者胜）：
      1. "regex" — 正则匹配
      2. "exact" — 精确匹配 compact form
      3. "keywords" — 所有关键词都出现 (AND)
      4. "negate" — 排除规则 (任一匹配则跳过)
    """
    patterns = _load_column_patterns()
    if not patterns:
        return None

    norm = normalize_header(raw_header).lower()
    compact = _normalize_key(raw_header)

    for entry in patterns:
        keywords = entry.get("keywords", [])
        exact_list = entry.get("exact", [])
        negate_words = entry.get("negate", [])
        regex_str = entry.get("regex", "")
        cn_str = entry.get("cn", "")
        extra_req = entry.get("extra_req", [])
        extra_req2 = entry.get("extra_req2", [])

        matched = False

        # ── Regex ──
        if regex_str:
            if re.search(regex_str, norm, re.IGNORECASE):
                matched = True

        # ── 精确匹配 ──
        if not matched and exact_list:
            for ex in exact_list:
                if ex.lower() == compact:
                    matched = True
                    break

        # ── 关键词 AND 匹配 ──
        if not matched and keywords:
            matched = True
            for kw in keywords:
                kw_lower = kw.lower()
                if kw_lower not in norm and kw_lower not in compact:
                    matched = False
                    break

        # ── 中文关键词备选 (cn: "增益|效率" → pipe-separated OR match) ──
        if not matched and cn_str:
            for cn_word in cn_str.split("|"):
                if cn_word and cn_word in raw_header:
                    matched = True
                    break

        if not matched:
            continue

        # ── extra_req: 额外必须条件 (ANY — 至少一个出现在 header 中) ──
        if extra_req:
            if not any(word.lower() in norm or word.lower() in compact or word in raw_header
                       for word in extra_req):
                continue
        if extra_req2:
            if not any(word.lower() in norm or word.lower() in compact or word in raw_header
                       for word in extra_req2):
                continue

        # ── 排除词 ──
        blocked = False
        for nkw in negate_words:
            if nkw.lower() in norm or nkw.lower() in compact or nkw in raw_header:
                blocked = True
                break
        if blocked:
            continue

        return entry["col_type"]

    return None


def reload_column_patterns():
    """强制重新加载 JSON 模式（对话框保存后调用）。"""
    global _COLUMN_PATTERNS
    _COLUMN_PATTERNS = None
    _load_column_patterns()


# ═══════════════════════════════════════════════════════════════
# 列类型分类注册表 — 所有 is_xxx_column 函数的单一数据源
#
# 新增列类型时只需:
#   1. 向 _COLUMN_CLASSIFIERS 添加一条规则
#   2. 向 _CLASSIFY_ORDER 添加 (col_type, 优先级)
#   3. 自动获得 classify_column() 聚合函数 + is_xxx_column() 包装器
# ═══════════════════════════════════════════════════════════════

def _col_classifier(must_contain: str = "", must_not_contain: str = "",
                    exact_words: tuple = (), keywords: tuple = (),
                    cn: str = "", extra_checks: tuple = ()) -> callable:
    """生成列头匹配器。

    匹配逻辑: (must_contain OR cn OR exact_words) AND keywords - must_not_contain

    Args:
        must_contain: 英文主关键词（子串匹配，忽略大小写）
        cn:           中文备选关键词（子串匹配，区分繁简）
        exact_words:  备选精确匹配（header 归一化后完全等于这些词时匹配）
        keywords:     所有关键词必须同时出现（AND 逻辑）
        must_not_contain: 排除关键词（任一匹配则拒绝，空格分隔）
        extra_checks: 额外自定义条件 (lambda header, header_lower → bool)
    """
    def _match(header: str) -> bool:
        lo = header.lower()
        # 主关键词检查: must_contain (英文) OR cn (中文, | 分隔) OR exact_words
        matched = False
        if must_contain and must_contain in lo:
            matched = True
        if not matched and cn:
            for cn_word in cn.split("|"):
                if cn_word and cn_word in header:
                    matched = True
                    break
        if not matched and exact_words and lo in exact_words:
            matched = True
        if must_contain and not matched:
            return False
        # must_not_contain: 排除词
        if must_not_contain:
            for w in must_not_contain.split():
                if w in lo:
                    return False
        # keywords: 所有关键词必须 AND 出现
        if keywords and not all(k in lo for k in keywords):
            return False
        # extra_checks: 额外自定义条件
        if extra_checks:
            return all(check(header, lo) for check in extra_checks)
        return True
    return _match


# 分类规则注册表: col_type → (match_fn, is_ratio_subtype)
_COLUMN_CLASSIFIERS: dict[str, tuple[callable, bool]] = {}

def _reg(col_type: str, **kw) -> None:
    _COLUMN_CLASSIFIERS[col_type] = (_col_classifier(**kw), False)

def _reg_fn(col_type: str, fn: callable, is_ratio: bool = False) -> None:
    _COLUMN_CLASSIFIERS[col_type] = (fn, is_ratio)


# ── 基础列类型 ──
_reg("frequency",     must_contain="frequency", cn="频率", exact_words=("freq", "f", "f(mhz)", "freq(mhz)"))
_reg("directivity",   must_contain="directivity", cn="方向性", exact_words=("dir", "d(dbi)"))


# ── Gain (需排除 LAG/Peak 变体) ──
def _match_gain(header: str) -> bool:
    h = _normalize_key(header)
    if any(w in h for w in ("average", "theta", "peakeirp")):
        return False
    # "Peak Gain" is gain, not peak_eirp
    if "peak" in h or "pk" in h:
        return True
    return h.startswith("gain") or h in ("g(dbi)", "pk")
_reg_fn("gain", _match_gain)
# ── 比率类型 (必须在基类之前, 否则会被 trp/nhprp_45 等误匹配) ──
_reg_fn("nhprp45_ratio",
    lambda h: ("nhprp" in h.lower() and "45" in h.lower() and ("ratio" in h.lower() or "比" in h or "/" in h or "／" in h)))
_reg_fn("nhprp30_ratio",
    lambda h: ("nhprp" in h.lower() and "30" in h.lower() and ("ratio" in h.lower() or "比" in h or "/" in h or "／" in h)))
_reg_fn("nhprp225_ratio",
    lambda h: ("nhprp" in h.lower() and ("22.5" in h or "22" in h) and ("ratio" in h.lower() or "比" in h or "/" in h or "／" in h)))
_reg_fn("uh_ratio",
    lambda h: (("uhprp" in h.lower() or ("upper" in h.lower() and "prp" in h.lower()) or "上半球" in h) and ("ratio" in h.lower() or "比" in h or "/" in h or "／" in h)))
_reg_fn("lh_ratio",
    lambda h: (("lhprp" in h.lower() or ("lower" in h.lower() and "prp" in h.lower()) or "下半球" in h) and ("ratio" in h.lower() or "比" in h or "/" in h or "／" in h)))


# ── TRP / NHPRP ──
def _match_trp(header: str) -> bool:
    h = header.lower()
    return ("trp" in h and "nhprp" not in h) or "total radiated power" in h or "tot. rad. pwr." in h
_reg_fn("trp", _match_trp)

_reg("nhprp_45",  must_contain="nhprp", keywords=("45",))
_reg("nhprp_30",  must_contain="nhprp", keywords=("30",))
_reg("nhprp_225", must_contain="nhprp", must_not_contain="45 30",
      extra_checks=(lambda h, lo: any(x in lo for x in ("22.5", "pi/8", "π/8")),))

# ── Peak EIRP ──
_reg("peak_eirp", must_contain="peakeirp", exact_words=("eirppeak", "pkgain", "peakgain"))

# ── AR ──
_reg("ar_single", must_contain="theta", extra_checks=(lambda h, lo: ("ar" in lo or "axial" in lo) and "~" not in lo,))
_reg("ar_range",  extra_checks=(lambda h, lo: ("ar" in lo or "axial" in lo) and "~" in lo,))

# ── PRP ──
_reg("uh_prp", keywords=("upper", "hem", "prp"))
_reg("lh_prp", keywords=("lower", "hem", "prp"))

# ── Power 统计 (min/max/avg 都含 "power") ──
_reg("max_power", must_contain="power", extra_checks=(lambda h, lo: ("maximum" in lo or "max" in lo) and "average" not in lo,))
_reg("min_power", must_contain="power", extra_checks=(lambda h, lo: ("minimum" in lo or "min" in lo) and "average" not in lo,))
_reg("avg_power", must_contain="power", extra_checks=(lambda h, lo: ("average" in lo or "avg" in lo),))

# ── Average Gain ──
_reg("avg_gain", must_contain="gain", extra_checks=(lambda h, lo: ("average" in lo or "avg" in lo) and "at" not in lo,))

# ── Boresight ──
_reg("boresight_phi",   must_contain="boresight", cn="瞄准角",
     extra_checks=(lambda h, lo: "phi" in lo or "φ" in h,))
_reg("boresight_theta", must_contain="boresight", cn="瞄准角",
    extra_checks=(lambda h, lo: any(x in lo for x in ("theta", "th.", "θ")),))

# ── XPI ──
_reg("xpi_boresight", must_contain="xpi", cn="交叉极化",
     extra_checks=(lambda h, lo: "boresight" in lo or "瞄准" in h or "瞄准角" in h,))
_reg("xpi_mean",      must_contain="xpi", cn="交叉极化",
     extra_checks=(lambda h, lo: "mean" in lo or "平均" in h,))
_reg("xpi_min",       must_contain="xpi", cn="交叉极化",
     extra_checks=(lambda h, lo: ("min" in lo or "最小" in h) and "mean" not in lo and "平均" not in h,))

# ── Total Efficiency / Mismatch ──
_reg("mismatch_loss_db", must_contain="mismatch", cn="失配",
     extra_checks=(lambda h, lo: "loss" in lo or "损耗" in h,))

# ── Phase Center ──
_reg_fn("pc_theta_mm",
    lambda h: ("pc" in h.lower() or "phase center" in h.lower() or "相位中心" in h) and ("theta" in h.lower() or "θ" in h))
_reg_fn("pc_phi_mm",
    lambda h: ("pc" in h.lower() or "phase center" in h.lower() or "相位中心" in h) and "phi" in h.lower())


# ── 效率 dB ──
def _match_efficiency_db(header: str) -> bool:
    """效率 dB — 必须在 efficiency_pct 之前匹配"""
    lo = header.lower()
    if ("efficiency" not in lo and "效率" not in header):
        return False
    if "total" in lo or "总" in header:
        return False
    return "db" in lo and "%" not in header and "％" not in header and "pct" not in lo
_reg_fn("efficiency_db", _match_efficiency_db)

_reg_fn("total_efficiency",
    lambda h: ("total" in h.lower() or "总" in h) and ("efficiency" in h.lower() or "效率" in h))

_reg("efficiency_pct", must_contain="efficiency", cn="效率", must_not_contain="total db 总")

# ── 圆极化 RHCP/LHCP (必须在 LAG 之前，避免被 lag_single 误匹配) ──
_reg_fn("rhcp_single",
    lambda h: ("rhcp" in h.lower() or "右旋圆极化" in h or "右旋" in h) and ("theta" in h.lower() or "θ" in h or "°" in h or "@" in h))
_reg_fn("cp_xpi_single",
    lambda h: ("cp" in h.lower() and "xpi" in h.lower()) or ("cp-xpi" in h.lower()) or ("交叉极化" in h and ("theta" in h.lower() or "θ" in h or "@" in h)))

# ── AR 轴比 (必须在 LAG 之前) ──
_reg_fn("ar_single",
    lambda h: ("ar" in h.lower() or "axial" in h.lower() or "轴比" in h) and ("theta" in h.lower() or "θ" in h or "@" in h) and "~" not in h and "～" not in h)
_reg_fn("ar_range",
    lambda h: ("ar" in h.lower() or "axial" in h.lower() or "轴比" in h) and ("~" in h or "～" in h or ("-" in h and ("theta" in h.lower() or "θ" in h))))

# ── LAG 单角度 / 范围 (委托给 lag_config 正则，此处用关键词兜底) ──
def _match_lag_single(header: str) -> bool:
    """LAG 单角度: Gain at Theta=30 / 增益@θ=30°"""
    from .lag_config import _RE_LAG_SINGLE, _RE_LAG_SINGLE_NO_PREFIX
    lo = header.lower()
    # 先排除范围 (含 ~, ～, 或 Theta=N-M 形式)
    if "~" in header or "～" in header:
        return False
    if _RE_LAG_RANGE.search(header) or _RE_LAG_RANGE_NO_PREFIX.search(header):
        return False
    if _RE_LAG_SINGLE.search(header) or _RE_LAG_SINGLE_NO_PREFIX.search(header):
        return True
    # 中文兜底: "增益@30°"
    if ("gain" in lo or "增益" in header) and ("theta" in lo or "θ" in lo or "°" in header):
        import re
        return bool(re.search(r"(?:theta|θ)\s*[=＝]\s*\d+", lo) or re.search(r"@\s*\d+", lo))
    return False
_reg_fn("lag_single", _match_lag_single)

def _match_lag_range(header: str) -> bool:
    """LAG 范围: Gain at Theta=0~70 / 增益@θ=0~70°"""
    from .lag_config import _RE_LAG_RANGE, _RE_LAG_RANGE_NO_PREFIX
    if _RE_LAG_RANGE.search(header) or _RE_LAG_RANGE_NO_PREFIX.search(header):
        return True
    lo = header.lower()
    if ("gain" in lo or "增益" in header) and ("theta" in lo or "θ" in lo or "°" in header):
        if "~" in header or "～" in header:
            return True
        # 检查 "Theta=0-70" 模式（排除 Theta=30 单角度）
        import re
        return bool(re.search(r"(?:theta|θ)\s*[=＝]\s*\d+.*[-~]\s*\d+", lo))
    return False
_reg_fn("lag_range", _match_lag_range)

# ── 波束参数 ──
_reg("theta_bw",     must_contain="beamwidth", cn="波束宽度",
     extra_checks=(lambda h, lo: ("3db" in lo or "3dB" in lo or "3ｄB" in lo) and ("theta" in lo or "θ" in lo),))
_reg("phi_bw",       must_contain="beamwidth", cn="波束宽度",
     extra_checks=(lambda h, lo: ("3db" in lo or "3dB" in lo or "3ｄB" in lo) and "phi" in lo,))
_reg("front_back_ratio", must_contain="front", cn="前后比",
     extra_checks=(lambda h, lo: ("back" in lo or "后" in h) and ("ratio" in lo or "比" in h or "/" in h),))

# ── 功率统计比率 ──
_reg("max_min_ratio",  must_contain="ratio", cn="峰均比|比率",
     extra_checks=(lambda h, lo: ("max" in lo or "最大" in h) and ("min" in lo or "最小" in h),))
_reg("max_avg_ratio",  must_contain="ratio", cn="比率",
     extra_checks=(lambda h, lo: ("max" in lo or "最大" in h) and ("avg" in lo or "average" in lo or "平均" in h),))
_reg("min_avg_ratio",  must_contain="ratio", cn="比率",
     extra_checks=(lambda h, lo: ("min" in lo or "最小" in h) and ("avg" in lo or "average" in lo or "平均" in h),))


# ── PRP_120 ──
_reg("prp_120",       must_contain="prp", cn="部分辐射功率", extra_checks=(lambda h, lo: "120" in h,))

# ── NHPRP 自定义角度 ──
_reg("nhprp_custom",  must_contain="nhprp", cn="近水平面", extra_checks=(lambda h, lo: "custom" in lo or "自定义" in h or "±" in h,))

# ── 半球比率 ──


# ── 比率类型 (必须在基类之前注册, 否则会被基类误匹配) ──


# ── NHPRP 比率 ──

_reg("nhprp30_ratio",  extra_checks=(lambda h, lo: "nhprp" in lo and "30" in lo and ("ratio" in lo or "比" in h or "/" in h),))
_reg("nhprp225_ratio", extra_checks=(lambda h, lo: "nhprp" in lo and ("22.5" in lo or "22" in lo) and ("ratio" in lo or "比" in h or "/" in h),))

# ── TIS (Total Isotropic Sensitivity) ──
_reg_fn("tis",
    lambda h: ("tis" in h.lower()) or ("total" in h.lower() and "isotropic" in h.lower() and "sensitivity" in h.lower()) or ("总" in h and "各向同性" in h and "灵敏度" in h))

# ── NHPIS ──
_reg("nhpis_45",      must_contain="nhpis", cn="近水平面灵敏度", keywords=("45",))
_reg("nhpis_30",      must_contain="nhpis", cn="近水平面灵敏度", keywords=("30",))
_reg("nhpis_225",     must_contain="nhpis", cn="近水平面灵敏度",
     extra_checks=(lambda h, lo: any(x in lo for x in ("22.5", "22", "pi/8", "π/8")),))
_reg("nhpis_custom",  must_contain="nhpis", cn="近水平面灵敏度", extra_checks=(lambda h, lo: "custom" in lo or "自定义" in h,))

# ── 半球 PIS ──
_reg("uh_pis",        must_contain="pis", cn="上半球灵敏度",
     extra_checks=(lambda h, lo: "upper" in lo or "uh" in lo or "上半球" in h,))
_reg("lh_pis",        must_contain="pis", cn="下半球灵敏度",
     extra_checks=(lambda h, lo: "lower" in lo or "lh" in lo or "下半球" in h,))
_reg("pis_120",       must_contain="pis", cn="部分灵敏度", extra_checks=(lambda h, lo: "120" in h,))

# ── NHPIS 比率 ──
_reg("nhpis45_ratio",  extra_checks=(lambda h, lo: "nhpis" in lo and "45" in lo and ("ratio" in lo or "比" in h or "/" in h),))
_reg("nhpis30_ratio",  extra_checks=(lambda h, lo: "nhpis" in lo and "30" in lo and ("ratio" in lo or "比" in h or "/" in h),))
_reg("nhpis225_ratio", extra_checks=(lambda h, lo: "nhpis" in lo and ("22.5" in lo or "22" in lo) and ("ratio" in lo or "比" in h or "/" in h),))

# ── TRP/NHPRP 中文补充 ──
# (trp 已有 must_contain="trp", 加中文)
# 更新 trp 匹配器以支持中文
_CN_TRP_MATCHER = _col_classifier(must_contain="trp", cn="总辐射功率|辐射功率|全向辐射",
    extra_checks=(lambda h, lo: "nhprp" not in lo and "nhpis" not in lo,))
_COLUMN_CLASSIFIERS["trp"] = (_CN_TRP_MATCHER, False)

_CN_NHPRP45_MATCHER = _col_classifier(must_contain="nhprp", cn="近水平面辐射功率", keywords=("45",))
_COLUMN_CLASSIFIERS["nhprp_45"] = (_CN_NHPRP45_MATCHER, False)

_CN_NHPRP30_MATCHER = _col_classifier(must_contain="nhprp", cn="近水平面辐射功率", keywords=("30",))
_COLUMN_CLASSIFIERS["nhprp_30"] = (_CN_NHPRP30_MATCHER, False)

_CN_PEAK_EIRP_MATCHER = _col_classifier(must_contain="eirp", cn="峰值|峰",
    exact_words=("eirppeak", "pkgain", "peakgain", "峰值eirp", "峰值增益"),
    extra_checks=(lambda h, lo: "peak" in lo or "pk" in lo or "峰" in h,))
_COLUMN_CLASSIFIERS["peak_eirp"] = (_CN_PEAK_EIRP_MATCHER, False)

# 更新 AR 匹配器支持中文


# 更新 PRP 匹配器支持中文
_CN_UH_PRP = _col_classifier(must_contain="prp", cn="上半球",
    extra_checks=(lambda h, lo: "upper" in lo or "uh" in lo or "上半球" in h,))
_COLUMN_CLASSIFIERS["uh_prp"] = (_CN_UH_PRP, False)

_CN_LH_PRP = _col_classifier(must_contain="prp", cn="下半球",
    extra_checks=(lambda h, lo: "lower" in lo or "lh" in lo or "下半球" in h,))
_COLUMN_CLASSIFIERS["lh_prp"] = (_CN_LH_PRP, False)

# 更新 Max/Min 功率支持中文
_CN_MAX_POWER = _col_classifier(must_contain="power", cn="功率",
    extra_checks=(lambda h, lo: ("maximum" in lo or "max" in lo or "最大" in h) and "average" not in lo and "平均" not in h,))
_COLUMN_CLASSIFIERS["max_power"] = (_CN_MAX_POWER, False)

_CN_MIN_POWER = _col_classifier(must_contain="power", cn="功率",
    extra_checks=(lambda h, lo: ("minimum" in lo or "min" in lo or "最小" in h) and "average" not in lo and "平均" not in h,))
_COLUMN_CLASSIFIERS["min_power"] = (_CN_MIN_POWER, False)

_CN_AVG_POWER = _col_classifier(must_contain="power", cn="功率",
    extra_checks=(lambda h, lo: ("average" in lo or "avg" in lo or "平均" in h),))
_COLUMN_CLASSIFIERS["avg_power"] = (_CN_AVG_POWER, False)

_CN_AVG_GAIN = _col_classifier(must_contain="gain", cn="增益",
    extra_checks=(lambda h, lo: ("average" in lo or "avg" in lo or "平均" in h) and "at" not in lo and "@" not in h,))
_COLUMN_CLASSIFIERS["avg_gain"] = (_CN_AVG_GAIN, False)

# ═══════════════════════════════════════════════════════════════
# 公开 API: 聚合分类器 + 向后兼容包装器
# ═══════════════════════════════════════════════════════════════

def classify_column(header: str) -> str | None:
    """统一列头分类入口。按优先级匹配所有已注册列类型。"""
    for col_type, (matcher, _is_ratio) in _COLUMN_CLASSIFIERS.items():
        try:
            if matcher(header):
                return col_type
        except Exception:
            continue
    return detect_ratio_column_type(header)


# ── 向后兼容的 is_xxx_column 函数 (委托给注册表) ──
def _make_is_fn(col_type: str):
    """为指定 col_type 生成 is_xxx_column() 包装器。"""
    matcher, _ = _COLUMN_CLASSIFIERS.get(col_type, (lambda h: False, False))
    return lambda header: matcher(header)

is_frequency_column          = _make_is_fn("frequency")
is_directivity_column        = _make_is_fn("directivity")
is_efficiency_column         = _make_is_fn("efficiency_pct")
is_efficiency_db_column      = _make_is_fn("efficiency_db")
is_gain_column               = _make_is_fn("gain")
is_trp_column                = _make_is_fn("trp")
is_nhprp_45_column           = _make_is_fn("nhprp_45")
is_nhprp_30_column           = _make_is_fn("nhprp_30")
is_nhprp_225_column          = _make_is_fn("nhprp_225")
is_nhprp_custom_column       = _make_is_fn("nhprp_custom")
is_peak_eirp_column          = _make_is_fn("peak_eirp")
is_ar_single_column          = _make_is_fn("ar_single")
is_ar_range_column           = _make_is_fn("ar_range")
is_uh_prp_column             = _make_is_fn("uh_prp")
is_lh_prp_column             = _make_is_fn("lh_prp")
is_prp_120_column            = _make_is_fn("prp_120")
is_max_power_column          = _make_is_fn("max_power")
is_min_power_column          = _make_is_fn("min_power")
is_avg_gain_column           = _make_is_fn("avg_gain")
is_avg_power_column          = _make_is_fn("avg_power")
is_max_min_ratio_column      = _make_is_fn("max_min_ratio")
is_max_avg_ratio_column      = _make_is_fn("max_avg_ratio")
is_min_avg_ratio_column      = _make_is_fn("min_avg_ratio")
is_boresight_phi_column      = _make_is_fn("boresight_phi")
is_boresight_theta_column    = _make_is_fn("boresight_theta")
is_theta_bw_column           = _make_is_fn("theta_bw")
is_phi_bw_column             = _make_is_fn("phi_bw")
is_front_back_ratio_column   = _make_is_fn("front_back_ratio")
is_xpi_boresight_column      = _make_is_fn("xpi_boresight")
is_xpi_mean_column           = _make_is_fn("xpi_mean")
is_xpi_min_column            = _make_is_fn("xpi_min")
is_cp_xpi_single_column      = _make_is_fn("cp_xpi_single")
is_rhcp_single_column        = _make_is_fn("rhcp_single")
is_total_efficiency_column   = _make_is_fn("total_efficiency")
is_mismatch_loss_column      = _make_is_fn("mismatch_loss_db")
is_pc_theta_column           = _make_is_fn("pc_theta_mm")
is_pc_phi_column             = _make_is_fn("pc_phi_mm")
is_uh_ratio_column           = _make_is_fn("uh_ratio")
is_lh_ratio_column           = _make_is_fn("lh_ratio")
is_nhprp45_ratio_column      = _make_is_fn("nhprp45_ratio")
is_nhprp30_ratio_column      = _make_is_fn("nhprp30_ratio")
is_nhprp225_ratio_column     = _make_is_fn("nhprp225_ratio")
is_tis_column                = _make_is_fn("tis")
is_nhpis_45_column           = _make_is_fn("nhpis_45")
is_nhpis_30_column           = _make_is_fn("nhpis_30")
is_nhpis_225_column          = _make_is_fn("nhpis_225")
is_nhpis_custom_column       = _make_is_fn("nhpis_custom")
is_uh_pis_column             = _make_is_fn("uh_pis")
is_lh_pis_column             = _make_is_fn("lh_pis")
is_pis_120_column            = _make_is_fn("pis_120")
is_nhpis45_ratio_column      = _make_is_fn("nhpis45_ratio")
is_nhpis30_ratio_column      = _make_is_fn("nhpis30_ratio")
is_nhpis225_ratio_column     = _make_is_fn("nhpis225_ratio")
is_lag_single_column         = _make_is_fn("lag_single")
is_lag_range_column          = _make_is_fn("lag_range")
is_xpi_boresight_column      = _make_is_fn("xpi_boresight")
is_xpi_mean_column           = _make_is_fn("xpi_mean")
is_xpi_min_column            = _make_is_fn("xpi_min")
is_total_efficiency_column   = _make_is_fn("total_efficiency")
is_mismatch_loss_column      = _make_is_fn("mismatch_loss_db")
is_pc_theta_column           = _make_is_fn("pc_theta_mm")
is_pc_phi_column             = _make_is_fn("pc_phi_mm")


def detect_ratio_column_type(header: str) -> str | None:
    """检测比率列类型，返回带 db/pct 后缀的 column type。

    Returns:
        None (非比率列) 或 column type 如 "nhprp45_ratio", "uh_ratio" 等。
    """
    h = header.lower()
    if "ratio" not in h:
        return None
    # 比率基点 → base type 映射
    ratio_bases = [
        (("nhprp4", "nhprp45", "nhprp+/-45", "nhprp+-45"), "nhprp45_ratio"),
        (("nhprp3", "nhprp30", "nhprp+/-30", "nhprp+-30"), "nhprp30_ratio"),
        (("nhprp2", "nhprp225", "nhprp22.5", "nhprp+/-22.5"), "nhprp225_ratio"),
        (("upper", "hem"), "uh_ratio"),
        (("lower", "hem"), "lh_ratio"),
    ]
    for keywords, base in ratio_bases:
        if all(k in h for k in keywords):
            if "%" in header or "pct" in h:
                return base + "_pct"
            if "db" in h:
                return base + "_db"
            return base
    return None


@dataclass
class ColumnInfo:
    """单列信息。"""
    col_letter: str          # 列字母，如 "B"
    col_index: int           # 1-based 列号
    raw_header: str          # 原始列头文本
    normalized_header: str   # 规范化列头
    col_type: str            # "frequency" | "directivity" | "efficiency_pct"
                             # | "efficiency_db" | "total_efficiency_pct"
                             # | "total_efficiency_db" | "gain" | "lag_single"
                             # | "lag_range" | "xpi_boresight" | "xpi_mean"
                             # | "xpi_min" | "mismatch_loss_db"
                             # | "pc_theta_mm" | "pc_phi_mm" | "unknown"


@dataclass
class SheetInfo:
    """一个工作表的完整结构信息。"""
    name: str
    header_row: int                    # 列头所在行号
    data_start_row: int                # 数据起始行号
    data_end_row: int                  # 数据结束行号
    columns: list[ColumnInfo] = field(default_factory=list)
    frequencies: list[float] = field(default_factory=list)
    lag_config: LagConfig = field(default_factory=LagConfig)
    ar_config: LagConfig = field(default_factory=LagConfig)  # AR 角度配置
    theta_range: str | None = None  # e.g., "0-110°"


def read_template(template_path: str) -> list[SheetInfo]:
    """读取输出模板，返回所有工作表信息。

    会自动跳过纯元数据的 Sheet（无 Frequency 列）。
    """
    wb = openpyxl.load_workbook(template_path, data_only=True)
    sheets: list[SheetInfo] = []

    for ws in wb.worksheets:
        info = _parse_sheet(ws)
        if info is not None:
            sheets.append(info)

    wb.close()
    return sheets


def _classify_by_builtin(raw: str, norm: str) -> str | None:
    """内置函数检测链（委托给 classify_column 聚合器）。

    所有 _reg() 注册的列类型自动生效，无需手动维护此函数。
    """
    return classify_column(raw)


def _parse_sheet(ws) -> SheetInfo | None:
    """解析单个 Sheet。返回 None 表示非天线数据 Sheet（无 Frequency 列）。"""
    name = ws.title
    max_row = ws.max_row or 100
    max_col = ws.max_column or 20

    # ---- 扫描行 ----
    header_row = None
    data_start_row = None
    data_end_row = max_row

    for row_idx in range(1, min(max_row + 1, 200)):
        row_values = [_cell_str(ws.cell(row_idx, c)) for c in range(1, max_col + 1)]

        # 寻找 Frequency 列所在行 → 列头行
        for c, val in enumerate(row_values):
            if is_frequency_column(val):
                header_row = row_idx
                break

        if header_row is not None:
            break

    if header_row is None:
        # 无 Frequency 列 → 不是天线数据 Sheet
        return None

    # ---- 解析列头 ----
    columns: list[ColumnInfo] = []
    lag_headers: list[str] = []
    ar_headers: list[str] = []

    for c in range(1, max_col + 1):
        raw = _cell_str(ws.cell(header_row, c))
        if not raw:
            continue
        norm = normalize_header(raw)
        col_letter = openpyxl.utils.get_column_letter(c)

        # 分类 — 内置函数优先，JSON 仅补充，最后比率列 + unknown
        builtin_type = _classify_by_builtin(raw, norm)
        if builtin_type is not None:
            ctype = builtin_type
        else:
            json_type = _classify_by_json_patterns(raw)
            if json_type is not None:
                ctype = json_type
            else:
                ratio_type = detect_ratio_column_type(raw)
                if ratio_type is not None:
                    ctype = ratio_type
                else:
                    ctype = "unknown"

        cinfo = ColumnInfo(
            col_letter=col_letter,
            col_index=c,
            raw_header=raw,
            normalized_header=norm,
            col_type=ctype,
        )
        columns.append(cinfo)

        # 收集 LAG / AR 列头用于解析
        if ctype in ("lag_single", "lag_range"):
            lag_headers.append(raw)
        if ctype in ("ar_single", "ar_range"):
            ar_headers.append(raw)

    # ---- 解析频点列表 ----
    data_start_row = header_row + 1
    frequencies: list[float] = []
    freq_col = None
    for cinfo in columns:
        if cinfo.col_type == "frequency":
            freq_col = cinfo.col_index
            break

    if freq_col:
        for r in range(data_start_row, max_row + 1):
            val = ws.cell(r, freq_col).value
            if val is None or str(val).strip() == "" or str(val).strip() == "-":
                data_end_row = r - 1
                break
            try:
                frequencies.append(float(val))
            except (ValueError, TypeError):
                data_end_row = r - 1
                break

    # ---- 解析 LAG / AR 配置 ----
    lag_config = LagConfig.from_template_headers(lag_headers)
    ar_config = LagConfig.from_ar_headers(ar_headers)

    # ---- 读取 θ 范围 ----
    theta_range = None
    # 在 header_row 前几行搜索 "θ Range"
    for r in range(1, header_row):
        for c in range(1, max_col + 1):
            v = _cell_str(ws.cell(r, c))
            if "θ" in v and ("range" in v.lower() or "step" in v.lower()):
                # 尝试读同一行后续值
                theta_range = _cell_str(ws.cell(r, c + 1))
                break

    return SheetInfo(
        name=name,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        columns=columns,
        frequencies=frequencies,
        lag_config=lag_config,
        ar_config=ar_config,
        theta_range=theta_range,
    )


def _cell_str(cell) -> str:
    """单元格值 → 字符串。"""
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()
