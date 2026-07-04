"""
多天线配置数据模型
===================
每个项目可包含多个天线，每个天线有独立的:
  - 数据文件
  - 测试模式
  - 天线参数配置 (LagConfig × 4)
  - 图表配置
  - SDT 后缀

支持从 Excel 导入/导出测试计划。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from src.lag_config import LagConfig


@dataclass
class AntennaConfig:
    """单个天线的完整配置。"""
    name: str = ""                              # 天线名称/标识
    data_files: list[str] = field(default_factory=list)
    test_mode: int = 0                          # 0=无源 1=有源发射 2=有源接收
    sdt_suffix: str = ""                        # SDT tag 后缀, 如 "_L1_amp"

    # 角度配置
    lag_config: LagConfig = field(default_factory=LagConfig)
    ar_lag_config: LagConfig = field(default_factory=LagConfig)
    rhcp_lag_config: LagConfig = field(default_factory=LagConfig)
    cpxpi_lag_config: LagConfig = field(default_factory=LagConfig)

    # 参数选择
    required_params: set = field(default_factory=set)
    extra_params: set = field(default_factory=set)

    # 图表选择 (存 key 列表, 如 ["pattern_3d_gain", "chart_gain_freq"])
    chart_keys: list[str] = field(default_factory=list)

    # 元数据
    metadata: dict = field(default_factory=dict)
    notes: str = ""

    def to_dict(self) -> dict:
        return {
            "name": self.name,
            "data_files": self.data_files,
            "test_mode": self.test_mode,
            "sdt_suffix": self.sdt_suffix,
            "lag_config": {
                "singles": self.lag_config.singles_sorted,
                "ranges": self.lag_config.ranges_sorted,
            },
            "ar_lag_config": {
                "singles": self.ar_lag_config.singles_sorted,
                "ranges": self.ar_lag_config.ranges_sorted,
            },
            "rhcp_lag_config": {
                "singles": self.rhcp_lag_config.singles_sorted,
                "ranges": self.rhcp_lag_config.ranges_sorted,
            },
            "cpxpi_lag_config": {
                "singles": self.cpxpi_lag_config.singles_sorted,
                "ranges": self.cpxpi_lag_config.ranges_sorted,
            },
            "required_params": list(self.required_params),
            "extra_params": list(self.extra_params),
            "chart_keys": self.chart_keys,
            "metadata": self.metadata,
            "notes": self.notes,
        }

    @classmethod
    def from_dict(cls, d: dict) -> AntennaConfig:
        return cls(
            name=d.get("name", ""),
            data_files=d.get("data_files", []),
            test_mode=d.get("test_mode", 0),
            sdt_suffix=d.get("sdt_suffix", ""),
            lag_config=LagConfig(
                single_angles=d.get("lag_config", {}).get("singles", []),
                ranges=d.get("lag_config", {}).get("ranges", []),
            ),
            ar_lag_config=LagConfig(
                single_angles=d.get("ar_lag_config", {}).get("singles", []),
                ranges=d.get("ar_lag_config", {}).get("ranges", []),
            ),
            rhcp_lag_config=LagConfig(
                single_angles=d.get("rhcp_lag_config", {}).get("singles", []),
                ranges=d.get("rhcp_lag_config", {}).get("ranges", []),
            ),
            cpxpi_lag_config=LagConfig(
                single_angles=d.get("cpxpi_lag_config", {}).get("singles", []),
                ranges=d.get("cpxpi_lag_config", {}).get("ranges", []),
            ),
            required_params=set(d.get("required_params", [])),
            extra_params=set(d.get("extra_params", [])),
            chart_keys=d.get("chart_keys", []),
            metadata=d.get("metadata", {}),
            notes=d.get("notes", ""),
        )


@dataclass
class MultiAntennaConfig:
    """项目多天线配置。"""
    project_name: str = ""
    customer: str = ""
    antennas: list[AntennaConfig] = field(default_factory=list)
    metadata: dict = field(default_factory=dict)

    def add_antenna(self, name: str = "", data_files: list[str] = None) -> AntennaConfig:
        cfg = AntennaConfig(name=name, data_files=data_files or [])
        self.antennas.append(cfg)
        return cfg

    def remove_antenna(self, index: int):
        if 0 <= index < len(self.antennas):
            self.antennas.pop(index)

    def get_antenna(self, name: str) -> AntennaConfig | None:
        for a in self.antennas:
            if a.name == name:
                return a
        return None

    def to_dict(self) -> dict:
        return {
            "project_name": self.project_name,
            "customer": self.customer,
            "metadata": self.metadata,
            "antennas": [a.to_dict() for a in self.antennas],
        }

    @classmethod
    def from_dict(cls, d: dict) -> MultiAntennaConfig:
        return cls(
            project_name=d.get("project_name", ""),
            customer=d.get("customer", ""),
            metadata=d.get("metadata", {}),
            antennas=[AntennaConfig.from_dict(a) for a in d.get("antennas", [])],
        )


# ═══════════════════════════════════════════════════════════════
# Excel 导入/导出
# ═══════════════════════════════════════════════════════════════

def export_to_excel(config: MultiAntennaConfig, output_path: str):
    """导出多天线配置到 Excel 测试计划。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    # 汇总 Sheet
    ws = wb.active
    ws.title = "汇总"
    headers = ["天线名称", "数据文件", "测试模式", "参数数", "图表数", "SDT后缀"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h); ws.cell(1, c).font = Font(bold=True)
    for i, ant in enumerate(config.antennas):
        r = i + 2
        ws.cell(r, 1, ant.name)
        ws.cell(r, 2, "; ".join(ant.data_files))
        ws.cell(r, 3, ["无源", "有源发射", "有源接收"][ant.test_mode])
        ws.cell(r, 4, len(ant.required_params))
        ws.cell(r, 5, len(ant.chart_keys))
        ws.cell(r, 6, ant.sdt_suffix)

    # 每个天线一个 Sheet
    for ant in config.antennas:
        safe_name = ant.name[:31].replace("/", "_").replace("\\", "_")
        ws = wb.create_sheet(safe_name)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60

        ws.cell(1, 1, "属性"); ws.cell(1, 2, "值")
        ws.cell(1, 1).font = Font(bold=True); ws.cell(1, 2).font = Font(bold=True)

        items = [
            ("天线名称", ant.name),
            ("数据文件", "; ".join(ant.data_files)),
            ("测试模式", ["无源", "有源发射", "有源接收"][ant.test_mode]),
            ("SDT 后缀", ant.sdt_suffix),
            ("", ""),
            ("Gain 单角度", ", ".join(f"{a}°" for a in ant.lag_config.singles_sorted)),
            ("Gain 范围", ", ".join(f"{l}~{h}°" for l, h in ant.lag_config.ranges_sorted)),
            ("AR 单角度", ", ".join(f"{a}°" for a in ant.ar_lag_config.singles_sorted)),
            ("", ""),
            ("天线参数", ", ".join(sorted(ant.required_params))),
            ("额外参数", ", ".join(sorted(ant.extra_params))),
            ("图表", ", ".join(ant.chart_keys)),
        ]
        for i, (k, v) in enumerate(items):
            ws.cell(i+2, 1, k); ws.cell(i+2, 2, str(v))

    wb.save(output_path)


def import_from_excel(excel_path: str) -> MultiAntennaConfig:
    """从 Excel 测试计划导入多天线配置。"""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    config = MultiAntennaConfig()

    # 读取汇总 Sheet
    if "汇总" in wb.sheetnames:
        ws = wb["汇总"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ant = AntennaConfig(
                    name=str(row[0]),
                    data_files=[f.strip() for f in str(row[1] or "").split(";") if f.strip()],
                    test_mode={"无源": 0, "有源发射": 1, "有源接收": 2}.get(str(row[2]), 0),
                    sdt_suffix=str(row[5] or ""),
                )
                config.antennas.append(ant)

    wb.close()
    return config


# ═══════════════════════════════════════════════════════════════
# 天线名称自动提取
# ═══════════════════════════════════════════════════════════════

def extract_antenna_name(filepath: str) -> str:
    """从文件名自动提取天线标识。"""
    import re
    name = Path(filepath).stem
    # 去掉常见后缀
    for suffix in ["_withAMP", "_noAMP", "_with_AMP", "_without_AMP",
                   "_withamp", "_noamp", "_merged", "_FinalSummary",
                   "_test", "_result", "_report"]:
        name = re.sub(re.escape(suffix), "", name, flags=re.IGNORECASE)
    # 去掉末尾数字/下划线
    name = re.sub(r'[\d_]+$', '', name).strip('_')
    return name or Path(filepath).stem
