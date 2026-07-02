"""
FinalSummary .xlsx → merged CSV 转换器
=====================================
将 FinalSummary 格式的 Excel 转换为项目标准 merged CSV，
供 MergedCSVParser 直接读取。一次转换 ~5 分钟，之后秒读。

GUI 和 CLI 共用同一入口: convert_fs_to_csv(src_path, out_path, progress_cb)
"""

from __future__ import annotations

import os
from collections.abc import Callable

import numpy as np
import openpyxl

from .raw_converter import _write_normal_csv


def convert_fs_to_csv(
    src_path: str,
    out_path: str | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> str:
    """将 FinalSummary .xlsx 转换为 merged CSV。

    Args:
        src_path: 源 FinalSummary .xlsx 路径
        out_path: 输出 .csv 路径, 默认在同目录生成 `{stem}_merged.csv`
        progress_callback: (current, total, message) 进度回调

    Returns:
        输出 .csv 文件路径
    """
    if out_path is None:
        stem = os.path.splitext(os.path.basename(src_path))[0]
        # strip trailing .json if present
        if stem.endswith('.json'):
            stem = stem[:-5]
        out_path = os.path.join(os.path.dirname(src_path), f"{stem}_merged.csv")

    def _report(cur, tot, msg):
        if progress_callback:
            progress_callback(cur, tot, msg)

    _report(0, 1, "打开 workbook...")
    wb = openpyxl.load_workbook(src_path, data_only=True, read_only=False)

    # 收集频点
    freqs: list[float] = []
    for sn in wb.sheetnames:
        try:
            freqs.append(float(sn))
        except ValueError:
            pass
    freqs.sort()
    n_freqs = len(freqs)

    # 探测结构
    sn0 = str(int(freqs[0])) if freqs[0] == int(freqs[0]) else str(freqs[0])
    ws0 = wb[sn0]

    theta_vals: list[float] = []
    theta_start = 0
    n_phi = 0
    n_theta = 0
    for r_idx, row in enumerate(ws0.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        vals = [v for v in row if v is not None]
        if vals and isinstance(vals[0], (int, float)):
            theta_start = r_idx
            for v in list(ws0.iter_rows(min_row=r_idx - 1, max_row=r_idx - 1, values_only=True))[0][1:]:
                if v is not None:
                    try:
                        theta_vals.append(float(v))
                    except (ValueError, TypeError):
                        pass
            for r2 in ws0.iter_rows(min_row=theta_start, max_row=theta_start + 400,
                                    min_col=1, max_col=1, values_only=True):
                v = r2[0]
                if v is None:
                    break
                try:
                    float(v)
                    n_phi += 1
                except (ValueError, TypeError):
                    break
            n_theta = len(theta_vals)
            break

    # 扫描 section 标签
    tp_start = pp_start = pp_phase_start = 0
    for r_idx, row in enumerate(ws0.iter_rows(min_row=theta_start + n_phi,
                                               max_row=ws0.max_row, max_col=3,
                                               values_only=True),
                                theta_start + n_phi):
        v = str(row[0]) if row[0] else ""
        if 'Phase' in v and 'Phi' not in v and tp_start == 0:
            tp_start = r_idx + 2
        if 'Phi Polarization' in v:
            pp_start = r_idx + 3
        if 'Phase' in v and pp_start > 0 and r_idx > pp_start:
            pp_phase_start = r_idx + 2
            break

    def _read_section(sec_start: int) -> np.ndarray | None:
        if sec_start <= 0:
            return None
        data = np.full((n_freqs, n_phi, n_theta), np.nan, dtype=np.float64)
        for fi, freq in enumerate(freqs):
            sn = str(int(freq)) if freq == int(freq) else str(freq)
            ws = wb[sn]
            for pi, row in enumerate(ws.iter_rows(min_row=sec_start,
                                                   max_row=sec_start + n_phi - 1,
                                                   min_col=2, max_col=1 + n_theta,
                                                   values_only=True)):
                if pi >= n_phi:
                    break
                for ti, v in enumerate(row):
                    if ti >= n_theta:
                        break
                    if v is not None:
                        try:
                            data[fi, pi, ti] = float(v)
                        except (ValueError, TypeError):
                            pass
            _report(fi + 1, n_freqs + 1, f"读取中... ({fi + 1}/{n_freqs})")
        return data

    # 检查并去掉 phi=360° 重复行（与 phi=0° 重合，保留 0-359）
    phi_col_a = []
    for r2 in ws0.iter_rows(min_row=theta_start, max_row=theta_start + n_phi,
                            min_col=1, max_col=1, values_only=True):
        v = r2[0]
        if v is None: break
        try: phi_col_a.append(float(v))
        except (ValueError, TypeError): break
    if phi_col_a and phi_col_a[-1] == 360.0:
        n_phi -= 1  # 去掉最后一个 phi=360°

    # 读 4 个 section
    _report(0, n_freqs, "读取 Theta LogMag...")
    tl = _read_section(theta_start)
    _report(0, n_freqs, "读取 Theta Phase...")
    tp = _read_section(tp_start) if tp_start > 0 else None
    _report(0, n_freqs, "读取 Phi LogMag...")
    pl = _read_section(pp_start) if pp_start > 0 else None
    _report(0, n_freqs, "读取 Phi Phase...")
    pp = _read_section(pp_phase_start) if pp_phase_start > 0 else None

    wb.close()

    # 写标准 merged CSV
    phi_vals = [float(i) for i in range(n_phi)]
    _report(n_freqs, n_freqs + 1, "写入 CSV...")
    _write_normal_csv(
        out_path,
        f"File Name: {os.path.basename(src_path)} (converted from FinalSummary)",
        freqs, theta_vals, phi_vals, tl, tp, pl, pp, None,
    )

    _report(n_freqs + 1, n_freqs + 1, "完成")
    return out_path
