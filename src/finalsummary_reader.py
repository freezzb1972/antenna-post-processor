"""
FinalSummary.xlsx 直接读取器 (v3)
=================================
实现 DataSource 接口，从 FinalSummary Excel 逐频点读取数据。

核心不变点：
  - 列 = Theta 坐标（一行连续数值）
  - 行 = Phi 坐标（列 A 数值递增）
  - Theta 表头行之前可能有若干描述行（名称/类型等）

自动探测结构，不硬编码任何行号/列数/phi 计数。
"""

from __future__ import annotations

from collections import OrderedDict
from typing import Dict, List, Optional, Union

import numpy as np
import openpyxl

from .datasource import DataSource


class _LRUDict(OrderedDict):
    """定长 LRU 缓存: 超过 maxsize 时自动淘汰最久未使用的条目。
    每次访问自动将条目标记为最近使用 (move_to_end)。
    """

    def __init__(self, maxsize: int = 128):
        super().__init__()
        self._maxsize = maxsize

    def __setitem__(self, key, value):
        super().__setitem__(key, value)
        if len(self) > self._maxsize:
            self.popitem(last=False)

    def __getitem__(self, key):
        value = super().__getitem__(key)
        self.move_to_end(key)
        return value


def _is_numeric(v) -> bool:
    """检查值是否可解释为数值（兼容 openpyxl read_only 返回字符串）。"""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v)
            return True
        except ValueError:
            return False
    return False


def _to_float(v):
    """将值转为 float（兼容 openpyxl read_only 返回字符串）。"""
    if v is None:
        return None
    if _is_numeric(v):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v)
        except ValueError:
            return None
    return None


class FinalSummarySource(DataSource):
    """从 FinalSummary.xlsx 逐频点读取天线测试数据。v3 完全自适应。"""

    def __init__(self, path: str):
        self._path = path
        self._wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

        # ---- 频点列表（数字命名的 sheet，排除 Cplx/AxR/Original 等汇总 sheet） ----
        self._freqs: List[float] = []
        for sn in self._wb.sheetnames:
            try:
                self._freqs.append(float(sn))
            except ValueError:
                pass
        if not self._freqs:
            self._wb.close()
            raise ValueError(f"在 {self._path} 中未找到数字命名的频点 Sheet")
        self._freqs.sort()

        # ---- 从第一个频率 sheet 探测结构（只读前 20 行） ----
        sn0 = _freq_sheet_name(self._freqs[0])
        ws0 = self._wb[sn0]

        self._theta_header_row, self._theta_start_row, self._n_phi, self._n_theta, self._data_type = \
            self._probe_structure(ws0)

        # ---- 读 theta 角度（流式，仅 header 行） ----
        self._theta: List[float] = []
        for row in ws0.iter_rows(min_row=self._theta_header_row, max_row=self._theta_header_row,
                                  values_only=True):
            for v in row[1:]:
                if v is not None:
                    try:
                        self._theta.append(float(v))
                    except (ValueError, TypeError):
                        pass

        # ---- 读 phi 角度（从数据区列 A，流式） ----
        self._phi: List[float] = []
        for row in ws0.iter_rows(min_row=self._theta_start_row,
                                 max_row=self._theta_start_row + self._n_phi - 1,
                                 min_col=1, max_col=1, values_only=True):
            v = row[0]
            if v is not None:
                try:
                    self._phi.append(float(v))
                except (ValueError, TypeError):
                    self._phi.append(float(len(self._phi)))

        # ---- 动态探测节结构：扫描标签行定位各 section ----
        self._has_phase = False
        self._theta_phase_start = 0
        self._has_phi_pol = False
        self._phi_pol_start = 0
        self._phi_phase_start = 0

        # 扫描列 A，收集所有节标签行号
        section_labels = self._scan_section_labels(ws0)

        # Theta Phase: Theta 振幅段后第一个 "Phase" 标签 → 数据从 label+2 开始
        after_amp = self._theta_start_row + self._n_phi
        theta_phase_label = section_labels.get('theta_phase_label')
        if theta_phase_label is not None:
            self._has_phase = True
            self._theta_phase_start = theta_phase_label + 2

        # Phi Power: "Phi Polarization" 标签 → 数据从 label+3 开始
        # (label → Power sub-label → Theta/Phi header → data)
        phi_pol_label = section_labels.get('phi_pol_label')
        if phi_pol_label is not None:
            self._has_phi_pol = True
            self._phi_pol_start = phi_pol_label + 3

        # Phi Phase: Phi 段后第二个 "Phase" 标签 → 数据从 label+2 开始
        phi_phase_label = section_labels.get('phi_phase_label')
        if phi_phase_label is not None:
            self._phi_phase_start = phi_phase_label + 2

        # ---- 缓存 (LRU: 最多缓存 512 个频点，覆盖宽频测试场景) ----
        self._cache: _LRUDict = _LRUDict(maxsize=512)

    @staticmethod
    def _probe_structure(ws) -> tuple:
        """只读前 20 行，探测工作表结构。

        不使用 ws.cell() — 全用 iter_rows() 流式读。
        不硬编码任何数字。

        Returns:
            (theta_header_row, theta_start_row, n_phi, n_theta, data_type)
        """
        # 一次性读前 30 行
        preview_rows = []
        theta_header_row = 3
        theta_col_count = 0
        data_type = "logmag"

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1
        ):
            preview_rows.append(row)
            col_a = row[0] if len(row) > 0 else None

            # 记录描述文本（用于数据类型判断）
            if col_a and isinstance(col_a, str):
                vl = col_a.lower()
                if "complex" in vl:
                    data_type = "complex"

            # Theta 表头行：col A 含 "theta/phi" 且后续列有连续数值
            if col_a is not None and isinstance(col_a, str) and "theta" in col_a.lower():
                # 跳过可能为空的 col B/C，找到第一个数值开始的位置
                numeric_count = 0
                for v in row[1:]:
                    if v is not None and _is_numeric(v):
                        numeric_count += 1
                    elif numeric_count > 0:
                        # 已经在数值序列中遇到非数值 → 停止
                        break
                    # 开头遇到 None/空值 → 继续（可能是格式占位列）
                if numeric_count >= 2:
                    theta_header_row = row_idx
                    theta_col_count = numeric_count
                    break

        theta_start_row = theta_header_row + 1

        # ---- 探测 n_phi —— 找振幅段真实边界（遇到空行/标签行停止） ----
        max_r = ws.max_row or 2000
        preview_phi_count = 0
        for row in ws.iter_rows(min_row=theta_start_row, max_row=min(theta_start_row + 9, max_r),
                                values_only=True):
            has_data = any(v is not None and _is_numeric(v) for v in row[1:])
            if has_data:
                preview_phi_count += 1
            elif preview_phi_count > 0:
                break
        if preview_phi_count >= 5:
            # 扫描找真实边界：第一个空行/文本行出现的位置
            phi_count = 0
            for row in ws.iter_rows(min_row=theta_start_row, max_row=max_r, values_only=True):
                col_a = row[0] if len(row) > 0 else None
                has_data = any(v is not None and _is_numeric(v) for v in row[1:])
                if has_data:
                    phi_count += 1
                else:
                    break  # 空行或标签行 = 节边界
        else:
            phi_count = preview_phi_count

        # 确定 n_theta：从 header 行的非空列数
        if theta_col_count == 0:
            # fallback: 检查 header 行的列数
            header_row_data = preview_rows[theta_header_row - 1]
            theta_col_count = sum(1 for v in header_row_data[1:] if v is not None)

        return theta_header_row, theta_start_row, int(phi_count), int(theta_col_count), data_type

    def _scan_section_labels(self, ws) -> dict:
        """动态扫描列 A，定位各 section 标签行号。

        兼容任意行数/列数的 FinalSummary 类文件，不依赖硬编码偏移。
        探测以下节标签：
          - theta_phase_label: 第一个 "Phase" (Theta 振幅段之后)
          - phi_pol_label:     "Phi Polarization"
          - phi_phase_label:   第二个 "Phase" (Phi 振幅段之后)

        Returns:
            dict with keys: theta_phase_label, phi_pol_label, phi_phase_label
            (values 为 int 行号或 None)
        """
        result = {
            'theta_phase_label': None,
            'phi_pol_label': None,
            'phi_phase_label': None,
        }
        phase_labels_found = []

        after_amp = self._theta_start_row + self._n_phi
        max_r = ws.max_row or 2000

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=1, values_only=True),
            start=1
        ):
            val = row[0]
            if val is None or not isinstance(val, str):
                continue
            vl = val.strip().lower()

            # "Phase" 标签
            if vl == 'phase':
                phase_labels_found.append(row_idx)
                # 第一个 Theta 振幅段之后的 Phase → theta_phase_label
                if result['theta_phase_label'] is None and row_idx > after_amp:
                    result['theta_phase_label'] = row_idx

            # "Phi Polarization" 标签
            if 'phi' in vl and 'polar' in vl:
                result['phi_pol_label'] = row_idx

        # Phi phase label: 在 Phi 段之后的第一个 Phase 标签
        if result['phi_pol_label'] is not None:
            for pr in phase_labels_found:
                if pr > result['phi_pol_label']:
                    result['phi_phase_label'] = pr
                    break

        return result

    @property
    def frequencies(self) -> List[float]:
        return list(self._freqs)

    @property
    def theta_angles(self) -> List[float]:
        return list(self._theta)

    @property
    def phi_angles(self) -> List[float]:
        return list(self._phi)

    def read_sections(self, freq_index: int) -> Dict[str, Optional[np.ndarray]]:
        freq = self._freqs[freq_index]

        if freq in self._cache:
            tl, pl, tp_data, pp_data = self._cache[freq]
        else:
            sn = _freq_sheet_name(freq)
            if sn not in self._wb.sheetnames:
                raise KeyError(f"Frequency {freq} MHz not found in {self._path}")

            ws = self._wb[sn]
            ntheta = self._n_theta
            nphi = self._n_phi

            # 读 Theta Pol 幅度（不做 clipping — CTIA/EMQuest 标准无此要求）
            tl = _read_matrix(ws, self._theta_start_row, nphi, ntheta)

            # 读 Theta Pol 相位（如有 Phase 段）
            tp_data = None
            if self._has_phase and self._theta_phase_start > 0:
                try:
                    tp_data = _read_matrix(ws, self._theta_phase_start, nphi, ntheta)
                except Exception:
                    tp_data = None

            # 读 Phi Pol 幅度（不做 clipping）
            if self._has_phi_pol and self._phi_pol_start > 0:
                pl = _read_matrix(ws, self._phi_pol_start, nphi, ntheta)
            else:
                pl = np.full_like(tl, -999.0)

            # 读 Phi Pol 相位（如有 Phase 段）
            pp_data = None
            if self._has_phase and self._phi_phase_start > 0:
                try:
                    pp_data = _read_matrix(ws, self._phi_phase_start, nphi, ntheta)
                except Exception:
                    pp_data = None

            self._cache[freq] = (tl, pl, tp_data, pp_data)

        return {
            "theta_logmag": tl,
            "theta_phase": tp_data,
            "phi_logmag": pl,
            "phi_phase": pp_data,
        }

    def close(self):
        if self._wb:
            self._wb.close()
            self._wb = None

    def __del__(self):
        try:
            self.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# 模块级工具
# ---------------------------------------------------------------------------

def _freq_sheet_name(freq: float) -> str:
    """频率值 → sheet 名："1154.0"→"1154" """
    return str(int(freq)) if freq == int(freq) else str(freq)


def _read_matrix_pandas(ws, start_row: int, n_rows: int, n_cols: int) -> np.ndarray:
    """Pandas 快速通道: 批量 numpy 转换代替逐值 float()。"""
    import pandas as pd

    rows_data = []
    end_row = start_row + n_rows - 1
    for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                             min_col=2, max_col=1 + n_cols, values_only=True):
        rows_data.append(list(row[:n_cols]) if row else [None] * n_cols)
        if len(rows_data) >= n_rows:
            break

    if not rows_data:
        return np.full((n_rows, n_cols), -999.0, dtype=np.float64)

    # pandas to_numpy 批量转换 (C 级别, 比 Python float() 快 3-5x)
    return pd.DataFrame(rows_data).to_numpy(dtype=np.float64, na_value=-999.0)[:n_rows, :n_cols]


def _read_matrix(ws, start_row: int, n_rows: int, n_cols: int) -> np.ndarray:
    """流式读取 n_rows × n_cols 矩阵，自动选择最快路径。"""
    # 小矩阵直接用 openpyxl（pandas 导入有开销）
    if n_rows * n_cols < 1000:
        data = np.full((n_rows, n_cols), -999.0, dtype=np.float64)
        rows = ws.iter_rows(min_row=start_row, max_row=start_row + n_rows - 1,
                             min_col=2, max_col=1 + n_cols, values_only=True)
        for pi, row in enumerate(rows):
            if pi >= n_rows:
                break
            for ti, v in enumerate(row):
                if ti >= n_cols:
                    break
                if v is None:
                    continue
                try:
                    data[pi, ti] = float(v)
                except (ValueError, TypeError):
                    pass
        return data

    # 大矩阵使用 pandas 批量读取
    try:
        return _read_matrix_pandas(ws, start_row, n_rows, n_cols)
    except Exception:
        pass

    # Fallback
    data = np.full((n_rows, n_cols), -999.0, dtype=np.float64)
    rows = ws.iter_rows(min_row=start_row, max_row=start_row + n_rows - 1,
                         min_col=2, max_col=1 + n_cols, values_only=True)
    for pi, row in enumerate(rows):
        if pi >= n_rows:
            break
        for ti, v in enumerate(row):
            if ti >= n_cols:
                break
            if v is None:
                continue
            try:
                data[pi, ti] = float(v)
            except (ValueError, TypeError):
                pass
    return data


